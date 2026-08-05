from __future__ import annotations

import json
import os
import re
import zipfile
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urlsplit, urlunsplit, parse_qsl
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

SITE_COL = "설비 사이트 코드"
PROCESS_COL = "공정 코드"
CUSTOMER_INPUT_COL = "고객 이름"
CUSTOMER_COL = "거래처"
INITIAL_COL = "이니셜"
SHORT_CODE_COL = "제품코드(5자리)"
SKU_COL = "제품 코드"
PRODUCT_NAME_COL = "제품 이름"
DEMAND_COL = "수요량(PCS)"
PLAN_COL = "계획 수량"
PLAN_COL_ALIASES = (PLAN_COL, "주문 대응 수량")

PROD_DATE_COL = "생산일자"
PROD_SITE_COL = "공장"
PROD_PROCESS_COL = "공정코드"
PROD_SKU_COL = "품목코드"
PROD_GOOD_QTY_COL = "샘플제외 양품수량"
PROD_DATE_COL_ALIASES = (PROD_DATE_COL, "기간", "date")
PROD_SKU_COL_ALIASES = (PROD_SKU_COL, "품명")
PROD_GOOD_QTY_COL_ALIASES = (PROD_GOOD_QTY_COL, "총_양품수량")
PROD_PRODUCT_NAME_COL = "품명"
PRODUCTION_PRODUCT_NAME_COL = "_생산제품명"

SITE_FILTER = "C관"
TARGET_PROCESSES = ("[10]사출조립", "[80]누수/규격검사")
INSPECTION_PROCESS = "[80]누수/규격검사"
INPUT_RE = re.compile(r"^수요정보\(전공정\)_(\d{8})\.xlsx$")
FALLBACK_INPUT_RE = re.compile(r"^생산유효도_공정별_(\d{8})(?:_[^.]*)?\.xlsx$")
PRODUCTION_FILE = "생산실적.xlsx"
PRODUCTION_SOURCE_ENV = "PRODUCTION_SOURCE"
PRODUCTION_API_URL_ENV = "PRODUCTION_API_URL"
PRODUCTION_API_TOKEN_ENV = "PRODUCTION_API_TOKEN"
PRODUCTION_API_KEY_ENV = "PRODUCTION_API_KEY"
PRODUCTION_API_KEY_HEADER_ENV = "PRODUCTION_API_KEY_HEADER"
PRODUCTION_API_DATE_FROM_PARAM_ENV = "PRODUCTION_API_DATE_FROM_PARAM"
PRODUCTION_API_DATE_TO_PARAM_ENV = "PRODUCTION_API_DATE_TO_PARAM"
PRODUCTION_API_TIMEOUT_ENV = "PRODUCTION_API_TIMEOUT"
PRODUCTION_API_FALLBACK_ENV = "PRODUCTION_API_FALLBACK_TO_EXCEL"
PRODUCTION_API_USER_AGENT_ENV = "PRODUCTION_API_USER_AGENT"
PRODUCTION_API_CHUNK_DAYS_ENV = "PRODUCTION_API_CHUNK_DAYS"
PRODUCTION_API_SAVE_RAW_ENV = "PRODUCTION_API_SAVE_RAW"
PRODUCTION_API_DATA_DIR_ENV = "PRODUCTION_API_DATA_DIR"
DEFAULT_PRODUCTION_API_URL = "https://plan.interojo.net/api/production-performance"
DEFAULT_PRODUCTION_API_CHUNK_DAYS = 7
DEFAULT_PRODUCTION_API_DATA_DIR = "data"
CLASSIFICATION_FILE = "제품명 기준 정보.xlsx"
CLASSIFICATION_SHEET = "분류정보"

SUMMARY_SHEET = "Summary"
CHANGE_ANALYSIS_SHEET = "증감분석"
KIND_ANALYSIS_SHEET = "종수_유효도분석"
DETAIL_SHEET = "Detail"
SHEET_NAME_ANALYSIS_SHEET = "시트이름_유효도분석"
MAJOR_CATEGORY_ANALYSIS_SHEET = "대분류_유효도분석"
TOP_INEFFECTIVE_SHEET = "Top_비유효생산"
TOP_REMAINING_SHEET = "Top_잔여필요"
REQUIRED_REPORT_SHEETS = {SUMMARY_SHEET, CHANGE_ANALYSIS_SHEET}

DATE_COL = "일자"
OUTPUT_PROCESS_COL = "공정"
OUTPUT_SKU_COL = "제품코드"
OUTPUT_PRODUCT_NAME_COL = "제품명"
SHEET_NAME_COL = "시트이름"
MAJOR_CATEGORY_COL = "대분류"
NEED_QTY_COL = "필요수량"
SHORTAGE_QTY_COL = "부족수량"
ACTUAL_QTY_COL = "실적수량"
EFFECTIVE_PRODUCTION_COL = "유효생산량"
INEFFECTIVE_PRODUCTION_COL = "비유효생산량"
REMAINING_NEED_COL = "잔여필요수량"
PRODUCTION_EFFECTIVENESS_COL = "생산유효도(%)"
MATCH_STATUS_COL = "매칭상태"

SUMMARY_NEED_COL = "필요수량 합계"
SUMMARY_SHORTAGE_COL = "부족수량 합계"
SUMMARY_ACTUAL_COL = "실적수량 합계"
SUMMARY_EFFECTIVE_COL = "유효생산량 합계"
SUMMARY_INEFFECTIVE_COL = "비유효생산량 합계"
SUMMARY_REMAINING_COL = "잔여필요수량 합계"
PRODUCTION_KIND_COL = "생산종수"
EFFECTIVE_KIND_COL = "유효생산종수"
INEFFECTIVE_KIND_COL = "비유효생산종수"
PRODUCTION_KIND_EFFECTIVENESS_COL = "생산종수 유효도(%)"
PRODUCTION_SKU_COL = "생산 SKU수"
EFFECTIVE_SKU_COL = "유효 SKU수"
INEFFECTIVE_SKU_COL = "비유효 SKU수"
NEEDED_SKU_COL = "필요 SKU수"
COVERED_SKU_COL = "커버 SKU수"
UNCOVERED_NEEDED_SKU_COL = "미생산 필요 SKU수"
EXCESS_SKU_COL = "과잉 SKU수"
SHORT_SKU_COL = "부족 SKU수"
SKU_EFFECTIVENESS_COL = "SKU 유효율(%)"
AVG_ACTUAL_PER_KIND_COL = "평균 실적/종"
AVG_EFFECTIVE_PER_KIND_COL = "평균 유효생산량/종"
AVG_INEFFECTIVE_PER_KIND_COL = "평균 비유효생산량/종"

DEMAND_EXISTS_COL = "_계획존재"
PRODUCTION_EXISTS_COL = "_실적존재"

PREV_DATE_COL = "전일"
PREV_EFFECTIVENESS_COL = "전일 생산유효도(%)"
EFFECTIVENESS_DELTA_COL = "증감(%p)"
DIRECTION_COL = "방향"
NEED_DELTA_COL = "필요수량 증감"
ACTUAL_DELTA_COL = "실적수량 증감"
EFFECTIVE_DELTA_COL = "유효생산량 증감"
INEFFECTIVE_DELTA_COL = "비유효생산량 증감"
REMAINING_DELTA_COL = "잔여필요수량 증감"
NO_PLAN_ACTUAL_COL = "계획없음 실적수량"
NO_ACTUAL_NEED_COL = "실적없음 필요수량"
MAIN_REASON_COL = "주요 요인"
TOP_INEFFECTIVE_SKU_COL = "Top 비유효 SKU"
TOP_REMAINING_SKU_COL = "Top 잔여필요 SKU"


def column_index(column_ref: str) -> int:
    index = 0
    for char in column_ref:
        index = index * 26 + ord(char.upper()) - ord("A") + 1
    return index - 1


def column_letters(cell_ref: str) -> str:
    return "".join(char for char in cell_ref if char.isalpha())


def read_shared_strings(xlsx: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(xlsx.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    return [
        "".join(text.text or "" for text in item.findall(".//a:t", NS))
        for item in root.findall("a:si", NS)
    ]


def cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", NS))

    value = cell.find("a:v", NS)
    if value is None:
        return ""

    text = value.text or ""
    if cell_type == "s" and text:
        return shared_strings[int(text)]
    return text


def read_xlsx_raw(path: Path) -> pd.DataFrame:
    """Read sheet1 raw values so quantity cells formatted as dates stay numeric."""
    with zipfile.ZipFile(path) as xlsx:
        shared_strings = read_shared_strings(xlsx)
        root = ET.fromstring(xlsx.read("xl/worksheets/sheet1.xml"))

    rows: list[list[str]] = []
    max_index = 0
    for row in root.findall("a:sheetData/a:row", NS):
        values_by_index: dict[int, str] = {}
        for cell in row.findall("a:c", NS):
            index = column_index(column_letters(cell.attrib["r"]))
            max_index = max(max_index, index)
            values_by_index[index] = cell_value(cell, shared_strings)
        rows.append([values_by_index.get(index, "") for index in range(max_index + 1)])

    if not rows:
        return pd.DataFrame()

    headers = rows[0]
    records = []
    for row in rows[1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        records.append(row[: len(headers)])
    return pd.DataFrame(records, columns=headers)


def has_report_sheets(path: Path) -> bool:
    try:
        with pd.ExcelFile(path) as workbook:
            return REQUIRED_REPORT_SHEETS.issubset(set(workbook.sheet_names))
    except Exception:
        return False


def find_input_files(root: Path) -> list[tuple[str, Path]]:
    files_by_date: dict[str, Path] = {}

    for path in root.glob("수요정보(전공정)_*.xlsx"):
        match = INPUT_RE.match(path.name)
        if match:
            files_by_date[match.group(1)] = path

    for path in root.glob("생산유효도_공정별_*.xlsx"):
        match = FALLBACK_INPUT_RE.match(path.name)
        if match and not has_report_sheets(path):
            files_by_date.setdefault(match.group(1), path)

    return sorted(files_by_date.items())


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() == "nan" else re.sub(r"\s+", " ", text)


def normalize_lookup_text(value: object) -> str:
    return clean_text(value).casefold()


def extract_base_code(value: object) -> str:
    match = re.match(r"([A-Z]\d{4})", clean_text(value).upper())
    return match.group(1) if match else ""


def load_classification_maps(root: Path) -> tuple[dict[str, str], dict[str, str]]:
    path = root / CLASSIFICATION_FILE
    if not path.exists():
        return {}, {}

    try:
        classification = pd.read_excel(path, sheet_name=CLASSIFICATION_SHEET)
    except Exception:
        return {}, {}

    code_map: dict[str, str] = {}
    name_map: dict[str, str] = {}
    code_columns = ("코드", "Q코드", "R코드", "U코드")

    for _, row in classification.iterrows():
        sheet_name = clean_text(row.get(SHEET_NAME_COL))
        if not sheet_name:
            continue

        for column in code_columns:
            code = clean_text(row.get(column)).upper()
            if code:
                code_map[code] = sheet_name

        product_name = normalize_lookup_text(row.get(OUTPUT_PRODUCT_NAME_COL))
        if product_name:
            name_map[product_name] = sheet_name

    return code_map, name_map


def infer_sheet_name(sku: object, product_name: object, code_map: dict[str, str], name_map: dict[str, str]) -> str:
    base_code = extract_base_code(sku)
    if base_code in code_map:
        return code_map[base_code]

    normalized_name = normalize_lookup_text(product_name)
    if normalized_name in name_map:
        return name_map[normalized_name]

    upper_name = clean_text(product_name).upper()
    if "HAPA" in upper_name:
        return "피피비(HAPA)"
    if "GEMHOUR" in upper_name:
        return "피피비(젬아워)"
    if "PIA" in upper_name or "BAGUMORE" in upper_name:
        return "PIA 종합"
    if "COFANCY" in upper_name:
        return "COFANCY"
    return "미분류"


def major_category(sheet_name: object, product_name: object) -> str:
    sheet = clean_text(sheet_name)
    upper_sheet = sheet.upper()
    upper_name = clean_text(product_name).upper()
    domestic_sheets = {
        "국내",
        "IRIS 2_국내",
        "O2O2_국내",
        "다비치",
        "렌즈미",
        "렌즈타운",
        "스타비젼2종",
    }
    pia_markers = ("PIA", "피피비", "HAPA", "GEMHOUR", "BAGUMORE")

    if sheet in domestic_sheets or "국내" in sheet:
        return "국내"
    if any(marker in upper_sheet or marker in upper_name for marker in pia_markers):
        return "PIA"
    return "기타해외"


def apply_classification(detail: pd.DataFrame, root: Path) -> pd.DataFrame:
    detail = detail.copy()
    code_map, name_map = load_classification_maps(root)
    detail[SHEET_NAME_COL] = [
        infer_sheet_name(sku, product_name, code_map, name_map)
        for sku, product_name in zip(detail[OUTPUT_SKU_COL], detail[OUTPUT_PRODUCT_NAME_COL])
    ]
    detail[MAJOR_CATEGORY_COL] = [
        major_category(sheet_name, product_name)
        for sheet_name, product_name in zip(detail[SHEET_NAME_COL], detail[OUTPUT_PRODUCT_NAME_COL])
    ]
    return detail


def normalize_process(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def to_number(series: pd.Series) -> pd.Series:
    normalized = series.astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(normalized, errors="coerce").fillna(0)


def calculate_rate(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    return numerator.div(denominator.where(denominator.ne(0))).fillna(0).mul(100).round(1)


def divide_by_kind(numerator: pd.Series, production_kind: pd.Series) -> pd.Series:
    return numerator.div(production_kind.where(production_kind.ne(0))).fillna(0).round(1)


def normalize_demand_columns(path: Path, df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for alias in PLAN_COL_ALIASES:
        if alias in df.columns:
            if alias != PLAN_COL:
                df = df.rename(columns={alias: PLAN_COL})
            return df
    raise ValueError(f"{path.name}: missing columns: {', '.join(PLAN_COL_ALIASES)} 중 하나")


def validate_demand_columns(path: Path, df: pd.DataFrame) -> None:
    required = {SITE_COL, PROCESS_COL, SKU_COL, PRODUCT_NAME_COL, DEMAND_COL, PLAN_COL}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{path.name}: missing columns: {', '.join(sorted(missing))}")


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def find_header_index(headers: list[object], target: str) -> int:
    normalized_target = normalize_header(target)
    for index, header in enumerate(headers):
        if normalize_header(header) == normalized_target:
            return index
    raise ValueError(f"missing header: {target}")


def find_optional_header_index(headers: list[object], target: str) -> int | None:
    try:
        return find_header_index(headers, target)
    except ValueError:
        return None


def find_process_column_index(
    top_headers: list[object], sub_headers: list[object], process: str, sub_header: str
) -> int:
    normalized_process = normalize_process(process)
    normalized_sub_header = normalize_header(sub_header)
    for index, top_header in enumerate(top_headers):
        if (
            normalize_process(top_header) == normalized_process
            and normalize_header(sub_headers[index]) == normalized_sub_header
        ):
            return index
    raise ValueError(f"missing process column: {process} / {sub_header}")


def join_unique_text(series: pd.Series) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for value in series:
        for part in str(value or "").split(";"):
            text = part.strip()
            if not text or text.lower() == "nan" or text in seen:
                continue
            seen.add(text)
            values.append(text)
    return "; ".join(values)


def is_wide_process_demand(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    first_row_headers = {normalize_header(value) for value in df.iloc[0].tolist()}
    process_headers = {normalize_process(value) for value in df.columns}
    return "설비 사이트 코드" in first_row_headers and bool(process_headers & set(TARGET_PROCESSES))


def prepare_wide_process_demand_data(
    date: str, path: Path, df: pd.DataFrame, match_on_product_name: bool = False
) -> pd.DataFrame:
    top_headers = df.columns.tolist()
    sub_headers = df.iloc[0].tolist()
    records = df.iloc[1:].copy()

    site_index = find_header_index(sub_headers, SITE_COL)
    customer_index = find_optional_header_index(sub_headers, CUSTOMER_INPUT_COL)
    initial_index = find_optional_header_index(sub_headers, INITIAL_COL)
    sku_index = find_header_index(sub_headers, "제품 코드")
    product_name_index = find_header_index(sub_headers, "수요 제품 이름")
    demand_index = find_header_index(sub_headers, "수요 수량")
    shortage_index = find_process_column_index(
        top_headers, sub_headers, INSPECTION_PROCESS, "생산 수량"
    )

    frames: list[pd.DataFrame] = []
    for process in TARGET_PROCESSES:
        process_qty_index = find_process_column_index(top_headers, sub_headers, process, "생산 수량")
        frame = pd.DataFrame(
            {
                SITE_COL: records.iloc[:, site_index],
                CUSTOMER_COL: records.iloc[:, customer_index] if customer_index is not None else "",
                INITIAL_COL: records.iloc[:, initial_index] if initial_index is not None else "",
                SKU_COL: records.iloc[:, sku_index],
                OUTPUT_PRODUCT_NAME_COL: records.iloc[:, product_name_index],
                "수요량": records.iloc[:, demand_index],
                NEED_QTY_COL: records.iloc[:, process_qty_index],
                SHORTAGE_QTY_COL: records.iloc[:, shortage_index],
            }
        )
        frame[DATE_COL] = date
        frame[PROCESS_COL] = process
        frames.append(frame)

    if not frames:
        raise ValueError(f"{path.name}: 대상 공정의 생산 수량 컬럼을 찾을 수 없습니다.")

    target = pd.concat(frames, ignore_index=True)
    target = target[target[SITE_COL].astype(str).str.contains(SITE_FILTER, na=False)].copy()
    if match_on_product_name:
        target[SKU_COL] = target[OUTPUT_PRODUCT_NAME_COL].astype(str).str.strip()
    else:
        target[SKU_COL] = target[SKU_COL].astype(str).str.strip()
    target = target[target[SKU_COL].ne("") & target[SKU_COL].ne("총합계")]
    target[CUSTOMER_COL] = target[CUSTOMER_COL].astype(str).str.strip()
    target[INITIAL_COL] = target[INITIAL_COL].astype(str).str.strip()
    target["수요량"] = to_number(target["수요량"])
    target[NEED_QTY_COL] = to_number(target[NEED_QTY_COL])
    target[SHORTAGE_QTY_COL] = to_number(target[SHORTAGE_QTY_COL])

    demand = (
        target.groupby([DATE_COL, PROCESS_COL, SKU_COL], as_index=False)
        .agg(
            **{
                OUTPUT_PRODUCT_NAME_COL: (OUTPUT_PRODUCT_NAME_COL, "first"),
                CUSTOMER_COL: (CUSTOMER_COL, join_unique_text),
                INITIAL_COL: (INITIAL_COL, join_unique_text),
                "수요량": ("수요량", "sum"),
                NEED_QTY_COL: (NEED_QTY_COL, "sum"),
                SHORTAGE_QTY_COL: (SHORTAGE_QTY_COL, "sum"),
            }
        )
        .sort_values([DATE_COL, PROCESS_COL, SKU_COL])
    )
    demand[DEMAND_EXISTS_COL] = True
    return demand


def prepare_demand_data(
    date: str, path: Path, match_on_product_name: bool = False
) -> pd.DataFrame:
    raw = read_xlsx_raw(path)
    if is_wide_process_demand(raw):
        return prepare_wide_process_demand_data(date, path, raw, match_on_product_name)

    df = normalize_demand_columns(path, raw)
    validate_demand_columns(path, df)

    target = df[
        df[SITE_COL].astype(str).str.contains(SITE_FILTER, na=False)
        & df[PROCESS_COL].map(normalize_process).isin(TARGET_PROCESSES)
    ].copy()
    if match_on_product_name:
        target[SKU_COL] = target[PRODUCT_NAME_COL].astype(str).str.strip()
    else:
        target[SKU_COL] = target[SKU_COL].astype(str).str.strip()
    target = target[target[SKU_COL].ne("") & target[SKU_COL].ne("총합계")]
    target[PROCESS_COL] = target[PROCESS_COL].map(normalize_process)
    target[CUSTOMER_COL] = (
        target[CUSTOMER_INPUT_COL].astype(str).str.strip()
        if CUSTOMER_INPUT_COL in target.columns
        else ""
    )
    target[INITIAL_COL] = (
        target[INITIAL_COL].astype(str).str.strip()
        if INITIAL_COL in target.columns
        else ""
    )
    target[DEMAND_COL] = to_number(target[DEMAND_COL])
    target[PLAN_COL] = to_number(target[PLAN_COL])
    target[SHORTAGE_QTY_COL] = target[PLAN_COL]
    target[DATE_COL] = date

    demand = (
        target.groupby([DATE_COL, PROCESS_COL, SKU_COL], as_index=False)
        .agg(
            **{
                OUTPUT_PRODUCT_NAME_COL: (PRODUCT_NAME_COL, "first"),
                CUSTOMER_COL: (CUSTOMER_COL, join_unique_text),
                INITIAL_COL: (INITIAL_COL, join_unique_text),
                "수요량": (DEMAND_COL, "sum"),
                NEED_QTY_COL: (PLAN_COL, "sum"),
                SHORTAGE_QTY_COL: (SHORTAGE_QTY_COL, "sum"),
            }
        )
        .sort_values([DATE_COL, PROCESS_COL, SKU_COL])
    )
    demand[DEMAND_EXISTS_COL] = True
    return demand


def build_demand(
    input_files: list[tuple[str, Path]], match_on_product_name: bool = False
) -> pd.DataFrame:
    frames = [prepare_demand_data(date, path, match_on_product_name) for date, path in input_files]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def source_display_name(source: Path | str) -> str:
    return source.name if isinstance(source, Path) else source


def resolve_alias_column(source: Path | str, df: pd.DataFrame, aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        if alias in df.columns:
            return alias
    raise ValueError(f"{source_display_name(source)}: missing columns: {', '.join(aliases)} 중 하나")


def resolve_production_columns(source: Path | str, df: pd.DataFrame) -> dict[str, str]:
    return {
        PROD_DATE_COL: resolve_alias_column(source, df, PROD_DATE_COL_ALIASES),
        PROD_SITE_COL: resolve_alias_column(source, df, (PROD_SITE_COL,)),
        PROD_PROCESS_COL: resolve_alias_column(source, df, (PROD_PROCESS_COL,)),
        PROD_SKU_COL: resolve_alias_column(source, df, PROD_SKU_COL_ALIASES),
        PROD_GOOD_QTY_COL: resolve_alias_column(source, df, PROD_GOOD_QTY_COL_ALIASES),
    }


def production_uses_product_name_key(path: Path) -> bool:
    df = pd.read_excel(path, nrows=0)
    columns = resolve_production_columns(path, df)
    return columns[PROD_SKU_COL] == PROD_PRODUCT_NAME_COL


def production_uses_product_name_key_from_df(source: Path | str, df: pd.DataFrame) -> bool:
    columns = resolve_production_columns(source, df)
    return columns[PROD_SKU_COL] == PROD_PRODUCT_NAME_COL


def should_use_production_api() -> bool:
    source = os.getenv(PRODUCTION_SOURCE_ENV, "").strip().lower()
    return source == "api" or bool(os.getenv(PRODUCTION_API_URL_ENV, "").strip())


def add_query_params(url: str, params: dict[str, str]) -> str:
    if not params:
        return url
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value})
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


def extract_api_rows(payload: object) -> list[dict[str, object]]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "items", "results", "rows", "orders"):
            value = payload.get(key)
            if isinstance(value, list):
                return value
    raise ValueError("생산실적 API 응답은 JSON 배열이거나 data/items/results/rows 배열을 포함해야 합니다.")


def normalize_production_api_process(value: object) -> str:
    text = clean_text(value)
    compact = re.sub(r"\D", "", text)
    if compact == "10":
        return "[10]사출조립"
    if compact == "80":
        return "[80]누수/규격검사"
    return text


def normalize_production_api_site(value: object) -> str:
    text = clean_text(value)
    if "C관" in text:
        return text
    code = re.sub(r"\D", "", text)
    return {
        "1": "A관",
        "01": "A관",
        "2": "C관",
        "02": "C관",
        "3": "S관",
        "03": "S관",
        "5": "5공장",
        "05": "5공장",
    }.get(code, text)


def normalize_production_api_frame(df: pd.DataFrame) -> pd.DataFrame:
    api_columns = {"pr_dt", "fac_cd", "gong_cd", "gd_cd", "pr_qty"}
    if not api_columns.issubset(set(df.columns)):
        return df

    normalized = df.copy()
    normalized[PROD_DATE_COL] = normalized["pr_dt"]
    normalized[PROD_SITE_COL] = normalized.get("fac_nm", normalized["fac_cd"]).map(
        normalize_production_api_site
    )
    normalized[PROD_PROCESS_COL] = normalized["gong_cd"].map(normalize_production_api_process)
    normalized[PROD_SKU_COL] = normalized["gd_cd"]
    normalized[PROD_PRODUCT_NAME_COL] = normalized.get("gd_nm", "")
    good_qty = to_number(normalized["pr_qty"])
    sample_qty = to_number(normalized["sample_qty"]) if "sample_qty" in normalized.columns else 0
    normalized[PROD_GOOD_QTY_COL] = (good_qty - sample_qty).clip(lower=0)
    return normalized


def parse_api_count(value: object) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(float(str(value).replace(",", "")))
    except (TypeError, ValueError):
        return None


def api_payload_is_truncated(payload: object, row_count: int) -> bool:
    if not isinstance(payload, dict):
        return False
    if payload.get("truncated") is True:
        return True

    total_count = parse_api_count(payload.get("total_count"))
    returned_count = parse_api_count(payload.get("returned_count"))
    if returned_count is None:
        returned_count = row_count
    return total_count is not None and total_count > returned_count


def normalize_api_valid_dates(valid_dates: set[str]) -> tuple[set[str], pd.Timestamp | None, pd.Timestamp | None]:
    if not valid_dates:
        return set(), None, None

    parsed = pd.to_datetime(pd.Series(list(valid_dates), dtype="string"), errors="coerce")
    parsed = parsed.dropna()
    if parsed.empty:
        return set(), None, None

    return set(parsed.dt.strftime("%Y%m%d")), parsed.min().normalize(), parsed.max().normalize()


def production_api_chunk_days() -> int:
    raw_value = os.getenv(PRODUCTION_API_CHUNK_DAYS_ENV, str(DEFAULT_PRODUCTION_API_CHUNK_DAYS))
    try:
        return max(1, int(raw_value))
    except ValueError:
        return DEFAULT_PRODUCTION_API_CHUNK_DAYS


def iter_date_chunks(
    start_date: pd.Timestamp, end_date: pd.Timestamp, chunk_days: int
) -> list[tuple[pd.Timestamp, pd.Timestamp]]:
    chunks: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    current = start_date
    while current <= end_date:
        chunk_end = min(current + pd.Timedelta(days=chunk_days - 1), end_date)
        chunks.append((current, chunk_end))
        current = chunk_end + pd.Timedelta(days=1)
    return chunks


def build_production_api_headers() -> dict[str, str]:
    headers = {
        "Accept": "application/json",
        "User-Agent": os.getenv(
            PRODUCTION_API_USER_AGENT_ENV,
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
        ),
    }
    token = os.getenv(PRODUCTION_API_TOKEN_ENV, "").strip()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    api_key = os.getenv(PRODUCTION_API_KEY_ENV, "").strip()
    if api_key:
        header_name = os.getenv(PRODUCTION_API_KEY_HEADER_ENV, "X-API-Key").strip() or "X-API-Key"
        headers[header_name] = api_key
    return headers


def request_production_api_payload(
    url: str, params: dict[str, str], headers: dict[str, str], timeout: float
) -> object:
    request = Request(add_query_params(url, params), headers=headers, method="GET")
    try:
        with urlopen(request, timeout=timeout) as response:
            charset = response.headers.get_content_charset() or "utf-8"
            return json.loads(response.read().decode(charset))
    except HTTPError as error:
        body = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"생산실적 API 호출 실패: HTTP {error.code} {body}") from error
    except URLError as error:
        raise RuntimeError(f"생산실적 API 연결 실패: {error.reason}") from error


def fetch_production_api_rows(
    url: str,
    headers: dict[str, str],
    timeout: float,
    date_from_param: str,
    date_to_param: str,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> list[dict[str, object]]:
    params: dict[str, str] = {"limit": "0"}
    if date_from_param:
        params[date_from_param] = start_date.strftime("%Y-%m-%d")
    if date_to_param:
        params[date_to_param] = end_date.strftime("%Y-%m-%d")

    payload = request_production_api_payload(url, params, headers, timeout)
    rows = extract_api_rows(payload)
    if not api_payload_is_truncated(payload, len(rows)):
        return rows

    if start_date >= end_date:
        raise RuntimeError(
            f"생산실적 API 응답이 {start_date:%Y-%m-%d} 단일 일자에서도 잘렸습니다."
        )

    mid_date = start_date + pd.Timedelta(days=(end_date - start_date).days // 2)
    return fetch_production_api_rows(
        url, headers, timeout, date_from_param, date_to_param, start_date, mid_date
    ) + fetch_production_api_rows(
        url,
        headers,
        timeout,
        date_from_param,
        date_to_param,
        mid_date + pd.Timedelta(days=1),
        end_date,
    )


def filter_production_api_dates(df: pd.DataFrame, valid_date_keys: set[str]) -> pd.DataFrame:
    if df.empty or not valid_date_keys or PROD_DATE_COL not in df.columns:
        return df
    date_keys = pd.to_datetime(df[PROD_DATE_COL], errors="coerce").dt.strftime("%Y%m%d")
    return df[date_keys.isin(valid_date_keys)].copy()


def should_save_production_api_raw() -> bool:
    value = os.getenv(PRODUCTION_API_SAVE_RAW_ENV, "1").strip().lower()
    return value not in {"0", "false", "no", "n"}


def save_production_api_data(
    root: Path,
    df: pd.DataFrame,
    start_date: pd.Timestamp | None,
    end_date: pd.Timestamp | None,
) -> Path:
    data_dir_value = os.getenv(PRODUCTION_API_DATA_DIR_ENV, DEFAULT_PRODUCTION_API_DATA_DIR).strip()
    data_dir = Path(data_dir_value or DEFAULT_PRODUCTION_API_DATA_DIR)
    if not data_dir.is_absolute():
        data_dir = root / data_dir
    data_dir.mkdir(parents=True, exist_ok=True)

    start_label = start_date.strftime("%Y%m%d") if start_date is not None else "all"
    end_label = end_date.strftime("%Y%m%d") if end_date is not None else "all"
    output_path = data_dir / f"생산실적_API_{start_label}_{end_label}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="생산실적_API", index=False)
    print(f"production_api_data={output_path}")
    return output_path


def load_production_api_data(valid_dates: set[str], root: Path | None = None) -> pd.DataFrame:
    url = os.getenv(PRODUCTION_API_URL_ENV, "").strip() or DEFAULT_PRODUCTION_API_URL
    if not url:
        raise ValueError(f"{PRODUCTION_API_URL_ENV} 환경변수가 필요합니다.")

    valid_date_keys, start_date, end_date = normalize_api_valid_dates(valid_dates)
    date_from_param = os.getenv(PRODUCTION_API_DATE_FROM_PARAM_ENV, "date_from").strip()
    date_to_param = os.getenv(PRODUCTION_API_DATE_TO_PARAM_ENV, "date_to").strip()
    headers = build_production_api_headers()
    timeout = float(os.getenv(PRODUCTION_API_TIMEOUT_ENV, "60"))

    rows: list[dict[str, object]] = []
    if start_date is not None and end_date is not None and date_from_param and date_to_param:
        for chunk_start, chunk_end in iter_date_chunks(
            start_date, end_date, production_api_chunk_days()
        ):
            rows.extend(
                fetch_production_api_rows(
                    url,
                    headers,
                    timeout,
                    date_from_param,
                    date_to_param,
                    chunk_start,
                    chunk_end,
                )
            )
    else:
        payload = request_production_api_payload(url, {"limit": "0"}, headers, timeout)
        rows = extract_api_rows(payload)

    df = filter_production_api_dates(
        normalize_production_api_frame(pd.DataFrame(rows)), valid_date_keys
    )
    if root is not None and should_save_production_api_raw():
        save_production_api_data(root, df, start_date, end_date)
    return df


def load_production_source(root: Path, valid_dates: set[str]) -> tuple[pd.DataFrame, Path | str]:
    production_path = root / PRODUCTION_FILE
    if should_use_production_api():
        try:
            return load_production_api_data(valid_dates, root), "production API"
        except Exception:
            if os.getenv(PRODUCTION_API_FALLBACK_ENV, "").strip() != "1":
                raise
            if not production_path.exists():
                raise

    if not production_path.exists():
        raise SystemExit(f"{PRODUCTION_FILE} 파일이 없습니다.")
    return pd.read_excel(production_path), production_path


def prepare_production_data(df: pd.DataFrame, source: Path | str, valid_dates: set[str]) -> pd.DataFrame:
    columns = resolve_production_columns(source, df)

    target = df[
        df[columns[PROD_SITE_COL]].astype(str).str.contains(SITE_FILTER, na=False)
        & df[columns[PROD_PROCESS_COL]].map(normalize_process).isin(TARGET_PROCESSES)
    ].copy()
    target[DATE_COL] = pd.to_datetime(target[columns[PROD_DATE_COL]], errors="coerce").dt.strftime(
        "%Y%m%d"
    )
    target = target[target[DATE_COL].isin(valid_dates)]
    target[PROCESS_COL] = target[columns[PROD_PROCESS_COL]].map(normalize_process)
    target[SKU_COL] = target[columns[PROD_SKU_COL]].astype(str).str.strip()
    target[PRODUCTION_PRODUCT_NAME_COL] = (
        target[PROD_PRODUCT_NAME_COL].astype(str).str.strip()
        if PROD_PRODUCT_NAME_COL in target.columns
        else ""
    )
    target[ACTUAL_QTY_COL] = to_number(target[columns[PROD_GOOD_QTY_COL]])
    target = target[target[SKU_COL].ne("") & target[SKU_COL].ne("nan")]

    production = (
        target.groupby([DATE_COL, PROCESS_COL, SKU_COL], as_index=False)
        .agg(
            **{
                ACTUAL_QTY_COL: (ACTUAL_QTY_COL, "sum"),
                PRODUCTION_PRODUCT_NAME_COL: (PRODUCTION_PRODUCT_NAME_COL, "first"),
            }
        )
        .sort_values([DATE_COL, PROCESS_COL, SKU_COL])
    )
    production[PRODUCTION_EXISTS_COL] = True
    return production


def build_detail(demand: pd.DataFrame, production: pd.DataFrame) -> pd.DataFrame:
    detail = demand.merge(production, on=[DATE_COL, PROCESS_COL, SKU_COL], how="outer")
    detail[DEMAND_EXISTS_COL] = detail[DEMAND_EXISTS_COL].eq(True)
    detail[PRODUCTION_EXISTS_COL] = detail[PRODUCTION_EXISTS_COL].eq(True)
    detail = detail[detail[DEMAND_EXISTS_COL] | detail[PRODUCTION_EXISTS_COL]].copy()

    detail[NEED_QTY_COL] = to_number(detail[NEED_QTY_COL])
    detail[SHORTAGE_QTY_COL] = to_number(detail[SHORTAGE_QTY_COL])
    detail[ACTUAL_QTY_COL] = to_number(detail[ACTUAL_QTY_COL])
    demand_product_name = detail[OUTPUT_PRODUCT_NAME_COL].fillna("").astype(str)
    if PRODUCTION_PRODUCT_NAME_COL in detail.columns:
        production_product_name = detail[PRODUCTION_PRODUCT_NAME_COL].fillna("").astype(str)
        detail[OUTPUT_PRODUCT_NAME_COL] = demand_product_name.where(
            demand_product_name.str.strip().ne("") & demand_product_name.str.lower().ne("nan"),
            production_product_name,
        )
    else:
        detail[OUTPUT_PRODUCT_NAME_COL] = demand_product_name
    detail[CUSTOMER_COL] = detail[CUSTOMER_COL].fillna("")
    detail[INITIAL_COL] = detail[INITIAL_COL].fillna("")

    detail[EFFECTIVE_PRODUCTION_COL] = detail[[NEED_QTY_COL, ACTUAL_QTY_COL]].min(axis=1)
    detail[INEFFECTIVE_PRODUCTION_COL] = (
        detail[ACTUAL_QTY_COL] - detail[EFFECTIVE_PRODUCTION_COL]
    ).clip(lower=0)
    detail[REMAINING_NEED_COL] = (detail[NEED_QTY_COL] - detail[ACTUAL_QTY_COL]).clip(lower=0)
    detail[PRODUCTION_EFFECTIVENESS_COL] = calculate_rate(
        detail[EFFECTIVE_PRODUCTION_COL], detail[ACTUAL_QTY_COL]
    )

    detail[MATCH_STATUS_COL] = "둘다없음"
    detail.loc[detail[DEMAND_EXISTS_COL] & detail[PRODUCTION_EXISTS_COL], MATCH_STATUS_COL] = "정상매칭"
    detail.loc[detail[DEMAND_EXISTS_COL] & ~detail[PRODUCTION_EXISTS_COL], MATCH_STATUS_COL] = "실적없음"
    detail.loc[~detail[DEMAND_EXISTS_COL] & detail[PRODUCTION_EXISTS_COL], MATCH_STATUS_COL] = "계획없음"
    detail = detail[detail[MATCH_STATUS_COL].ne("둘다없음")].copy()

    detail = detail.rename(columns={PROCESS_COL: OUTPUT_PROCESS_COL, SKU_COL: OUTPUT_SKU_COL})
    columns = [
        DATE_COL,
        OUTPUT_PROCESS_COL,
        OUTPUT_SKU_COL,
        OUTPUT_PRODUCT_NAME_COL,
        CUSTOMER_COL,
        INITIAL_COL,
        SHORTAGE_QTY_COL,
        NEED_QTY_COL,
        ACTUAL_QTY_COL,
        EFFECTIVE_PRODUCTION_COL,
        INEFFECTIVE_PRODUCTION_COL,
        REMAINING_NEED_COL,
        PRODUCTION_EFFECTIVENESS_COL,
        MATCH_STATUS_COL,
    ]
    return detail[columns].sort_values([DATE_COL, OUTPUT_PROCESS_COL, OUTPUT_SKU_COL])


def summarize_by_process(detail: pd.DataFrame) -> pd.DataFrame:
    summary = (
        detail.groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(
            **{
                SUMMARY_SHORTAGE_COL: (SHORTAGE_QTY_COL, "sum"),
                SUMMARY_NEED_COL: (NEED_QTY_COL, "sum"),
                SUMMARY_ACTUAL_COL: (ACTUAL_QTY_COL, "sum"),
                SUMMARY_EFFECTIVE_COL: (EFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_INEFFECTIVE_COL: (INEFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_REMAINING_COL: (REMAINING_NEED_COL, "sum"),
            }
        )
        .sort_values([OUTPUT_PROCESS_COL, DATE_COL])
    )
    production_kinds = detail[detail[ACTUAL_QTY_COL].gt(0)].copy()
    production_kinds["_제품코드5"] = production_kinds[OUTPUT_SKU_COL].astype(str).str.slice(0, 5)
    production_kind_summary = (
        production_kinds.groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{PRODUCTION_KIND_COL: ("_제품코드5", "nunique")})
    )
    effective_kind_summary = (
        production_kinds[production_kinds[EFFECTIVE_PRODUCTION_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{EFFECTIVE_KIND_COL: ("_제품코드5", "nunique")})
    )
    production_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{PRODUCTION_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    effective_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(0) & detail[EFFECTIVE_PRODUCTION_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{EFFECTIVE_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    needed_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{NEEDED_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    covered_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(0) & detail[ACTUAL_QTY_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{COVERED_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    excess_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(detail[NEED_QTY_COL]) & detail[ACTUAL_QTY_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{EXCESS_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    short_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(detail[ACTUAL_QTY_COL]) & detail[NEED_QTY_COL].gt(0)]
        .groupby([DATE_COL, OUTPUT_PROCESS_COL], as_index=False)
        .agg(**{SHORT_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    summary = summary.merge(production_kind_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(effective_kind_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(production_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(effective_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(needed_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(covered_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(excess_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary = summary.merge(short_sku_summary, on=[DATE_COL, OUTPUT_PROCESS_COL], how="left")
    summary[PRODUCTION_KIND_COL] = summary[PRODUCTION_KIND_COL].fillna(0).astype(int)
    summary[EFFECTIVE_KIND_COL] = summary[EFFECTIVE_KIND_COL].fillna(0).astype(int)
    summary[INEFFECTIVE_KIND_COL] = (
        summary[PRODUCTION_KIND_COL] - summary[EFFECTIVE_KIND_COL]
    ).clip(lower=0)
    summary[PRODUCTION_KIND_EFFECTIVENESS_COL] = calculate_rate(
        summary[EFFECTIVE_KIND_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_ACTUAL_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_ACTUAL_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_EFFECTIVE_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_EFFECTIVE_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_INEFFECTIVE_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_INEFFECTIVE_COL], summary[PRODUCTION_KIND_COL]
    )
    sku_count_columns = [
        PRODUCTION_SKU_COL,
        EFFECTIVE_SKU_COL,
        NEEDED_SKU_COL,
        COVERED_SKU_COL,
        EXCESS_SKU_COL,
        SHORT_SKU_COL,
    ]
    for column in sku_count_columns:
        summary[column] = summary[column].fillna(0).astype(int)
    summary[INEFFECTIVE_SKU_COL] = (
        summary[PRODUCTION_SKU_COL] - summary[EFFECTIVE_SKU_COL]
    ).clip(lower=0)
    summary[UNCOVERED_NEEDED_SKU_COL] = (
        summary[NEEDED_SKU_COL] - summary[COVERED_SKU_COL]
    ).clip(lower=0)
    summary[SKU_EFFECTIVENESS_COL] = calculate_rate(
        summary[EFFECTIVE_SKU_COL], summary[PRODUCTION_SKU_COL]
    )
    summary[PRODUCTION_EFFECTIVENESS_COL] = calculate_rate(
        summary[SUMMARY_EFFECTIVE_COL], summary[SUMMARY_ACTUAL_COL]
    )
    return summary[
        [
            DATE_COL,
            OUTPUT_PROCESS_COL,
            SUMMARY_SHORTAGE_COL,
            SUMMARY_NEED_COL,
            SUMMARY_ACTUAL_COL,
            SUMMARY_EFFECTIVE_COL,
            SUMMARY_INEFFECTIVE_COL,
            SUMMARY_REMAINING_COL,
            PRODUCTION_KIND_COL,
            EFFECTIVE_KIND_COL,
            INEFFECTIVE_KIND_COL,
            PRODUCTION_KIND_EFFECTIVENESS_COL,
            PRODUCTION_SKU_COL,
            EFFECTIVE_SKU_COL,
            INEFFECTIVE_SKU_COL,
            NEEDED_SKU_COL,
            COVERED_SKU_COL,
            UNCOVERED_NEEDED_SKU_COL,
            EXCESS_SKU_COL,
            SHORT_SKU_COL,
            SKU_EFFECTIVENESS_COL,
            AVG_ACTUAL_PER_KIND_COL,
            AVG_EFFECTIVE_PER_KIND_COL,
            AVG_INEFFECTIVE_PER_KIND_COL,
            PRODUCTION_EFFECTIVENESS_COL,
        ]
    ]


def summarize_by_classification(detail: pd.DataFrame, category_col: str) -> pd.DataFrame:
    dimensions = [DATE_COL, OUTPUT_PROCESS_COL, category_col]
    summary = (
        detail.groupby(dimensions, as_index=False)
        .agg(
            **{
                SUMMARY_SHORTAGE_COL: (SHORTAGE_QTY_COL, "sum"),
                SUMMARY_NEED_COL: (NEED_QTY_COL, "sum"),
                SUMMARY_ACTUAL_COL: (ACTUAL_QTY_COL, "sum"),
                SUMMARY_EFFECTIVE_COL: (EFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_INEFFECTIVE_COL: (INEFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_REMAINING_COL: (REMAINING_NEED_COL, "sum"),
            }
        )
        .sort_values(dimensions)
    )

    production_kinds = detail[detail[ACTUAL_QTY_COL].gt(0)].copy()
    production_kinds["_제품코드5"] = production_kinds[OUTPUT_SKU_COL].astype(str).str.slice(0, 5)
    production_kind_summary = (
        production_kinds.groupby(dimensions, as_index=False)
        .agg(**{PRODUCTION_KIND_COL: ("_제품코드5", "nunique")})
    )
    effective_kind_summary = (
        production_kinds[production_kinds[EFFECTIVE_PRODUCTION_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{EFFECTIVE_KIND_COL: ("_제품코드5", "nunique")})
    )
    production_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{PRODUCTION_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    effective_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(0) & detail[EFFECTIVE_PRODUCTION_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{EFFECTIVE_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    needed_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{NEEDED_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    covered_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(0) & detail[ACTUAL_QTY_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{COVERED_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    excess_sku_summary = (
        detail[detail[ACTUAL_QTY_COL].gt(detail[NEED_QTY_COL]) & detail[ACTUAL_QTY_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{EXCESS_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )
    short_sku_summary = (
        detail[detail[NEED_QTY_COL].gt(detail[ACTUAL_QTY_COL]) & detail[NEED_QTY_COL].gt(0)]
        .groupby(dimensions, as_index=False)
        .agg(**{SHORT_SKU_COL: (OUTPUT_SKU_COL, "nunique")})
    )

    for frame in (
        production_kind_summary,
        effective_kind_summary,
        production_sku_summary,
        effective_sku_summary,
        needed_sku_summary,
        covered_sku_summary,
        excess_sku_summary,
        short_sku_summary,
    ):
        summary = summary.merge(frame, on=dimensions, how="left")

    summary[PRODUCTION_KIND_COL] = summary[PRODUCTION_KIND_COL].fillna(0).astype(int)
    summary[EFFECTIVE_KIND_COL] = summary[EFFECTIVE_KIND_COL].fillna(0).astype(int)
    summary[INEFFECTIVE_KIND_COL] = (
        summary[PRODUCTION_KIND_COL] - summary[EFFECTIVE_KIND_COL]
    ).clip(lower=0)
    summary[PRODUCTION_KIND_EFFECTIVENESS_COL] = calculate_rate(
        summary[EFFECTIVE_KIND_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_ACTUAL_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_ACTUAL_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_EFFECTIVE_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_EFFECTIVE_COL], summary[PRODUCTION_KIND_COL]
    )
    summary[AVG_INEFFECTIVE_PER_KIND_COL] = divide_by_kind(
        summary[SUMMARY_INEFFECTIVE_COL], summary[PRODUCTION_KIND_COL]
    )

    sku_count_columns = [
        PRODUCTION_SKU_COL,
        EFFECTIVE_SKU_COL,
        NEEDED_SKU_COL,
        COVERED_SKU_COL,
        EXCESS_SKU_COL,
        SHORT_SKU_COL,
    ]
    for column in sku_count_columns:
        summary[column] = summary[column].fillna(0).astype(int)
    summary[INEFFECTIVE_SKU_COL] = (
        summary[PRODUCTION_SKU_COL] - summary[EFFECTIVE_SKU_COL]
    ).clip(lower=0)
    summary[UNCOVERED_NEEDED_SKU_COL] = (
        summary[NEEDED_SKU_COL] - summary[COVERED_SKU_COL]
    ).clip(lower=0)
    summary[SKU_EFFECTIVENESS_COL] = calculate_rate(
        summary[EFFECTIVE_SKU_COL], summary[PRODUCTION_SKU_COL]
    )
    summary[PRODUCTION_EFFECTIVENESS_COL] = calculate_rate(
        summary[SUMMARY_EFFECTIVE_COL], summary[SUMMARY_ACTUAL_COL]
    )

    return summary[
        [
            DATE_COL,
            OUTPUT_PROCESS_COL,
            category_col,
            SUMMARY_SHORTAGE_COL,
            SUMMARY_NEED_COL,
            SUMMARY_ACTUAL_COL,
            SUMMARY_EFFECTIVE_COL,
            SUMMARY_INEFFECTIVE_COL,
            SUMMARY_REMAINING_COL,
            PRODUCTION_KIND_COL,
            EFFECTIVE_KIND_COL,
            INEFFECTIVE_KIND_COL,
            PRODUCTION_KIND_EFFECTIVENESS_COL,
            PRODUCTION_SKU_COL,
            EFFECTIVE_SKU_COL,
            INEFFECTIVE_SKU_COL,
            NEEDED_SKU_COL,
            COVERED_SKU_COL,
            UNCOVERED_NEEDED_SKU_COL,
            EXCESS_SKU_COL,
            SHORT_SKU_COL,
            SKU_EFFECTIVENESS_COL,
            AVG_ACTUAL_PER_KIND_COL,
            AVG_EFFECTIVE_PER_KIND_COL,
            AVG_INEFFECTIVE_PER_KIND_COL,
            PRODUCTION_EFFECTIVENESS_COL,
        ]
    ]


def most_common_text(series: pd.Series) -> str:
    modes = series.dropna().astype(str).mode()
    return modes.iloc[0] if not modes.empty else ""


def attach_major_category(sheet_name_analysis: pd.DataFrame, detail: pd.DataFrame) -> pd.DataFrame:
    major_by_sheet = (
        detail.groupby(SHEET_NAME_COL, as_index=False)
        .agg(**{MAJOR_CATEGORY_COL: (MAJOR_CATEGORY_COL, most_common_text)})
    )
    result = sheet_name_analysis.merge(major_by_sheet, on=SHEET_NAME_COL, how="left")
    columns = result.columns.tolist()
    columns.remove(MAJOR_CATEGORY_COL)
    insert_at = columns.index(SHEET_NAME_COL)
    columns.insert(insert_at, MAJOR_CATEGORY_COL)
    return result[columns]


def build_top_sheet(detail: pd.DataFrame, value_col: str) -> pd.DataFrame:
    group_columns = [OUTPUT_PROCESS_COL, OUTPUT_SKU_COL, OUTPUT_PRODUCT_NAME_COL]
    if SHEET_NAME_COL in detail.columns:
        group_columns.insert(1, SHEET_NAME_COL)
    if MAJOR_CATEGORY_COL in detail.columns:
        group_columns.insert(1, MAJOR_CATEGORY_COL)

    top = (
        detail.groupby(group_columns, as_index=False)
        .agg(
            **{
                CUSTOMER_COL: (CUSTOMER_COL, join_unique_text),
                INITIAL_COL: (INITIAL_COL, join_unique_text),
                SUMMARY_SHORTAGE_COL: (SHORTAGE_QTY_COL, "sum"),
                SUMMARY_NEED_COL: (NEED_QTY_COL, "sum"),
                SUMMARY_ACTUAL_COL: (ACTUAL_QTY_COL, "sum"),
                SUMMARY_EFFECTIVE_COL: (EFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_INEFFECTIVE_COL: (INEFFECTIVE_PRODUCTION_COL, "sum"),
                SUMMARY_REMAINING_COL: (REMAINING_NEED_COL, "sum"),
            }
        )
    )
    top[PRODUCTION_EFFECTIVENESS_COL] = calculate_rate(
        top[SUMMARY_EFFECTIVE_COL], top[SUMMARY_ACTUAL_COL]
    )
    return top.sort_values(value_col, ascending=False).head(30)


def format_top_skus(frame: pd.DataFrame, value_col: str, limit: int = 3) -> str:
    rows = frame.sort_values(value_col, ascending=False).head(limit)
    parts = []
    for _, row in rows.iterrows():
        value = row[value_col]
        if value <= 0:
            continue
        parts.append(f"{row[OUTPUT_SKU_COL]} {value:,.0f}")
    return "; ".join(parts)


def describe_change(
    effectiveness_delta: float,
    effective_delta: float,
    actual_delta: float,
    ineffective_delta: float,
    no_plan_actual_delta: float,
) -> str:
    if effectiveness_delta == 0:
        return "생산유효도 변동 없음"

    reasons: list[str] = []
    if effectiveness_delta > 0:
        if effective_delta > 0:
            reasons.append("유효생산량 증가")
        if actual_delta < 0:
            reasons.append("실적수량 감소로 분모 축소")
        if ineffective_delta < 0:
            reasons.append("비유효생산량 감소")
        if no_plan_actual_delta < 0:
            reasons.append("계획없음 실적 감소")
        if not reasons:
            reasons.append("실적 대비 유효생산 비중 상승")
    else:
        if effective_delta < 0:
            reasons.append("유효생산량 감소")
        if actual_delta > 0 and effective_delta <= 0:
            reasons.append("실적 증가가 필요 SKU와 매칭되지 않음")
        if ineffective_delta > 0:
            reasons.append("비유효생산량 증가")
        if no_plan_actual_delta > 0:
            reasons.append("계획없음 실적 증가")
        if not reasons:
            reasons.append("실적 대비 유효생산 비중 하락")
    return ", ".join(reasons)


def build_change_analysis(summary: pd.DataFrame, detail: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    for process, process_summary in summary.groupby(OUTPUT_PROCESS_COL):
        process_summary = process_summary.sort_values(DATE_COL).reset_index(drop=True)
        for index in range(1, len(process_summary)):
            previous = process_summary.loc[index - 1]
            current = process_summary.loc[index]
            previous_date = str(previous[DATE_COL])
            current_date = str(current[DATE_COL])

            previous_detail = detail[
                detail[DATE_COL].astype(str).eq(previous_date)
                & detail[OUTPUT_PROCESS_COL].eq(process)
            ]
            current_detail = detail[
                detail[DATE_COL].astype(str).eq(current_date)
                & detail[OUTPUT_PROCESS_COL].eq(process)
            ]

            effectiveness_delta = (
                current[PRODUCTION_EFFECTIVENESS_COL] - previous[PRODUCTION_EFFECTIVENESS_COL]
            )
            need_delta = current[SUMMARY_NEED_COL] - previous[SUMMARY_NEED_COL]
            actual_delta = current[SUMMARY_ACTUAL_COL] - previous[SUMMARY_ACTUAL_COL]
            effective_delta = current[SUMMARY_EFFECTIVE_COL] - previous[SUMMARY_EFFECTIVE_COL]
            ineffective_delta = current[SUMMARY_INEFFECTIVE_COL] - previous[SUMMARY_INEFFECTIVE_COL]
            remaining_delta = current[SUMMARY_REMAINING_COL] - previous[SUMMARY_REMAINING_COL]

            current_no_plan_actual = current_detail.loc[
                current_detail[MATCH_STATUS_COL].eq("계획없음"), ACTUAL_QTY_COL
            ].sum()
            previous_no_plan_actual = previous_detail.loc[
                previous_detail[MATCH_STATUS_COL].eq("계획없음"), ACTUAL_QTY_COL
            ].sum()
            current_no_actual_need = current_detail.loc[
                current_detail[MATCH_STATUS_COL].eq("실적없음"), NEED_QTY_COL
            ].sum()

            rows.append(
                {
                    DATE_COL: current_date,
                    PREV_DATE_COL: previous_date,
                    OUTPUT_PROCESS_COL: process,
                    PREV_EFFECTIVENESS_COL: previous[PRODUCTION_EFFECTIVENESS_COL],
                    PRODUCTION_EFFECTIVENESS_COL: current[PRODUCTION_EFFECTIVENESS_COL],
                    EFFECTIVENESS_DELTA_COL: round(effectiveness_delta, 1),
                    DIRECTION_COL: "상승"
                    if effectiveness_delta > 0
                    else "하락"
                    if effectiveness_delta < 0
                    else "변동없음",
                    NEED_DELTA_COL: need_delta,
                    ACTUAL_DELTA_COL: actual_delta,
                    EFFECTIVE_DELTA_COL: effective_delta,
                    INEFFECTIVE_DELTA_COL: ineffective_delta,
                    REMAINING_DELTA_COL: remaining_delta,
                    NO_PLAN_ACTUAL_COL: current_no_plan_actual,
                    NO_ACTUAL_NEED_COL: current_no_actual_need,
                    MAIN_REASON_COL: describe_change(
                        effectiveness_delta,
                        effective_delta,
                        actual_delta,
                        ineffective_delta,
                        current_no_plan_actual - previous_no_plan_actual,
                    ),
                    TOP_INEFFECTIVE_SKU_COL: format_top_skus(
                        current_detail, INEFFECTIVE_PRODUCTION_COL
                    ),
                    TOP_REMAINING_SKU_COL: format_top_skus(current_detail, REMAINING_NEED_COL),
                }
            )

    return pd.DataFrame(rows)


def build_kind_analysis(summary: pd.DataFrame) -> pd.DataFrame:
    return summary[
        [
            DATE_COL,
            OUTPUT_PROCESS_COL,
            PRODUCTION_KIND_COL,
            EFFECTIVE_KIND_COL,
            INEFFECTIVE_KIND_COL,
            PRODUCTION_KIND_EFFECTIVENESS_COL,
            PRODUCTION_SKU_COL,
            EFFECTIVE_SKU_COL,
            INEFFECTIVE_SKU_COL,
            NEEDED_SKU_COL,
            COVERED_SKU_COL,
            UNCOVERED_NEEDED_SKU_COL,
            EXCESS_SKU_COL,
            SHORT_SKU_COL,
            SKU_EFFECTIVENESS_COL,
            PRODUCTION_EFFECTIVENESS_COL,
            SUMMARY_ACTUAL_COL,
            SUMMARY_EFFECTIVE_COL,
            SUMMARY_INEFFECTIVE_COL,
            AVG_ACTUAL_PER_KIND_COL,
            AVG_EFFECTIVE_PER_KIND_COL,
            AVG_INEFFECTIVE_PER_KIND_COL,
        ]
    ].copy()


def add_process_line_chart(
    ws, summary: pd.DataFrame, value_col_name: str, title: str, anchor: str
) -> None:
    if summary.empty:
        return
    chart = LineChart()
    chart.title = title
    chart.height = 8
    chart.width = 18
    chart.y_axis.title = title
    chart.x_axis.title = DATE_COL
    chart.y_axis.numFmt = '0.0"%"'

    value_col = list(summary.columns).index(value_col_name) + 1
    for process in sorted(summary[OUTPUT_PROCESS_COL].unique()):
        row_indexes = summary.index[summary[OUTPUT_PROCESS_COL].eq(process)].tolist()
        if not row_indexes:
            continue
        start_row = row_indexes[0] + 2
        end_row = row_indexes[-1] + 2
        values = Reference(ws, min_col=value_col, min_row=start_row, max_row=end_row)
        xvalues = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
        chart.series.append(Series(values, xvalues=xvalues, title=process))
    ws.add_chart(chart, anchor)


def add_process_bar_chart(
    ws, summary: pd.DataFrame, value_col_name: str, title: str, anchor: str
) -> None:
    if summary.empty:
        return
    chart = BarChart()
    chart.title = title
    chart.height = 8
    chart.width = 18
    chart.y_axis.title = title
    chart.x_axis.title = DATE_COL

    value_col = list(summary.columns).index(value_col_name) + 1
    for process in sorted(summary[OUTPUT_PROCESS_COL].unique()):
        row_indexes = summary.index[summary[OUTPUT_PROCESS_COL].eq(process)].tolist()
        if not row_indexes:
            continue
        start_row = row_indexes[0] + 2
        end_row = row_indexes[-1] + 2
        values = Reference(ws, min_col=value_col, min_row=start_row, max_row=end_row)
        xvalues = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
        chart.series.append(Series(values, xvalues=xvalues, title=process))
    ws.add_chart(chart, anchor)


def add_summary_charts(ws, summary: pd.DataFrame) -> None:
    chart_summary = summary.reset_index(drop=True)
    add_process_line_chart(
        ws,
        chart_summary,
        PRODUCTION_EFFECTIVENESS_COL,
        "일자별/공정별 생산유효도 추이",
        "L2",
    )
    add_process_bar_chart(ws, chart_summary, SUMMARY_INEFFECTIVE_COL, "공정별 비유효생산량", "L20")


def add_kind_analysis_charts(ws, kind_analysis: pd.DataFrame) -> None:
    chart_data = kind_analysis.reset_index(drop=True)
    add_process_bar_chart(ws, chart_data, PRODUCTION_KIND_COL, "일자별/공정별 생산종수", "L2")
    add_process_line_chart(
        ws,
        chart_data,
        PRODUCTION_KIND_EFFECTIVENESS_COL,
        "일자별/공정별 생산종수 유효도",
        "L20",
    )
    add_process_line_chart(
        ws,
        chart_data,
        PRODUCTION_EFFECTIVENESS_COL,
        "일자별/공정별 생산유효도",
        "L38",
    )
    if PRODUCTION_SKU_COL in chart_data.columns:
        add_process_bar_chart(ws, chart_data, PRODUCTION_SKU_COL, "일자별/공정별 생산 SKU수", "L56")
    if SKU_EFFECTIVENESS_COL in chart_data.columns:
        add_process_line_chart(
            ws,
            chart_data,
            SKU_EFFECTIVENESS_COL,
            "일자별/공정별 SKU 유효율",
            "L74",
        )


def format_workbook(workbook) -> None:
    rate_columns = {
        PRODUCTION_EFFECTIVENESS_COL,
        PREV_EFFECTIVENESS_COL,
        PRODUCTION_KIND_EFFECTIVENESS_COL,
        SKU_EFFECTIVENESS_COL,
    }
    average_columns = {
        AVG_ACTUAL_PER_KIND_COL,
        AVG_EFFECTIVE_PER_KIND_COL,
        AVG_INEFFECTIVE_PER_KIND_COL,
    }
    integer_keywords = (
        "수량",
        "생산량",
        "필요",
        "실적",
        "종수",
        "SKU수",
    )

    for ws in workbook.worksheets:
        if ws.max_row == 0:
            continue

        header = [cell.value for cell in ws[1]]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2E8F0")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        for col_index, column_name in enumerate(header, start=1):
            if column_name in rate_columns:
                number_format = '0.0"%"'
            elif column_name in average_columns:
                number_format = "#,##0.0"
            elif column_name and any(keyword in str(column_name) for keyword in integer_keywords):
                number_format = "#,##0"
            else:
                number_format = None

            if number_format:
                for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
                    row[0].number_format = number_format

        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(length + 2, 10), 34
            )


def write_report(
    output_path: Path,
    summary: pd.DataFrame,
    change_analysis: pd.DataFrame,
    kind_analysis: pd.DataFrame,
    sheet_name_analysis: pd.DataFrame,
    major_category_analysis: pd.DataFrame,
    detail: pd.DataFrame,
    top_ineffective: pd.DataFrame,
    top_remaining: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
        change_analysis.to_excel(writer, sheet_name=CHANGE_ANALYSIS_SHEET, index=False)
        kind_analysis.to_excel(writer, sheet_name=KIND_ANALYSIS_SHEET, index=False)
        sheet_name_analysis.to_excel(writer, sheet_name=SHEET_NAME_ANALYSIS_SHEET, index=False)
        major_category_analysis.to_excel(
            writer, sheet_name=MAJOR_CATEGORY_ANALYSIS_SHEET, index=False
        )
        detail.to_excel(writer, sheet_name=DETAIL_SHEET, index=False)
        top_ineffective.to_excel(writer, sheet_name=TOP_INEFFECTIVE_SHEET, index=False)
        top_remaining.to_excel(writer, sheet_name=TOP_REMAINING_SHEET, index=False)
        add_summary_charts(writer.book[SUMMARY_SHEET], summary)
        add_kind_analysis_charts(writer.book[KIND_ANALYSIS_SHEET], kind_analysis)
        format_workbook(writer.book)


def timestamped_report_path(root: Path, latest_date: str) -> Path:
    timestamp = pd.Timestamp.now().strftime("%H%M%S")
    return root / f"생산유효도_공정별_{latest_date}_{timestamp}.xlsx"


def main() -> None:
    root = Path(__file__).parent
    input_files = find_input_files(root)
    if not input_files:
        raise SystemExit("수요정보(전공정)_YYYYMMDD.xlsx 파일이 필요합니다.")

    input_dates = {date for date, _ in input_files}
    production_df, production_source = load_production_source(root, input_dates)
    match_on_product_name = production_uses_product_name_key_from_df(
        production_source, production_df
    )
    demand = build_demand(input_files, match_on_product_name)
    valid_dates = set(demand[DATE_COL].unique())
    production = prepare_production_data(production_df, production_source, valid_dates)
    detail = build_detail(demand, production)
    detail = apply_classification(detail, root)
    summary = summarize_by_process(detail)
    change_analysis = build_change_analysis(summary, detail)
    kind_analysis = build_kind_analysis(summary)
    sheet_name_analysis = attach_major_category(
        summarize_by_classification(detail, SHEET_NAME_COL), detail
    )
    major_category_analysis = summarize_by_classification(detail, MAJOR_CATEGORY_COL)
    top_ineffective = build_top_sheet(detail, SUMMARY_INEFFECTIVE_COL)
    top_remaining = build_top_sheet(detail, SUMMARY_REMAINING_COL)

    latest_date = max(date for date, _ in input_files)
    output_path = root / f"생산유효도_공정별_{latest_date}.xlsx"
    input_paths = {path.resolve() for _, path in input_files}
    if output_path.exists() and (
        output_path.resolve() in input_paths or not has_report_sheets(output_path)
    ):
        output_path = timestamped_report_path(root, latest_date)

    try:
        write_report(
            output_path,
            summary,
            change_analysis,
            kind_analysis,
            sheet_name_analysis,
            major_category_analysis,
            detail,
            top_ineffective,
            top_remaining,
        )
    except PermissionError:
        output_path = timestamped_report_path(root, latest_date)
        write_report(
            output_path,
            summary,
            change_analysis,
            kind_analysis,
            sheet_name_analysis,
            major_category_analysis,
            detail,
            top_ineffective,
            top_remaining,
        )

    print(f"output={output_path.name}")
    print(f"demand_dates={input_files[0][0]}~{input_files[-1][0]}")
    print(f"production_source={source_display_name(production_source)}")
    print(f"match_key={'제품명' if match_on_product_name else '제품코드'}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
