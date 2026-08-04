# -*- coding: utf-8 -*-
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[2]
WORK_DIR = BASE_DIR / ".codex_tmp" / "pcode_template"
OUTPUT_DIR = BASE_DIR / "outputs" / "pcode_template"
INPUT_JSON = WORK_DIR / "pcode_auto_rows.json"
OUTPUT_PATH = OUTPUT_DIR / "전체_P코드_수식형_거래처별_양식.xlsx"

MASTER_SHEET = "품목마스터"
DEMAND_SHEET = "수요_집계"
WIP_SHEET = "공정재고_집계"
FINISHED_SHEET = "완제품재고_집계"

MASTER_HEADERS = ["거래처그룹", "제품분류", "생산코드", "분리코드", "사출코드", "제품명", "파워"]
DEMAND_HEADERS = ["거래처그룹", "생산코드", "이니셜", "오더수량", "납기일", "제품부족수량", "사출부족수량"]
WIP_HEADERS = ["생산코드", "사출재고", "분리재고", "검사접착재고", "누수규격검사재고", "공정재고 합계"]
FINISHED_HEADERS = ["생산코드", "완제품재고", "재고변화", "DOI기준오더", "DOI", "재고비율", "재고신호", "대응판단"]
BASE_CUSTOMER_HEADERS = ["생산코드", "분리코드", "사출코드", "제품명", "파워"]
TRAILING_CUSTOMER_HEADERS = [
    "오더합계",
    "납기일",
    "제품부족수량",
    "사출부족수량",
    "사출재고",
    "분리재고",
    "검사접착재고",
    "누수규격검사재고",
    "공정재고 합계",
    "완제품재고",
    "재고변화",
    "DOI기준오더",
    "DOI",
    "재고비율",
    "재고신호",
    "대응판단",
    "상태",
]
FIRST_DATA_ROW = 6
HEADER_ROW = 5

NAVY = "172554"
LIGHT_BLUE = "DBEAFE"
GRAY = "F8FAFC"
LIGHT_AMBER = "FEF3C7"
TEXT = "0F172A"
MUTED = "64748B"
BORDER = "CBD5E1"

thin_border = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)


def make_cell(ws, value, fill=None, font=None, align=None, border=False, number_format=None):
    cell = WriteOnlyCell(ws, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = thin_border
    if number_format:
        cell.number_format = number_format
    return cell


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "nat", "null", "<na>"} else text


def numeric(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def parse_due_date(value):
    text = clean_text(value)
    if not text or text == "-":
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def safe_sheet_name(raw, used):
    invalid = "\\/*?:[]"
    name = "".join("_" if ch in invalid else ch for ch in str(raw or "거래처 미지정")).strip()
    name = " ".join(name.split()) or "거래처 미지정"
    name = name[:31]
    base = name
    idx = 2
    while name in used:
        suffix = f"_{idx}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used.add(name)
    return name


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_title(ws, title, subtitle, col_count):
    title_row = [
        make_cell(
            ws,
            title if i == 0 else "",
            fill=PatternFill("solid", fgColor=NAVY),
            font=Font(name="맑은 고딕", bold=True, color="FFFFFF", size=15),
            align=Alignment(horizontal="left"),
        )
        for i in range(col_count)
    ]
    subtitle_row = [
        make_cell(
            ws,
            subtitle if i == 0 else "",
            fill=PatternFill("solid", fgColor=GRAY),
            font=Font(name="맑은 고딕", color=MUTED, size=10),
            align=Alignment(wrap_text=True),
        )
        for i in range(col_count)
    ]
    ws.append(title_row)
    ws.append(subtitle_row)


def styled_header(ws, headers):
    fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    font = Font(name="맑은 고딕", bold=True, color=TEXT)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append([make_cell(ws, h, fill=fill, font=font, align=align, border=True) for h in headers])


def index_match_formula(sheet_name, return_col, lookup_col, code_cell, fallback="0"):
    return (
        f'=IFERROR(INDEX(\'{sheet_name}\'!${return_col}${FIRST_DATA_ROW}:${return_col}${{last_row}},'
        f'MATCH({code_cell},\'{sheet_name}\'!${lookup_col}${FIRST_DATA_ROW}:${lookup_col}${{last_row}},0)),{fallback})'
    )


def build_source_rows(rows):
    master_rows = []
    demand_rows = []
    wip_rows = []
    finished_rows = []
    customer_rows = {}
    customer_initials = {}

    for rec in rows:
        customer = clean_text(rec.get("거래처그룹")) or "거래처 미지정"
        product_code = clean_text(rec.get("생산코드"))
        master_row = [
            customer,
            clean_text(rec.get("제품분류")) or "기타",
            product_code,
            clean_text(rec.get("분리코드")),
            clean_text(rec.get("사출코드")),
            clean_text(rec.get("제품명")),
            clean_text(rec.get("파워")),
        ]
        master_rows.append(master_row)
        customer_rows.setdefault(customer, []).append(master_row[2:])

        order_qty = numeric(rec.get("오더수량1"))
        product_shortage = numeric(rec.get("제품부족수량"))
        injection_shortage = numeric(rec.get("사출부족수량"))
        initial = clean_text(rec.get("이니셜"))
        due = parse_due_date(rec.get("납기일"))
        if order_qty or product_shortage or injection_shortage or initial or due:
            demand_rows.append([customer, product_code, initial, order_qty, due, product_shortage, injection_shortage])
            if order_qty and initial:
                customer_initials.setdefault(customer, set()).add(initial)

        wip_rows.append(
            [
                product_code,
                numeric(rec.get("사출재고")),
                numeric(rec.get("분리재고")),
                numeric(rec.get("검사접착재고")),
                numeric(rec.get("누수규격검사재고")),
                "",
            ]
        )
        finished_rows.append(
            [
                product_code,
                numeric(rec.get("완제품재고")),
                numeric(rec.get("재고변화")),
                numeric(rec.get("DOI기준오더")),
                numeric(rec.get("DOI")),
                clean_text(rec.get("재고비율")),
                clean_text(rec.get("재고신호")),
                clean_text(rec.get("대응판단")),
            ]
        )

    customers = sorted(customer_rows)
    customer_initials = {customer: sorted(values) for customer, values in customer_initials.items()}
    return master_rows, demand_rows, wip_rows, finished_rows, customers, customer_rows, customer_initials


def append_source_sheet(wb, name, title, subtitle, headers, rows, widths, formula_cols=None, number_cols=None, date_cols=None):
    ws = wb.create_sheet(name)
    append_title(ws, title, subtitle, len(headers))
    ws.append([])
    styled_header(ws, headers)
    formula_cols = formula_cols or {}
    number_cols = set(number_cols or [])
    date_cols = set(date_cols or [])
    for row_idx, row_values in enumerate(rows, start=FIRST_DATA_ROW):
        row = []
        for col_idx, value in enumerate(row_values, start=1):
            if col_idx in formula_cols:
                value = formula_cols[col_idx](row_idx)
            fmt = None
            if col_idx in number_cols:
                fmt = "#,##0"
            elif col_idx in date_cols:
                fmt = "yyyy-mm-dd"
            row.append(make_cell(ws, value, font=Font(name="맑은 고딕", size=10, color=TEXT), number_format=fmt))
        ws.append(row)
    if rows:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(rows)}"
    ws.freeze_panes = "A6"
    set_widths(ws, widths)
    return ws


def customer_formula_row(ws, master_row, row_idx, headers, initial_headers, ranges):
    values = list(master_row)
    for initial in initial_headers:
        escaped_initial = str(initial).replace('"', '""')
        values.append(
            f'=SUMIFS(\'{DEMAND_SHEET}\'!$D${FIRST_DATA_ROW}:$D${ranges["demand_last"]},'
            f'\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
            f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx},'
            f'\'{DEMAND_SHEET}\'!$C${FIRST_DATA_ROW}:$C${ranges["demand_last"]},"{escaped_initial}")'
        )

    order_start = len(BASE_CUSTOMER_HEADERS) + 1
    order_end = order_start + len(initial_headers) - 1
    trailing_start = len(BASE_CUSTOMER_HEADERS) + len(initial_headers) + 1
    col = {header: get_column_letter(headers.index(header) + 1) for header in headers}

    if initial_headers:
        values.append(f"=SUM({get_column_letter(order_start)}{row_idx}:{get_column_letter(order_end)}{row_idx})")
    else:
        values.append(
            f'=SUMIFS(\'{DEMAND_SHEET}\'!$D${FIRST_DATA_ROW}:$D${ranges["demand_last"]},'
            f'\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
            f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx})'
        )
    values.append(
        f'=IF(COUNTIFS(\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
        f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx})=0,"",'
        f'MINIFS(\'{DEMAND_SHEET}\'!$E${FIRST_DATA_ROW}:$E${ranges["demand_last"]},'
        f'\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
        f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx}))'
    )
    values.append(
        f'=SUMIFS(\'{DEMAND_SHEET}\'!$F${FIRST_DATA_ROW}:$F${ranges["demand_last"]},'
        f'\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
        f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx})'
    )
    values.append(
        f'=SUMIFS(\'{DEMAND_SHEET}\'!$G${FIRST_DATA_ROW}:$G${ranges["demand_last"]},'
        f'\'{DEMAND_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["demand_last"]},$B$3,'
        f'\'{DEMAND_SHEET}\'!$B${FIRST_DATA_ROW}:$B${ranges["demand_last"]},$A{row_idx})'
    )

    for source_col in ["B", "C", "D", "E", "F"]:
        values.append(
            f'=IFERROR(INDEX(\'{WIP_SHEET}\'!${source_col}${FIRST_DATA_ROW}:${source_col}${ranges["wip_last"]},'
            f'MATCH($A{row_idx},\'{WIP_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["wip_last"]},0)),0)'
        )
    for source_col, fallback in [("B", "0"), ("C", "0"), ("D", "0"), ("E", "0"), ("F", '""'), ("G", '""'), ("H", '""')]:
        values.append(
            f'=IFERROR(INDEX(\'{FINISHED_SHEET}\'!${source_col}${FIRST_DATA_ROW}:${source_col}${ranges["finished_last"]},'
            f'MATCH($A{row_idx},\'{FINISHED_SHEET}\'!$A${FIRST_DATA_ROW}:$A${ranges["finished_last"]},0)),{fallback})'
        )
    values.append(
        f'=IF({col["사출부족수량"]}{row_idx}>0,"사출부족",'
        f'IF({col["제품부족수량"]}{row_idx}>0,"제품부족",'
        f'IF(AND({col["오더합계"]}{row_idx}=0,{col["공정재고 합계"]}{row_idx}>0),"수요없음공정재고",'
        f'IF(AND({col["오더합계"]}{row_idx}=0,{col["완제품재고"]}{row_idx}>0),"수요없음완제품재고",'
        f'IF(AND(ISNUMBER({col["DOI"]}{row_idx}),{col["DOI"]}{row_idx}>180),"DOI높음",'
        f'IF({col["오더합계"]}{row_idx}>0,"수요있음","수요없음"))))))'
    )

    row = []
    for col_idx, value in enumerate(values, start=1):
        header = headers[col_idx - 1]
        fmt = None
        if header == "납기일":
            fmt = "yyyy-mm-dd"
        elif header == "DOI":
            fmt = "0.0"
        elif header not in {"생산코드", "분리코드", "사출코드", "제품명", "파워", "재고비율", "재고신호", "대응판단", "상태"}:
            fmt = "#,##0"
        row.append(make_cell(ws, value, font=Font(name="맑은 고딕", size=10, color=TEXT), number_format=fmt))
    return row


def append_customer_sheet(wb, customer, rows, initial_headers, ranges, used_names):
    sheet_name = safe_sheet_name(customer, used_names)
    order_headers = [f"오더수량({initial})" for initial in initial_headers]
    headers = [*BASE_CUSTOMER_HEADERS, *order_headers, *TRAILING_CUSTOMER_HEADERS]
    ws = wb.create_sheet(sheet_name)
    append_title(ws, f"{customer} 품목 현황", "품목코드는 고정값이며 수요, 공정재고, 완제품 재고는 집계 시트를 조회하는 수식입니다.", len(headers))
    ws.append(
        [
            make_cell(ws, "조회 거래처", fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="맑은 고딕", bold=True), border=True),
            make_cell(ws, customer, fill=PatternFill("solid", fgColor=GRAY), font=Font(name="맑은 고딕", bold=True), border=True),
            *[make_cell(ws, "") for _ in range(len(headers) - 2)],
        ]
    )
    ws.append([])
    styled_header(ws, headers)
    for idx, master_row in enumerate(rows, start=FIRST_DATA_ROW):
        ws.append(customer_formula_row(ws, master_row, idx, headers, initial_headers, ranges))
    if rows:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(rows)}"
    ws.freeze_panes = "F6"
    width_count = len(headers)
    default_widths = [18, 18, 18, 34, 10] + [15] * len(order_headers) + [13, 13, 15, 15, 13, 13, 15, 18, 15, 15, 14, 14, 10, 12, 12, 32, 15]
    set_widths(ws, default_widths[:width_count])
    return sheet_name


def build_workbook():
    payload = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    rows = payload["rows"]
    master_rows, demand_rows, wip_rows, finished_rows, customers, customer_rows, customer_initials = build_source_rows(rows)
    ranges = {
        "demand_last": HEADER_ROW + max(len(demand_rows), 1),
        "wip_last": HEADER_ROW + max(len(wip_rows), 1),
        "finished_last": HEADER_ROW + max(len(finished_rows), 1),
    }

    wb = Workbook(write_only=True)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    guide = wb.create_sheet("안내")
    append_title(
        guide,
        "전체 P코드 수식형 관리 양식",
        "품목마스터는 고정값, 수요/재고/DOI는 집계 시트를 조회하는 수식으로 구성했습니다.",
        4,
    )
    guide.append([])
    styled_header(guide, ["구분", "내용", "행수/비고", ""])
    guide_rows = [
        ["품목마스터", "품목리스트(대량) 기준으로 전체 P코드를 고정값으로 보관합니다.", f"{len(master_rows):,}개 P코드", ""],
        ["수요_집계", "오더수량, 납기일, 부족수량, 사출부족수량의 계산 원천입니다.", f"{len(demand_rows):,}행", ""],
        ["공정재고_집계", "사출/분리/검사접착/누수규격 재고와 공정재고 합계 원천입니다.", f"{len(wip_rows):,}행", ""],
        ["완제품재고_집계", "완제품재고, 재고변화, DOI, 신호, 대응판단 원천입니다.", payload.get("finished_goods_stock", ""), ""],
        ["거래처별 시트", "품목코드/제품명은 고정값이고 나머지는 SUMIFS, INDEX/MATCH 수식입니다.", f"{len(customers):,}개 시트", ""],
    ]
    for row in guide_rows:
        guide.append([make_cell(guide, v, font=Font(name="맑은 고딕", size=10), border=True) for v in row])
    set_widths(guide, [26, 86, 45, 12])

    append_source_sheet(
        wb,
        MASTER_SHEET,
        "품목마스터",
        "품목리스트(대량) 기준 전체 P코드 고정값입니다.",
        MASTER_HEADERS,
        master_rows,
        [20, 13, 18, 18, 18, 36, 10],
    )
    append_source_sheet(
        wb,
        DEMAND_SHEET,
        "수요 집계",
        "거래처/생산코드/이니셜 기준 수요 계산 원천입니다.",
        DEMAND_HEADERS,
        demand_rows or [["", "", "", 0, None, 0, 0]],
        [20, 18, 18, 14, 13, 15, 15],
        number_cols={4, 6, 7},
        date_cols={5},
    )
    append_source_sheet(
        wb,
        WIP_SHEET,
        "공정재고 집계",
        "생산코드 기준 공정재고 계산 원천입니다.",
        WIP_HEADERS,
        wip_rows,
        [18, 13, 13, 15, 18, 15],
        formula_cols={6: lambda r: f"=SUM(B{r}:E{r})"},
        number_cols={2, 3, 4, 5, 6},
    )
    append_source_sheet(
        wb,
        FINISHED_SHEET,
        "완제품재고 집계",
        "생산코드 기준 완제품 재고 변화와 DOI 원천입니다.",
        FINISHED_HEADERS,
        finished_rows,
        [18, 15, 14, 14, 10, 12, 12, 36],
        number_cols={2, 3, 4},
    )

    used_names = {"안내", MASTER_SHEET, DEMAND_SHEET, WIP_SHEET, FINISHED_SHEET}
    sheet_map = []
    for customer in customers:
        sheet_name = append_customer_sheet(
            wb,
            customer,
            customer_rows[customer],
            customer_initials.get(customer, []),
            ranges,
            used_names,
        )
        sheet_map.append((customer, sheet_name, len(customer_rows[customer]), len(customer_initials.get(customer, []))))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH, len(master_rows), len(demand_rows), len(customers), sheet_map


def verify(path: Path):
    wb = load_workbook(path, read_only=True, data_only=False)
    customer_sheet = wb.worksheets[5]
    checks = {
        "sheets": len(wb.sheetnames),
        "master_headers": [wb[MASTER_SHEET].cell(HEADER_ROW, c).value for c in range(1, len(MASTER_HEADERS) + 1)],
        "demand_headers": [wb[DEMAND_SHEET].cell(HEADER_ROW, c).value for c in range(1, len(DEMAND_HEADERS) + 1)],
        "customer_sheet": customer_sheet.title,
        "customer_a6": customer_sheet["A6"].value,
        "customer_order_formula": customer_sheet.cell(FIRST_DATA_ROW, 6).value,
        "customer_status_formula": customer_sheet.cell(FIRST_DATA_ROW, customer_sheet.max_column).value,
    }
    wb.close()
    return checks


if __name__ == "__main__":
    path, rows, demand_rows, customers, sheet_map = build_workbook()
    checks = verify(path)
    print(
        json.dumps(
            {
                "output": str(path),
                "rows": rows,
                "demand_rows": demand_rows,
                "customers": customers,
                "checks": checks,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
