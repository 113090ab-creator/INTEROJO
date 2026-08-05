# -*- coding: utf-8 -*-
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[2]
WORK_DIR = BASE_DIR / ".codex_tmp" / "pcode_template"
OUTPUT_DIR = BASE_DIR / "outputs" / "pcode_template"
INPUT_JSON = WORK_DIR / "pcode_auto_rows.json"

AUTO_OUTPUT = OUTPUT_DIR / "전체_P코드_자동조회_거래처별_양식_신제품반영.xlsx"
FORMULA_OUTPUT = OUTPUT_DIR / "전체_P코드_수식형_거래처별_양식_신제품반영.xlsx"

FIRST_DATA_ROW = 6
HEADER_ROW = 5

NAVY = "172554"
LIGHT_BLUE = "DBEAFE"
GRAY = "F8FAFC"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED = "FEE2E2"
TEXT = "0F172A"
MUTED = "64748B"
BORDER = "CBD5E1"

thin_border = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)

NEW_PRODUCT_PATTERNS = ("BAGUMORE", "Burn Sugar", "Viva Boom", "중국_축고정", "축고정")


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "nat", "null", "<na>"} else text


def numeric(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def parse_due_date(value):
    text = clean_text(value)
    if not text or text == "-":
        return ""
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except ValueError:
        return text


def is_new_product(product_name):
    name = clean_text(product_name)
    lowered = name.lower()
    return any(pattern.lower() in lowered for pattern in NEW_PRODUCT_PATTERNS if pattern.isascii()) or any(
        pattern in name for pattern in NEW_PRODUCT_PATTERNS if not pattern.isascii()
    )


def doi_status(row):
    if row["신제품여부"] == "Y":
        return "신제품"
    finished_stock = numeric(row.get("완제품재고"))
    order_total = numeric(row.get("오더수량1")) + numeric(row.get("오더수량2"))
    shortage = numeric(row.get("제품부족수량"))
    injection_shortage = numeric(row.get("사출부족수량"))
    doi = numeric(row.get("DOI"))
    if finished_stock <= 0 and (order_total > 0 or shortage > 0 or injection_shortage > 0):
        return "재고없음"
    if finished_stock > 0 and order_total <= 0:
        return "오더없음"
    if doi > 0:
        return "계산됨"
    return "확인필요"


def base_status(row):
    if row.get("DOI상태") == "신제품":
        return ""
    order_total = numeric(row.get("오더수량1")) + numeric(row.get("오더수량2"))
    shortage = numeric(row.get("제품부족수량"))
    injection_shortage = numeric(row.get("사출부족수량"))
    process_stock = (
        numeric(row.get("사출재고"))
        + numeric(row.get("분리재고"))
        + numeric(row.get("검사접착재고"))
        + numeric(row.get("누수규격검사재고"))
    )
    doi = numeric(row.get("DOI"))
    if injection_shortage > 0:
        return "사출부족"
    if shortage > 0:
        return "제품부족"
    if order_total == 0 and process_stock > 0:
        return "수요없음재고"
    if doi > 0 and doi < 7:
        return "DOI주의"
    if order_total > 0:
        return "수요있음"
    return "수요없음"


def enrich_rows(rows):
    enriched = []
    for source in rows:
        row = dict(source)
        row["신제품여부"] = "Y" if is_new_product(row.get("제품명")) else ""
        row["DOI상태"] = doi_status(row)
        row["상태"] = base_status(row)
        enriched.append(row)
    return enriched


def make_cell(ws, value, fill=None, font=None, align=None, border=False, number_format=None):
    cell = WriteOnlyCell(ws, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = thin_border
    if number_format:
        cell.number_format = number_format
    return cell


def styled_header(ws, headers):
    fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    font = Font(name="맑은 고딕", bold=True, color=TEXT)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append([make_cell(ws, h, fill=fill, font=font, align=align, border=True) for h in headers])


def append_title(ws, title, subtitle, col_count):
    ws.append(
        [
            make_cell(
                ws,
                title if idx == 0 else "",
                fill=PatternFill("solid", fgColor=NAVY),
                font=Font(name="맑은 고딕", bold=True, color="FFFFFF", size=15),
            )
            for idx in range(col_count)
        ]
    )
    ws.append(
        [
            make_cell(
                ws,
                subtitle if idx == 0 else "",
                fill=PatternFill("solid", fgColor=GRAY),
                font=Font(name="맑은 고딕", color=MUTED, size=10),
                align=Alignment(wrap_text=True),
            )
            for idx in range(col_count)
        ]
    )


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def safe_sheet_name(raw, used):
    invalid = "\\/*?:[]"
    name = "".join("_" if ch in invalid else ch for ch in clean_text(raw) or "거래처 미지정")
    name = " ".join(name.split())[:31] or "거래처 미지정"
    base = name
    idx = 2
    while name in used:
        suffix = f"_{idx}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used.add(name)
    return name


def append_data_row(ws, values, headers):
    row = []
    for header, value in zip(headers, values):
        fmt = None
        if header in {"오더수량1", "오더수량2", "오더합계", "제품부족수량", "사출부족수량", "사출재고", "분리재고", "검사접착재고", "누수규격검사재고", "공정재고 합계", "공정재고합계", "완제품재고", "재고변화", "DOI기준오더"}:
            fmt = "#,##0"
        elif header == "DOI":
            fmt = "0.0"
        elif header == "납기일":
            fmt = "yyyy-mm-dd"
        row.append(make_cell(ws, value, font=Font(name="맑은 고딕", size=10, color=TEXT), number_format=fmt))
    ws.append(row)


def formula_status(row_idx, doi_status_col, order_col, product_shortage_col, injection_shortage_col, process_stock_col, doi_col):
    return (
        f'=IF({doi_status_col}{row_idx}="신제품","",'
        f'IF({injection_shortage_col}{row_idx}>0,"사출부족",'
        f'IF({product_shortage_col}{row_idx}>0,"제품부족",'
        f'IF(AND({order_col}{row_idx}=0,{process_stock_col}{row_idx}>0),"수요없음재고",'
        f'IF(AND(ISNUMBER({doi_col}{row_idx}),{doi_col}{row_idx}>0,{doi_col}{row_idx}<7),"DOI주의",'
        f'IF({order_col}{row_idx}>0,"수요있음","수요없음"))))))'
    )


def build_auto_workbook(rows, customers, customer_rows):
    headers = [
        "거래처그룹",
        "제품분류",
        "생산코드",
        "분리코드",
        "사출코드",
        "제품명",
        "파워",
        "오더수량1",
        "오더수량2",
        "오더합계",
        "제품부족수량",
        "사출부족수량",
        "사출재고",
        "분리재고",
        "검사접착재고",
        "누수규격검사재고",
        "공정재고합계",
        "완제품재고",
        "재고변화",
        "DOI기준오더",
        "DOI",
        "DOI상태",
        "신제품여부",
        "상태",
        "납기일",
        "이니셜",
        "재고신호",
        "기존상태",
        "비고",
    ]
    wb = Workbook(write_only=True)

    guide = wb.create_sheet("안내")
    append_title(guide, "전체 P코드 자동조회 거래처별 양식", "신제품은 DOI상태만 표시하고 상태는 빈칸 처리합니다.", 4)
    guide.append([])
    styled_header(guide, ["구분", "내용", "건수", "비고"])
    guide_rows = [
        ["전체 P코드", "전체코드_입력 기준", len(rows), ""],
        ["거래처 시트", "거래처그룹별 FILTER 조회", len(customers), ""],
        ["신제품 기준", ", ".join(NEW_PRODUCT_PATTERNS), sum(1 for row in rows if row["신제품여부"] == "Y"), "상태 빈칸"],
    ]
    for row in guide_rows:
        guide.append([make_cell(guide, value, border=True, font=Font(name="맑은 고딕", size=10)) for value in row])
    set_widths(guide, [24, 80, 15, 20])

    input_ws = wb.create_sheet("전체코드_입력")
    append_title(input_ws, "전체코드 입력", "거래처별 시트는 이 시트의 값을 자동 조회합니다.", len(headers))
    input_ws.append([])
    input_ws.append([])
    styled_header(input_ws, headers)
    for idx, row in enumerate(rows, start=FIRST_DATA_ROW):
        values = [
            row.get("거래처그룹", ""),
            row.get("제품분류", ""),
            row.get("생산코드", ""),
            row.get("분리코드", ""),
            row.get("사출코드", ""),
            row.get("제품명", ""),
            row.get("파워", ""),
            numeric(row.get("오더수량1")),
            numeric(row.get("오더수량2")),
            f"=SUM(H{idx}:I{idx})",
            numeric(row.get("제품부족수량")),
            numeric(row.get("사출부족수량")),
            numeric(row.get("사출재고")),
            numeric(row.get("분리재고")),
            numeric(row.get("검사접착재고")),
            numeric(row.get("누수규격검사재고")),
            f"=SUM(M{idx}:P{idx})",
            numeric(row.get("완제품재고")),
            numeric(row.get("재고변화")),
            numeric(row.get("DOI기준오더")),
            numeric(row.get("DOI")),
            row.get("DOI상태", ""),
            row.get("신제품여부", ""),
            f'=IF(V{idx}="신제품","",IF(L{idx}>0,"사출부족",IF(K{idx}>0,"제품부족",IF(AND(J{idx}=0,Q{idx}>0),"수요없음재고",IF(AND(ISNUMBER(U{idx}),U{idx}>0,U{idx}<7),"DOI주의",IF(J{idx}>0,"수요있음","수요없음"))))))',
            parse_due_date(row.get("납기일")),
            row.get("이니셜", ""),
            row.get("재고신호", ""),
            row.get("기존상태", ""),
            row.get("비고", ""),
        ]
        append_data_row(input_ws, values, headers)
    input_ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(rows)}"
    input_ws.freeze_panes = "F6"
    set_widths(input_ws, [20, 13, 18, 18, 18, 34, 10, 12, 12, 13, 15, 15, 13, 13, 15, 18, 15, 15, 14, 14, 10, 13, 10, 14, 13, 16, 12, 20, 26])

    customer_headers = headers[2:]
    last_row = FIRST_DATA_ROW + len(rows) - 1
    used = {"안내", "전체코드_입력"}
    for customer in customers:
        sheet_name = safe_sheet_name(customer, used)
        ws = wb.create_sheet(sheet_name)
        append_title(ws, f"{customer} 자동조회", "전체코드_입력 시트에서 해당 거래처 행만 조회합니다.", len(customer_headers))
        ws.append([make_cell(ws, "조회 거래처", fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="맑은 고딕", bold=True)), make_cell(ws, customer)])
        ws.append([])
        styled_header(ws, customer_headers)
        ws.append([f'=FILTER(\'전체코드_입력\'!$C${FIRST_DATA_ROW}:$AC${last_row},\'전체코드_입력\'!$A${FIRST_DATA_ROW}:$A${last_row}=$B$3,"")'])
        ws.freeze_panes = "F6"
        set_widths(ws, [18, 18, 18, 34, 10, 12, 12, 13, 15, 15, 13, 13, 15, 18, 15, 15, 14, 14, 10, 13, 10, 14, 13, 16, 12, 20, 26])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(AUTO_OUTPUT)


def unique_ordered(rows, key_func):
    seen = set()
    out = []
    for row in rows:
        key = key_func(row)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def append_source_sheet(wb, name, title, subtitle, headers, matrix, widths):
    ws = wb.create_sheet(name)
    append_title(ws, title, subtitle, len(headers))
    ws.append([])
    ws.append([])
    styled_header(ws, headers)
    for row in matrix:
        append_data_row(ws, row, headers)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(matrix)}"
    ws.freeze_panes = "A6"
    set_widths(ws, widths)


def build_formula_workbook(rows, customers, customer_rows, customer_initials):
    master_headers = ["거래처그룹", "제품분류", "생산코드", "분리코드", "사출코드", "제품명", "파워", "신제품여부"]
    demand_headers = ["거래처그룹", "생산코드", "이니셜", "오더수량", "납기일", "제품부족수량", "사출부족수량"]
    wip_headers = ["생산코드", "사출재고", "분리재고", "검사접착재고", "누수규격검사재고", "공정재고 합계"]
    finished_headers = ["생산코드", "완제품재고", "재고변화", "DOI기준오더", "DOI", "DOI상태", "신제품여부", "재고비율", "재고신호", "대응판단"]

    master_rows = unique_ordered(
        rows,
        lambda row: (row.get("거래처그룹", ""), row.get("생산코드", "")),
    )
    master_matrix = [
        [
            row.get("거래처그룹", ""),
            row.get("제품분류", ""),
            row.get("생산코드", ""),
            row.get("분리코드", ""),
            row.get("사출코드", ""),
            row.get("제품명", ""),
            row.get("파워", ""),
            row.get("신제품여부", ""),
        ]
        for row in master_rows
    ]
    demand_matrix = [
        [
            row.get("거래처그룹", ""),
            row.get("생산코드", ""),
            row.get("이니셜", ""),
            numeric(row.get("오더수량1")),
            parse_due_date(row.get("납기일")),
            numeric(row.get("제품부족수량")),
            numeric(row.get("사출부족수량")),
        ]
        for row in rows
        if numeric(row.get("오더수량1")) or numeric(row.get("제품부족수량")) or numeric(row.get("사출부족수량")) or clean_text(row.get("이니셜"))
    ]
    pcode_rows = unique_ordered(rows, lambda row: row.get("생산코드", ""))
    wip_matrix = [
        [
            row.get("생산코드", ""),
            numeric(row.get("사출재고")),
            numeric(row.get("분리재고")),
            numeric(row.get("검사접착재고")),
            numeric(row.get("누수규격검사재고")),
            f"=SUM(B{FIRST_DATA_ROW + idx}:E{FIRST_DATA_ROW + idx})",
        ]
        for idx, row in enumerate(pcode_rows)
    ]
    finished_matrix = [
        [
            row.get("생산코드", ""),
            numeric(row.get("완제품재고")),
            numeric(row.get("재고변화")),
            numeric(row.get("DOI기준오더")),
            numeric(row.get("DOI")),
            row.get("DOI상태", ""),
            row.get("신제품여부", ""),
            row.get("재고비율", ""),
            row.get("재고신호", ""),
            row.get("대응판단", ""),
        ]
        for row in pcode_rows
    ]

    wb = Workbook(write_only=True)
    guide = wb.create_sheet("안내")
    append_title(guide, "전체 P코드 수식형 거래처별 양식", "신제품은 DOI상태만 표시하고 상태는 빈칸 처리합니다.", 4)
    guide.append([])
    styled_header(guide, ["구분", "내용", "건수", "비고"])
    for row in [
        ["품목마스터", "거래처별 고정 P코드 기준", len(master_matrix), ""],
        ["수요_집계", "오더/납기/부족 수식 원천", len(demand_matrix), ""],
        ["신제품 기준", ", ".join(NEW_PRODUCT_PATTERNS), sum(1 for row in rows if row["신제품여부"] == "Y"), "상태 빈칸"],
    ]:
        guide.append([make_cell(guide, value, border=True, font=Font(name="맑은 고딕", size=10)) for value in row])
    set_widths(guide, [24, 80, 15, 20])

    append_source_sheet(wb, "품목마스터", "품목마스터", "전체 P코드 고정값입니다.", master_headers, master_matrix, [20, 13, 18, 18, 18, 34, 10, 10])
    append_source_sheet(wb, "수요_집계", "수요 집계", "오더/부족/납기 원천입니다.", demand_headers, demand_matrix, [20, 18, 18, 14, 13, 15, 15])
    append_source_sheet(wb, "공정재고_집계", "공정재고 집계", "P코드별 공정재고 원천입니다.", wip_headers, wip_matrix, [18, 13, 13, 15, 18, 15])
    append_source_sheet(wb, "완제품재고_집계", "완제품재고 집계", "P코드별 DOI/신제품 상태 원천입니다.", finished_headers, finished_matrix, [18, 15, 14, 14, 10, 13, 10, 12, 12, 36])

    demand_last = HEADER_ROW + max(len(demand_matrix), 1)
    wip_last = HEADER_ROW + max(len(wip_matrix), 1)
    finished_last = HEADER_ROW + max(len(finished_matrix), 1)

    used = {"안내", "품목마스터", "수요_집계", "공정재고_집계", "완제품재고_집계"}
    base_headers = ["생산코드", "분리코드", "사출코드", "제품명", "파워"]
    trailing = [
        "오더합계",
        "납기일",
        "제품부족수량",
        "사출부족수량",
        "사출재고",
        "분리재고",
        "검사접착재고",
        "누수규격검사재고",
        "공정재고 합계",
        "완제품재고",
        "재고변화",
        "DOI기준오더",
        "DOI",
        "DOI상태",
        "신제품여부",
        "재고비율",
        "재고신호",
        "대응판단",
        "상태",
    ]
    for customer in customers:
        rows_for_customer = customer_rows[customer]
        initials = sorted(customer_initials.get(customer, []))
        order_headers = [f"오더수량({initial})" for initial in initials]
        headers = [*base_headers, *order_headers, *trailing]
        ws = wb.create_sheet(safe_sheet_name(customer, used))
        append_title(ws, f"{customer} 품목 현황", "품목마스터와 집계 시트를 조회하는 수식형 양식입니다.", len(headers))
        ws.append([make_cell(ws, "조회 거래처", fill=PatternFill("solid", fgColor=LIGHT_BLUE), font=Font(name="맑은 고딕", bold=True)), make_cell(ws, customer)])
        ws.append([])
        styled_header(ws, headers)
        for idx, master_row in enumerate(rows_for_customer, start=FIRST_DATA_ROW):
            pcode = master_row.get("생산코드", "")
            row_values = [pcode, master_row.get("분리코드", ""), master_row.get("사출코드", ""), master_row.get("제품명", ""), master_row.get("파워", "")]
            for initial in initials:
                row_values.append(
                    f'=SUMIFS(\'수요_집계\'!$D${FIRST_DATA_ROW}:$D${demand_last},\'수요_집계\'!$A${FIRST_DATA_ROW}:$A${demand_last},$B$3,\'수요_집계\'!$B${FIRST_DATA_ROW}:$B${demand_last},$A{idx},\'수요_집계\'!$C${FIRST_DATA_ROW}:$C${demand_last},"{initial}")'
                )
            order_start = 6
            order_end = 5 + len(initials)
            if initials:
                order_total = f"=SUM({get_column_letter(order_start)}{idx}:{get_column_letter(order_end)}{idx})"
            else:
                order_total = f'=SUMIFS(\'수요_집계\'!$D${FIRST_DATA_ROW}:$D${demand_last},\'수요_집계\'!$A${FIRST_DATA_ROW}:$A${demand_last},$B$3,\'수요_집계\'!$B${FIRST_DATA_ROW}:$B${demand_last},$A{idx})'
            row_values.extend(
                [
                    order_total,
                    f'=IFERROR(AGGREGATE(15,6,\'수요_집계\'!$E${FIRST_DATA_ROW}:$E${demand_last}/((\'수요_집계\'!$A${FIRST_DATA_ROW}:$A${demand_last}=$B$3)*(\'수요_집계\'!$B${FIRST_DATA_ROW}:$B${demand_last}=$A{idx})*(\'수요_집계\'!$E${FIRST_DATA_ROW}:$E${demand_last}>0)),1),"")',
                    f'=SUMIFS(\'수요_집계\'!$F${FIRST_DATA_ROW}:$F${demand_last},\'수요_집계\'!$A${FIRST_DATA_ROW}:$A${demand_last},$B$3,\'수요_집계\'!$B${FIRST_DATA_ROW}:$B${demand_last},$A{idx})',
                    f'=SUMIFS(\'수요_집계\'!$G${FIRST_DATA_ROW}:$G${demand_last},\'수요_집계\'!$A${FIRST_DATA_ROW}:$A${demand_last},$B$3,\'수요_집계\'!$B${FIRST_DATA_ROW}:$B${demand_last},$A{idx})',
                ]
            )
            for source_col in ["B", "C", "D", "E", "F"]:
                row_values.append(f'=IFERROR(INDEX(\'공정재고_집계\'!${source_col}${FIRST_DATA_ROW}:${source_col}${wip_last},MATCH($A{idx},\'공정재고_집계\'!$A${FIRST_DATA_ROW}:$A${wip_last},0)),0)')
            for source_col, fallback in [("B", "0"), ("C", "0"), ("D", "0"), ("E", "0"), ("F", '""'), ("G", '""'), ("H", '""'), ("I", '""'), ("J", '""')]:
                row_values.append(f'=IFERROR(INDEX(\'완제품재고_집계\'!${source_col}${FIRST_DATA_ROW}:${source_col}${finished_last},MATCH($A{idx},\'완제품재고_집계\'!$A${FIRST_DATA_ROW}:$A${finished_last},0)),{fallback})')
            columns = {header: get_column_letter(pos + 1) for pos, header in enumerate(headers)}
            row_values.append(
                formula_status(
                    idx,
                    columns["DOI상태"],
                    columns["오더합계"],
                    columns["제품부족수량"],
                    columns["사출부족수량"],
                    columns["공정재고 합계"],
                    columns["DOI"],
                )
            )
            append_data_row(ws, row_values, headers)
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{HEADER_ROW + len(rows_for_customer)}"
        ws.freeze_panes = "F6"
        set_widths(ws, [18, 18, 18, 34, 10] + [15] * len(order_headers) + [13, 13, 15, 15, 13, 13, 15, 18, 15, 15, 14, 14, 10, 13, 10, 12, 12, 32, 14])

    wb.save(FORMULA_OUTPUT)


def main():
    payload = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    rows = enrich_rows(payload["rows"])
    customers = sorted({clean_text(row.get("거래처그룹")) or "거래처 미지정" for row in rows})
    customer_rows = {customer: [] for customer in customers}
    customer_initials = {customer: set() for customer in customers}
    for row in rows:
        customer = clean_text(row.get("거래처그룹")) or "거래처 미지정"
        customer_rows[customer].append(row)
        initial = clean_text(row.get("이니셜"))
        if initial and numeric(row.get("오더수량1")):
            customer_initials[customer].add(initial)

    build_auto_workbook(rows, customers, customer_rows)
    build_formula_workbook(rows, customers, customer_rows, customer_initials)

    for output in [AUTO_OUTPUT, FORMULA_OUTPUT]:
        wb = load_workbook(output, read_only=True, data_only=False)
        print(
            json.dumps(
                {
                    "output": str(output),
                    "sheets": len(wb.sheetnames),
                    "first_sheets": wb.sheetnames[:8],
                },
                ensure_ascii=False,
            )
        )
        wb.close()


if __name__ == "__main__":
    main()
