# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "outputs" / "c_site_input_template_20260805"
OUTPUT_PATH = OUTPUT_DIR / "C관_고정품목_입력형_자동확인_양식.xlsx"

MASTER_SOURCE = BASE_DIR / ".codex_tmp" / "c_site_pcode_power_doi_criteria.csv"
DEMAND_SOURCE = BASE_DIR / "cloud_snapshots" / "shortage_snapshot.csv.gz"
WIP_SOURCE = BASE_DIR / "ODV_WIP_20260506.xlsx"

MASTER_SHEET = "품목마스터_C관"
DEMAND_SHEET = "수요정보_입력"
SAMPLE_SHEET = "샘플가능수량_입력"
WIP_SHEET = "WIP_입력"
RESULT_SHEET = "C관_자동확인"
SUMMARY_SHEET = "요약"

MAX_DEMAND_ROWS = 10000
MAX_WIP_ROWS = 40000
MAX_SAMPLE_ROWS = 65000
HEADER_ROW = 1
DATA_ROW = 2

NAVY = "172554"
BLUE = "2563EB"
LIGHT_BLUE = "DBEAFE"
GRAY = "F8FAFC"
LIGHT_GRAY = "F1F5F9"
RED = "DC2626"
LIGHT_RED = "FEE2E2"
YELLOW = "FEF3C7"
GREEN = "DCFCE7"
BORDER = "CBD5E1"
TEXT = "0F172A"
MUTED = "64748B"

thin_border = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "nat", "null", "<na>"} else text


def numeric(value: object) -> float:
    return float(pd.to_numeric(value, errors="coerce") if not pd.isna(pd.to_numeric(value, errors="coerce")) else 0)


def normalize_code(value: object) -> str:
    return clean_text(value).replace(" ", "").upper()


def style_sheet_base(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def write_header(ws, headers: list[str], row: int = HEADER_ROW):
    fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = Font(name="맑은 고딕", bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row].height = 24
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def apply_body_style(ws, min_row: int, max_row: int, max_col: int):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10, color=TEXT)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")


def set_widths(ws, widths: list[float]):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def master_ranges(master_len: int) -> dict[str, str]:
    last = master_len + 1
    return {
        "p": f"'{MASTER_SHEET}'!$A$2:$A${last}",
        "r": f"'{MASTER_SHEET}'!$B$2:$B${last}",
        "q": f"'{MASTER_SHEET}'!$C$2:$C${last}",
    }


def build_master() -> pd.DataFrame:
    master = pd.read_csv(MASTER_SOURCE, low_memory=False)
    cols = [
        "생산코드",
        "사출코드",
        "분리코드",
        "제품명",
        "파워",
        "거래처그룹",
        "거래처",
        "이니셜목록",
        "기준등급",
        "기준DOI하한",
        "기준DOI상한",
        "신제품여부",
        "DOI상태",
        "완제품재고",
        "현재DOI",
    ]
    for col in cols:
        if col not in master.columns:
            master[col] = ""
    master = master[cols].copy()
    for col in ["생산코드", "사출코드", "분리코드"]:
        master[col] = master[col].map(normalize_code)
    for col in ["제품명", "거래처그룹", "거래처", "이니셜목록", "기준등급", "신제품여부", "DOI상태"]:
        master[col] = master[col].map(clean_text)
    for col in ["기준DOI하한", "기준DOI상한", "완제품재고", "현재DOI"]:
        master[col] = pd.to_numeric(master[col], errors="coerce").fillna(0)
    master = master.drop_duplicates(subset=["생산코드"], keep="first")
    return master.sort_values(["거래처그룹", "제품명", "파워", "생산코드"]).reset_index(drop=True)


def build_demand_prefill() -> pd.DataFrame:
    if not DEMAND_SOURCE.exists():
        return pd.DataFrame()
    demand = pd.read_csv(DEMAND_SOURCE, compression="gzip", low_memory=False)
    if demand.empty:
        return demand
    site_col = demand.columns[0]
    demand = demand[demand[site_col].fillna("").astype(str).str.contains("C관", na=False)].copy()
    return demand


def build_wip_prefill() -> pd.DataFrame:
    if not WIP_SOURCE.exists():
        return pd.DataFrame()
    return pd.read_excel(WIP_SOURCE, sheet_name="WIP")


def build_sample_prefill() -> pd.DataFrame:
    if not WIP_SOURCE.exists():
        return pd.DataFrame(columns=["품목코드", "품명", "샘플 신청 가능 수량"])
    sample = pd.read_excel(WIP_SOURCE, sheet_name="Sheet1")
    if sample.empty:
        return pd.DataFrame(columns=["품목코드", "품명", "샘플 신청 가능 수량"])
    code = sample.iloc[:, 2].map(normalize_code)
    name = sample.iloc[:, 3].map(clean_text)
    qty = pd.to_numeric(sample.iloc[:, 8], errors="coerce").fillna(0)
    sample_df = pd.DataFrame({"품목코드": code, "품명": name, "샘플 신청 가능 수량": qty})
    sample_df = sample_df[sample_df["품목코드"].str.startswith("P", na=False)].copy()
    sample_df = (
        sample_df.groupby("품목코드", as_index=False)
        .agg({"품명": "first", "샘플 신청 가능 수량": "sum"})
        .sort_values("품목코드")
    )
    return sample_df


def code_match_formula(row: int, code_cell: str, r_cell: str | None, q_cell: str | None, ranges: dict[str, str]) -> str:
    p_rng, r_rng, q_rng = ranges["p"], ranges["r"], ranges["q"]
    if r_cell and q_cell:
        return (
            f'=IF(COUNTA({code_cell}{row},{r_cell}{row},{q_cell}{row})=0,"",'
            f'IFERROR(XLOOKUP({code_cell}{row},{p_rng},{p_rng}),'
            f'IFERROR(XLOOKUP({code_cell}{row},{r_rng},{p_rng}),'
            f'IFERROR(XLOOKUP({code_cell}{row},{q_rng},{p_rng}),'
            f'IFERROR(XLOOKUP({r_cell}{row},{r_rng},{p_rng}),'
            f'IFERROR(XLOOKUP({q_cell}{row},{q_rng},{p_rng}),"코드미매칭"))))))'
        )
    return (
        f'=IF({code_cell}{row}="","",'
        f'IFERROR(XLOOKUP({code_cell}{row},{p_rng},{p_rng}),'
        f'IFERROR(XLOOKUP({code_cell}{row},{r_rng},{p_rng}),'
        f'IFERROR(XLOOKUP({code_cell}{row},{q_rng},{p_rng}),"코드미매칭"))))'
    )


def create_guide_sheet(wb: Workbook, master_len: int, demand_len: int, wip_len: int, sample_len: int):
    ws = wb.create_sheet("안내")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "C관 고정품목 입력형 자동확인 양식"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")
    ws["A2"] = "수요정보, WIP, 샘플 가능수량만 갱신하면 C관 품목별 오더/부족/공정재고/샘플 가능수량/상태를 자동 확인합니다."
    ws["A2"].font = Font(name="맑은 고딕", color=MUTED)
    ws.merge_cells("A2:F2")

    rows = [
        ["시트", "역할", "사용 방법", "현재 포함 행수"],
        [MASTER_SHEET, "고정 기준", "C관 대상 생산코드/사출코드/분리코드/제품명/DOI 기준. 직접 수정하지 않는 기준표입니다.", master_len],
        [DEMAND_SHEET, "수요 입력", "APS 수요정보를 B열부터 붙여넣습니다. A열 생산코드_매칭은 수식이므로 지우지 않습니다.", demand_len],
        [WIP_SHEET, "WIP 입력", "ODV WIP의 WIP 시트를 B열부터 붙여넣습니다. 제품 코드, WH_NAME, 총 재공 수량을 사용합니다.", wip_len],
        [SAMPLE_SHEET, "샘플 입력", "LOT 원본은 무거우므로 품목코드별 샘플 신청 가능 수량 합계 형태로 입력합니다.", sample_len],
        [RESULT_SHEET, "자동 확인", "품목코드는 고정값이고 수요/WIP/샘플 값은 입력 시트를 조회합니다. 이 시트만 보면 됩니다.", master_len],
    ]
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = Font(name="맑은 고딕", bold=r_idx == 4, color=TEXT)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_idx == 4 else "FFFFFF")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    set_widths(ws, [24, 18, 95, 16, 14, 14])
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 26


def create_master_sheet(wb: Workbook, master: pd.DataFrame):
    ws = wb.create_sheet(MASTER_SHEET)
    style_sheet_base(ws)
    headers = master.columns.tolist()
    write_header(ws, headers)
    for row in master.itertuples(index=False):
        ws.append(list(row))
    apply_body_style(ws, DATA_ROW, len(master) + 1, len(headers))
    set_widths(ws, [18, 18, 18, 42, 10, 16, 22, 28, 10, 12, 12, 10, 12, 14, 12])
    for col_idx in [10, 11, 14]:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(master) + 1):
            for c in cell:
                c.number_format = "#,##0"
    for cell in ws.iter_cols(min_col=15, max_col=15, min_row=2, max_row=len(master) + 1):
        for c in cell:
            c.number_format = "0.0"


def create_demand_sheet(wb: Workbook, demand: pd.DataFrame, ranges: dict[str, str]):
    ws = wb.create_sheet(DEMAND_SHEET)
    style_sheet_base(ws)
    headers = ["생산코드_매칭", *demand.columns.tolist()]
    write_header(ws, headers)
    for values in demand.itertuples(index=False):
        ws.append(["", *list(values)])
    for row_idx in range(2, MAX_DEMAND_ROWS + 2):
        ws.cell(row_idx, 1).value = code_match_formula(row_idx, "F", "S", "T", ranges)
    apply_body_style(ws, 2, max(MAX_DEMAND_ROWS + 1, len(demand) + 1), len(headers))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["R"].width = 32
    ws.column_dimensions["S"].width = 18
    ws.column_dimensions["T"].width = 18
    for col in ["L", "M", "N", "O", "P", "Q"]:
        for cell in ws[col][1:MAX_DEMAND_ROWS + 1]:
            cell.number_format = "#,##0"
    for cell in ws["A"]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)


def create_wip_sheet(wb: Workbook, wip: pd.DataFrame, ranges: dict[str, str]):
    ws = wb.create_sheet(WIP_SHEET)
    style_sheet_base(ws)
    headers = ["생산코드_매칭", *wip.columns.tolist()]
    write_header(ws, headers)
    for values in wip.itertuples(index=False):
        ws.append(["", *list(values)])
    for row_idx in range(2, MAX_WIP_ROWS + 2):
        ws.cell(row_idx, 1).value = code_match_formula(row_idx, "J", None, None, ranges)
    apply_body_style(ws, 2, max(MAX_WIP_ROWS + 1, len(wip) + 1), len(headers))
    set_widths(ws, [18, 28, 20, 18, 18, 14, 14, 14, 14, 18, 14, 14, 14, 12, 12, 14, 14, 12, 12, 18, 12, 18, 12, 16, 16])
    for cell in ws["A"]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    for cell in ws["H"][1:MAX_WIP_ROWS + 1]:
        cell.number_format = "#,##0"


def create_sample_sheet(wb: Workbook, sample: pd.DataFrame, ranges: dict[str, str]):
    ws = wb.create_sheet(SAMPLE_SHEET)
    style_sheet_base(ws)
    headers = ["생산코드_매칭", "품목코드", "품명", "샘플 신청 가능 수량"]
    write_header(ws, headers)
    for row in sample.itertuples(index=False):
        ws.append(["", row[0], row[1], row[2]])
    for row_idx in range(2, MAX_SAMPLE_ROWS + 2):
        ws.cell(row_idx, 1).value = code_match_formula(row_idx, "B", None, None, ranges)
    apply_body_style(ws, 2, max(MAX_SAMPLE_ROWS + 1, len(sample) + 1), len(headers))
    set_widths(ws, [18, 20, 36, 18])
    for cell in ws["A"]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    for cell in ws["D"][1:MAX_SAMPLE_ROWS + 1]:
        cell.number_format = "#,##0"


def create_result_sheet(wb: Workbook, master: pd.DataFrame):
    ws = wb.create_sheet(RESULT_SHEET)
    style_sheet_base(ws)
    headers = [
        "생산코드",
        "사출코드",
        "분리코드",
        "제품명",
        "파워",
        "거래처그룹",
        "기준등급",
        "DOI하한",
        "DOI상한",
        "신제품여부",
        "완제품재고",
        "현재DOI",
        "DOI상태",
        "오더수량",
        "납기일",
        "부족수량",
        "사출부족수량",
        "샘플출고가능수량",
        "사출재고",
        "분리재고",
        "검사접착재고",
        "누수규격검사재고",
        "공정재고합계",
        "부족대비공정재고율",
        "상태",
        "확인메모",
    ]
    write_header(ws, headers)
    demand_last = MAX_DEMAND_ROWS + 1
    wip_last = MAX_WIP_ROWS + 1
    sample_last = MAX_SAMPLE_ROWS + 1
    for idx, row in enumerate(master.itertuples(index=False), start=2):
        # master columns: 생산코드, 사출코드, 분리코드, 제품명, 파워, 거래처그룹, 거래처, 이니셜목록,
        # 기준등급, 기준DOI하한, 기준DOI상한, 신제품여부, DOI상태, 완제품재고, 현재DOI
        static = [
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[8],
            row[9],
            row[10],
            row[11],
            row[13],
            row[14],
            row[12],
        ]
        for col, value in enumerate(static, start=1):
            ws.cell(idx, col, value)
        ws.cell(idx, 14, f"=SUMIFS('{DEMAND_SHEET}'!$L$2:$L${demand_last},'{DEMAND_SHEET}'!$A$2:$A${demand_last},$A{idx})")
        ws.cell(
            idx,
            15,
            f'=IF(COUNTIFS(\'{DEMAND_SHEET}\'!$A$2:$A${demand_last},$A{idx})=0,"-",IFERROR(TEXT(MINIFS(\'{DEMAND_SHEET}\'!$G$2:$G${demand_last},\'{DEMAND_SHEET}\'!$A$2:$A${demand_last},$A{idx}),"yyyy-mm-dd"),"-"))',
        )
        ws.cell(idx, 16, f"=SUMIFS('{DEMAND_SHEET}'!$M$2:$M${demand_last},'{DEMAND_SHEET}'!$A$2:$A${demand_last},$A{idx})")
        ws.cell(idx, 17, f"=SUMIFS('{DEMAND_SHEET}'!$N$2:$N${demand_last},'{DEMAND_SHEET}'!$A$2:$A${demand_last},$A{idx})")
        ws.cell(idx, 18, f"=SUMIFS('{SAMPLE_SHEET}'!$D$2:$D${sample_last},'{SAMPLE_SHEET}'!$A$2:$A${sample_last},$A{idx})")
        ws.cell(idx, 19, f'=SUMIFS(\'{WIP_SHEET}\'!$H$2:$H${wip_last},\'{WIP_SHEET}\'!$A$2:$A${wip_last},$A{idx},\'{WIP_SHEET}\'!$Y$2:$Y${wip_last},"*사출*")')
        ws.cell(idx, 20, f'=SUMIFS(\'{WIP_SHEET}\'!$H$2:$H${wip_last},\'{WIP_SHEET}\'!$A$2:$A${wip_last},$A{idx},\'{WIP_SHEET}\'!$Y$2:$Y${wip_last},"*분리*")')
        ws.cell(idx, 21, f'=SUMIFS(\'{WIP_SHEET}\'!$H$2:$H${wip_last},\'{WIP_SHEET}\'!$A$2:$A${wip_last},$A{idx},\'{WIP_SHEET}\'!$Y$2:$Y${wip_last},"*검사접착*")+SUMIFS(\'{WIP_SHEET}\'!$H$2:$H${wip_last},\'{WIP_SHEET}\'!$A$2:$A${wip_last},$A{idx},\'{WIP_SHEET}\'!$Y$2:$Y${wip_last},"*검사/접착*")')
        ws.cell(idx, 22, f'=SUMIFS(\'{WIP_SHEET}\'!$H$2:$H${wip_last},\'{WIP_SHEET}\'!$A$2:$A${wip_last},$A{idx},\'{WIP_SHEET}\'!$Y$2:$Y${wip_last},"*누수*")')
        ws.cell(idx, 23, f"=SUM(S{idx}:V{idx})")
        ws.cell(idx, 24, f'=IF(P{idx}>0,W{idx}/P{idx},"")')
        ws.cell(
            idx,
            25,
            f'=IF(J{idx}="Y","",IF(Q{idx}>0,"사출부족",IF(P{idx}>0,"제품부족",IF(N{idx}>0,"수요있음",IF(W{idx}>0,"수요없음+공정재고",IF(R{idx}>0,"수요없음+샘플가능","수요없음"))))))',
        )
        ws.cell(
            idx,
            26,
            f'=IF(J{idx}="Y","신제품: 판단 보류",IF(Q{idx}>0,"사출 생산 필요",IF(P{idx}>0,IF(W{idx}+R{idx}>=P{idx},"부족 있음: 공정/샘플 대체 가능성 확인","제품 생산 필요"),IF(N{idx}=0,IF(W{idx}+R{idx}>0,"수요 없음 재고/샘플 확인",""),""))))',
        )

    last = len(master) + 1
    apply_body_style(ws, 2, last, len(headers))
    ws.freeze_panes = "F2"
    set_widths(ws, [18, 18, 18, 42, 10, 16, 10, 10, 10, 10, 14, 10, 12, 14, 13, 14, 15, 17, 13, 13, 15, 17, 15, 16, 18, 38])
    for col in range(8, 25):
        letter = get_column_letter(col)
        for cell in ws[letter][1:last]:
            if col == 24:
                cell.number_format = "0.0%"
            elif col == 12:
                cell.number_format = "0.0"
            elif col == 15:
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.number_format = "#,##0"
    status_range = f"Y2:Y{last}"
    ws.conditional_formatting.add(status_range, FormulaRule(formula=['Y2="사출부족"'], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=['Y2="제품부족"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=TEXT, bold=True)))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=['Y2="수요없음+공정재고"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=TEXT)))


def create_summary_sheet(wb: Workbook, master_len: int):
    ws = wb.create_sheet(SUMMARY_SHEET)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "C관 자동확인 요약"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")
    last = master_len + 1
    kpis = [
        ["지표", "값"],
        ["고정 품목수", f"=COUNTA('{RESULT_SHEET}'!$A$2:$A${last})"],
        ["오더수량", f"=SUM('{RESULT_SHEET}'!$N$2:$N${last})"],
        ["부족수량", f"=SUM('{RESULT_SHEET}'!$P$2:$P${last})"],
        ["사출부족수량", f"=SUM('{RESULT_SHEET}'!$Q$2:$Q${last})"],
        ["공정재고합계", f"=SUM('{RESULT_SHEET}'!$W$2:$W${last})"],
        ["샘플출고가능수량", f"=SUM('{RESULT_SHEET}'!$R$2:$R${last})"],
    ]
    for r_idx, row in enumerate(kpis, start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = Font(name="맑은 고딕", bold=r_idx == 3, color=TEXT)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_idx == 3 else "FFFFFF")
            cell.border = thin_border
            if c_idx == 2 and r_idx > 3:
                cell.number_format = "#,##0"
    status_rows = [
        ["상태", "품목수", "부족수량", "사출부족수량"],
        ["사출부족", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A12)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A12,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A12,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["제품부족", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A13)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A13,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A13,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["수요있음", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A14)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A14,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A14,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["수요없음+공정재고", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A15)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A15,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A15,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["수요없음+샘플가능", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A16)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A16,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A16,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["수요없음", f'=COUNTIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A17)', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A17,\'{RESULT_SHEET}\'!$P$2:$P${last})', f'=SUMIF(\'{RESULT_SHEET}\'!$Y$2:$Y${last},A17,\'{RESULT_SHEET}\'!$Q$2:$Q${last})'],
        ["신제품 판단보류", f'=COUNTIF(\'{RESULT_SHEET}\'!$J$2:$J${last},"Y")', "", ""],
    ]
    for r_idx, row in enumerate(status_rows, start=11):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = Font(name="맑은 고딕", bold=r_idx == 11, color=TEXT)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if r_idx == 11 else "FFFFFF")
            cell.border = thin_border
            if c_idx > 1 and r_idx > 11:
                cell.number_format = "#,##0"
    set_widths(ws, [24, 18, 18, 18, 16, 16])


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    master = build_master()
    demand = build_demand_prefill()
    wip = build_wip_prefill()
    sample = build_sample_prefill()

    wb = Workbook()
    wb.remove(wb.active)
    ranges = master_ranges(len(master))
    create_guide_sheet(wb, len(master), len(demand), len(wip), len(sample))
    create_master_sheet(wb, master)
    create_demand_sheet(wb, demand, ranges)
    create_sample_sheet(wb, sample, ranges)
    create_wip_sheet(wb, wip, ranges)
    create_result_sheet(wb, master)
    create_summary_sheet(wb, len(master))
    wb.save(OUTPUT_PATH)

    check_wb = load_workbook(OUTPUT_PATH, read_only=True, data_only=False)
    result = check_wb[RESULT_SHEET]
    print(
        {
            "output": str(OUTPUT_PATH),
            "sheets": check_wb.sheetnames,
            "master_rows": len(master),
            "demand_prefill_rows": len(demand),
            "sample_prefill_rows": len(sample),
            "wip_prefill_rows": len(wip),
            "result_a2": result["A2"].value,
            "result_n2_formula": result["N2"].value,
            "result_y2_formula": result["Y2"].value,
        }
    )
    check_wb.close()


if __name__ == "__main__":
    main()
