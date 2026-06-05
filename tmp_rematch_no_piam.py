from __future__ import annotations

import difflib
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd


POWER_PATTERN = re.compile(r"([+-]\d{1,2}(?:\.\d{1,2})?)")
MATCH_COLUMNS = [
    "Sheet1매칭여부",
    "Sheet1제품코드",
    "Sheet1제품이름",
    "Sheet1파워",
    "Sheet1재공건수",
    "Sheet1기초재고합계",
    "Sheet1사용재고합계",
    "Sheet1샘플가능수량합계",
    "매칭근거",
    "공통토큰",
]
STOP_TOKENS = set(
    "pia kr uv 1d 1day day lens contact clear color colorlens "
    "10p 30p 30pack pack 중사각 원형 소 중 대 m d a r q y new".split()
)


def format_power(value: object) -> str:
    text = str(value).strip()
    match = POWER_PATTERN.search(text)
    if match:
        number = float(match.group(1))
    elif re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        number = float(text)
    else:
        return ""
    return ("+" if number > 0 else "-") + f"{abs(number):05.2f}"


def code_suffix(value: object) -> str:
    text = str(value).upper()
    match = POWER_PATTERN.search(text)
    if not match:
        return ""
    suffix = text[match.end() :]
    suffix = re.sub(r"[^A-Z가-힣0-9]+", "", suffix)
    return re.sub(r"\d+$", "", suffix)


def tokens(value: object) -> list[str]:
    raw = re.findall(r"[A-Za-z가-힣0-9]+", str(value).lower())
    return [token for token in raw if len(token) >= 3 and token not in STOP_TOKENS and not token.isdigit()]


def name_score(left: object, right: object) -> tuple[int, float, str]:
    left_tokens = set(tokens(left))
    right_tokens = set(tokens(right))
    common = left_tokens & right_tokens
    seq = difflib.SequenceMatcher(None, " ".join(tokens(left)), " ".join(tokens(right))).ratio()
    return len(common), seq, ",".join(sorted(common))


def main() -> None:
    source = Path(sys.argv[1])
    output = source.with_name(source.stem + "_no_piam.xlsx")

    production = pd.read_excel(source, sheet_name="생산현황").drop(columns=MATCH_COLUMNS, errors="ignore")
    sheet1_all = pd.read_excel(source, sheet_name="Sheet1")
    exclude_mask = sheet1_all["제품이름"].astype(str).str.contains("PIA_M", case=False, na=False)
    sheet1 = sheet1_all[~exclude_mask].copy()

    sheet1["Sheet1파워"] = sheet1["제품코드"].map(format_power)
    sheet1["Sheet1suffix"] = sheet1["제품코드"].map(code_suffix)
    for column in ["기초재고", "사용재고", "샘플 신청 가능 수량"]:
        sheet1[column] = pd.to_numeric(sheet1[column], errors="coerce").fillna(0)

    sheet1_summary = (
        sheet1.groupby(["제품코드", "제품이름", "Sheet1파워", "Sheet1suffix"], as_index=False)
        .agg(
            {
                "재공코드": "count",
                "기초재고": "sum",
                "사용재고": "sum",
                "샘플 신청 가능 수량": "sum",
            }
        )
        .rename(
            columns={
                "재공코드": "Sheet1재공건수",
                "기초재고": "Sheet1기초재고합계",
                "사용재고": "Sheet1사용재고합계",
                "샘플 신청 가능 수량": "Sheet1샘플가능수량합계",
            }
        )
    )

    production["생산파워정규"] = production["파워"].map(format_power)
    production["생산suffix"] = production["품목코드"].map(code_suffix)

    results: list[dict[str, object]] = []
    for _, production_row in production.iterrows():
        candidates = sheet1_summary[sheet1_summary["Sheet1파워"].eq(production_row["생산파워정규"])].copy()
        if production_row["생산suffix"]:
            candidates = candidates[candidates["Sheet1suffix"].eq(production_row["생산suffix"])].copy()
            reason = "Sheet1 PIA_M 제외 후 파워+제품코드 suffix 일치"
            min_common = 0
        else:
            reason = "Sheet1 PIA_M 제외 후 파워+제품명 토큰 보조"
            min_common = 2

        best: tuple[tuple[float, float, float], pd.Series, str] | None = None
        for _, candidate in candidates.iterrows():
            common_count, seq_score, common_tokens = name_score(production_row["제품명"], candidate["제품이름"])
            valid = common_count >= min_common or (not production_row["생산suffix"] and seq_score >= 0.60)
            rank = (
                float(common_count),
                float(seq_score),
                float(candidate["Sheet1샘플가능수량합계"] or 0),
            )
            if valid and (best is None or rank > best[0]):
                best = (rank, candidate, common_tokens)

        if best is None:
            results.append(
                {
                    "Sheet1매칭여부": "N",
                    "Sheet1제품코드": "",
                    "Sheet1제품이름": "",
                    "Sheet1파워": "",
                    "Sheet1재공건수": 0,
                    "Sheet1기초재고합계": 0,
                    "Sheet1사용재고합계": 0,
                    "Sheet1샘플가능수량합계": 0,
                    "매칭근거": "PIA_M 제외 후 매칭 후보 없음",
                    "공통토큰": "",
                }
            )
            continue

        candidate = best[1]
        results.append(
            {
                "Sheet1매칭여부": "Y",
                "Sheet1제품코드": candidate["제품코드"],
                "Sheet1제품이름": candidate["제품이름"],
                "Sheet1파워": candidate["Sheet1파워"],
                "Sheet1재공건수": int(candidate["Sheet1재공건수"]),
                "Sheet1기초재고합계": float(candidate["Sheet1기초재고합계"]),
                "Sheet1사용재고합계": float(candidate["Sheet1사용재고합계"]),
                "Sheet1샘플가능수량합계": float(candidate["Sheet1샘플가능수량합계"]),
                "매칭근거": reason,
                "공통토큰": best[2],
            }
        )

    match_df = pd.DataFrame(results)
    production_output = production.drop(columns=["생산파워정규", "생산suffix"])
    unmatched_df = pd.concat([production_output, match_df], axis=1)
    unmatched_df = unmatched_df[unmatched_df["Sheet1매칭여부"].eq("N")]
    summary_df = pd.DataFrame(
        [
            {"항목": "생산현황 행수", "값": len(production_output)},
            {"항목": "Sheet1 전체 행수", "값": len(sheet1_all)},
            {"항목": "Sheet1 PIA_M 제외 행수", "값": int(exclude_mask.sum())},
            {"항목": "Sheet1 매칭 후보 행수", "값": len(sheet1)},
            {"항목": "매칭 행수", "값": int(match_df["Sheet1매칭여부"].eq("Y").sum())},
            {"항목": "미매칭 행수", "값": int(match_df["Sheet1매칭여부"].eq("N").sum())},
            {
                "항목": "매칭 기준",
                "값": "Sheet1 제품이름에 PIA_M 포함 행 제외 후, 생산현황 제품명+파워와 Sheet1 제품이름+제품코드 파워/suffix 매칭",
            },
        ]
    )

    workbook = openpyxl.load_workbook(source)
    production_sheet = workbook["생산현황"]

    for column_idx in range(production_sheet.max_column, 0, -1):
        if production_sheet.cell(row=1, column=column_idx).value in MATCH_COLUMNS:
            production_sheet.delete_cols(column_idx)

    start_column = production_sheet.max_column + 1
    for offset, column_name in enumerate(match_df.columns):
        production_sheet.cell(row=1, column=start_column + offset, value=column_name)
    for row_idx, values in enumerate(match_df.itertuples(index=False, name=None), start=2):
        for offset, value in enumerate(values):
            production_sheet.cell(row=row_idx, column=start_column + offset, value=value)

    for sheet_name in ["매칭요약", "미매칭목록"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    summary_sheet = workbook.create_sheet("매칭요약")
    for column_idx, column_name in enumerate(summary_df.columns, start=1):
        summary_sheet.cell(row=1, column=column_idx, value=column_name)
    for row_idx, row in enumerate(summary_df.itertuples(index=False, name=None), start=2):
        for column_idx, value in enumerate(row, start=1):
            summary_sheet.cell(row=row_idx, column=column_idx, value=value)

    unmatched_sheet = workbook.create_sheet("미매칭목록")
    for column_idx, column_name in enumerate(unmatched_df.columns, start=1):
        unmatched_sheet.cell(row=1, column=column_idx, value=column_name)
    for row_idx, row in enumerate(unmatched_df.itertuples(index=False, name=None), start=2):
        for column_idx, value in enumerate(row, start=1):
            unmatched_sheet.cell(row=row_idx, column=column_idx, value=None if pd.isna(value) else value)

    workbook.save(output)
    print(output)
    print(
        "excluded_pia_m",
        int(exclude_mask.sum()),
        "matched",
        int(match_df["Sheet1매칭여부"].eq("Y").sum()),
        "unmatched",
        int(match_df["Sheet1매칭여부"].eq("N").sum()),
    )


if __name__ == "__main__":
    main()
