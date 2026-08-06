import hashlib
import json
import os
import pickle
import re
import shutil
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import requests
except ImportError:  # pragma: no cover - handled as a runtime configuration issue
    requests = None

try:
    import effective_production_report as effective_report
except ImportError:  # pragma: no cover - deployed without optional production-effectiveness module
    effective_report = None

st.set_page_config(page_title="생산현황", layout="wide")

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_WORKSPACE_ROOT = BASE_DIR / ".uploaded_workspaces"
LATEST_UPLOAD_SESSION_FILE = UPLOAD_WORKSPACE_ROOT / "latest_session.txt"
UPLOAD_SIGNATURE_FILE = "upload_signature.txt"
CLOUD_SNAPSHOT_DIR = BASE_DIR / "cloud_snapshots"
DISPLAY_TZ = ZoneInfo("Asia/Seoul")
ORDER_NO_COL = "수주번호"
LEADJI_REQUIRED_QTY_COL = "[45]하이드레이션/전면검사 필요수량"
LEADJI_REQUIRED_DUE_COL = "[45]하이드레이션/전면검사 납기일"
ADHESION_REQUIRED_QTY_COL = "[55]접착/멸균 필요수량"
ADHESION_REQUIRED_DUE_COL = "[55]접착/멸균 납기일"
LEADJI_COMPLETED_STOCK_COL = "누수규격검사 창고"
DEMAND_QTY_COL = "수요수량"
SEPARATION_REQUIRED_QTY_COL = "분리생산필요수량"
SEPARATION_REQUIRED_DUE_COL = "분리납기일"
DEMAND_DATA_SHEET_NAME = "Sheet1"
REWORK_SHEET_NAMES = ("재작업", "재작업리스트")
PRODUCTION_STATUS_SHEET_NAME = "생산현황"
REWORK_PRODUCTION_FILE_STEM_PREFIX = "shortage_production"
REWORK_PRODUCTION_DEMAND_COLUMNS = (
    "거래처",
    "이니셜",
    "품목코드",
    "R코드",
    "Q코드",
    "제품명",
    "파워",
    "납기일",
    "부족수량",
)

WAREHOUSE_MAP = {
    "사출창고": "사출창고",
    "분리창고": "분리창고",
    "검사접착": "검사접착창고",
    "검사접착재작업": "검사접착재작업창고",
    "누수규격검사": "누수규격검사 창고",
}
TARGET_WAREHOUSES = list(WAREHOUSE_MAP.keys())
TABLE_STYLE_CELL_LIMIT = 12000
CACHE_MAX_ENTRIES = 64
APP_CACHE_VERSION = "20260804-performance-v3"
PLAN_API_BASE_URL_DEFAULT = "https://plan.interojo.net"
PLAN_API_KEY_ENV = "PLAN_API_KEY"
PLAN_API_BASE_URL_ENV = "PLAN_API_BASE_URL"
PLAN_API_TIMEOUT_SECONDS = 120
PLAN_API_DEFAULT_ROW_LIMIT = 0
PLAN_API_CACHE_TTL_SECONDS = 300
LOCAL_CACHE_DIR = BASE_DIR / ".local_cache"
PLAN_API_DISK_CACHE_DIR = LOCAL_CACHE_DIR / "plan_api"
PLAN_API_KEY_LOCAL_CACHE_FILE = LOCAL_CACHE_DIR / "plan_api_key.txt"
ALL_ITEM_STATUS_DISK_CACHE_DIR = LOCAL_CACHE_DIR / "all_item_status"
ALL_ITEM_FLOW_STATUS_DISK_CACHE_DIR = LOCAL_CACHE_DIR / "all_item_flow_status"
FINISHED_GOODS_STOCK_DISK_CACHE_DIR = LOCAL_CACHE_DIR / "finished_goods_stock"
APS_PLAN_ENDPOINT = "/api/aps-plan"
APS_WIP_ENDPOINT = "/api/aps-wip"
APS_PLAN_META_ENDPOINT = "/api/aps-plan/meta"
PRODUCTION_PERFORMANCE_ENDPOINT = "/api/production-performance"
APS_PLAN_SHORTAGE_OPERATIONS = ("10", "20", "45", "55", "80")
APS_PLAN_FLOW_OPERATIONS = ("10", "80")
EFFECTIVE_PRODUCTION_OPERATIONS = ("10", "80")
EFFECTIVE_PRODUCTION_PROCESS_ORDER = ("[10]사출조립", "[80]누수/규격검사")
EFFECTIVE_PRODUCTION_CATEGORY_ORDER = ("국내", "PIA", "기타해외")
EFFECTIVE_PRODUCTION_DEFAULT_SITE = "C관"
EFFECTIVE_PRODUCTION_DEFAULT_DAYS = 60
EFFECTIVE_PRODUCTION_SOURCE_DIR = BASE_DIR / "effective_production_sources"
EFFECTIVE_PRODUCTION_DOI_CRITERIA_FILE = "c_site_pcode_power_doi_criteria.csv"
EFFECTIVE_SAMPLE_AVAILABLE_REFERENCE_FILE = "effective_sample_available_reference.csv"
EFFECTIVE_OVERPRODUCTION_ACTION_ORDER = ("생산자제", "생산조정", "계획초과 확인", "모니터링")
EFFECTIVE_SAMPLE_AVAILABLE_COL = "샘플 신청가능수량"
ITEM_LIST_BULK_ENDPOINT = "/api/item-list-bulk"
ALL_ITEM_MASTER_SHEET = "생성가능_P코드"
ALL_ITEM_SNAPSHOT_FILE = "all_item_status_snapshot.csv.gz"
CODE_MISMATCH_SNAPSHOT_FILE = "code_mismatch_snapshot.csv.gz"
FINISHED_GOODS_STOCK_UPLOAD_FILE = "완제품_재고변화_uploaded.xlsx"
FINISHED_GOODS_STOCK_SHEET_HINTS = ("전체 품목코드 재고", "품목코드 변화 조회결과")
USE_FINISHED_GOODS_STOCK_CHANGE = True
ALL_ITEM_DOWNLOAD_COLUMNS = [
    "사이트코드",
    "제품대분류",
    "거래처그룹",
    "거래처",
    "이니셜",
    "신규분류",
    "제품명",
    "파워",
    "납기일",
    "사출코드",
    "분리코드",
    "생산코드",
    "오더수량",
    "실수요수량",
    "총수요수량",
    "생산부족수량",
    "사출부족수량",
    "사출창고",
    "분리창고",
    "검사접착창고",
    "누수규격검사",
    "공정재고합계",
    "완제품재고",
    "재고변화",
    "DOI기준오더",
    "DOI",
    "재고비율",
    "신호",
    "재고대응판단",
    "초과재고수량",
    "부족수량",
    "샘플 신청가능수량",
    "판단",
    "상태",
    "코드매칭상태",
]
ALL_ITEM_NUMERIC_COLUMNS = [
    "오더수량",
    "실수요수량",
    "총수요수량",
    "생산부족수량",
    "사출부족수량",
    "사출창고",
    "분리창고",
    "검사접착창고",
    "누수규격검사",
    "공정재고합계",
    "완제품재고",
    "재고변화",
    "DOI기준오더",
    "DOI",
    "초과재고수량",
    "부족수량",
    "샘플 신청가능수량",
]
ALL_ITEM_FLOW_DISPLAY_COLUMNS = [
    "이니셜",
    "제품명",
    "오더수량",
    "납기일",
    "부족수량",
    "사출부족수량",
    "공정재고",
    "완제품재고",
    "재고변화",
    "DOI",
    "신호",
    "재고대응판단",
]
ALL_ITEM_FLOW_POWER_DETAIL_COLUMNS = [
    "사이트코드",
    "제품대분류",
    "거래처그룹",
    "거래처",
    "이니셜",
    "신규분류",
    "제품명",
    "파워",
    "생산코드",
    "오더수량",
    "납기일",
    "생산부족수량",
    "사출부족수량",
    "공정재고합계",
    "사출창고",
    "분리창고",
    "검사접착창고",
    "누수규격검사",
    "완제품재고",
    "재고변화",
    "DOI기준오더",
    "DOI",
    "재고비율",
    "신호",
    "재고대응판단",
]
ALL_ITEM_FLOW_CUSTOMER_ORDER = [
    "PIA",
    "OPHTALMIC",
    "Sincere",
    "HEARTS/TopTrend",
    "OPTICAL SUPPLIES",
    "MAXVUE/OPTIMAX",
    "ALENSA",
    "FEEL GOOD",
    "CHINA/IRIS",
    "MG Medical",
    "T-Garden",
    "HAPA/PPB",
    "Alcon",
    "from-eyes",
    "EYEQUE",
    "ESSILOR",
    "국내",
    "기타 거래처",
    "거래처 미지정",
]
ALL_ITEM_STATUS_OPTIONS = [
    "전체",
    "주의 필요",
    "수요 있음",
    "수요 없음",
    "수요 없음 + 공정재고 있음",
    "수요 대비 재고 초과",
    "수요 없음 + 샘플 신청가능수량 있음",
    "재고 있음",
    "코드미매칭",
]
POWER_VALUE_PATTERN = re.compile(r"([+-]\d{1,2}(?:\.\d{1,2})?)")
UNCLASSIFIED_SHEET_CATEGORY = "미분류"
INVALID_CATEGORY_VALUES = {"", "-", "nan", "none", "nat", "null", "na", "<na>"}
REWORK_AVAILABLE_QTY_COL = "재작업가능"
INITIAL_ORDER_MAP_COL = "이니셜별오더수량"
DEMAND_DETAIL_ROWS_COL = "수요상세목록"
ROW_DETAIL_MARKER = "__ROW_DETAIL__"
SITE_GROUP_ORDER = ["A관", "C관", "S관"]
SITE_GROUP_API_PARAMS = {"A관": "A", "C관": "C", "S관": "S"}

CUSTOMER_EXACT_CATEGORY_RULES = {
    "PIA Co.,Ltd.": "PIA 종합",
    "PIA Corporation": "PIA 종합",
    "INTEROJO CHINA CO., LTD": "중국(IRIS)",
    "MG Medical Group": "MG MEDICAL",
    "SINCERE Co.,Ltd": "Sincere",
    "T-garden": "T-Garden",
    "CROSSBIRD LIMITED (Feel Good Contacts)": "Feel Good",
    "(주)피피비스튜디오스": "피피비(HAPA)",
}

CUSTOMER_CATEGORY_RULES = {
    "PIA 종합": ["PIA", "PIA CO", "PIA CO.,LTD", "PIA CORPORATION"],
    "중국(IRIS)": ["IRIS", "CHINA", "중국"],
    "Freedom수출": ["FREEDOM"],
    "렌즈미": ["렌즈미", "LENSME", "LENS ME"],
    "피피비(HAPA)": ["HAPA", "피피비", "PPB"],
    "Layala": ["LAYALA"],
    "ANW": ["ANW"],
    "T-Garden": ["T-GARDEN", "TGARDEN", "T GARDEN"],
    "Sincere": ["SINCERE"],
    "Feel Good": ["FEEL GOOD", "FEELGOOD"],
    "국내": ["국내", "KOREA", "인터로조", "클라렌", "CLALEN", "LENSVERY", "LENS VERY"],
    "MG MEDICAL": ["MG MEDICAL", "MG MEDICAL GROUP"],
}

PRODUCT_CATEGORY_RULES = {
    "피피비(HAPA)": ["HAPA", "PPB"],
    "Sincere_2Week": ["SINCERE 2WEEK", "SINCERE_2WEEK", "2WEEK", "2-WEEK"],
    "T-Garden": ["T-Garden", "TGarden", "T Garden"],
    "Feel Good": ["Feel Good", "FEELGOOD", "FGC", "comfi"],
    "Freedom수출": ["Freedom", "Freedom380"],
    "1-DAY_Metha": ["Metha_Daily", "1-DAY Metha", "BL Metha"],
    "1-DAY_58": ["1-DAY_58", "1DAY58", "1-Day_Contakt", "1-Day Contakt"],
    "중국(IRIS)": ["^IRIS_"],
    "Layala": ["Layala"],
    "ANW": ["ANW"],
    "Sincere": ["Sincere"],
    "렌즈미": ["렌즈미", "LENSME", "AKMA"],
    "국내": ["PIA_KR", "Clalen", "클라렌", "Lensvery", "Lens Very"],
    "MG MEDICAL": ["MG M_"],
    "PIA 종합": ["PIA", "feliamo", "Lilmoon", "MOLAK"],
}

SINCERE_2WEEK_RULES = {
    "Sincere_2Week": ["SINCERE 2WEEK", "SINCERE_2WEEK", "2WEEK", "2-WEEK"]
}

COLUMN_LABEL_ALIASES = {
    "사출창고": "사출 재고",
    "분리창고": "분리 재고",
    "검사접착창고": "검사접착 재고",
    "검사접착재작업창고": "검사접착재작업 재고",
    "누수규격검사 창고": "누수규격 재고",
    "공정재고합계": "공정재고",
    "사출창고 합계": "사출 재고",
    "분리창고 합계": "분리 재고",
    "검사접착창고 합계": "검사접착 재고",
    "검사접착재작업창고 합계": "검사접착재작업 재고",
    "누수규격검사창고 합계": "누수규격 재고",
    "공정재고 합계": "공정재고",
    "사출 부족수량": "사출부족",
    "사출부족수량": "사출부족",
    "생산부족수량": "부족수량",
    "사출생산필요수량": "사출필요",
    "분리생산필요수량": "분리필요",
    "[45]하이드레이션/전면검사 필요수량": "45필요",
    "[55]접착/멸균 필요수량": "55필요",
    "수요수량": "수요",
    "생산필요수량": "생산필요",
    "최소납기일": "생산 최소 납기일",
    "재작업가능": "재작업가능",
    "비고": "비고",
    "확인구분": "확인구분",
    "리스크구분": "리스크",
    "제품군키": "제품군",
    "재고수량": "재고",
    "현재수요수량": "현재수요",
    "초과수량": "초과",
    "오더수량": "오더수량",
    "완제품재고": "완제품재고",
    "제품대분류": "제품분류",
    "거래처그룹": "거래처",
    "DOI": "DOI",
    "신호": "신호",
    "수요코드수": "수요코드",
    "제품명 예시": "제품명",
    "이니셜 예시": "이니셜",
    "재공코드 예시": "재공코드",
    "LOT 예시": "LOT",
}


def inject_dashboard_theme() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
        .stApp {
            background: #FFFFFF;
            color: #111827;
            font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        }
        [data-testid="stAppViewContainer"] > .main {
            background: linear-gradient(180deg, #FFFFFF 0%, #F4F6F8 100%);
        }
        [data-testid="stHeader"] {
            background: rgba(255, 255, 255, 0.94);
            backdrop-filter: blur(8px);
        }
        .main .block-container,
        .block-container {
            max-width: 100%;
            padding-left: 2rem;
            padding-right: 2rem;
            padding-top: 24px;
            padding-bottom: 44px;
        }
        [data-testid="stSidebar"] {
            background: #FFFFFF;
            border-right: 1px solid #E5E7EB;
        }
        [data-testid="stSidebar"] * {
            font-family: Inter, sans-serif;
        }
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
            color: #334155;
        }
        .sidebar-brand {
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 10px 0 18px;
            margin-bottom: 10px;
            border-bottom: 1px solid #E5E7EB;
        }
        .sidebar-brand-icon {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 30px;
            height: 30px;
            border-radius: 8px;
            background: #EEF2FF;
            color: #1A2B5E;
        }
        .sidebar-brand-title {
            color: #111827;
            font-size: 19px;
            font-weight: 850;
            letter-spacing: 0;
        }
        .sidebar-divider {
            height: 1px;
            background: #E5E7EB;
            margin: 16px 0;
        }
        [data-testid="stSidebar"] [role="radiogroup"] label {
            border-radius: 8px;
            padding: 6px 8px;
            margin-bottom: 3px;
        }
        h1, h2, h3 {
            color: #111827;
            letter-spacing: 0;
        }
        h1 {
            font-size: 28px;
            font-weight: 800;
            margin-bottom: 8px;
        }
        h2, h3 {
            font-weight: 800;
        }
        .dashboard-hero {
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            background: linear-gradient(135deg, #FFFFFF 0%, #F4F6F8 100%);
            box-shadow: 0 14px 40px rgba(15, 23, 42, 0.08);
            padding: 22px 24px;
            margin-bottom: 18px;
            border-left: 5px solid #1A2B5E;
        }
        .dashboard-hero-title {
            color: #1A2B5E;
            font-size: 30px;
            font-weight: 850;
            line-height: 1.25;
            margin: 0 0 6px;
        }
        .dashboard-hero-subtitle {
            color: #64748B;
            font-size: 14px;
            margin: 0;
        }
        .sidebar-section-title {
            color: #1A2B5E;
            font-size: 14px;
            font-weight: 850;
            padding: 4px 0 9px;
            border-bottom: 1px solid #E5E7EB;
            margin-bottom: 12px;
        }
        [data-testid="stCaptionContainer"] {
            color: #6b7280;
        }
        [data-testid="stAlert"] {
            border-radius: 8px;
            border: 1px solid #D7DBE8;
            background: #F7F8FB;
            box-shadow: 0 4px 14px rgba(26, 43, 94, 0.06);
        }
        [data-testid="stDataFrame"] {
            border: 1px solid #E5E7EB;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 10px 30px rgba(15, 23, 42, 0.06);
            background: #ffffff;
        }
        [data-testid="stExpander"] {
            border: 1px solid #E5E7EB;
            border-radius: 12px;
            background: #ffffff;
            box-shadow: 0 8px 22px rgba(15, 23, 42, 0.05);
        }
        [data-testid="stExpanderToggleIcon"],
        [data-testid="stExpander"] summary [data-testid="stIconMaterial"],
        [data-testid="stExpander"] summary .material-icons,
        [data-testid="stExpander"] summary .material-icons-outlined,
        [data-testid="stExpander"] summary .material-symbols-rounded {
            font-size: 0 !important;
            width: 0 !important;
            min-width: 0 !important;
            margin: 0 !important;
            overflow: hidden !important;
        }
        [data-testid="stTextInput"] input {
            border-radius: 8px;
            border-color: #D1D5DB;
            background: #ffffff;
            color: #111827;
        }
        [data-testid="stTextInput"] input:focus {
            border-color: #1A2B5E;
            box-shadow: 0 0 0 2px rgba(26, 43, 94, 0.10);
        }
        div[data-testid="stMetric"] {
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            padding: 15px 16px;
            border-left: 4px solid #1A2B5E;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.07);
        }
        div[data-testid="stMetric"] label {
            color: #64748B !important;
            font-weight: 700;
        }
        div[data-testid="stMetric"] [data-testid="stMetricValue"] {
            color: #374151;
            font-weight: 850;
            white-space: nowrap;
            overflow: visible;
            text-overflow: unset;
            font-size: clamp(22px, 1.7vw, 32px);
        }
        .ops-kpi-card {
            min-height: 108px;
            border-radius: 14px;
            background: #FFFFFF;
            border: 1px solid #E5E7EB;
            border-left: 5px solid #2563EB;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.07);
            padding: 16px 18px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            overflow: visible;
        }
        .ops-kpi-card.risk {
            border-left-color: #DC2626;
            background: linear-gradient(180deg, #FFFFFF 0%, #FFF7F7 100%);
        }
        .ops-kpi-card.stock {
            border-left-color: #2563EB;
            background: linear-gradient(180deg, #FFFFFF 0%, #F8FAFF 100%);
        }
        .kpi-label {
            color: #64748B;
            font-size: 13px;
            font-weight: 800;
            line-height: 1.25;
        }
        .kpi-value {
            white-space: nowrap;
            overflow: visible;
            text-overflow: unset;
            font-size: clamp(24px, 2vw, 36px);
            font-weight: 850;
            line-height: 1.1;
            letter-spacing: 0;
            color: #374151;
        }
        .ops-kpi-card.risk .kpi-value {
            color: #DC2626;
        }
        .ops-kpi-card.stock .kpi-value {
            color: #1A2B5E;
        }
        .stButton > button,
        [data-testid="stDownloadButton"] button {
            border-radius: 8px;
            border: 1px solid #1A2B5E;
            background: #1A2B5E;
            color: #FFFFFF;
            font-weight: 700;
            box-shadow: 0 6px 16px rgba(26, 43, 94, 0.16);
        }
        .stButton > button p,
        [data-testid="stDownloadButton"] button p {
            color: #FFFFFF !important;
        }
        .stButton > button:hover,
        [data-testid="stDownloadButton"] button:hover {
            border-color: #233A7A;
            color: #FFFFFF;
            background: #233A7A;
        }
        .stButton > button:disabled,
        [data-testid="stDownloadButton"] button:disabled {
            background: #1A2B5E;
            border-color: #1A2B5E;
            color: #FFFFFF;
            opacity: 0.72;
        }
        .stButton > button:disabled p,
        [data-testid="stDownloadButton"] button:disabled p {
            color: #FFFFFF !important;
        }
        [data-testid="stSegmentedControl"] {
            background: #F4F6F8;
            border: 1px solid #E5E7EB;
            border-radius: 10px;
            padding: 4px;
        }
        .dashboard-section-header {
            display: flex;
            align-items: center;
            gap: 10px;
            margin: 22px 0 10px;
        }
        .dashboard-section-header h3 {
            margin: 0;
            font-size: 20px;
            line-height: 1.2;
        }
        .dashboard-count-badge {
            display: inline-flex;
            align-items: center;
            border-radius: 999px;
            background: #EEF2FF;
            color: #1A2B5E;
            padding: 4px 9px;
            font-size: 12px;
            font-weight: 800;
        }
        .dashboard-section-subtle {
            color: #6b7280;
            font-size: 13px;
            margin-top: -4px;
            margin-bottom: 10px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def find_excel_files(base_dir: Path) -> tuple[Path, Path]:
    xlsx_files = [p for p in base_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    if len(xlsx_files) < 2:
        raise FileNotFoundError("xlsx 파일 2개(재고/수요)가 필요합니다.")

    odv_candidates = [p for p in xlsx_files if p.stem.upper().startswith("ODV_WIP")]
    if not odv_candidates:
        odv_candidates = [p for p in xlsx_files if "ODV" in p.stem.upper() and "WIP" in p.stem.upper()]

    if odv_candidates:
        inv_path = max(odv_candidates, key=lambda p: p.stat().st_mtime)
    else:
        inventory_candidates = [p for p in xlsx_files if "재고" in p.name]
        inv_path = max(inventory_candidates, key=lambda p: p.stat().st_size) if inventory_candidates else max(
            xlsx_files, key=lambda p: p.stat().st_size
        )

    demand_candidates = [p for p in xlsx_files if p != inv_path]
    demand_named = [p for p in demand_candidates if "수요" in p.name]
    if demand_named:
        demand_candidates = demand_named

    full_process_candidates = [p for p in demand_candidates if "전공정" in p.stem]
    if full_process_candidates:
        dem_path = max(full_process_candidates, key=lambda p: p.stat().st_mtime)
    else:
        dem_path = max(demand_candidates, key=lambda p: p.stat().st_mtime)

    return inv_path, dem_path


def find_demand_update_file(base_dir: Path) -> Path | None:
    xlsx_files = [p for p in base_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    if not xlsx_files:
        return None

    exact = [p for p in xlsx_files if p.name == "수요정보(전공정).xlsx"]
    if exact:
        return max(exact, key=lambda p: p.stat().st_mtime)

    normalized = lambda s: str(s).replace(" ", "")
    full_process = [p for p in xlsx_files if "수요정보(전공정)" in normalized(p.stem)]
    if full_process:
        return max(full_process, key=lambda p: p.stat().st_mtime)

    demand_info = [p for p in xlsx_files if "수요정보" in normalized(p.stem)]
    if demand_info:
        return max(demand_info, key=lambda p: p.stat().st_mtime)

    demand_like = [p for p in xlsx_files if "수요" in normalized(p.stem)]
    if demand_like:
        return max(demand_like, key=lambda p: p.stat().st_mtime)

    return None


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        return sheet_name in wb.sheetnames
    finally:
        wb.close()


def normalize_excel_sheet_name(value: object) -> str:
    return str(value).replace(" ", "").strip()


def find_workbook_sheet_name(wb, preferred_names: tuple[str, ...] | list[str], fallback_first: bool = True) -> str | None:
    sheet_names = list(wb.sheetnames)
    for preferred in preferred_names:
        preferred_norm = normalize_excel_sheet_name(preferred)
        for sheet_name in sheet_names:
            if normalize_excel_sheet_name(sheet_name) == preferred_norm:
                return sheet_name
    return sheet_names[0] if fallback_first and sheet_names else None


def find_all_item_master_file(base_dir: Path) -> Path | None:
    search_dirs = [base_dir]
    if base_dir.resolve() != BASE_DIR.resolve():
        search_dirs.append(BASE_DIR)

    candidates: list[Path] = []
    for search_dir in search_dirs:
        for path in search_dir.glob("*.xlsx"):
            if path.name.startswith("~$") or path in candidates:
                continue
            if workbook_has_sheet(path, ALL_ITEM_MASTER_SHEET):
                candidates.append(path)

    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def workbook_has_any_sheet(path: Path, sheet_names: tuple[str, ...]) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        normalized_wb_names = {normalize_excel_sheet_name(name) for name in wb.sheetnames}
        return any(normalize_excel_sheet_name(name) in normalized_wb_names for name in sheet_names)
    finally:
        wb.close()


def find_finished_goods_stock_file(base_dir: Path) -> Path | None:
    if not USE_FINISHED_GOODS_STOCK_CHANGE:
        return None

    search_dirs = [base_dir]
    if base_dir.resolve() != BASE_DIR.resolve():
        search_dirs.append(BASE_DIR)
    downloads_dir = Path.home() / "Downloads"
    if downloads_dir.exists():
        search_dirs.append(downloads_dir)

    candidates: list[Path] = []
    for search_dir in unique_existing_paths(search_dirs):
        for path in search_dir.glob("*.xlsx"):
            if path.name.startswith("~$") or path in candidates:
                continue
            normalized_name = path.name.replace(" ", "")
            is_staged_upload = path.name == FINISHED_GOODS_STOCK_UPLOAD_FILE
            is_stock_change = "재고변화" in normalized_name and "품목코드" in normalized_name
            is_lot_stock = "LOT" in normalized_name.upper() and "품목코드" in normalized_name and "재고" in normalized_name
            if not (is_staged_upload or is_stock_change or is_lot_stock):
                continue
            candidates.append(path)

    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_leadji_order_status_file(base_dir: Path) -> Path | None:
    xlsx_files = [p for p in base_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    if not xlsx_files:
        return None

    normalized = lambda s: str(s).replace(" ", "")
    candidates = [p for p in xlsx_files if normalized(p.name) == "리드지발주현황.xlsx"]
    if not candidates:
        candidates = [p for p in xlsx_files if "리드지" in normalized(p.stem) and "발주현황" in normalized(p.stem)]
    if not candidates:
        candidates = [p for p in xlsx_files if "발주현황" in normalized(p.stem)]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def unique_existing_paths(paths: list[Path | None]) -> list[Path]:
    result: list[Path] = []
    seen: set[Path] = set()
    for path in paths:
        if path is None or not path.exists():
            continue
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        result.append(path)
    return result


def find_rework_production_source_file(base_dir: Path) -> Path | None:
    search_dirs = [base_dir]
    try:
        if base_dir.resolve() != BASE_DIR.resolve():
            search_dirs.append(BASE_DIR)
    except OSError:
        search_dirs.append(BASE_DIR)

    candidates: list[Path] = []
    for search_dir in unique_existing_paths(search_dirs):
        for path in search_dir.glob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            normalized_stem = path.stem.replace(" ", "").lower()
            if not normalized_stem.startswith(REWORK_PRODUCTION_FILE_STEM_PREFIX):
                continue
            if workbook_has_any_sheet(path, (PRODUCTION_STATUS_SHEET_NAME,)):
                candidates.append(path)

    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def build_files_refresh_key(paths: list[Path]) -> str:
    parts: list[str] = []
    for path in paths:
        stat = path.stat()
        parts.append(f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}")
    return "|".join(parts)


def get_file_updated_datetime(path: Path) -> datetime:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        modified = wb.properties.modified
        wb.close()
        if isinstance(modified, datetime):
            if modified.tzinfo is None:
                return modified.replace(tzinfo=ZoneInfo("UTC")).astimezone(DISPLAY_TZ)
            return modified.astimezone(DISPLAY_TZ)
    except Exception:
        pass
    return datetime.fromtimestamp(path.stat().st_mtime, tz=DISPLAY_TZ)


def get_latest_files_updated_at(paths: list[Path]) -> str:
    if not paths:
        return "-"
    refresh_key = build_files_refresh_key(paths)
    return get_latest_files_updated_at_cached(refresh_key, tuple(str(path) for path in paths))


@st.cache_data(show_spinner=False)
def get_latest_files_updated_at_cached(refresh_key: str, path_strs: tuple[str, ...]) -> str:
    _ = refresh_key
    paths = [Path(path_str) for path_str in path_strs]
    existing_paths = [path for path in paths if path.exists()]
    if not existing_paths:
        return "-"
    latest_dt = max(get_file_updated_datetime(path) for path in existing_paths)
    return latest_dt.strftime("%Y-%m-%d %H:%M:%S")


def get_streamlit_or_env_secret(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name, "")
    except Exception:
        value = ""
    if value is None or str(value).strip() == "":
        value = os.environ.get(name, default)
    if value is None or str(value).strip() == "":
        value = read_local_streamlit_secret(name, default)
    return str(value).strip()


def get_plan_api_base_url() -> str:
    return get_streamlit_or_env_secret(PLAN_API_BASE_URL_ENV, PLAN_API_BASE_URL_DEFAULT).rstrip("/")


def get_user_home_candidates() -> list[Path]:
    candidates = [Path.home()]
    for env_name in ("USERPROFILE", "HOME"):
        env_value = os.environ.get(env_name, "")
        if env_value:
            candidates.append(Path(env_value))
    return unique_existing_paths(candidates)


def get_local_secret_file_candidates() -> list[Path]:
    paths = [
        BASE_DIR / ".streamlit" / "secrets.toml",
        Path.cwd() / ".streamlit" / "secrets.toml",
    ]
    for home in get_user_home_candidates():
        paths.append(home / "Documents" / "GitHub" / "INTEROJO" / ".streamlit" / "secrets.toml")
    return unique_existing_paths(paths)


def read_local_streamlit_secret(name: str, default: str = "") -> str:
    pattern = re.compile(rf"^\s*{re.escape(name)}\s*=\s*(.+?)\s*$")
    for path in get_local_secret_file_candidates():
        if not path.exists():
            continue
        for encoding in ("utf-8-sig", "utf-8", "cp949"):
            try:
                lines = path.read_text(encoding=encoding).splitlines()
            except UnicodeDecodeError:
                continue
            except OSError:
                break
            for line in lines:
                match = pattern.match(line)
                if not match:
                    continue
                value = match.group(1).strip().strip("\"'")
                if value:
                    return value
            break
    return default


def extract_plan_api_key_from_text(text: str) -> str:
    key_labels = ("X-API-Key", "PLAN_API_KEY", "API_KEY", "API Key", "API키", "인증키", "키")
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        value_scope = line
        if any(label.lower() in line.lower() for label in key_labels):
            parts = re.split(r"[:=]\s*", line, maxsplit=1)
            value_scope = parts[1].strip() if len(parts) == 2 else line
        candidates = re.findall(r"[A-Za-z0-9][A-Za-z0-9._~+/=-]{15,}", value_scope)
        for candidate in candidates:
            token = candidate.strip().strip("\"'")
            lowered = token.lower()
            if lowered.startswith(("http", "api_key", "x-api-key", "plan_api_key")):
                continue
            if "@" in token:
                continue
            return token
    return ""


def find_local_plan_api_key_file() -> Path | None:
    home_candidates = get_user_home_candidates()
    search_dirs = unique_existing_paths(
        [
            BASE_DIR,
            LOCAL_CACHE_DIR,
            BASE_DIR / ".local_cache",
            BASE_DIR / ".streamlit",
            Path.cwd(),
            Path.cwd() / ".local_cache",
            Path.cwd() / ".streamlit",
            *[
                path
                for home in home_candidates
                for path in (
                    home / "Downloads",
                    home / "Desktop",
                    home / "Documents",
                    home / "Documents" / "GitHub" / "INTEROJO",
                    home / "Documents" / "GitHub" / "INTEROJO" / ".local_cache",
                    home / "Documents" / "GitHub" / "INTEROJO" / ".streamlit",
                )
            ],
        ]
    )
    if PLAN_API_KEY_LOCAL_CACHE_FILE.exists():
        return PLAN_API_KEY_LOCAL_CACHE_FILE
    candidates: list[Path] = []
    for path in get_local_secret_file_candidates():
        if path.exists() and path not in candidates:
            candidates.append(path)
    patterns = (
        "API_KEY*.txt",
        "API_KEY_*.txt",
        "*API*KEY*.txt",
        "*api*key*.txt",
        "*API*키*.txt",
        "*키*안내*.txt",
        "*KEY*안내*.txt",
    )
    for search_dir in search_dirs:
        for pattern in patterns:
            for path in search_dir.glob(pattern):
                if path.is_file() and path not in candidates:
                    candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def read_local_plan_api_key() -> str:
    path = find_local_plan_api_key_file()
    if path is None:
        return ""
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        except OSError:
            return ""
        token = extract_plan_api_key_from_text(text)
        if token:
            return token
    return ""


def get_plan_api_key() -> str:
    session_key = get_session_value("plan_api_key_input", "")
    if session_key is not None and str(session_key).strip():
        return str(session_key).strip()
    configured_key = get_streamlit_or_env_secret(PLAN_API_KEY_ENV, "")
    if configured_key:
        return configured_key
    return read_local_plan_api_key()


def get_plan_api_key_source_label() -> str:
    session_key = get_session_value("plan_api_key_input", "")
    if session_key is not None and str(session_key).strip():
        return "화면 입력"
    if get_streamlit_or_env_secret(PLAN_API_KEY_ENV, ""):
        return "환경변수/Secrets"
    if read_local_plan_api_key():
        return "로컬 API 키 안내 파일"
    return ""


def get_session_value(key: str, default: object = None) -> object:
    try:
        return st.session_state.get(key, default)
    except Exception:
        return default


def set_session_value(key: str, value: object) -> None:
    try:
        st.session_state[key] = value
    except Exception:
        pass


def get_plan_api_refresh_nonce() -> int:
    value = get_session_value("plan_api_refresh_nonce", 0)
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def is_plan_api_configured() -> bool:
    return requests is not None and bool(get_plan_api_key())


def is_plan_api_enabled() -> bool:
    return bool(get_session_value("use_plan_api_data_mode", False)) and is_plan_api_configured()


def build_plan_api_refresh_key() -> str:
    if not is_plan_api_enabled():
        return "plan-api:disabled"
    key_hash = hashlib.sha256(get_plan_api_key().encode("utf-8")).hexdigest()[:12]
    source_updated_at = get_plan_api_updated_at()
    return f"plan-api:{get_plan_api_base_url()}:{key_hash}:{source_updated_at}:{get_plan_api_refresh_nonce()}"


def build_local_cache_hash(*parts: object) -> str:
    payload = json.dumps(parts, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_plan_api_disk_cache_path(
    base_url: str,
    endpoint: str,
    params_tuple: tuple[tuple[str, object], ...],
    api_key_hash: str,
    source_updated_at: str,
) -> Path:
    cache_hash = build_local_cache_hash(
        APP_CACHE_VERSION,
        base_url.rstrip("/"),
        endpoint,
        params_tuple,
        api_key_hash,
        source_updated_at,
    )
    return PLAN_API_DISK_CACHE_DIR / f"{cache_hash}.pkl.gz"


def read_plan_api_disk_cache(cache_path: Path) -> pd.DataFrame | None:
    if not cache_path.exists():
        return None
    try:
        cached = pd.read_pickle(cache_path, compression="gzip")
    except Exception:
        return None
    return cached if isinstance(cached, pd.DataFrame) else None


def write_plan_api_disk_cache(cache_path: Path, df: pd.DataFrame) -> None:
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = cache_path.with_suffix(cache_path.suffix + f".{uuid.uuid4().hex}.tmp")
        df.to_pickle(temp_path, compression="gzip")
        temp_path.replace(cache_path)
    except Exception:
        pass


def normalize_api_column_key(value: object) -> str:
    return re.sub(r"[\s_./()\-\[\]:]+", "", str(value).strip().lower())


def pick_api_column(columns: list[str], candidates: list[str]) -> str | None:
    exact = pick_first_existing_column(columns, candidates)
    if exact is not None:
        return exact
    normalized = {normalize_api_column_key(col): col for col in columns}
    for candidate in candidates:
        matched = normalized.get(normalize_api_column_key(candidate))
        if matched is not None:
            return matched
    return None


def normalize_api_records(payload: object) -> list[dict[str, object]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if not isinstance(payload, dict):
        return []

    for key in ("data", "rows", "orders", "items", "result", "results", "records", "list"):
        value = payload.get(key)
        if isinstance(value, list):
            return [row for row in value if isinstance(row, dict)]
        if isinstance(value, dict):
            nested = normalize_api_records(value)
            if nested:
                return nested

    for value in payload.values():
        if isinstance(value, list) and all(isinstance(row, dict) for row in value[:5]):
            return value
    return []


@st.cache_data(show_spinner=False, ttl=PLAN_API_CACHE_TTL_SECONDS, max_entries=CACHE_MAX_ENTRIES)
def fetch_plan_api_dataframe_cached(
    base_url: str,
    endpoint: str,
    params_tuple: tuple[tuple[str, object], ...],
    api_key_hash: str,
    source_updated_at: str,
    refresh_nonce: int,
) -> tuple[pd.DataFrame, str]:
    _ = api_key_hash, refresh_nonce
    if requests is None:
        return pd.DataFrame(), "requests 패키지가 설치되어 있지 않습니다."

    api_key = get_plan_api_key()
    if not api_key:
        return pd.DataFrame(), f"{PLAN_API_KEY_ENV}가 설정되어 있지 않습니다."

    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    params = {str(key): value for key, value in params_tuple if value not in (None, "")}
    disk_cache_path = build_plan_api_disk_cache_path(base_url, endpoint, params_tuple, api_key_hash, source_updated_at)
    if source_updated_at != "-":
        cached_df = read_plan_api_disk_cache(disk_cache_path)
        if cached_df is not None:
            return cached_df, ""

    try:
        response = requests.get(
            url,
            params=params,
            headers={"X-API-Key": api_key},
            timeout=PLAN_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        return pd.DataFrame(), str(exc)

    if isinstance(payload, dict) and payload.get("truncated") is True:
        returned_count = payload.get("returned_count", "?")
        total_count = payload.get("total_count") or payload.get("source_total_count") or "?"
        return pd.DataFrame(), f"API response was truncated ({returned_count}/{total_count}); ignored partial API data."

    records = normalize_api_records(payload)
    if not records:
        return pd.DataFrame(), "API 응답에서 행 데이터를 찾지 못했습니다."
    df = pd.DataFrame.from_records(records)
    if source_updated_at != "-":
        write_plan_api_disk_cache(disk_cache_path, df)
    return df, ""


def fetch_plan_api_dataframe_direct(
    base_url: str,
    endpoint: str,
    params_tuple: tuple[tuple[str, object], ...],
    api_key: str,
    api_key_hash: str,
    source_updated_at: str,
) -> tuple[pd.DataFrame, str]:
    if requests is None:
        return pd.DataFrame(), "requests package is not installed."
    if not api_key:
        return pd.DataFrame(), f"{PLAN_API_KEY_ENV} is not configured."

    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    params = {str(key): value for key, value in params_tuple if value not in (None, "")}
    disk_cache_path = build_plan_api_disk_cache_path(base_url, endpoint, params_tuple, api_key_hash, source_updated_at)
    if source_updated_at != "-":
        cached_df = read_plan_api_disk_cache(disk_cache_path)
        if cached_df is not None:
            return cached_df, ""

    try:
        response = requests.get(
            url,
            params=params,
            headers={"X-API-Key": api_key},
            timeout=PLAN_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        return pd.DataFrame(), str(exc)

    if isinstance(payload, dict) and payload.get("truncated") is True:
        returned_count = payload.get("returned_count", "?")
        total_count = payload.get("total_count") or payload.get("source_total_count") or "?"
        return pd.DataFrame(), f"API response was truncated ({returned_count}/{total_count}); ignored partial API data."

    records = normalize_api_records(payload)
    if not records:
        return pd.DataFrame(), "API response did not contain data rows."
    df = pd.DataFrame.from_records(records)
    if source_updated_at != "-":
        write_plan_api_disk_cache(disk_cache_path, df)
    return df, ""


@st.cache_data(show_spinner=False, ttl=PLAN_API_CACHE_TTL_SECONDS, max_entries=CACHE_MAX_ENTRIES)
def fetch_plan_api_meta_cached(
    base_url: str,
    endpoint: str,
    api_key_hash: str,
    refresh_nonce: int,
) -> tuple[dict[str, object], str]:
    _ = api_key_hash, refresh_nonce
    if requests is None:
        return {}, "requests 패키지가 설치되어 있지 않습니다."

    api_key = get_plan_api_key()
    if not api_key:
        return {}, f"{PLAN_API_KEY_ENV}가 설정되어 있지 않습니다."

    url = f"{base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    try:
        response = requests.get(
            url,
            headers={"X-API-Key": api_key},
            timeout=PLAN_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception as exc:
        return {}, str(exc)
    return payload if isinstance(payload, dict) else {}, ""


def read_plan_api_dataframe(endpoint: str, params: dict[str, object] | None = None) -> tuple[pd.DataFrame, str]:
    if not is_plan_api_enabled():
        return pd.DataFrame(), "API 자동조회가 꺼져 있습니다."
    api_key_hash = hashlib.sha256(get_plan_api_key().encode("utf-8")).hexdigest()[:12]
    params_tuple = tuple(sorted((params or {}).items(), key=lambda item: item[0]))
    source_updated_at = get_plan_api_updated_at()
    return fetch_plan_api_dataframe_cached(
        get_plan_api_base_url(),
        endpoint,
        params_tuple,
        api_key_hash,
        source_updated_at,
        get_plan_api_refresh_nonce(),
    )


def read_aps_plan_operations_dataframe(
    operations: tuple[str, ...],
    site_filter: str = "전체",
) -> tuple[pd.DataFrame, str]:
    if not is_plan_api_enabled():
        return pd.DataFrame(), "API 자동조회가 꺼져 있습니다."
    api_key = get_plan_api_key()
    api_key_hash = hashlib.sha256(api_key.encode("utf-8")).hexdigest()[:12]
    base_url = get_plan_api_base_url()
    source_updated_at = get_plan_api_updated_at()
    frames: list[pd.DataFrame] = []
    errors: list[str] = []
    site_param = build_plan_api_site_param(site_filter)

    def fetch_operation(operation: str) -> tuple[str, pd.DataFrame, str]:
        params: dict[str, object] = {
            "limit": PLAN_API_DEFAULT_ROW_LIMIT,
            "oper": operation,
        }
        if site_param:
            params["site"] = site_param
        params_tuple = tuple(sorted(params.items(), key=lambda item: item[0]))
        frame, error = fetch_plan_api_dataframe_direct(
            base_url,
            APS_PLAN_ENDPOINT,
            params_tuple,
            api_key,
            api_key_hash,
            source_updated_at,
        )
        return operation, frame, error

    max_workers = max(1, min(len(operations), 5))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(fetch_operation, operation) for operation in operations]
        for future in as_completed(futures):
            operation, frame, error = future.result()
            if error:
                errors.append(f"oper={operation}: {error}")
                continue
            if not frame.empty:
                frames.append(frame)
    if frames:
        return pd.concat(frames, ignore_index=True, sort=False), ""
    return pd.DataFrame(), "; ".join(errors)


def find_meta_value(payload: dict[str, object], candidates: list[str]) -> str:
    if not payload:
        return ""
    normalized_candidates = {normalize_api_column_key(candidate) for candidate in candidates}
    for key, value in payload.items():
        if normalize_api_column_key(key) in normalized_candidates:
            text = str(value).strip()
            if text and text.lower() not in INVALID_CATEGORY_VALUES:
                return text
        if isinstance(value, dict):
            nested = find_meta_value(value, candidates)
            if nested:
                return nested
    return ""


def get_plan_api_updated_at() -> str:
    if not is_plan_api_configured():
        return "-"
    api_key_hash = hashlib.sha256(get_plan_api_key().encode("utf-8")).hexdigest()[:12]
    payload, error = fetch_plan_api_meta_cached(
        get_plan_api_base_url(),
        APS_PLAN_META_ENDPOINT,
        api_key_hash,
        get_plan_api_refresh_nonce(),
    )
    if error:
        return "-"
    value = find_meta_value(
        payload,
        [
            "updated_at",
            "last_updated_at",
            "last_refreshed_at",
            "lastRefreshedAt",
            "lastUpdateAt",
            "source_refreshed_at",
            "sourceRefreshedAt",
            "refreshed_at",
            "load_dt",
            "LOAD_DT",
            "snapshot_at",
            "generated_at",
            "created_at",
            "원장 마지막 갱신시각",
            "마지막 갱신시각",
            "추출시각",
            "기준시각",
            "APS실행일시",
            "실행시각",
        ],
    )
    if not value:
        return "-"
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    return value


def render_plan_api_status() -> None:
    if requests is None:
        st.warning("API 자동조회 불가: requests 패키지가 설치되어 있지 않습니다.")
        return
    if not get_plan_api_key():
        st.caption(f"API 자동조회 비활성: {PLAN_API_KEY_ENV} 설정 필요")
        return
    st.caption(f"API 기준: {get_plan_api_base_url()}")
    updated_at = get_plan_api_updated_at()
    if updated_at != "-":
        st.caption(f"APS API 갱신시각: {updated_at}")


def get_data_updated_at(base_dir: Path) -> str:
    try:
        inv_path, dem_path = find_excel_files(base_dir)
    except Exception:
        inv_path = None
        dem_path = find_demand_update_file(base_dir)
    ref_path = find_product_name_reference_file(base_dir)
    rework_source_path = find_rework_production_source_file(base_dir)

    return get_latest_files_updated_at(unique_existing_paths([inv_path, dem_path, ref_path, rework_source_path]))


def get_wip_updated_at(base_dir: Path) -> str:
    try:
        inv_path, _ = find_excel_files(base_dir)
    except Exception:
        inv_path = None
    return get_latest_files_updated_at(unique_existing_paths([inv_path]))


def get_local_demand_updated_at(base_dir: Path) -> str:
    try:
        _, dem_path = find_excel_files(base_dir)
    except Exception:
        dem_path = find_demand_update_file(base_dir)
    return get_latest_files_updated_at(unique_existing_paths([dem_path]))


def get_aps_or_file_updated_at(file_updated_at: str) -> str:
    if is_plan_api_enabled():
        api_updated_at = get_plan_api_updated_at()
        if api_updated_at != "-":
            return api_updated_at
    return file_updated_at


def format_reference_timestamp(value: str, fallback: str = "확인 불가") -> str:
    text = clean_text_value(value)
    return text if text and text != "-" else fallback


def render_sidebar_reference_dates(data_base_dir: Path, source_label: str) -> None:
    api_configured = is_plan_api_configured()
    api_updated_at = get_plan_api_updated_at() if api_configured else "-"
    if not api_configured:
        api_label = "미반영 (API 키 없음)"
    elif not is_plan_api_enabled():
        api_label = f"미반영 (자동조회 꺼짐, 갱신 {format_reference_timestamp(api_updated_at)})"
    else:
        api_label = f"반영 ({format_reference_timestamp(api_updated_at)})"

    st.markdown('<div class="sidebar-section-title">반영 기준일자</div>', unsafe_allow_html=True)
    st.caption(f"APS API 수요: {api_label}")
    st.caption(f"WIP 파일: {format_reference_timestamp(get_wip_updated_at(data_base_dir))}")
    st.caption(f"로컬 수요 파일: {format_reference_timestamp(get_local_demand_updated_at(data_base_dir))}")
    st.caption(f"적용 데이터: {source_label}")


def get_all_item_updated_at(base_dir: Path) -> str:
    try:
        inv_path, dem_path = find_excel_files(base_dir)
    except Exception:
        inv_path = None
        dem_path = find_demand_update_file(base_dir)
    paths = unique_existing_paths(
        [
            inv_path,
            dem_path,
            find_product_name_reference_file(base_dir),
            find_all_item_master_file(base_dir),
            find_finished_goods_stock_file(base_dir),
        ]
    )
    return get_latest_files_updated_at(paths)


def get_leadji_status_updated_at(base_dir: Path) -> str:
    try:
        inv_path, dem_path = find_excel_files(base_dir)
    except Exception:
        inv_path = None
        dem_path = find_demand_update_file(base_dir)
    paths = unique_existing_paths(
        [inv_path, dem_path, find_product_name_reference_file(base_dir), find_leadji_order_status_file(base_dir)]
    )
    return get_latest_files_updated_at(paths)


def get_leadji_order_updated_at(base_dir: Path) -> str:
    order_path = find_leadji_order_status_file(base_dir)
    if order_path is None:
        return "-"
    stat = order_path.stat()
    refresh_key = f"{order_path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    return get_leadji_order_updated_at_cached(refresh_key, str(order_path))


@st.cache_data(show_spinner=False)
def get_leadji_order_updated_at_cached(refresh_key: str, order_path_str: str) -> str:
    _ = refresh_key
    order_path = Path(order_path_str)
    if not order_path.exists():
        return "-"
    return get_file_updated_datetime(order_path).strftime("%Y-%m-%d %H:%M:%S")


def get_or_create_upload_session_id() -> str:
    key = "upload_session_id"
    if key not in st.session_state:
        st.session_state[key] = uuid.uuid4().hex
    return str(st.session_state[key])


def is_valid_uploaded_workspace(path: Path) -> bool:
    return (
        path.is_dir()
        and (path / "ODV_WIP_uploaded.xlsx").exists()
        and (path / "수요정보(전공정).xlsx").exists()
    )


def get_latest_uploaded_workspace() -> Path | None:
    try:
        session_id = LATEST_UPLOAD_SESSION_FILE.read_text(encoding="utf-8").strip()
    except OSError:
        session_id = ""
    if session_id:
        candidate = UPLOAD_WORKSPACE_ROOT / session_id
        if is_valid_uploaded_workspace(candidate):
            return candidate

    if not UPLOAD_WORKSPACE_ROOT.exists():
        return None
    candidates = [p for p in UPLOAD_WORKSPACE_ROOT.iterdir() if is_valid_uploaded_workspace(p)]
    if not candidates:
        return None

    def latest_xlsx_mtime(path: Path) -> float:
        return max((p.stat().st_mtime for p in path.glob("*.xlsx")), default=0.0)

    return max(candidates, key=latest_xlsx_mtime)


def mark_latest_uploaded_workspace(session_id: str) -> None:
    try:
        UPLOAD_WORKSPACE_ROOT.mkdir(parents=True, exist_ok=True)
        LATEST_UPLOAD_SESSION_FILE.write_text(session_id, encoding="utf-8")
    except OSError:
        pass


def write_upload_workspace_signature(session_dir: Path, upload_signature: str) -> None:
    try:
        (session_dir / UPLOAD_SIGNATURE_FILE).write_text(upload_signature, encoding="utf-8")
    except OSError:
        pass


def hash_file_contents(path: Path) -> str:
    try:
        return hashlib.md5(path.read_bytes()).hexdigest()
    except OSError:
        return "-"


def build_staged_upload_signature(base_dir: Path) -> str:
    inv_staged = base_dir / "ODV_WIP_uploaded.xlsx"
    dem_staged = base_dir / "수요정보(전공정).xlsx"
    ref_staged = base_dir / "제품명 기준 정보.xlsx"
    finished_goods_staged = base_dir / FINISHED_GOODS_STOCK_UPLOAD_FILE
    if not inv_staged.exists() or not dem_staged.exists():
        return ""
    return "|".join(
        [
            hash_file_contents(inv_staged),
            hash_file_contents(dem_staged),
            hash_file_contents(ref_staged) if ref_staged.exists() else "-",
            hash_file_contents(finished_goods_staged) if finished_goods_staged.exists() else "-",
            "staged-upload",
        ]
    )


def read_upload_workspace_signature(base_dir: Path) -> str:
    try:
        signature = (base_dir / UPLOAD_SIGNATURE_FILE).read_text(encoding="utf-8").strip()
    except OSError:
        signature = ""
    if signature:
        return signature

    signature = build_staged_upload_signature(base_dir)
    if signature:
        write_upload_workspace_signature(base_dir, signature)
    return signature


def stage_uploaded_data_files(
    base_dir: Path,
    inventory_file,
    demand_file,
    reference_file=None,
    finished_goods_stock_file=None,
) -> Path:
    session_id = get_or_create_upload_session_id()
    session_dir = UPLOAD_WORKSPACE_ROOT / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    inv_bytes = bytes(inventory_file.getbuffer())
    dem_bytes = bytes(demand_file.getbuffer())
    ref_bytes = bytes(reference_file.getbuffer()) if reference_file is not None else None
    finished_goods_bytes = (
        bytes(finished_goods_stock_file.getbuffer()) if finished_goods_stock_file is not None else None
    )

    local_ref = find_product_name_reference_file(base_dir) if reference_file is None else None
    local_ref_sig = ""
    if local_ref is not None and local_ref.exists():
        stat = local_ref.stat()
        local_ref_sig = f"{local_ref.name}:{stat.st_size}:{stat.st_mtime_ns}"
    local_finished_goods = find_finished_goods_stock_file(base_dir) if finished_goods_stock_file is None else None
    local_finished_goods_sig = ""
    if local_finished_goods is not None and local_finished_goods.exists():
        stat = local_finished_goods.stat()
        local_finished_goods_sig = f"{local_finished_goods.name}:{stat.st_size}:{stat.st_mtime_ns}"

    upload_signature = "|".join(
        [
            hashlib.md5(inv_bytes).hexdigest(),
            hashlib.md5(dem_bytes).hexdigest(),
            hashlib.md5(ref_bytes).hexdigest() if ref_bytes is not None else "-",
            local_ref_sig,
            hashlib.md5(finished_goods_bytes).hexdigest() if finished_goods_bytes is not None else "-",
            local_finished_goods_sig,
        ]
    )
    signature_key = f"upload_workspace_signature_{session_id}"
    inv_staged = session_dir / "ODV_WIP_uploaded.xlsx"
    dem_staged = session_dir / "수요정보(전공정).xlsx"
    ref_staged = session_dir / "제품명 기준 정보.xlsx"
    finished_goods_staged = session_dir / FINISHED_GOODS_STOCK_UPLOAD_FILE
    if (
        st.session_state.get(signature_key) == upload_signature
        and inv_staged.exists()
        and dem_staged.exists()
        and (reference_file is not None or ref_staged.exists() or local_ref is None)
        and (
            finished_goods_stock_file is not None
            or finished_goods_staged.exists()
            or local_finished_goods is None
        )
    ):
        write_upload_workspace_signature(session_dir, upload_signature)
        mark_latest_uploaded_workspace(session_id)
        return session_dir

    for old_xlsx in session_dir.glob("*.xlsx"):
        old_xlsx.unlink(missing_ok=True)

    inv_staged.write_bytes(inv_bytes)
    dem_staged.write_bytes(dem_bytes)
    ref_dst = ref_staged
    if reference_file is not None:
        ref_dst.write_bytes(ref_bytes if ref_bytes is not None else b"")
    else:
        if local_ref is not None and local_ref.exists():
            shutil.copy2(local_ref, ref_dst)
    if finished_goods_stock_file is not None:
        finished_goods_staged.write_bytes(finished_goods_bytes if finished_goods_bytes is not None else b"")
    elif local_finished_goods is not None and local_finished_goods.exists():
        shutil.copy2(local_finished_goods, finished_goods_staged)

    st.session_state[signature_key] = upload_signature
    write_upload_workspace_signature(session_dir, upload_signature)
    mark_latest_uploaded_workspace(session_id)
    return session_dir


def sync_plan_api_data_mode() -> bool:
    api_configured = is_plan_api_configured()
    if api_configured:
        set_session_value("use_plan_api_data_mode", True)
    set_session_value("plan_api_key_available", api_configured)
    if not api_configured and get_session_value("use_plan_api_data_mode", False):
        set_session_value("use_plan_api_data_mode", False)
    return api_configured


def select_data_source(base_dir: Path) -> tuple[Path, str, str]:
    st.subheader("데이터 소스")
    api_key_source = get_plan_api_key_source_label()
    if api_key_source:
        st.caption(f"API 키 적용: {api_key_source}")
    else:
        st.caption("API 키를 자동으로 찾지 못했습니다. API_KEY_안내.txt의 X-API-Key를 아래에 붙여넣고 Enter를 누르세요.")
        st.text_input(
            "API Key",
            type="password",
            key="plan_api_key_input",
            placeholder="X-API-Key 붙여넣기",
            help="입력값은 현재 앱 세션에서만 사용합니다. 영구 적용은 PLAN_API_KEY 환경변수나 Streamlit secrets에 등록하세요.",
        )

    api_configured = sync_plan_api_data_mode()
    use_api = st.toggle(
        "APS API 자동조회 사용",
        key="use_plan_api_data_mode",
        disabled=not api_configured,
    )
    if use_api:
        render_plan_api_status()
        if api_configured:
            if st.button("APS API 새로고침", key="refresh_plan_api_data", use_container_width=True):
                set_session_value("plan_api_refresh_nonce", get_plan_api_refresh_nonce() + 1)
                st.cache_data.clear()
                st.cache_resource.clear()
                st.rerun()
            st.caption("APS 수요는 API 전체조회, WIP는 API 응답이 잘리거나 실패하면 기존 WIP 파일 기준으로 계산합니다.")
            api_updated_at = get_plan_api_updated_at()
            updated_at = api_updated_at if api_updated_at != "-" else get_data_updated_at(base_dir)
            return base_dir, "APS API + 로컬 기준정보", updated_at
    elif not api_configured:
        st.caption("API 키 미설정: 기존 파일 기준으로 표시합니다.")

    use_uploaded = st.toggle("업로드 파일 사용", value=False, key="use_uploaded_data_mode")
    inv_file = None
    dem_file = None
    ref_file = None
    finished_goods_stock_file = None
    if use_uploaded:
        inv_file = st.file_uploader("재고 파일(.xlsx)", type=["xlsx"], key="upload_inventory_xlsx")
        dem_file = st.file_uploader("수요 파일(.xlsx)", type=["xlsx"], key="upload_demand_xlsx")
        ref_file = st.file_uploader(
            "기준정보 파일(.xlsx, 선택)",
            type=["xlsx"],
            key="upload_reference_xlsx",
            help="미업로드 시 로컬의 '제품명 기준 정보.xlsx'를 사용합니다.",
        )
        finished_goods_stock_file = st.file_uploader(
            "완제품 재고 변화 파일(.xlsx, 선택)",
            type=["xlsx"],
            key="upload_finished_goods_stock_xlsx",
            help="미업로드 시 로컬/최근 업로드의 완제품 재고 변화 파일을 사용합니다.",
        )
    else:
        st.caption("현재 설정: 로컬 폴더의 파일 사용")

    if not use_uploaded:
        return base_dir, "로컬 파일", get_data_updated_at(base_dir)

    if inv_file is None and dem_file is None:
        latest_workspace = get_latest_uploaded_workspace()
        if latest_workspace is not None:
            st.warning("새 업로드 파일이 없어 최근 업로드 파일을 다시 사용합니다.")
            updated_at = get_data_updated_at(latest_workspace)
            return latest_workspace, f"최근 업로드 파일 ({latest_workspace.name})", updated_at

    if inv_file is None or dem_file is None:
        st.info("업로드 모드에서는 재고/수요 파일 2개 업로드가 필요합니다.")
        st.stop()

    workspace_dir = stage_uploaded_data_files(base_dir, inv_file, dem_file, ref_file, finished_goods_stock_file)
    uploaded_names = [inv_file.name, dem_file.name]
    if finished_goods_stock_file is not None:
        uploaded_names.append(finished_goods_stock_file.name)
    source_label = f"업로드 파일 ({', '.join(uploaded_names)})"
    updated_at = get_data_updated_at(workspace_dir)
    return workspace_dir, source_label, updated_at


def resolve_data_source_from_state(base_dir: Path) -> tuple[Path, str, str]:
    sync_plan_api_data_mode()
    if is_plan_api_enabled():
        api_updated_at = get_plan_api_updated_at()
        updated_at = api_updated_at if api_updated_at != "-" else get_data_updated_at(base_dir)
        return base_dir, "APS API + 로컬 기준정보", updated_at

    use_uploaded = bool(get_session_value("use_uploaded_data_mode", False))
    if not use_uploaded:
        return base_dir, "로컬 파일", get_data_updated_at(base_dir)

    inv_file = get_session_value("upload_inventory_xlsx")
    dem_file = get_session_value("upload_demand_xlsx")
    ref_file = get_session_value("upload_reference_xlsx")
    finished_goods_stock_file = get_session_value("upload_finished_goods_stock_xlsx")

    if inv_file is None and dem_file is None:
        latest_workspace = get_latest_uploaded_workspace()
        if latest_workspace is not None:
            updated_at = get_data_updated_at(latest_workspace)
            return latest_workspace, f"최근 업로드 파일 ({latest_workspace.name})", updated_at
        return base_dir, "업로드 파일 대기", get_data_updated_at(base_dir)

    if inv_file is None or dem_file is None:
        return base_dir, "업로드 파일 대기", get_data_updated_at(base_dir)

    workspace_dir = stage_uploaded_data_files(base_dir, inv_file, dem_file, ref_file, finished_goods_stock_file)
    uploaded_names = [inv_file.name, dem_file.name]
    if finished_goods_stock_file is not None:
        uploaded_names.append(finished_goods_stock_file.name)
    source_label = f"업로드 파일 ({', '.join(uploaded_names)})"
    updated_at = get_data_updated_at(workspace_dir)
    return workspace_dir, source_label, updated_at


def is_streamlit_cloud_runtime() -> bool:
    return (
        bool(os.environ.get("STREAMLIT_CLOUD"))
        or bool(os.environ.get("STREAMLIT_SHARING_MODE"))
        or Path("/mount/src").exists()
    )


def should_use_cloud_snapshots(data_base_dir: Path) -> bool:
    try:
        is_default_source = Path(data_base_dir).resolve() == BASE_DIR.resolve()
    except OSError:
        is_default_source = False
    live_data_override = os.environ.get("INTEROJO_USE_LIVE_DATA", "").strip().lower()
    return (
        is_default_source
        and not is_plan_api_enabled()
        and live_data_override not in {"1", "true", "yes", "on"}
        and (CLOUD_SNAPSHOT_DIR / "shortage_snapshot.csv.gz").exists()
    )


def build_cloud_snapshot_refresh_key(*names: str) -> str:
    parts: list[str] = []
    for name in names:
        path = CLOUD_SNAPSHOT_DIR / name
        try:
            stat = path.stat()
            parts.append(f"{name}:{stat.st_size}:{stat.st_mtime_ns}")
        except OSError:
            parts.append(f"{name}:missing")
    return "|".join(parts)


@st.cache_data(show_spinner=False)
def read_cloud_snapshot_csv(name: str, refresh_key: str) -> pd.DataFrame:
    _ = refresh_key
    path = CLOUD_SNAPSHOT_DIR / name
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, encoding="utf-8-sig", compression="infer")


def load_cloud_snapshot_csv(name: str) -> pd.DataFrame:
    refresh_key = build_cloud_snapshot_refresh_key(name)
    return read_cloud_snapshot_csv(name, refresh_key)


def get_cloud_snapshot_meta_value(key: str, default: str = "-") -> str:
    meta = load_cloud_snapshot_csv("snapshot_meta.csv")
    if meta.empty or not {"key", "value"}.issubset(meta.columns):
        return default
    values = meta.loc[meta["key"].astype(str) == key, "value"]
    if values.empty:
        return default
    return str(values.iloc[0])


def parse_updated_at_value(value: object) -> datetime | None:
    text = str(value or "").strip()
    if not text or text == "-":
        return None
    try:
        return datetime.strptime(text[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=DISPLAY_TZ)
    except ValueError:
        return None


def is_cloud_snapshot_fresh(meta_key: str, live_updated_at: str) -> bool:
    cloud_dt = parse_updated_at_value(get_cloud_snapshot_meta_value(meta_key, "-"))
    live_dt = parse_updated_at_value(live_updated_at)
    if cloud_dt is None or live_dt is None:
        return False
    return cloud_dt.timestamp() + 1 >= live_dt.timestamp()


def load_cloud_shortage_snapshot() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    return (
        load_cloud_snapshot_csv("shortage_snapshot.csv.gz"),
        load_cloud_snapshot_csv("shortage_file_info.csv.gz"),
        load_cloud_snapshot_csv("process_map.csv.gz"),
    )


def load_cloud_inventory_risk_snapshot() -> pd.DataFrame:
    return load_cloud_snapshot_csv("inventory_risk_snapshot.csv.gz")


def load_cloud_all_item_status_snapshot() -> tuple[pd.DataFrame, pd.DataFrame]:
    return (
        load_cloud_snapshot_csv(ALL_ITEM_SNAPSHOT_FILE),
        load_cloud_snapshot_csv(CODE_MISMATCH_SNAPSHOT_FILE),
    )


def build_all_item_status_disk_cache_path(refresh_key: str) -> Path:
    cache_hash = build_local_cache_hash(APP_CACHE_VERSION, "all_item_status", refresh_key)
    return ALL_ITEM_STATUS_DISK_CACHE_DIR / f"{cache_hash}.pkl.gz"


def read_all_item_status_disk_cache(refresh_key: str) -> tuple[pd.DataFrame, pd.DataFrame] | None:
    cache_path = build_all_item_status_disk_cache_path(refresh_key)
    if not cache_path.exists():
        return None
    try:
        cached = pd.read_pickle(cache_path, compression="gzip")
    except Exception:
        return None
    if not isinstance(cached, dict):
        return None
    all_items = cached.get("all_items")
    code_mismatch = cached.get("code_mismatch")
    if not isinstance(all_items, pd.DataFrame) or not isinstance(code_mismatch, pd.DataFrame):
        return None
    if DEMAND_DETAIL_ROWS_COL not in all_items.columns:
        return None
    return all_items, code_mismatch


def write_all_item_status_disk_cache(
    refresh_key: str,
    all_items: pd.DataFrame,
    code_mismatch: pd.DataFrame,
) -> None:
    if all_items.empty:
        return
    cache_path = build_all_item_status_disk_cache_path(refresh_key)
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = cache_path.with_suffix(cache_path.suffix + f".{uuid.uuid4().hex}.tmp")
        pd.to_pickle(
            {"all_items": all_items, "code_mismatch": code_mismatch},
            temp_path,
            compression="gzip",
        )
        temp_path.replace(cache_path)
    except Exception:
        pass


def build_all_item_flow_status_disk_cache_path(refresh_key: str, site_filter: str) -> Path:
    cache_hash = build_local_cache_hash(APP_CACHE_VERSION, "all_item_flow_status", refresh_key, site_filter)
    return ALL_ITEM_FLOW_STATUS_DISK_CACHE_DIR / f"{cache_hash}.pkl.gz"


def read_all_item_flow_status_disk_cache(refresh_key: str, site_filter: str) -> pd.DataFrame | None:
    cache_path = build_all_item_flow_status_disk_cache_path(refresh_key, site_filter)
    if not cache_path.exists():
        return None
    try:
        cached = pd.read_pickle(cache_path, compression="gzip")
    except Exception:
        return None
    if not isinstance(cached, pd.DataFrame):
        return None
    if DEMAND_DETAIL_ROWS_COL not in cached.columns:
        return None
    return cached


def write_all_item_flow_status_disk_cache(refresh_key: str, site_filter: str, all_items: pd.DataFrame) -> None:
    if all_items.empty:
        return
    cache_path = build_all_item_flow_status_disk_cache_path(refresh_key, site_filter)
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = cache_path.with_suffix(cache_path.suffix + f".{uuid.uuid4().hex}.tmp")
        all_items.to_pickle(temp_path, compression="gzip")
        temp_path.replace(cache_path)
    except Exception:
        pass


def load_cloud_leadji_status_snapshot() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    return (
        load_cloud_snapshot_csv("leadji_shortage_snapshot.csv.gz"),
        load_cloud_snapshot_csv("leadji_info.csv.gz"),
        load_cloud_snapshot_csv("leadji_stock.csv.gz"),
        load_cloud_snapshot_csv("leadji_order.csv.gz"),
    )


def pick_first_existing_column(columns: list[str], candidates: list[str]) -> str | None:
    for col in candidates:
        if col in columns:
            return col
    return None


def parse_mixed_excel_date(series: pd.Series) -> pd.Series:
    """Parse mixed date inputs safely, including Excel serial dates."""
    text = series.astype(str).str.strip()
    invalid_tokens = {"", "nan", "none", "nat"}
    cleaned = series.where(~text.str.lower().isin(invalid_tokens), pd.NA)

    # First pass: strings/datetime objects.
    parsed = pd.to_datetime(cleaned, errors="coerce")

    # Second pass: Excel serial numbers (days since 1899-12-30).
    # IMPORTANT: only treat true numeric-like cells as serials.
    # Datetime64 values can be converted to large integers (ns) by to_numeric,
    # which would overwrite valid dates with NaT if we don't filter first.
    obj = cleaned.astype("object")
    numeric_like = obj.map(lambda v: isinstance(v, (int, float)) and not isinstance(v, bool))
    numeric_text = text.str.fullmatch(r"[+-]?\d+(?:\.\d+)?").fillna(False)
    numeric_mask_source = numeric_like | numeric_text

    numeric = pd.to_numeric(obj.where(numeric_mask_source), errors="coerce")
    numeric_mask = numeric.notna() & (numeric > 0)
    if numeric_mask.any():
        parsed.loc[numeric_mask] = pd.to_datetime(
            numeric.loc[numeric_mask], unit="D", origin="1899-12-30", errors="coerce"
        )

    return parsed


def parse_mixed_numeric(series: pd.Series) -> pd.Series:
    """Parse mixed numeric inputs safely (number/string/comma text)."""
    text = series.astype(str).str.strip()
    invalid_tokens = {"", "nan", "none", "nat", "-"}
    normalized = text.where(~text.str.lower().isin(invalid_tokens), pd.NA)

    # Support accounting-style negatives like "(1,234)".
    normalized = normalized.str.replace(r"^\((.*)\)$", r"-\1", regex=True)
    normalized = normalized.str.replace(",", "", regex=False)
    normalized = normalized.str.replace("\u00a0", "", regex=False).str.replace(" ", "", regex=False)

    return pd.to_numeric(normalized, errors="coerce").fillna(0)


def join_unique_text_values(series: pd.Series) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for value in series:
        for item in str(value).split(","):
            text = item.strip()
            if not text or text.lower() in {"nan", "none", "null", "nat", "<na>"}:
                continue
            if text not in seen:
                seen.add(text)
                values.append(text)
    return ", ".join(values) if values else "-"


def is_inspection_rework_wip_code(value: object) -> bool:
    code = str(value).strip().upper()
    if not code or code.lower() in INVALID_CATEGORY_VALUES:
        return False
    return bool(re.search(r"-C(?:\d*)$", code))


def canonicalize_warehouse_label(raw_label: str) -> str:
    label = str(raw_label).strip()
    if not label or label.lower() == "nan":
        return ""

    if label in WAREHOUSE_MAP:
        return label

    display_to_key = {display: key for key, display in WAREHOUSE_MAP.items()}
    if label in display_to_key:
        return display_to_key[label]

    normalized = normalize_process_to_warehouse(label)
    if normalized is None:
        return label.replace(" ", "")

    return display_to_key.get(normalized, normalized)


def build_inventory_df(inv: pd.DataFrame) -> pd.DataFrame:
    inv = inv.copy()
    inv.columns = [str(c).strip() for c in inv.columns]
    columns = inv.columns.tolist()
    output_columns = ["품목코드", "창고", "재공코드", "재고량"]
    if inv.empty or not columns:
        return pd.DataFrame(columns=output_columns)

    def column_as_series(column_name: str) -> pd.Series:
        selected = inv[column_name]
        if isinstance(selected, pd.DataFrame):
            return selected.iloc[:, 0]
        return selected

    qty_col = pick_api_column(columns, ["총 재공 수량", "재공수량", "재고수량", "WIP_QTY", "WIP수량", "QTY", "재고량"])
    item_col = pick_api_column(columns, ["제품 코드", "ITEM_ID", "ITEM_CODE", "ITEM_CD", "제품코드", "품목코드"])
    warehouse_col = pick_api_column(
        columns,
        ["WH_NAME", "창고명", "공정(버퍼)", "공정", "버퍼 코드", "BUFFER_CODE", "제품위치(창고)", "PROP02", "창고"],
    )
    wip_code_col = pick_api_column(columns, ["재공 코드", "재공코드", "WIP_CODE", "WIP ID", "WIP_ID", "수요ID"])

    # Fallback for unknown layouts
    if qty_col is None:
        qty_col = columns[6] if len(columns) > 6 else columns[0]
    if item_col is None:
        item_col = columns[8] if len(columns) > 8 else (columns[1] if len(columns) > 1 else columns[0])
    if warehouse_col is None:
        warehouse_col = (
            columns[23]
            if len(columns) > 23
            else (columns[10] if len(columns) > 10 else (columns[5] if len(columns) > 5 else columns[0]))
        )
    if wip_code_col is None:
        wip_code_col = columns[3] if len(columns) > 3 else item_col

    inv_df = pd.DataFrame(
        {
            "품목코드": column_as_series(item_col).astype(str).str.strip(),
            "창고": column_as_series(warehouse_col).astype(str).str.strip().map(canonicalize_warehouse_label),
            "재공코드": column_as_series(wip_code_col).astype(str).str.strip(),
            "재고량": parse_mixed_numeric(column_as_series(qty_col)),
        }
    )
    rework_mask = (inv_df["창고"] == "검사접착") & inv_df["재공코드"].map(is_inspection_rework_wip_code)
    inv_df.loc[rework_mask, "창고"] = "검사접착재작업"

    inv_df = inv_df[(inv_df["품목코드"] != "") & (inv_df["품목코드"].str.lower() != "nan")]
    inv_df = inv_df[(inv_df["창고"] != "") & (inv_df["창고"].str.lower() != "nan")]
    return inv_df


def normalize_process_to_warehouse(process_label: str) -> str | None:
    label = str(process_label).replace(" ", "")
    if "사출" in label:
        return "사출창고"
    if "분리" in label:
        return "분리창고"
    if "검사접착" in label or ("검사" in label and "접착" in label) or "접착" in label or "멸균" in label:
        return "검사접착창고"
    if "누수" in label or "규격검사" in label:
        return "누수규격검사 창고"
    return None


def extract_demand_header_info(dem_path: Path) -> tuple[
    dict[str, str], dict[str, int], list[int], list[int], dict[str, int], dict[int, str]
]:
    try:
        wb = openpyxl.load_workbook(dem_path, read_only=True, data_only=True)
    except Exception:
        return {}, {}, [], [], {}, {}
    try:
        demand_sheet = find_workbook_sheet_name(wb, [DEMAND_DATA_SHEET_NAME])
        if demand_sheet is None:
            return {}, {}, [], [], {}, {}
        ws = wb[demand_sheet]
        rows = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        header_values = [list(row) for row in rows]
    finally:
        wb.close()

    if len(header_values) < 2:
        return {}, {}, [], [], {}, {}
    max_len = max(len(row) for row in header_values)
    header_values = [row + [None] * (max_len - len(row)) for row in header_values]
    header_rows = pd.DataFrame(header_values)

    top_row = header_rows.iloc[0]
    second_row = header_rows.iloc[1]
    header_labels = {int(idx): str(value).strip() for idx, value in second_row.items()}

    code_map: dict[str, str] = {}
    warehouse_qty_col_indices: dict[str, int] = {}
    qty_col_indices: list[int] = []
    total_qty_col_indices: list[int] = []
    process_qty_col_indices: dict[str, int] = {}

    for idx, column_name in second_row.items():
        if "생산 수량" not in str(column_name):
            continue

        idx = int(idx)
        qty_col_indices.append(idx)

        top_label = str(top_row.iloc[idx]).strip()
        if not top_label or top_label.lower() == "nan":
            continue
        process_qty_col_indices[top_label.replace(" ", "")] = idx
        if "총합계" in top_label:
            total_qty_col_indices.append(idx)
            continue

        warehouse_name = normalize_process_to_warehouse(top_label)
        if not warehouse_name:
            continue

        match = re.search(r"\[(.*?)\]", top_label)
        extracted_code = match.group(1).strip() if match else top_label
        code_map[warehouse_name] = extracted_code
        warehouse_qty_col_indices[warehouse_name] = idx

    return (
        code_map,
        warehouse_qty_col_indices,
        qty_col_indices,
        total_qty_col_indices,
        process_qty_col_indices,
        header_labels,
    )


def pick_first_existing_column_index(header_labels: dict[int, str], candidates: list[str]) -> int | None:
    for candidate in candidates:
        for idx, label in header_labels.items():
            if label == candidate:
                return idx
    return None


def build_demand_read_plan(
    header_labels: dict[int, str],
    warehouse_qty_col_indices: dict[str, int],
    qty_col_indices: list[int],
    total_qty_col_indices: list[int],
    process_qty_col_indices: dict[str, int],
) -> dict[str, object]:
    column_count = len(header_labels)

    def existing_idx(idx: int | None) -> int | None:
        return idx if idx is not None and 0 <= idx < column_count else None

    site_col_idx = pick_first_existing_column_index(
        header_labels,
        ["설비 사이트 코드", "설비사이트코드", "사이트코드"],
    )
    customer_col_idx = pick_first_existing_column_index(
        header_labels,
        ["고객 이름", "고객이름", "거래처"],
    )
    order_no_col_idx = pick_first_existing_column_index(
        header_labels,
        ["수주번호", "수주 번호", "오더번호", "오더 번호"],
    )
    initial_col_idx = pick_first_existing_column_index(header_labels, ["이니셜"])
    demand_item_col_idx = pick_first_existing_column_index(
        header_labels,
        ["제품 코드", "제품코드", "품목코드", "ITEM_ID"],
    )
    demand_name_col_idx = pick_first_existing_column_index(
        header_labels,
        ["수요 제품 이름", "수요제품이름", "제품명"],
    )
    demand_qty_idx = pick_first_existing_column_index(
        header_labels,
        ["수요 수량", "수요수량"],
    )

    leak_qty_idx = existing_idx(warehouse_qty_col_indices.get("누수규격검사 창고"))
    leak_due_idx = existing_idx(leak_qty_idx + 1 if leak_qty_idx is not None else None)
    separation_qty_idx = existing_idx(warehouse_qty_col_indices.get("분리창고"))
    separation_due_idx = existing_idx(separation_qty_idx + 1 if separation_qty_idx is not None else None)
    adhesion_qty_idx = existing_idx(warehouse_qty_col_indices.get("검사접착창고"))
    adhesion_due_idx = existing_idx(adhesion_qty_idx + 1 if adhesion_qty_idx is not None else None)

    leadji_qty_idx: int | None = None
    for process_label, idx in process_qty_col_indices.items():
        if "[45]" in process_label and ("하이드레이션" in process_label or "전면검사" in process_label):
            leadji_qty_idx = existing_idx(idx)
            break
    if leadji_qty_idx is None:
        for process_label, idx in process_qty_col_indices.items():
            if "하이드레이션" in process_label or "전면검사" in process_label:
                leadji_qty_idx = existing_idx(idx)
                break
    leadji_due_idx = existing_idx(leadji_qty_idx + 1 if leadji_qty_idx is not None else None)

    selected_qty_idx = existing_idx(warehouse_qty_col_indices.get("사출창고"))
    if selected_qty_idx is None:
        selected_qty_idx = pick_first_existing_column_index(
            header_labels,
            ["사출조립 생산수량", "사출조립생산수량", "사출조립 생산 수량"],
        )
    if selected_qty_idx is None:
        selected_qty_idx = existing_idx(5)
    inj_due_idx = existing_idx(selected_qty_idx + 1 if selected_qty_idx is not None else None)

    usecols: set[int] = set()
    for idx in [
        site_col_idx if site_col_idx is not None else existing_idx(0),
        customer_col_idx if customer_col_idx is not None else existing_idx(1),
        order_no_col_idx if order_no_col_idx is not None else existing_idx(2),
        initial_col_idx if initial_col_idx is not None else existing_idx(3),
        demand_item_col_idx if demand_item_col_idx is not None else existing_idx(4),
        demand_name_col_idx if demand_name_col_idx is not None else existing_idx(5),
        demand_qty_idx if demand_qty_idx is not None else existing_idx(6),
        leak_qty_idx,
        leak_due_idx,
        separation_qty_idx,
        separation_due_idx,
        adhesion_qty_idx,
        adhesion_due_idx,
        leadji_qty_idx,
        leadji_due_idx,
        selected_qty_idx,
        inj_due_idx,
    ]:
        if idx is not None:
            usecols.add(idx)

    if leak_qty_idx is None:
        if total_qty_col_indices:
            idx = existing_idx(total_qty_col_indices[-1])
            if idx is not None:
                usecols.add(idx)
        else:
            usecols.update(idx for idx in (existing_idx(i) for i in qty_col_indices) if idx is not None)

    return {
        "header_labels": header_labels,
        "site_col_idx": site_col_idx if site_col_idx is not None else existing_idx(0),
        "customer_col_idx": customer_col_idx if customer_col_idx is not None else existing_idx(1),
        "order_no_col_idx": order_no_col_idx if order_no_col_idx is not None else existing_idx(2),
        "initial_col_idx": initial_col_idx if initial_col_idx is not None else existing_idx(3),
        "demand_item_col_idx": demand_item_col_idx if demand_item_col_idx is not None else existing_idx(4),
        "demand_name_col_idx": demand_name_col_idx if demand_name_col_idx is not None else existing_idx(5),
        "demand_qty_idx": demand_qty_idx if demand_qty_idx is not None else existing_idx(6),
        "leak_qty_idx": leak_qty_idx,
        "leak_due_idx": leak_due_idx,
        "separation_qty_idx": separation_qty_idx,
        "separation_due_idx": separation_due_idx,
        "adhesion_qty_idx": adhesion_qty_idx,
        "adhesion_due_idx": adhesion_due_idx,
        "leadji_qty_idx": leadji_qty_idx,
        "leadji_due_idx": leadji_due_idx,
        "selected_qty_idx": selected_qty_idx,
        "inj_due_idx": inj_due_idx,
        "qty_col_indices": [i for i in qty_col_indices if existing_idx(i) is not None],
        "total_qty_col_indices": [i for i in total_qty_col_indices if existing_idx(i) is not None],
        "usecols": sorted(usecols),
    }


def build_dashboard_cache_path(source_path: Path, prefix: str, *key_parts: object) -> Path:
    try:
        stat = source_path.stat()
        file_key = f"{source_path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    except OSError:
        file_key = source_path.name
    raw_key = "|".join([prefix, file_key, *(str(part) for part in key_parts)])
    cache_key = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()[:24]
    return source_path.resolve().parent / ".dashboard_cache" / f"{prefix}_{cache_key}.pkl"


def read_pickle_cache(cache_path: Path) -> object | None:
    try:
        if cache_path.exists():
            with cache_path.open("rb") as f:
                return pickle.load(f)
    except Exception:
        return None
    return None


def write_pickle_cache(cache_path: Path, value: object) -> None:
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        with cache_path.open("wb") as f:
            pickle.dump(value, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception:
        pass


def read_inventory_excel_subset(inv_path: Path) -> pd.DataFrame:
    cache_path = build_dashboard_cache_path(inv_path, "inventory_subset", "v4")
    cached = read_pickle_cache(cache_path)
    if isinstance(cached, pd.DataFrame):
        return cached.copy()

    try:
        wb = openpyxl.load_workbook(inv_path, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()

    try:
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        columns: list[str] = []
        header_found = False
        qty_col = item_col = warehouse_col = wip_code_col = None
        for row_idx, row in enumerate(rows):
            columns = [str(c).strip() if c is not None else "" for c in row]
            if not any(columns):
                continue

            qty_col = pick_first_existing_column(columns, ["총 재공 수량", "WIP_QTY", "재고량"])
            item_col = pick_first_existing_column(columns, ["제품 코드", "ITEM_ID", "제품코드", "품목코드"])
            warehouse_col = pick_first_existing_column(columns, ["WH_NAME", "창고명", "버퍼 코드", "제품위치(창고)", "PROP02", "창고"])
            wip_code_col = pick_first_existing_column(columns, ["재공 코드", "재공코드", "WIP_CODE", "WIP ID", "WIP_ID"])
            if qty_col is not None and item_col is not None and warehouse_col is not None:
                header_found = True
                break
            if row_idx >= 50:
                break

        if not header_found or not columns:
            return pd.DataFrame()

        qty_idx = columns.index(qty_col) if qty_col in columns else (6 if len(columns) > 6 else 0)
        item_idx = columns.index(item_col) if item_col in columns else (8 if len(columns) > 8 else (1 if len(columns) > 1 else 0))
        warehouse_idx = (
            columns.index(warehouse_col)
            if warehouse_col in columns
            else (
                23
                if len(columns) > 23
                else (10 if len(columns) > 10 else (5 if len(columns) > 5 else 0))
            )
        )
        wip_code_idx = columns.index(wip_code_col) if wip_code_col in columns else (3 if len(columns) > 3 else item_idx)
        optional_cols = [
            "LOT_NO",
            "사용가능한 날짜",
            "생성 일시",
            "수정 일시",
            "BASE_SALES_CD",
            "재공 제품 구분",
        ]
        optional_indices = {columns.index(col) for col in optional_cols if col in columns}
        usecols = sorted({qty_idx, item_idx, warehouse_idx, wip_code_idx, *optional_indices})
        selected_columns = [columns[i] for i in usecols]

        records: list[list[object]] = []
        for row in rows:
            selected = [row[i] if i < len(row) else None for i in usecols]
            if any(value is not None and str(value).strip() != "" for value in selected):
                records.append(selected)

        result = pd.DataFrame.from_records(records, columns=selected_columns)
        write_pickle_cache(cache_path, result)
        return result
    finally:
        wb.close()


def read_demand_excel_subset(dem_path: Path, usecols: list[int]) -> pd.DataFrame:
    if not usecols:
        return pd.DataFrame()
    try:
        wb = openpyxl.load_workbook(dem_path, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()
    try:
        demand_sheet = find_workbook_sheet_name(wb, [DEMAND_DATA_SHEET_NAME])
    finally:
        wb.close()
    if demand_sheet is None:
        return pd.DataFrame()

    cache_path = build_dashboard_cache_path(
        dem_path,
        "demand_subset",
        "v3",
        f"sheet:{demand_sheet}",
        ",".join(map(str, usecols)),
    )
    cached = read_pickle_cache(cache_path)
    if isinstance(cached, pd.DataFrame):
        return cached.copy()

    dem = pd.read_excel(dem_path, sheet_name=demand_sheet, header=None, skiprows=2, usecols=usecols)
    if len(dem.columns) == len(usecols):
        dem.columns = usecols
    write_pickle_cache(cache_path, dem)
    return dem


def map_demand_code_to_process_code(demand_code: str, process_prefix: str) -> str:
    code = str(demand_code).strip()
    if not code or code.lower() == "nan":
        return code

    letter_pattern = re.match(r"^P(\d{4})([A-Z])(.*)$", code)
    if letter_pattern:
        return f"{process_prefix}{letter_pattern.group(1)}{letter_pattern.group(3)}"
    if code.startswith("P"):
        return f"{process_prefix}{code[1:]}"
    if code[0] in {"Q", "R"} and len(code) > 1:
        return f"{process_prefix}{code[1:]}"
    return code


def normalize_rework_match_value(value: object) -> str:
    text = str(value).strip().upper()
    if not text or text.lower() in INVALID_CATEGORY_VALUES:
        return ""
    return text


def parse_single_numeric_value(value: object) -> float:
    return float(parse_mixed_numeric(pd.Series([value])).iat[0])


def build_empty_rework_meta() -> dict[str, object]:
    return {
        "sheet": "-",
        "source_path": "",
        "match_scope": "none",
        "initial_col": "",
        "product_col": "",
        "quantity_col": "",
        "sheet_columns": [],
        "source_rows": 0,
        "source_qty_total": 0.0,
        "note_col": "",
        "note_count": 0,
        "remarks_by_key": {},
        "demand_qty_by_key": {},
        "remarks_by_demand_key": {},
        "demand_key_count": 0,
    }


def normalize_rework_date_value(value: object) -> str:
    text = str(value).strip()
    if not text or text.lower() in INVALID_CATEGORY_VALUES:
        return ""

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%Y-%m-%d")
    return text


def build_rework_demand_key(
    key_type: str,
    customer: object,
    initial: object,
    product_code: object,
    r_code: object,
    q_code: object,
    due_date: object,
) -> tuple[str, ...]:
    customer_key = normalize_rework_match_value(customer)
    initial_key = normalize_rework_match_value(initial)
    product_key = normalize_rework_match_value(product_code)
    r_key = normalize_rework_match_value(r_code)
    q_key = normalize_rework_match_value(q_code)
    due_key = normalize_rework_date_value(due_date)

    if key_type == "full":
        return (key_type, customer_key, initial_key, product_key, r_key, q_key, due_key)
    if key_type == "customer_item_due":
        return (key_type, customer_key, initial_key, product_key, due_key)
    if key_type == "item_due":
        return (key_type, initial_key, product_key, due_key)
    return (key_type, initial_key, product_key)


def build_rework_demand_key_variants(
    customer: object,
    initial: object,
    product_code: object,
    r_code: object,
    q_code: object,
    due_date: object,
) -> list[tuple[str, ...]]:
    return [
        build_rework_demand_key("full", customer, initial, product_code, r_code, q_code, due_date),
        build_rework_demand_key("customer_item_due", customer, initial, product_code, r_code, q_code, due_date),
        build_rework_demand_key("item_due", customer, initial, product_code, r_code, q_code, due_date),
    ]


def build_rework_demand_key_series(df: pd.DataFrame, key_type: str) -> pd.Series:
    def column_values(column_name: str) -> pd.Series:
        if column_name in df.columns:
            return df[column_name]
        return pd.Series([""] * len(df), index=df.index)

    return pd.Series(
        [
            build_rework_demand_key(key_type, customer, initial, product_code, r_code, q_code, due_date)
            for customer, initial, product_code, r_code, q_code, due_date in zip(
                column_values("거래처"),
                column_values("이니셜"),
                column_values("품목코드"),
                column_values("R코드"),
                column_values("Q코드"),
                column_values("납기일"),
            )
        ],
        index=df.index,
        dtype="object",
    )


def format_rework_match_sample_key(key: object) -> str:
    if isinstance(key, tuple) and key:
        key_type = key[0]
        if key_type == "full" and len(key) >= 7:
            _, customer, initial, product_code, _, _, due_date = key
            parts = [customer, initial, product_code, due_date]
        elif key_type == "customer_item_due" and len(key) >= 5:
            _, customer, initial, product_code, due_date = key
            parts = [customer, initial, product_code, due_date]
        elif key_type == "item_due" and len(key) >= 4:
            _, initial, product_code, due_date = key
            parts = [initial, product_code, due_date]
        else:
            parts = [str(part) for part in key[1:]]
        return " / ".join(str(part) for part in parts if str(part).strip())
    return str(key).strip()


def apply_rework_flags_to_demand_rows(
    demand_df: pd.DataFrame,
    rework_item_qty_map: dict[tuple[str, str], float],
    rework_meta: dict[str, object],
) -> tuple[pd.DataFrame, list[str], float]:
    result = demand_df.copy()
    rework_demand_qty_by_key = rework_meta.get("demand_qty_by_key", {})
    rework_remarks_by_demand_key = rework_meta.get("remarks_by_demand_key", {})
    if not isinstance(rework_demand_qty_by_key, dict):
        rework_demand_qty_by_key = {}
    if not isinstance(rework_remarks_by_demand_key, dict):
        rework_remarks_by_demand_key = {}

    if rework_demand_qty_by_key:
        demand_key_set = set(rework_demand_qty_by_key.keys())
        rework_available_mask = pd.Series(False, index=result.index)
        rework_available_qty = pd.Series(0.0, index=result.index)
        rework_note = pd.Series("", index=result.index, dtype="object")
        rework_sample_keys = pd.Series([None] * len(result), index=result.index, dtype="object")

        for key_type in ["full", "customer_item_due", "item_due"]:
            demand_keys = build_rework_demand_key_series(result, key_type)
            match_mask = (~rework_available_mask) & demand_keys.isin(demand_key_set)
            if not match_mask.any():
                continue
            matched_keys = demand_keys.loc[match_mask]
            rework_available_mask.loc[match_mask] = True
            rework_available_qty.loc[match_mask] = matched_keys.map(
                lambda key: rework_demand_qty_by_key.get(key, 0.0)
            ).fillna(0.0)
            rework_note.loc[match_mask] = matched_keys.map(lambda key: rework_remarks_by_demand_key.get(key, "")).fillna(
                ""
            )
            rework_sample_keys.loc[match_mask] = matched_keys.to_numpy(dtype=object)
    else:
        rework_lookup_initials = result.get("이니셜", pd.Series([""] * len(result), index=result.index)).map(
            normalize_rework_match_value
        )
        rework_lookup_item_codes = result.get("품목코드", pd.Series([""] * len(result), index=result.index)).map(
            normalize_rework_match_value
        )
        rework_lookup_keys = pd.Series(
            list(zip(rework_lookup_initials, rework_lookup_item_codes)),
            index=result.index,
        )
        rework_item_only_keys = pd.Series(
            list(zip([""] * len(rework_lookup_item_codes), rework_lookup_item_codes)),
            index=result.index,
        )
        rework_key_set = set(rework_item_qty_map.keys())
        rework_remarks_by_key = rework_meta.get("remarks_by_key", {})
        if not isinstance(rework_remarks_by_key, dict):
            rework_remarks_by_key = {}
        rework_exact_match_mask = rework_lookup_keys.isin(rework_key_set)
        rework_item_only_match_mask = rework_item_only_keys.isin(rework_key_set)
        rework_available_mask = rework_exact_match_mask | rework_item_only_match_mask
        rework_exact_qty = rework_lookup_keys.map(lambda key: rework_item_qty_map.get(key, 0.0)).fillna(0.0)
        rework_item_only_qty = rework_item_only_keys.map(lambda key: rework_item_qty_map.get(key, 0.0)).fillna(0.0)
        rework_available_qty = rework_exact_qty.where(rework_exact_match_mask, rework_item_only_qty)
        rework_exact_note = rework_lookup_keys.map(lambda key: rework_remarks_by_key.get(key, "")).fillna("")
        rework_item_only_note = rework_item_only_keys.map(lambda key: rework_remarks_by_key.get(key, "")).fillna("")
        rework_note = rework_exact_note.where(rework_exact_note.astype(str).str.strip().ne(""), rework_item_only_note)
        rework_sample_keys = rework_lookup_keys.where(rework_exact_match_mask, rework_item_only_keys)

    result[REWORK_AVAILABLE_QTY_COL] = rework_available_qty
    result["재작업"] = rework_available_mask.map({True: "재작업", False: ""})
    result["비고"] = rework_note.where(
        rework_note.astype(str).str.strip().ne(""),
        result["재작업"],
    )
    result["비고"] = clean_display_text_series(result["비고"])
    matched_keys = sorted(
        {
            format_rework_match_sample_key(key)
            for key in rework_sample_keys[rework_available_mask].tolist()
            if format_rework_match_sample_key(key)
        }
    )
    matched_qty_total = float(rework_available_qty[rework_available_mask].sum())
    return result, matched_keys, matched_qty_total


def find_rework_quantity_column_index(header: list[str]) -> tuple[int | None, str]:
    normalized = [str(label).replace(" ", "").strip() for label in header]
    preferred_names = ["이동요청수량", "재작업가능수량", "재작업수량", "가능수량"]
    for preferred in preferred_names:
        if preferred in normalized:
            idx = normalized.index(preferred)
            return idx, header[idx]

    move_request_indices = [idx for idx, label in enumerate(normalized) if label == "이동요청"]
    if len(move_request_indices) >= 2:
        idx = move_request_indices[-1]
        return idx, f"{header[idx]}(수량)"

    fallback_names = ["잔량", "전산재고"]
    for fallback in fallback_names:
        if fallback in normalized:
            idx = normalized.index(fallback)
            return idx, header[idx]

    return None, ""


def find_rework_note_column_index(header: list[str]) -> tuple[int | None, str]:
    normalized = [str(label).replace(" ", "").strip().upper() for label in header]
    preferred_names = ["비고", "비고란", "메모", "참고", "REMARK", "REMARKS", "NOTE", "NOTES", "COMMENT", "COMMENTS"]
    for preferred in preferred_names:
        if preferred in normalized:
            idx = normalized.index(preferred)
            return idx, header[idx]
    return None, ""


def normalize_rework_note_value(value: object) -> str:
    text = str(value).strip()
    return "" if text.lower() in INVALID_CATEGORY_VALUES else text


def clean_display_text_series(series: pd.Series) -> pd.Series:
    text = series.astype(str).str.strip()
    return text.mask(text.str.lower().isin(INVALID_CATEGORY_VALUES), "")


def merge_rework_note(existing: str, note: str) -> str:
    note = normalize_rework_note_value(note)
    if not note:
        return existing
    parts = [part.strip() for part in str(existing).split(" / ") if part.strip()]
    if note not in parts:
        parts.append(note)
    return " / ".join(parts)


def find_rework_header_info(ws) -> tuple[list[str], object, str | None, str | None, int | None, str, int | None, str]:
    rows = ws.iter_rows(values_only=True)
    last_header: list[str] = []
    last_product_col: str | None = None
    last_initial_col: str | None = None
    last_quantity_idx: int | None = None
    last_quantity_col = ""
    last_note_idx: int | None = None
    last_note_col = ""
    product_candidates = ["제품코드", "제품 코드", "품목코드", "품목 코드", "생산코드", "생산 코드"]

    for row_idx, row in enumerate(rows):
        header = [str(c).strip() if c is not None else "" for c in row]
        product_col = pick_first_existing_column(header, product_candidates)
        initial_col = pick_first_existing_column(header, ["이니셜"])
        quantity_idx, quantity_col = find_rework_quantity_column_index(header)
        note_idx, note_col = find_rework_note_column_index(header)
        if product_col is not None:
            return header, rows, initial_col, product_col, quantity_idx, quantity_col, note_idx, note_col

        if any(header):
            last_header = header
            last_product_col = product_col
            last_initial_col = initial_col
            last_quantity_idx = quantity_idx
            last_quantity_col = quantity_col
            last_note_idx = note_idx
            last_note_col = note_col
        if row_idx >= 30:
            break

    return (
        last_header,
        rows,
        last_initial_col,
        last_product_col,
        last_quantity_idx,
        last_quantity_col,
        last_note_idx,
        last_note_col,
    )


def find_production_status_rework_header_info(ws) -> tuple[list[str], int] | None:
    expected = list(REWORK_PRODUCTION_DEMAND_COLUMNS)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
        values = list(row)
        demand_values = values[1 : 1 + len(expected)]
        demand_values = demand_values + [None] * (len(expected) - len(demand_values))
        header = [str(value).strip() if value is not None else "" for value in demand_values]
        if header == expected:
            return header, row_idx
    return None


def read_rework_item_keys_from_production_status_file(
    source_path: Path,
) -> tuple[dict[tuple[str, str], float], dict[str, object]]:
    meta = build_empty_rework_meta()
    meta["source_path"] = str(source_path)
    try:
        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    except Exception:
        return {}, meta

    try:
        source_sheet = find_workbook_sheet_name(wb, [PRODUCTION_STATUS_SHEET_NAME], fallback_first=False)
        if source_sheet is None:
            return {}, meta
        ws = wb[source_sheet]
        header_info = find_production_status_rework_header_info(ws)
        if header_info is None:
            meta["sheet"] = f"{source_path.name} / {source_sheet}"
            return {}, meta

        preview_cols, header_row_idx = header_info
        field_indices = {column_name: idx + 1 for idx, column_name in enumerate(preview_cols)}
        qty_by_key: dict[tuple[str, str], float] = {}
        demand_qty_by_key: dict[tuple[str, ...], float] = {}
        source_qty_total = 0.0
        source_rows = 0

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            values = {
                column_name: row[field_idx] if field_idx < len(row) else ""
                for column_name, field_idx in field_indices.items()
            }
            initial = normalize_rework_match_value(values.get("이니셜", ""))
            product_code = normalize_rework_match_value(values.get("품목코드", ""))
            if not product_code:
                continue

            qty = parse_single_numeric_value(values.get("부족수량", 0))
            source_qty_total += qty
            source_rows += 1

            item_key = (initial, product_code)
            qty_by_key[item_key] = qty_by_key.get(item_key, 0.0) + qty

            for demand_key in build_rework_demand_key_variants(
                values.get("거래처", ""),
                values.get("이니셜", ""),
                values.get("품목코드", ""),
                values.get("R코드", ""),
                values.get("Q코드", ""),
                values.get("납기일", ""),
            ):
                demand_qty_by_key[demand_key] = demand_qty_by_key.get(demand_key, 0.0) + qty

        meta.update(
            {
                "sheet": f"{source_path.name} / {source_sheet}",
                "match_scope": "production_demand",
                "initial_col": "이니셜",
                "product_col": "품목코드",
                "quantity_col": "부족수량",
                "sheet_columns": preview_cols,
                "source_rows": source_rows,
                "source_qty_total": float(source_qty_total),
                "demand_qty_by_key": demand_qty_by_key,
                "demand_key_count": int(
                    len([key for key in demand_qty_by_key if isinstance(key, tuple) and key[:1] == ("full",)])
                ),
            }
        )
        return qty_by_key, meta
    finally:
        wb.close()


def read_rework_item_keys_from_demand_file(dem_path: Path) -> tuple[dict[tuple[str, str], float], dict[str, object]]:
    empty_meta = build_empty_rework_meta()
    production_source = find_rework_production_source_file(dem_path.parent)
    if production_source is not None:
        production_qty_map, production_meta = read_rework_item_keys_from_production_status_file(production_source)
        production_demand_map = production_meta.get("demand_qty_by_key", {})
        if production_qty_map or (isinstance(production_demand_map, dict) and production_demand_map):
            return production_qty_map, production_meta

    try:
        wb = openpyxl.load_workbook(dem_path, read_only=True, data_only=True)
    except Exception:
        return {}, empty_meta

    try:
        rework_sheet = find_workbook_sheet_name(wb, REWORK_SHEET_NAMES, fallback_first=False)
        if rework_sheet is None:
            return {}, empty_meta

        ws = wb[rework_sheet]
        (
            preview_cols,
            rows,
            initial_col,
            product_col,
            quantity_idx,
            quantity_col,
            note_idx,
            note_col,
        ) = find_rework_header_info(ws)
        meta = {
            "sheet": rework_sheet,
            "source_path": str(dem_path),
            "match_scope": "rework_sheet",
            "initial_col": initial_col or "",
            "product_col": product_col or "",
            "quantity_col": quantity_col or "",
            "note_col": note_col or "",
            "sheet_columns": preview_cols,
            "source_rows": 0,
            "source_qty_total": 0.0,
            "note_count": 0,
            "remarks_by_key": {},
            "demand_qty_by_key": {},
            "remarks_by_demand_key": {},
            "demand_key_count": 0,
        }
        if product_col is None:
            return {}, meta

        initial_idx = preview_cols.index(initial_col) if initial_col is not None else None
        product_idx = preview_cols.index(product_col)
        qty_by_key: dict[tuple[str, str], float] = {}
        remarks_by_key: dict[tuple[str, str], str] = {}
        for row in rows:
            initial = (
                normalize_rework_match_value(row[initial_idx] if initial_idx < len(row) else "")
                if initial_idx is not None
                else ""
            )
            product_code = normalize_rework_match_value(row[product_idx] if product_idx < len(row) else "")
            if product_code:
                qty = (
                    parse_single_numeric_value(row[quantity_idx])
                    if quantity_idx is not None and quantity_idx < len(row)
                    else 0.0
                )
                key = (initial, product_code)
                qty_by_key[key] = qty_by_key.get(key, 0.0) + qty
                if note_idx is not None and note_idx < len(row):
                    remarks_by_key[key] = merge_rework_note(
                        remarks_by_key.get(key, ""),
                        normalize_rework_note_value(row[note_idx]),
                    )

        meta["source_rows"] = int(len(qty_by_key))
        meta["source_qty_total"] = float(sum(qty_by_key.values()))
        remarks_by_key = {key: note for key, note in remarks_by_key.items() if note}
        meta["remarks_by_key"] = remarks_by_key
        meta["note_count"] = int(len(remarks_by_key))
        return qty_by_key, meta
    finally:
        wb.close()


def is_power_column(column_name: str) -> bool:
    return "파워" in str(column_name) or str(column_name).strip().lower() == "power"


def format_power_value(value: object) -> str:
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "nat", "-"}:
        return "-"

    match = POWER_VALUE_PATTERN.search(text)
    numeric_text = match.group(1) if match else text.replace(",", "")
    try:
        numeric_value = float(numeric_text)
    except (TypeError, ValueError):
        return "-"

    sign = "+" if numeric_value > 0 else "-"
    return f"{sign}{abs(numeric_value):05.2f}"


def extract_power_from_code(item_code: str) -> str:
    code = str(item_code).strip()
    match = POWER_VALUE_PATTERN.search(code)
    return format_power_value(match.group(1)) if match else "-"


def extract_power_key_from_code(item_code: str) -> str:
    code = str(item_code).strip().upper()
    variant_match = re.match(r"^[PQRSTU]\d{4}[A-Z]?(.*)$", code)
    if variant_match and variant_match.group(1):
        return variant_match.group(1)
    matches = POWER_VALUE_PATTERN.findall(code)
    if not matches:
        return "-"
    return "|".join(format_power_value(match) for match in matches)


def clean_text_value(value: object) -> str:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    return "" if text.lower() in INVALID_CATEGORY_VALUES else text


def clean_initial_value(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""
    parts = [part.strip() for part in re.split(r"[,;\n\r]+", text)]
    cleaned: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if not part:
            continue
        token_key = normalize_keyword_key(part)
        if token_key in {"이니셜", "initial", "initialcode"}:
            continue
        if part not in seen:
            seen.add(part)
            cleaned.append(part)
    return ", ".join(cleaned)


def normalize_site_group(value: object) -> str:
    text = clean_text_value(value).upper()
    compact = re.sub(r"[\s_./()\-]+", "", text)
    for site in SITE_GROUP_ORDER:
        site_key = site.upper()
        site_letter = site_key[0]
        if site_key in compact or re.match(rf"^{site_letter}(관|동|공장|$)", compact):
            return site
    return "기타" if text else "미지정"


def build_plan_api_site_param(site_filter: object) -> str:
    text = clean_text_value(site_filter)
    if not text or text == "전체":
        return ""
    site_group = normalize_site_group(text)
    return SITE_GROUP_API_PARAMS.get(site_group, text)


def clean_sheet_category(value: object) -> str:
    text = clean_text_value(value)
    if text == UNCLASSIFIED_SHEET_CATEGORY:
        return ""
    return text


def normalize_lookup_key(value: object) -> str:
    text = clean_text_value(value).lower()
    return re.sub(r"\s+", " ", text).strip()


def normalize_keyword_key(value: object) -> str:
    text = clean_text_value(value).lower()
    return re.sub(r"[\s_\-./()]+", "", text)


def normalize_flow_link_key_value(value: object) -> str:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "nat", "null", "<na>"} else text


def normalize_flow_link_key_columns(df: pd.DataFrame, key_cols: list[str]) -> pd.DataFrame:
    normalized = df.copy()
    for col in key_cols:
        if col in normalized.columns:
            normalized[col] = normalized[col].map(normalize_flow_link_key_value)
    return normalized


def match_keyword_category(text: object, rules: dict[str, list[str]]) -> tuple[str, str]:
    normalized_text = normalize_keyword_key(text)
    if not normalized_text:
        return UNCLASSIFIED_SHEET_CATEGORY, "분류 기준 없음"

    for category, keywords in rules.items():
        for keyword in keywords:
            raw_keyword = clean_text_value(keyword)
            startswith_match = raw_keyword.startswith("^")
            keyword_body = raw_keyword[1:] if startswith_match else raw_keyword
            normalized_keyword = normalize_keyword_key(keyword_body)
            if not normalized_keyword:
                continue
            matched = (
                normalized_text.startswith(normalized_keyword)
                if startswith_match
                else normalized_keyword in normalized_text
            )
            if matched:
                return category, keyword_body

    return UNCLASSIFIED_SHEET_CATEGORY, "분류 기준 없음"


def match_exact_customer_category(customer: object) -> tuple[str, str]:
    customer_key = normalize_lookup_key(customer)
    if not customer_key:
        return UNCLASSIFIED_SHEET_CATEGORY, "분류 기준 없음"

    for rule_customer, category in CUSTOMER_EXACT_CATEGORY_RULES.items():
        if customer_key == normalize_lookup_key(rule_customer):
            return category, rule_customer

    return UNCLASSIFIED_SHEET_CATEGORY, "분류 기준 없음"


def is_english_customer_name(customer: object) -> bool:
    text = clean_text_value(customer)
    if not text:
        return False
    return text.isascii() and bool(re.search(r"[A-Za-z]", text))


def classify_sheet_with_reason(row: pd.Series) -> tuple[str, str]:
    customer = row.get("거래처", "")
    product_name = row.get("제품명", "")
    combined_text = f"{clean_text_value(customer)} {clean_text_value(product_name)}"

    category, keyword = match_keyword_category(combined_text, SINCERE_2WEEK_RULES)
    if category != UNCLASSIFIED_SHEET_CATEGORY:
        return category, f"거래처/제품명에 {keyword} 포함"

    category, matched_customer = match_exact_customer_category(customer)
    if category != UNCLASSIFIED_SHEET_CATEGORY:
        return category, f"거래처명 완전 일치: {matched_customer}"

    category, keyword = match_keyword_category(customer, CUSTOMER_CATEGORY_RULES)
    if category != UNCLASSIFIED_SHEET_CATEGORY:
        return category, f"거래처명 키워드 매칭: {keyword}"

    category, keyword = match_keyword_category(product_name, PRODUCT_CATEGORY_RULES)
    if category != UNCLASSIFIED_SHEET_CATEGORY:
        return category, f"제품명 보조 키워드 매칭: {keyword}"

    if is_english_customer_name(customer):
        return "기타 해외", "영문 거래처명 기준 기타 해외 분류"

    return UNCLASSIFIED_SHEET_CATEGORY, "거래처명 기준 분류 불가"


def classify_sheet(row: pd.Series) -> str:
    category, _ = classify_sheet_with_reason(row)
    return category


def find_product_name_reference_file(base_dir: Path) -> Path | None:
    candidates = [
        p
        for p in base_dir.glob("*.xlsx")
        if not p.name.startswith("~$") and ("제품명" in p.stem and "기준" in p.stem)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_product_reference_maps(base_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return {}, {}

    ref = pd.read_excel(ref_path, sheet_name=0)
    ref.columns = [str(c).strip() for c in ref.columns]

    code_col = "제품명코드" if "제품명코드" in ref.columns else ref.columns[0]
    name_col = "제품명" if "제품명" in ref.columns else ref.columns[1]
    group_col = "분류요약" if "분류요약" in ref.columns else None
    if group_col is None and "판매제품군" in ref.columns:
        group_col = "판매제품군"
    if group_col is None and "생산제품군" in ref.columns:
        group_col = "생산제품군"

    selected_cols = [code_col, name_col] + ([group_col] if group_col is not None else [])
    ref_df = ref[selected_cols].copy()
    ref_df[code_col] = ref_df[code_col].astype(str).str.strip()
    ref_df[name_col] = ref_df[name_col].astype(str).str.strip()
    if group_col is not None:
        ref_df[group_col] = ref_df[group_col].astype(str).str.strip()
    ref_df = ref_df[
        ref_df[code_col].str.startswith("P")
        & (ref_df[code_col].str.lower() != "nan")
        & (ref_df[name_col] != "")
        & (ref_df[name_col].str.lower() != "nan")
    ]
    ref_df["코드5"] = ref_df[code_col].str[:5]
    ref_df = ref_df.drop_duplicates(subset=["코드5"], keep="first")
    name_map = ref_df.set_index("코드5")[name_col].to_dict()

    if group_col is None:
        group_map: dict[str, str] = {}
    else:
        group_df = ref_df[(ref_df[group_col] != "") & (ref_df[group_col].str.lower() != "nan")]
        group_map = group_df.set_index("코드5")[group_col].to_dict()

    return name_map, group_map


def find_reference_sheet_with_columns(
    ref_path: Path, sheet_names: list[str], required_columns: set[str], preferred_name: str | None = None
) -> str | None:
    if preferred_name:
        normalized = preferred_name.replace(" ", "")
        by_name = next((s for s in sheet_names if str(s).replace(" ", "") == normalized), None)
        if by_name is not None:
            return by_name

    for sheet_name in sheet_names:
        try:
            preview = pd.read_excel(ref_path, sheet_name=sheet_name, nrows=0)
        except Exception:
            continue
        cols = {str(c).strip() for c in preview.columns}
        if required_columns.issubset(cols):
            return sheet_name
    return None


def load_bom_base_code_maps(base_dir: Path) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, str]]:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return {}, {}, {}, {}

    sheet_names = pd.ExcelFile(ref_path).sheet_names
    bom_sheet = find_reference_sheet_with_columns(
        ref_path, sheet_names, {"SALES_ITEM_CD", "FROM_ITEM_ID"}, preferred_name="BOM정보"
    )
    if bom_sheet is None:
        return {}, {}, {}, {}

    use_cols = ["SALES_ITEM_CD", "TO_ITEM_ID", "FROM_ITEM_ID", "SEQ"]
    bom = pd.read_excel(ref_path, sheet_name=bom_sheet, usecols=lambda c: str(c).strip() in set(use_cols))
    bom.columns = [str(c).strip() for c in bom.columns]
    if not {"SALES_ITEM_CD", "FROM_ITEM_ID"}.issubset(bom.columns):
        return {}, {}, {}, {}

    bom["SALES_ITEM_CD"] = bom["SALES_ITEM_CD"].astype(str).str.strip()
    bom["FROM_ITEM_ID"] = bom["FROM_ITEM_ID"].astype(str).str.strip()
    if "SEQ" in bom.columns:
        bom["SEQ"] = pd.to_numeric(bom["SEQ"], errors="coerce").fillna(9999)
    else:
        bom["SEQ"] = 9999

    bom = bom[
        (bom["SALES_ITEM_CD"] != "")
        & (bom["SALES_ITEM_CD"].str.lower() != "nan")
        & (bom["FROM_ITEM_ID"] != "")
        & (bom["FROM_ITEM_ID"].str.lower() != "nan")
    ].copy()
    if bom.empty:
        return {}, {}, {}, {}

    # Exact TO_ITEM_ID mapping (authoritative when available).
    if "TO_ITEM_ID" in bom.columns:
        bom["TO_ITEM_ID"] = bom["TO_ITEM_ID"].astype(str).str.strip()
        exact = bom[
            (bom["TO_ITEM_ID"] != "")
            & (bom["TO_ITEM_ID"].str.lower() != "nan")
            & bom["FROM_ITEM_ID"].str.match(r"^[QR].+", na=False)
        ].copy()
        exact = exact.sort_values(["TO_ITEM_ID", "SEQ"], ascending=[True, True]).drop_duplicates(
            subset=["TO_ITEM_ID"], keep="first"
        )
    else:
        exact = pd.DataFrame(columns=["TO_ITEM_ID", "FROM_ITEM_ID"])

    q_exact_map: dict[str, str] = {}
    r_exact_map: dict[str, str] = {}
    if not exact.empty:
        for to_code, from_code in exact[["TO_ITEM_ID", "FROM_ITEM_ID"]].itertuples(index=False):
            from_code = str(from_code).strip()
            if from_code.startswith("Q"):
                q_exact_map[to_code] = from_code
                if len(from_code) > 1:
                    r_exact_map[to_code] = "R" + from_code[1:]
            elif from_code.startswith("R"):
                r_exact_map[to_code] = from_code
                if len(from_code) > 1:
                    q_exact_map[to_code] = "Q" + from_code[1:]

    bom["SALES_CODE5"] = bom["SALES_ITEM_CD"].str[:5]
    bom["FROM_CODE5"] = bom["FROM_ITEM_ID"].str[:5]
    bom = bom[bom["SALES_CODE5"].str.match(r"^[PQRSTU]\d{4}$", na=False)]
    bom = bom[bom["FROM_CODE5"].str.match(r"^[PQRSTU]\d{4}$", na=False)]
    if bom.empty:
        return {}, {}, r_exact_map, q_exact_map

    bom = bom.sort_values(["SALES_CODE5", "SEQ"], ascending=[True, True])

    q_df = bom[bom["FROM_CODE5"].str.startswith("Q")].drop_duplicates(subset=["SALES_CODE5"], keep="first")
    r_df = bom[bom["FROM_CODE5"].str.startswith("R")].drop_duplicates(subset=["SALES_CODE5"], keep="first")
    q_base_map = q_df.set_index("SALES_CODE5")["FROM_CODE5"].to_dict()
    r_base_map = r_df.set_index("SALES_CODE5")["FROM_CODE5"].to_dict()

    # If BOM has only Q mapping for a sales code, derive R base from the same numeric part.
    for sales_code5, q_code5 in q_base_map.items():
        if sales_code5 not in r_base_map and str(q_code5).startswith("Q") and len(str(q_code5)) >= 5:
            r_base_map[sales_code5] = "R" + str(q_code5)[1:5]

    return r_base_map, q_base_map, r_exact_map, q_exact_map


def load_sheet2_group_map(base_dir: Path) -> dict[str, str]:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return {}

    sheet_names = pd.ExcelFile(ref_path).sheet_names
    if len(sheet_names) < 2:
        return {}

    sheet_name = find_reference_sheet_with_columns(
        ref_path, sheet_names, {"코드", "시트이름"}, preferred_name="분류정보"
    )
    if sheet_name is None:
        return {}

    sheet2 = pd.read_excel(ref_path, sheet_name=sheet_name)
    sheet2.columns = [str(c).strip() for c in sheet2.columns]
    if "코드" not in sheet2.columns or "시트이름" not in sheet2.columns:
        return {}

    df = sheet2[["코드", "시트이름"]].copy()
    df["코드"] = df["코드"].astype(str).str.strip()
    df["시트이름"] = df["시트이름"].astype(str).str.strip()
    df = df[
        df["코드"].str.startswith("P")
        & (df["코드"].str.lower() != "nan")
        & (df["시트이름"] != "")
        & (df["시트이름"].str.lower() != "nan")
    ]
    df["코드5"] = df["코드"].str[:5]
    df = df.drop_duplicates(subset=["코드5"], keep="first")
    return df.set_index("코드5")["시트이름"].to_dict()


def load_rq_code_maps(base_dir: Path) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return {}, {}, {}

    sheet_names = pd.ExcelFile(ref_path).sheet_names
    if len(sheet_names) < 2:
        return {}, {}, {}

    sheet_name = find_reference_sheet_with_columns(
        ref_path, sheet_names, {"코드", "Q코드", "R코드"}, preferred_name="분류정보"
    )
    if sheet_name is None:
        return {}, {}, {}

    sheet2 = pd.read_excel(ref_path, sheet_name=sheet_name)
    sheet2.columns = [str(c).strip() for c in sheet2.columns]
    required = {"코드", "Q코드", "R코드"}
    if not required.issubset(sheet2.columns):
        return {}, {}, {}

    name_col = "제품명" if "제품명" in sheet2.columns else None

    df = sheet2.copy()
    for col in ["코드", "Q코드", "R코드"]:
        df[col] = df[col].astype(str).str.strip()
    if name_col:
        df[name_col] = df[name_col].astype(str).str.strip()

    df = df[
        df["코드"].str.startswith("P")
        & (df["코드"].str.lower() != "nan")
        & (df["Q코드"].str.lower() != "nan")
        & (df["R코드"].str.lower() != "nan")
        & (df["Q코드"] != "")
        & (df["R코드"] != "")
    ]

    df["코드5"] = df["코드"].str[:5]
    code5_df = df.drop_duplicates(subset=["코드5"], keep="first")

    q_map = code5_df.set_index("코드5")["Q코드"].to_dict()
    r_map = code5_df.set_index("코드5")["R코드"].to_dict()
    if name_col:
        r_name_df = df[(df[name_col] != "") & (df[name_col].str.lower() != "nan")].copy()
        r_name_df["R코드5"] = r_name_df["R코드"].str[:5]
        r_name_df = r_name_df[(r_name_df["R코드5"] != "") & (r_name_df["R코드5"].str.lower() != "nan")]
        r_name_df = r_name_df.drop_duplicates(subset=["R코드5", name_col], keep="first")
        r_name_df = r_name_df.drop_duplicates(subset=["R코드5"], keep="first")
        r_name_map = r_name_df.set_index("R코드5")[name_col].to_dict()
    else:
        r_name_map = {}

    # Fallback: enrich R코드5 -> 제품명 from sheet0(제품명정보).
    # 제품명코드가 P/Q/R/T/U + 4자리인 경우 모두 R + 4자리 키로 정규화한다.
    try:
        sheet1 = pd.read_excel(ref_path, sheet_name=sheet_names[0])
        sheet1.columns = [str(c).strip() for c in sheet1.columns]
        if len(sheet1.columns) >= 2:
            code_col = "제품명코드" if "제품명코드" in sheet1.columns else sheet1.columns[0]
            name1_col = "제품명" if "제품명" in sheet1.columns else sheet1.columns[1]
            fb = sheet1[[code_col, name1_col]].copy()
            fb[code_col] = fb[code_col].astype(str).str.strip()
            fb[name1_col] = fb[name1_col].astype(str).str.strip()
            fb = fb[
                (fb[code_col] != "")
                & (fb[code_col].str.lower() != "nan")
                & (fb[name1_col] != "")
                & (fb[name1_col].str.lower() != "nan")
            ]
            fb["코드5"] = fb[code_col].str[:5]
            fb = fb[fb["코드5"].str.match(r"^[PQRSTU]\d{4}$", na=False)]
            fb["R코드5"] = "R" + fb["코드5"].str[-4:]
            fb = fb.drop_duplicates(subset=["R코드5", name1_col], keep="first")
            fb = fb.drop_duplicates(subset=["R코드5"], keep="first")
            for k, v in fb.set_index("R코드5")[name1_col].to_dict().items():
                if k not in r_name_map:
                    r_name_map[k] = v
    except Exception:
        pass
    return r_map, q_map, r_name_map


def load_leadji_process_maps(base_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return {}, {}

    sheet_names = pd.ExcelFile(ref_path).sheet_names
    if len(sheet_names) < 3:
        return {}, {}

    leadji_sheet = next((s for s in sheet_names if s.replace(" ", "") == "리드지정보"), sheet_names[2])
    leadji = pd.read_excel(ref_path, sheet_name=leadji_sheet)
    leadji.columns = [str(c).strip() for c in leadji.columns]

    prod_col = "생산" if "생산" in leadji.columns else (leadji.columns[3] if len(leadji.columns) > 3 else None)
    q_col = "분리" if "분리" in leadji.columns else (leadji.columns[9] if len(leadji.columns) > 9 else None)
    r_col = "사출" if "사출" in leadji.columns else (leadji.columns[21] if len(leadji.columns) > 21 else None)
    if prod_col is None or q_col is None or r_col is None:
        return {}, {}

    df = leadji[[prod_col, q_col, r_col]].copy()
    for col in [prod_col, q_col, r_col]:
        df[col] = df[col].astype(str).str.strip()
        df.loc[df[col].str.lower() == "nan", col] = ""

    df = df[df[prod_col].str.startswith("P")]
    if df.empty:
        return {}, {}

    df["코드5"] = df[prod_col].str[:5]
    df = df[(df["코드5"] != "") & (df["코드5"].str.lower() != "nan")]

    def normalize_to_code(code: str, prefix: str) -> str:
        v = str(code).strip()
        if not v or v.lower() == "nan":
            return ""
        if v.startswith(prefix):
            return v
        if v.startswith("P"):
            return f"{prefix}{v[1:]}"
        return v

    df["Q정규"] = df[q_col].map(lambda x: normalize_to_code(x, "Q"))
    df["R정규"] = df[r_col].map(lambda x: normalize_to_code(x, "R"))

    q_df = df[df["Q정규"] != ""].drop_duplicates(subset=["코드5"], keep="first")
    r_df = df[df["R정규"] != ""].drop_duplicates(subset=["코드5"], keep="first")
    q_map = q_df.set_index("코드5")["Q정규"].to_dict()
    r_map = r_df.set_index("코드5")["R정규"].to_dict()
    return r_map, q_map


def merge_mapped_base_code(inferred_code: str, mapped_base_code: str, prefix: str) -> str:
    inferred = str(inferred_code).strip()
    mapped = str(mapped_base_code).strip()

    if not mapped or mapped.lower() == "nan":
        return inferred
    if not inferred or inferred.lower() == "nan":
        return mapped

    if inferred.startswith(prefix) and mapped.startswith(prefix) and len(inferred) >= 5 and len(mapped) >= 5:
        return mapped[:5] + inferred[5:]
    return mapped


def iter_inventory_code_candidates(process_code: str) -> list[str]:
    code = str(process_code).strip()
    if not code or code.lower() == "nan":
        return []

    candidates = [code]
    bul_match = re.match(r"^(.*BUL)\d+$", code, flags=re.IGNORECASE)
    if bul_match:
        candidates.append(bul_match.group(1))

    # 순서를 유지한 unique
    return list(dict.fromkeys(candidates))


def lookup_stock_qty(stock_map: dict[str, float], process_code: str) -> float:
    for candidate in iter_inventory_code_candidates(process_code):
        qty = stock_map.get(candidate)
        if qty is not None:
            return float(qty)
    return 0.0


def lookup_stock_qty_from_candidates(stock_map: dict[str, float], process_codes: list[str]) -> float:
    for process_code in process_codes:
        qty = lookup_stock_qty(stock_map, process_code)
        if qty:
            return qty
    return 0.0


def resolve_process_code_for_stock(stock_map: dict[str, float], process_code: str) -> str:
    code = str(process_code).strip()
    if not code or code.lower() == "nan":
        return code
    for candidate in iter_inventory_code_candidates(code):
        if candidate in stock_map:
            return candidate
    return code


def normalize_leadji_code_key(value: object) -> str:
    code = str(value).strip().upper()
    if code in {"", "NAN", "NONE", "-", "NULL"}:
        return ""
    code = re.sub(r"\s+", "", code)
    matched = re.match(r"^([A-Z]{2}\d{4})", code)
    return matched.group(1) if matched else code


def build_data_refresh_key(base_dir: Path) -> str:
    inv_path, dem_path = find_excel_files(base_dir)
    ref_path = find_product_name_reference_file(base_dir)
    rework_source_path = find_rework_production_source_file(base_dir)

    paths = [inv_path, dem_path]
    if ref_path is not None:
        paths.append(ref_path)
    if rework_source_path is not None:
        paths.append(rework_source_path)

    parts = [f"app:{APP_CACHE_VERSION}"]
    upload_signature = read_upload_workspace_signature(base_dir)
    if upload_signature:
        parts.append(f"upload:{upload_signature}")
    for p in paths:
        stat = p.stat()
        parts.append(f"{p.name}:{stat.st_size}:{stat.st_mtime_ns}")
    return "|".join(parts)


def build_api_shortage_refresh_key(base_dir: Path) -> str:
    parts = [f"api-shortage:{APP_CACHE_VERSION}", build_plan_api_refresh_key()]
    try:
        inv_path, _ = find_excel_files(base_dir)
        stat = inv_path.stat()
        parts.append(f"wip:{inv_path.name}:{stat.st_size}:{stat.st_mtime_ns}")
    except Exception as exc:
        parts.append(f"wip-error:{exc}")

    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        parts.append("reference:missing")
    else:
        stat = ref_path.stat()
        parts.append(f"reference:{ref_path.name}:{stat.st_size}:{stat.st_mtime_ns}")
    rework_source_path = find_rework_production_source_file(base_dir)
    if rework_source_path is None:
        parts.append("rework-production:missing")
    else:
        stat = rework_source_path.stat()
        parts.append(f"rework-production:{rework_source_path.name}:{stat.st_size}:{stat.st_mtime_ns}")
    return "|".join(parts)


def build_reference_refresh_key(base_dir: Path) -> str:
    ref_path = find_product_name_reference_file(base_dir)
    if ref_path is None:
        return "-"
    stat = ref_path.stat()
    return f"{ref_path.name}:{stat.st_size}:{stat.st_mtime_ns}"


def build_leadji_order_refresh_key(base_dir: Path) -> str:
    order_path = find_leadji_order_status_file(base_dir)
    if order_path is None:
        return "-"
    stat = order_path.stat()
    return f"{order_path.name}:{stat.st_size}:{stat.st_mtime_ns}"


def build_leadji_status_refresh_key(base_dir: Path) -> str:
    parts = [f"leadji-status:{APP_CACHE_VERSION}"]
    try:
        parts.append(build_data_refresh_key(base_dir))
    except Exception as exc:
        parts.append(f"data-error:{exc}")
    parts.append(f"reference:{build_reference_refresh_key(base_dir)}")
    parts.append(f"leadji-order:{build_leadji_order_refresh_key(base_dir)}")
    return "|".join(parts)


def build_all_item_refresh_key(base_dir: Path) -> str:
    parts = [f"all-items:{APP_CACHE_VERSION}"]
    parts.append(build_plan_api_refresh_key())
    try:
        parts.append(build_data_refresh_key(base_dir))
    except Exception as exc:
        parts.append(f"data-error:{exc}")
    parts.append(f"reference:{build_reference_refresh_key(base_dir)}")

    master_path = find_all_item_master_file(base_dir)
    if master_path is None:
        parts.append("all-item-master:missing")
    else:
        stat = master_path.stat()
        parts.append(f"{master_path.name}:{stat.st_size}:{stat.st_mtime_ns}")

    finished_goods_stock_path = find_finished_goods_stock_file(base_dir)
    if finished_goods_stock_path is None:
        parts.append("finished-goods-stock:missing")
    else:
        stat = finished_goods_stock_path.stat()
        parts.append(f"{finished_goods_stock_path.name}:{stat.st_size}:{stat.st_mtime_ns}")

    return "|".join(parts)


def parse_sequence_priority(value: object) -> float:
    if value is None:
        return 9999.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 9999.0


def get_row_value(row: tuple[object, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()


def load_bom_maps_streaming(
    ref_path_str: str, reference_refresh_key: str, bom_sheet: str
) -> tuple[dict[str, str], dict[str, str], dict[str, str], dict[str, str]]:
    _ = reference_refresh_key
    bom_r_base_map: dict[str, str] = {}
    bom_q_base_map: dict[str, str] = {}
    bom_r_exact_map: dict[str, str] = {}
    bom_q_exact_map: dict[str, str] = {}

    cache_key = hashlib.sha256(
        f"bom-stream-v1|{reference_refresh_key}|{bom_sheet}".encode("utf-8")
    ).hexdigest()[:24]
    cache_path = Path(ref_path_str).resolve().parent / ".dashboard_cache" / f"bom_maps_{cache_key}.pkl"
    try:
        if cache_path.exists():
            with cache_path.open("rb") as f:
                cached = pickle.load(f)
            if isinstance(cached, tuple) and len(cached) == 4:
                return cached
    except Exception:
        pass

    try:
        wb = openpyxl.load_workbook(ref_path_str, read_only=True, data_only=True)
    except Exception:
        return bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map

    try:
        if bom_sheet not in wb.sheetnames:
            return bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map

        ws = wb[bom_sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        except StopIteration:
            return bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map

        idx_map = {name: headers.index(name) for name in ["SALES_ITEM_CD", "TO_ITEM_ID", "FROM_ITEM_ID", "SEQ"] if name in headers}
        sales_idx = idx_map.get("SALES_ITEM_CD")
        to_idx = idx_map.get("TO_ITEM_ID")
        from_idx = idx_map.get("FROM_ITEM_ID")
        seq_idx = idx_map.get("SEQ")
        if sales_idx is None or from_idx is None:
            return bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map

        code5_pattern = re.compile(r"^[PQRSTU]\d{4}$")
        exact_best: dict[str, tuple[float, str]] = {}
        q_base_best: dict[str, tuple[float, str]] = {}
        r_base_best: dict[str, tuple[float, str]] = {}

        for row in rows:
            from_code = get_row_value(row, from_idx)
            if not from_code or from_code.lower() == "nan":
                continue

            seq = parse_sequence_priority(row[seq_idx] if seq_idx is not None and seq_idx < len(row) else None)

            if to_idx is not None:
                to_code = get_row_value(row, to_idx)
                if to_code and to_code.lower() != "nan" and re.match(r"^[QR].+", from_code):
                    current = exact_best.get(to_code)
                    if current is None or seq < current[0]:
                        exact_best[to_code] = (seq, from_code)

            sales_code = get_row_value(row, sales_idx)
            if not sales_code or sales_code.lower() == "nan":
                continue

            sales_code5 = sales_code[:5]
            from_code5 = from_code[:5]
            if not code5_pattern.match(sales_code5) or not code5_pattern.match(from_code5):
                continue

            if from_code5.startswith("Q"):
                current = q_base_best.get(sales_code5)
                if current is None or seq < current[0]:
                    q_base_best[sales_code5] = (seq, from_code5)
            elif from_code5.startswith("R"):
                current = r_base_best.get(sales_code5)
                if current is None or seq < current[0]:
                    r_base_best[sales_code5] = (seq, from_code5)

        for to_code, (_, from_code) in exact_best.items():
            if from_code.startswith("Q"):
                bom_q_exact_map[to_code] = from_code
                if len(from_code) > 1:
                    bom_r_exact_map[to_code] = "R" + from_code[1:]
            elif from_code.startswith("R"):
                bom_r_exact_map[to_code] = from_code
                if len(from_code) > 1:
                    bom_q_exact_map[to_code] = "Q" + from_code[1:]

        bom_q_base_map = {sales_code5: code5 for sales_code5, (_, code5) in q_base_best.items()}
        bom_r_base_map = {sales_code5: code5 for sales_code5, (_, code5) in r_base_best.items()}
        for sales_code5, q_code5 in bom_q_base_map.items():
            if sales_code5 not in bom_r_base_map and q_code5.startswith("Q") and len(q_code5) >= 5:
                bom_r_base_map[sales_code5] = "R" + q_code5[1:5]

        result = (bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map)
        try:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            with cache_path.open("wb") as f:
                pickle.dump(result, f, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception:
            pass
        return result
    finally:
        wb.close()


def load_reference_maps_bundle(
    base_dir: Path, reference_refresh_key: str
) -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
]:
    _ = reference_refresh_key
    ref_path = find_product_name_reference_file(base_dir)
    empty_bundle = ({}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {})
    if ref_path is None:
        return empty_bundle
    cache_key = hashlib.sha256(f"reference-bundle-v3|{reference_refresh_key}".encode("utf-8")).hexdigest()[:24]
    cache_path = ref_path.resolve().parent / ".dashboard_cache" / f"reference_bundle_{cache_key}.pkl"
    cached = read_pickle_cache(cache_path)
    if isinstance(cached, tuple) and len(cached) == 13 and all(isinstance(part, dict) for part in cached):
        return cached

    try:
        xls = pd.ExcelFile(ref_path)
    except Exception:
        return empty_bundle

    sheet_names = xls.sheet_names
    if not sheet_names:
        return empty_bundle

    def parse_sheet(sheet_name: str, usecols=None) -> pd.DataFrame:
        try:
            df = xls.parse(sheet_name=sheet_name, usecols=usecols)
        except Exception:
            return pd.DataFrame()
        df.columns = [str(c).strip() for c in df.columns]
        return df

    def find_sheet(required_columns: set[str], preferred_name: str | None = None) -> str | None:
        if preferred_name:
            normalized = preferred_name.replace(" ", "")
            by_name = next((s for s in sheet_names if str(s).replace(" ", "") == normalized), None)
            if by_name is not None:
                return by_name

        for sheet_name in sheet_names:
            try:
                preview = xls.parse(sheet_name=sheet_name, nrows=0)
            except Exception:
                continue
            cols = {str(c).strip() for c in preview.columns}
            if required_columns.issubset(cols):
                return sheet_name
        return None

    product_name_map: dict[str, str] = {}
    product_group_map: dict[str, str] = {}
    sheet2_group_map: dict[str, str] = {}
    r_ref_map: dict[str, str] = {}
    q_ref_map: dict[str, str] = {}
    r_name_map: dict[str, str] = {}
    bom_r_base_map: dict[str, str] = {}
    bom_q_base_map: dict[str, str] = {}
    bom_r_exact_map: dict[str, str] = {}
    bom_q_exact_map: dict[str, str] = {}
    leadji_r_map: dict[str, str] = {}
    leadji_q_map: dict[str, str] = {}
    leadji_u_map: dict[str, str] = {}

    # 1) 제품명/분류 맵 + R코드명 fallback 기반 시트
    try:
        sheet0_preview = xls.parse(sheet_name=sheet_names[0], nrows=0)
        sheet0_cols = [str(c).strip() for c in sheet0_preview.columns]
    except Exception:
        sheet0_cols = []

    if sheet0_cols:
        code_col = "제품명코드" if "제품명코드" in sheet0_cols else sheet0_cols[0]
        name_col = "제품명" if "제품명" in sheet0_cols else (sheet0_cols[1] if len(sheet0_cols) > 1 else sheet0_cols[0])
        group_col = "분류요약" if "분류요약" in sheet0_cols else None
        if group_col is None and "판매제품군" in sheet0_cols:
            group_col = "판매제품군"
        if group_col is None and "생산제품군" in sheet0_cols:
            group_col = "생산제품군"
        sheet0_selected_cols = {code_col, name_col}
        if group_col is not None:
            sheet0_selected_cols.add(group_col)
        sheet0 = parse_sheet(sheet_names[0], usecols=lambda c: str(c).strip() in sheet0_selected_cols)
    else:
        sheet0 = pd.DataFrame()

    if not sheet0.empty and len(sheet0.columns) >= 2:
        code_col = "제품명코드" if "제품명코드" in sheet0.columns else sheet0.columns[0]
        name_col = "제품명" if "제품명" in sheet0.columns else sheet0.columns[1]
        group_col = "분류요약" if "분류요약" in sheet0.columns else None
        if group_col is None and "판매제품군" in sheet0.columns:
            group_col = "판매제품군"
        if group_col is None and "생산제품군" in sheet0.columns:
            group_col = "생산제품군"

        selected_cols = [code_col, name_col] + ([group_col] if group_col is not None else [])
        ref_df = sheet0[selected_cols].copy()
        ref_df[code_col] = ref_df[code_col].astype(str).str.strip()
        ref_df[name_col] = ref_df[name_col].astype(str).str.strip()
        if group_col is not None:
            ref_df[group_col] = ref_df[group_col].astype(str).str.strip()

        ref_df = ref_df[
            ref_df[code_col].str.startswith("P")
            & (ref_df[code_col].str.lower() != "nan")
            & (ref_df[name_col] != "")
            & (ref_df[name_col].str.lower() != "nan")
        ]
        ref_df["코드5"] = ref_df[code_col].str[:5]
        ref_df = ref_df.drop_duplicates(subset=["코드5"], keep="first")
        product_name_map = ref_df.set_index("코드5")[name_col].to_dict()

        if group_col is not None:
            group_df = ref_df[(ref_df[group_col] != "") & (ref_df[group_col].str.lower() != "nan")]
            product_group_map = group_df.set_index("코드5")[group_col].to_dict()

    # 2) 분류정보 시트 기반 (시트분류 + R/Q 맵 + R코드명 우선)
    group_sheet = find_sheet({"코드", "시트이름"}, preferred_name="분류정보")
    rq_sheet = find_sheet({"코드", "Q코드", "R코드"}, preferred_name="분류정보")
    classification_cols = {"코드", "시트이름", "Q코드", "R코드", "제품명"}
    classification_usecols = lambda c: str(c).strip() in classification_cols
    group_df_source = parse_sheet(group_sheet, usecols=classification_usecols) if group_sheet else pd.DataFrame()
    rq_df_source = group_df_source if (rq_sheet and group_sheet and rq_sheet == group_sheet) else (
        parse_sheet(rq_sheet, usecols=classification_usecols) if rq_sheet else pd.DataFrame()
    )

    if not group_df_source.empty and {"코드", "시트이름"}.issubset(group_df_source.columns):
        s2 = group_df_source[["코드", "시트이름"]].copy()
        s2["코드"] = s2["코드"].astype(str).str.strip()
        s2["시트이름"] = s2["시트이름"].astype(str).str.strip()
        s2 = s2[
            s2["코드"].str.startswith("P")
            & (s2["코드"].str.lower() != "nan")
            & (s2["시트이름"] != "")
            & (s2["시트이름"].str.lower() != "nan")
        ]
        s2["코드5"] = s2["코드"].str[:5]
        s2 = s2.drop_duplicates(subset=["코드5"], keep="first")
        sheet2_group_map = s2.set_index("코드5")["시트이름"].to_dict()

    if not rq_df_source.empty and {"코드", "Q코드", "R코드"}.issubset(rq_df_source.columns):
        rq = rq_df_source.copy()
        for col in ["코드", "Q코드", "R코드"]:
            rq[col] = rq[col].astype(str).str.strip()
        name_col = "제품명" if "제품명" in rq.columns else None
        if name_col:
            rq[name_col] = rq[name_col].astype(str).str.strip()

        rq = rq[
            rq["코드"].str.startswith("P")
            & (rq["코드"].str.lower() != "nan")
            & (rq["Q코드"].str.lower() != "nan")
            & (rq["R코드"].str.lower() != "nan")
            & (rq["Q코드"] != "")
            & (rq["R코드"] != "")
        ]
        rq["코드5"] = rq["코드"].str[:5]
        code5_df = rq.drop_duplicates(subset=["코드5"], keep="first")
        q_ref_map = code5_df.set_index("코드5")["Q코드"].to_dict()
        r_ref_map = code5_df.set_index("코드5")["R코드"].to_dict()

        if name_col:
            r_name_df = rq[(rq[name_col] != "") & (rq[name_col].str.lower() != "nan")].copy()
            r_name_df["R코드5"] = r_name_df["R코드"].str[:5]
            r_name_df = r_name_df[(r_name_df["R코드5"] != "") & (r_name_df["R코드5"].str.lower() != "nan")]
            r_name_df = r_name_df.drop_duplicates(subset=["R코드5", name_col], keep="first")
            r_name_df = r_name_df.drop_duplicates(subset=["R코드5"], keep="first")
            r_name_map = r_name_df.set_index("R코드5")[name_col].to_dict()

    # 2-b) sheet0 기반 R코드명 fallback
    if not sheet0.empty and len(sheet0.columns) >= 2:
        code_col = "제품명코드" if "제품명코드" in sheet0.columns else sheet0.columns[0]
        name_col = "제품명" if "제품명" in sheet0.columns else sheet0.columns[1]
        fb = sheet0[[code_col, name_col]].copy()
        fb[code_col] = fb[code_col].astype(str).str.strip()
        fb[name_col] = fb[name_col].astype(str).str.strip()
        fb = fb[
            (fb[code_col] != "")
            & (fb[code_col].str.lower() != "nan")
            & (fb[name_col] != "")
            & (fb[name_col].str.lower() != "nan")
        ]
        fb["코드5"] = fb[code_col].str[:5]
        fb = fb[fb["코드5"].str.match(r"^[PQRSTU]\d{4}$", na=False)]
        fb["R코드5"] = "R" + fb["코드5"].str[-4:]
        fb = fb.drop_duplicates(subset=["R코드5", name_col], keep="first")
        fb = fb.drop_duplicates(subset=["R코드5"], keep="first")
        for key, value in fb.set_index("R코드5")[name_col].to_dict().items():
            if key not in r_name_map:
                r_name_map[key] = value

    # 3) BOM 기반 매핑
    bom_sheet = find_sheet({"SALES_ITEM_CD", "FROM_ITEM_ID"}, preferred_name="BOM정보")
    if bom_sheet is not None:
        bom_r_base_map, bom_q_base_map, bom_r_exact_map, bom_q_exact_map = load_bom_maps_streaming(
            str(ref_path), reference_refresh_key, bom_sheet
        )

    # 4) 리드지 공정 맵
    if len(sheet_names) >= 3:
        leadji_sheet = next((s for s in sheet_names if s.replace(" ", "") == "리드지정보"), sheet_names[2])
        try:
            leadji_preview = xls.parse(sheet_name=leadji_sheet, nrows=0)
            leadji_cols = [str(c).strip() for c in leadji_preview.columns]
        except Exception:
            leadji_cols = []
        if leadji_cols:
            prod_col = "생산" if "생산" in leadji_cols else (leadji_cols[3] if len(leadji_cols) > 3 else None)
            q_col = "분리" if "분리" in leadji_cols else (leadji_cols[9] if len(leadji_cols) > 9 else None)
            r_col = "사출" if "사출" in leadji_cols else (leadji_cols[21] if len(leadji_cols) > 21 else None)
            u_col = "외주" if "외주" in leadji_cols else ("U코드" if "U코드" in leadji_cols else None)
            selected_cols = {c for c in [prod_col, q_col, r_col, u_col] if c is not None}
            leadji = (
                parse_sheet(leadji_sheet, usecols=lambda c: str(c).strip() in selected_cols)
                if selected_cols
                else pd.DataFrame()
            )
            if not leadji.empty:
                leadji = leadji.rename(
                    columns={
                        prod_col: "생산",
                        q_col: "분리",
                        r_col: "사출",
                        u_col: "외주",
                    }
                )
        else:
            leadji = pd.DataFrame()
        if not leadji.empty:
            prod_col = "생산" if "생산" in leadji.columns else (leadji.columns[3] if len(leadji.columns) > 3 else None)
            q_col = "분리" if "분리" in leadji.columns else (leadji.columns[9] if len(leadji.columns) > 9 else None)
            r_col = "사출" if "사출" in leadji.columns else (leadji.columns[21] if len(leadji.columns) > 21 else None)
            u_col = "외주" if "외주" in leadji.columns else ("U코드" if "U코드" in leadji.columns else None)
            selected_leadji_cols = [c for c in [prod_col, q_col, r_col, u_col] if c is not None]
            if prod_col is not None and selected_leadji_cols:
                ldf = leadji[selected_leadji_cols].copy()
                for col in selected_leadji_cols:
                    ldf[col] = ldf[col].astype(str).str.strip()
                    ldf.loc[ldf[col].str.lower() == "nan", col] = ""
                ldf = ldf[ldf[prod_col].str.startswith("P")]
                if not ldf.empty:
                    ldf["코드5"] = ldf[prod_col].str[:5]
                    ldf = ldf[(ldf["코드5"] != "") & (ldf["코드5"].str.lower() != "nan")]

                    def normalize_to_code(code: str, prefix: str) -> str:
                        value = str(code).strip()
                        if not value or value.lower() == "nan":
                            return ""
                        if value.startswith(prefix):
                            return value
                        if value.startswith("P"):
                            return f"{prefix}{value[1:]}"
                        return value

                    if q_col is not None and q_col in ldf.columns:
                        ldf["Q정규"] = ldf[q_col].map(lambda x: normalize_to_code(x, "Q"))
                        q_df = ldf[ldf["Q정규"] != ""].drop_duplicates(subset=["코드5"], keep="first")
                        leadji_q_map = q_df.set_index("코드5")["Q정규"].to_dict()
                    if r_col is not None and r_col in ldf.columns:
                        ldf["R정규"] = ldf[r_col].map(lambda x: normalize_to_code(x, "R"))
                        r_df = ldf[ldf["R정규"] != ""].drop_duplicates(subset=["코드5"], keep="first")
                        leadji_r_map = r_df.set_index("코드5")["R정규"].to_dict()
                    if u_col is not None and u_col in ldf.columns:
                        ldf["U정규"] = ldf[u_col].map(lambda x: normalize_to_code(x, "U"))
                        u_df = ldf[ldf["U정규"] != ""].drop_duplicates(subset=["코드5"], keep="first")
                        leadji_u_map = u_df.set_index("코드5")["U정규"].to_dict()

    result = (
        product_name_map,
        product_group_map,
        sheet2_group_map,
        r_ref_map,
        q_ref_map,
        r_name_map,
        bom_r_base_map,
        bom_q_base_map,
        bom_r_exact_map,
        bom_q_exact_map,
        leadji_r_map,
        leadji_q_map,
        leadji_u_map,
    )
    write_pickle_cache(cache_path, result)
    return result


def summarize_unique(values: pd.Series, head_count: int = 1) -> str:
    uniq = [v for v in values.astype(str).str.strip().tolist() if v and v.lower() != "nan"]
    # 순서를 유지한 unique
    uniq = list(dict.fromkeys(uniq))
    if not uniq:
        return "-"
    if len(uniq) <= head_count:
        return ", ".join(uniq)
    return f"{', '.join(uniq[:head_count])} 외 {len(uniq) - head_count}"


def format_pill_label(option: str, value_map: dict[str, float]) -> str:
    value = float(value_map.get(option, 0))
    return f"{option} ({value:,.0f})"


def _multi_pill_previous_key(key: str) -> str:
    return f"{key}__previous_selection"


def _as_pill_selection_list(selection: object) -> list[str]:
    if selection is None:
        return []
    if isinstance(selection, str):
        raw_selection = [selection]
    else:
        try:
            raw_selection = list(selection)  # type: ignore[arg-type]
        except TypeError:
            raw_selection = [selection]

    selections: list[str] = []
    for item in raw_selection:
        label = str(item).strip()
        if label and label not in selections:
            selections.append(label)
    return selections


def normalize_multi_pill_selection(
    selection: object,
    previous_selection: object = None,
    all_option: str = "전체",
) -> list[str]:
    current = _as_pill_selection_list(selection)
    previous = _as_pill_selection_list(previous_selection)
    if not current:
        return [all_option]
    if all_option in current and len(current) > 1:
        if all_option not in previous:
            return [all_option]
        return [value for value in current if value != all_option]
    return current


def prepare_multi_pill_state(key: str, options: list[str], all_option: str = "전체") -> None:
    valid_options = {str(option).strip() for option in options}
    previous_key = _multi_pill_previous_key(key)
    selection = normalize_multi_pill_selection(
        st.session_state.get(key, [all_option]),
        st.session_state.get(previous_key, [all_option]),
        all_option,
    )
    selection = [value for value in selection if value in valid_options]
    if not selection:
        selection = [all_option]
    st.session_state[key] = selection
    st.session_state[previous_key] = selection


def sync_multi_pill_state(key: str, all_option: str = "전체") -> None:
    previous_key = _multi_pill_previous_key(key)
    selection = normalize_multi_pill_selection(
        st.session_state.get(key),
        st.session_state.get(previous_key, [all_option]),
        all_option,
    )
    st.session_state[key] = selection
    st.session_state[previous_key] = selection


def finalize_multi_pill_selection(key: str, selection: object, all_option: str = "전체") -> tuple[str, ...]:
    previous_key = _multi_pill_previous_key(key)
    normalized = normalize_multi_pill_selection(
        selection,
        st.session_state.get(previous_key, [all_option]),
        all_option,
    )
    st.session_state[previous_key] = normalized
    return tuple(normalized)


def is_specific_pill_selection(selection: tuple[str, ...], all_option: str = "전체") -> bool:
    return bool(selection) and all_option not in selection


def build_thousand_separator_config(df: pd.DataFrame) -> dict[str, st.column_config.NumberColumn]:
    config: dict[str, st.column_config.NumberColumn] = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            config[col] = st.column_config.NumberColumn(format="%,.0f")
    return config


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def format_numeric_columns_for_display(df: pd.DataFrame) -> pd.DataFrame:
    display_df = df.copy()
    if "비고" in display_df.columns:
        display_df["비고"] = clean_display_text_series(display_df["비고"])

    for col in display_df.columns:
        if is_power_column(col):
            display_df[col] = display_df[col].map(format_power_value)
            continue

        if str(col).upper() == "DOI":
            series = parse_mixed_numeric(display_df[col])
            display_df[col] = series.map(lambda x: "" if pd.isna(x) or x <= 0 else f"{x:,.1f}")
            continue

        force_integer_display = any(token in str(col) for token in ["수량", "부족", "재고", "창고", "발주", "합계"])
        numeric_dtype = pd.api.types.is_numeric_dtype(display_df[col])
        numeric_like = numeric_dtype or force_integer_display or infer_numeric_like_series(display_df[col])
        if not numeric_like:
            continue

        series = parse_mixed_numeric(display_df[col])
        non_null = series.dropna()
        if non_null.empty and not numeric_dtype:
            continue
        is_integer_like = force_integer_display or non_null.empty or ((non_null % 1) == 0).all()

        if is_integer_like:
            display_df[col] = series.round(0).map(lambda x: "" if pd.isna(x) else f"{x:,.0f}")
        else:
            display_df[col] = series.map(lambda x: "" if pd.isna(x) else f"{x:,.2f}")

    return display_df


def limit_dataframe_for_display(
    df: pd.DataFrame, limit: int | None = None
) -> tuple[pd.DataFrame, bool]:
    _ = limit
    return df, False


def caption_limited_rows(total_rows: int, displayed_rows: int) -> None:
    _ = total_rows, displayed_rows


def move_columns_to_end(columns: list[str], trailing_columns: list[str]) -> list[str]:
    trailing = [col for col in trailing_columns if col in columns]
    return [col for col in columns if col not in trailing] + trailing


def infer_numeric_like_series(series: pd.Series) -> bool:
    sample = series.astype(str).str.replace(",", "", regex=False).str.strip()
    sample = sample[~sample.str.lower().isin({"", "nan", "none"})].head(200)
    if sample.empty:
        return False
    numeric_mask = sample.str.fullmatch(r"[+-]?\d+(?:\.\d+)?").fillna(False)
    return bool(float(numeric_mask.mean()) >= 0.85)


def pick_fixed_column_width_px(column_name: str, max_length: int, numeric_like: bool) -> int:
    if numeric_like:
        return int(max(90, min(145, 24 + max_length * 7)))

    long_text_columns = {"제품명", "R코드 제품명", "리드지명", "제품명 예시", "분류 판단 근거"}
    medium_text_columns = {"품목코드", "R코드", "Q코드", "U코드", "생산코드", "리드지코드", "P코드 예시"}
    status_columns = {"상태", "확인구분"}
    date_columns = {"납기일", "입고예상일자", "생산 최소 납기일", "최소납기일"}

    if column_name in long_text_columns:
        return int(max(240, min(380, 28 + max_length * 7)))
    if column_name in medium_text_columns:
        return int(max(120, min(170, 24 + max_length * 7)))
    if column_name in {"거래처그룹", "제품대분류"}:
        return 132
    if column_name == "DOI":
        return 92
    if column_name == "신호":
        return 112
    if column_name == "확인구분":
        return 150
    if column_name in {"재작업", "비고"}:
        return int(max(120, min(220, 28 + max_length * 7)))
    if column_name in status_columns:
        return 96
    if column_name in date_columns:
        return 118
    return int(max(92, min(145, 24 + max_length * 7)))


def build_auto_column_config(
    df: pd.DataFrame, columns: list[str], source_df: pd.DataFrame | None = None
) -> dict[str, st.column_config.Column]:
    config: dict[str, st.column_config.Column] = {}
    for col in columns:
        if col not in df.columns:
            continue
        col_series = df[col].astype(str)
        length_series = col_series.map(len)
        p90_len = int(length_series.quantile(0.90)) if not length_series.empty else 0
        max_len = max(len(str(col)), p90_len)

        numeric_like = False
        if source_df is not None and col in source_df.columns:
            numeric_like = pd.api.types.is_numeric_dtype(source_df[col]) or infer_numeric_like_series(col_series)
        else:
            numeric_like = infer_numeric_like_series(col_series)

        width_px = pick_fixed_column_width_px(col, max_len, numeric_like)
        config[col] = st.column_config.Column(
            label=COLUMN_LABEL_ALIASES.get(col, col),
            width=width_px,
        )
    return config


def render_dashboard_kpi(label: str, value: str, variant: str = "stock") -> None:
    st.markdown(
        f"""
        <div class="ops-kpi-card {variant}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def style_operational_table(display_df: pd.DataFrame, source_df: pd.DataFrame | None = None):
    if display_df.empty:
        return display_df.style
    if len(display_df) * max(len(display_df.columns), 1) > TABLE_STYLE_CELL_LIMIT:
        return display_df

    source = source_df if source_df is not None else display_df
    styler = display_df.style
    numeric_cols: list[str] = []
    for col in display_df.columns:
        source_col = source[col] if col in source.columns else display_df[col]
        if pd.api.types.is_numeric_dtype(source_col) or infer_numeric_like_series(display_df[col]):
            numeric_cols.append(col)

    if numeric_cols:
        styler = styler.set_properties(subset=numeric_cols, **{"text-align": "right"})

    text_cols = [c for c in ["제품명", "R코드 제품명", "리드지명", "제품명 예시", "비고"] if c in display_df.columns]
    if text_cols:
        styler = styler.set_properties(subset=text_cols, **{"text-align": "left"})

    shortage_cols = [c for c in display_df.columns if "부족" in c]
    for col in shortage_cols:
        source_col = source[col] if col in source.columns else display_df[col]
        shortage_numeric = parse_mixed_numeric(source_col)
        shortage_style = shortage_numeric.map(
            lambda v: "color: #DC2626; font-weight: 850;" if pd.notna(v) and abs(float(v)) > 0 else "color: #6B7280;"
        )
        styler = styler.apply(lambda _, style=shortage_style: style, axis=0, subset=[col])

    if "상태" in display_df.columns:
        styler = styler.set_properties(subset=["상태"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: (
                "background-color: #FEE2E2; color: #B91C1C; font-weight: 850;"
                if str(v).strip() == "부족"
                else "background-color: #FEF3C7; color: #92400E; font-weight: 850;"
                if str(v).strip() == "확인필요"
                else "background-color: #DCFCE7; color: #166534; font-weight: 850;"
                if str(v).strip() == "정상"
                else "background-color: #EEF2FF; color: #1A2B5E; font-weight: 800;"
            ),
            subset=["상태"],
        )
    if "확인구분" in display_df.columns:
        styler = styler.set_properties(subset=["확인구분"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: (
                "background-color: #FEE2E2; color: #B91C1C; font-weight: 850;"
                if str(v).strip() == "최종부족/재고없음"
                else "background-color: #FEF3C7; color: #92400E; font-weight: 850;"
                if str(v).strip() == "공정재고 확인"
                else "background-color: #EEF2FF; color: #1A2B5E; font-weight: 850;"
                if str(v).strip() == "사출필요"
                else "background-color: #FED7AA; color: #C2410C; font-weight: 850;"
                if str(v).strip() == "재작업가능"
                else ""
            ),
            subset=["확인구분"],
        )
    if "리스크구분" in display_df.columns:
        styler = styler.set_properties(subset=["리스크구분"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: (
                "background-color: #FEE2E2; color: #B91C1C; font-weight: 850;"
                if str(v).strip() == "현재수요 제품군 없음"
                else "background-color: #FEF3C7; color: #92400E; font-weight: 850;"
                if str(v).strip() == "동일제품 타도수 재고"
                else "background-color: #FFEDD5; color: #C2410C; font-weight: 850;"
                if str(v).strip() == "수요초과 재고"
                else "background-color: #DCFCE7; color: #166534; font-weight: 800;"
                if str(v).strip() == "수요코드 직접매칭"
                else ""
            ),
            subset=["리스크구분"],
        )
    rework_note_style_cols = [col for col in ["재작업"] if col in display_df.columns]
    if rework_note_style_cols:
        styler = styler.set_properties(subset=rework_note_style_cols, **{"text-align": "left"})
        styler = styler.map(
            lambda v: (
                "background-color: #FED7AA; color: #C2410C; font-weight: 850; "
                "border-radius: 999px;"
                if str(v).strip() and str(v).strip().lower() not in INVALID_CATEGORY_VALUES
                else ""
            ),
            subset=rework_note_style_cols,
        )

    return styler


@st.cache_data(show_spinner=False, max_entries=32)
def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "data") -> bytes:
    safe_sheet = sanitize_excel_sheet_name(sheet_name, "data")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=safe_sheet)
    output.seek(0)
    return output.getvalue()


def sanitize_excel_sheet_name(value: object, fallback: str = "data") -> str:
    text = clean_text_value(value) or clean_text_value(fallback) or "data"
    safe_sheet = re.sub(r"[\[\]\*:/\\?]", "_", text)
    safe_sheet = re.sub(r"[\x00-\x1f]", "_", safe_sheet)
    safe_sheet = re.sub(r"\s+", " ", safe_sheet).strip().strip("'")
    safe_sheet = re.sub(r"_+", "_", safe_sheet).strip(" _")
    if not safe_sheet:
        safe_sheet = clean_text_value(fallback) or "data"
    return safe_sheet[:31]


def unique_excel_sheet_name(value: object, used_names: set[str], fallback: str = "data") -> str:
    base = sanitize_excel_sheet_name(value, fallback)
    sheet_name = base
    suffix = 2
    while sheet_name.casefold() in used_names:
        suffix_text = f"_{suffix}"
        sheet_name = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(sheet_name.casefold())
    return sheet_name


def format_excel_worksheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    header_font = Font(bold=True, color="111827")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for col_idx, column_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample = [str(column_name)]
        if not df.empty:
            sample.extend(df[column_name].head(200).astype(str).tolist())
        width = min(max(max((len(text) for text in sample), default=8) + 2, 10), 32)
        ws.column_dimensions[letter].width = width

        numeric_like = pd.api.types.is_numeric_dtype(df[column_name]) or infer_numeric_like_series(df[column_name])
        if numeric_like:
            number_format = "#,##0.0" if str(column_name) == "DOI" else "#,##0"
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for item in cell:
                    item.number_format = number_format
                    item.alignment = Alignment(horizontal="right")


def dataframes_to_excel_bytes(sheets: list[tuple[str, pd.DataFrame]]) -> bytes:
    output = BytesIO()
    sheet_items = sheets if sheets else [("data", pd.DataFrame())]
    used_names: set[str] = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheet_items:
            safe_sheet = unique_excel_sheet_name(sheet_name, used_names)
            df.to_excel(writer, index=False, sheet_name=safe_sheet)
            format_excel_worksheet(writer.sheets[safe_sheet], df)
    output.seek(0)
    return output.getvalue()


def render_lazy_excel_download_button(
    label: str,
    df: pd.DataFrame,
    sheet_name: str,
    file_name: str,
    key: str,
) -> None:
    prepare_key = f"{key}_prepare"
    if st.button("엑셀 파일 생성", key=f"{key}_prepare_button", width="content"):
        st.session_state[prepare_key] = True
    if not st.session_state.get(prepare_key, False):
        st.caption("다운로드가 필요할 때만 엑셀 파일을 생성합니다.")
        return
    st.download_button(
        label,
        data=dataframe_to_excel_bytes(df, sheet_name=sheet_name),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        width="content",
    )


def render_lazy_excel_download_builder(
    label: str,
    data_builder,
    sheet_name: str,
    file_name: str,
    key: str,
    source_key: str = "",
) -> None:
    prepare_key = f"{key}_prepare"
    data_key = f"{key}_excel_data"
    file_key = f"{key}_excel_file_name"
    source_state_key = f"{key}_excel_source_key"
    if st.session_state.get(source_state_key) != source_key:
        st.session_state.pop(prepare_key, None)
        st.session_state.pop(data_key, None)
        st.session_state.pop(file_key, None)
        st.session_state[source_state_key] = source_key
    if st.button("엑셀 파일 생성", key=f"{key}_prepare_button", width="content"):
        st.session_state[prepare_key] = True
        df = data_builder()
        st.session_state[data_key] = dataframe_to_excel_bytes(df, sheet_name=sheet_name)
        st.session_state[file_key] = file_name
    if not st.session_state.get(prepare_key, False) or data_key not in st.session_state:
        st.caption("다운로드가 필요할 때만 엑셀 파일을 생성합니다.")
        return
    st.download_button(
        label,
        data=st.session_state[data_key],
        file_name=st.session_state.get(file_key, file_name),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        width="content",
    )


def split_query_terms(query: str) -> list[str]:
    return [term.strip() for term in str(query).split(",") if term.strip()]


def filter_with_terms(df: pd.DataFrame, column: str, query: str) -> pd.DataFrame:
    terms = split_query_terms(query)
    if not terms:
        return df
    pattern = "|".join(re.escape(term) for term in terms)
    return df[df[column].astype(str).str.contains(pattern, case=False, na=False)]


def filter_with_terms_any(df: pd.DataFrame, columns: list[str], query: str) -> pd.DataFrame:
    terms = split_query_terms(query)
    if not terms:
        return df

    pattern = "|".join(re.escape(term) for term in terms)
    mask = pd.Series(False, index=df.index)
    for col in columns:
        mask = mask | df[col].astype(str).str.contains(pattern, case=False, na=False)
    return df[mask]


def filter_display_table_with_query(df: pd.DataFrame, query: str) -> pd.DataFrame:
    if df.empty:
        return df
    search_columns = df.columns.tolist()
    if not search_columns:
        return df
    return filter_with_terms_any(df, search_columns, query)


def normalize_warehouse_name(value: str) -> str:
    return re.sub(r"\s+", "", str(value)).strip().lower()


def find_warehouse_column(columns: list[str], candidates: list[str]) -> str | None:
    normalized_map = {normalize_warehouse_name(col): col for col in columns}
    for candidate in candidates:
        matched = normalized_map.get(normalize_warehouse_name(candidate))
        if matched is not None:
            return matched
    return None


def style_leadji_shortage_table(display_df: pd.DataFrame, source_df: pd.DataFrame):
    if display_df.empty:
        return display_df.style

    styler = display_df.style
    if "우선순위" in display_df.columns:
        styler = styler.set_properties(subset=["우선순위"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: (
                "background-color: #FEE2E2; color: #B91C1C; font-weight: 800;"
                if str(v).strip() == "긴급"
                else "background-color: #FEF3C7; color: #92400E; font-weight: 800;"
                if str(v).strip() == "확인필요"
                else "background-color: #F1F5F9; color: #475569; font-weight: 700;"
            ),
            subset=["우선순위"],
        )
    if "리드지부족" in display_df.columns:
        styler = styler.set_properties(subset=["리드지부족"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: "color: #d00000; font-weight: 700;" if str(v).strip() not in {"", "-", "nan", "None"} else "",
            subset=["리드지부족"],
        )
    if "리드지부족수량" in display_df.columns and "리드지부족수량" in source_df.columns:
        shortage_numeric = parse_mixed_numeric(source_df["리드지부족수량"])
        shortage_style = shortage_numeric.map(
            lambda v: "color: #DC2626; font-weight: 850;" if pd.notna(v) and v < 0 else "color: #6B7280;"
        )
        styler = styler.apply(lambda _: shortage_style, axis=0, subset=["리드지부족수량"])
    if "상태" in display_df.columns:
        styler = styler.set_properties(subset=["상태"], **{"text-align": "center"})
        styler = styler.map(
            lambda v: (
                "background-color: #FEE2E2; color: #B91C1C; font-weight: 800;"
                if str(v).strip() == "입고일 미확인"
                else "background-color: #FEF3C7; color: #92400E; font-weight: 800;"
                if str(v).strip() == "발주부족"
                else "background-color: #EEF2FF; color: #1A2B5E; font-weight: 800;"
                if str(v).strip() in {"입고 예정", "입고 예정+의뢰"}
                else "background-color: #D1FAE5; color: #047857; font-weight: 800;"
                if str(v).strip() == "구매의뢰"
                else "background-color: #F1F5F9; color: #475569; font-weight: 700;"
            ),
            subset=["상태"],
        )
    return styler


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def add_rq_group_columns(df: pd.DataFrame) -> pd.DataFrame:
    enriched = df.copy()
    if "R코드" not in enriched.columns:
        enriched["R코드"] = enriched["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "R"))
    if "Q코드" not in enriched.columns:
        enriched["Q코드"] = enriched["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "Q"))
    if "U코드" not in enriched.columns:
        enriched["U코드"] = ""
    if "R코드 제품명" not in enriched.columns:
        enriched["R코드 제품명"] = enriched.get("제품명", "-")
    enriched["R코드5"] = enriched["R코드"].astype(str).str[:5]
    enriched["Q코드5"] = enriched["Q코드"].astype(str).str[:5]
    enriched["P코드5"] = enriched["품목코드"].astype(str).str[:5]
    enriched["RQ그룹"] = enriched["R코드"].astype(str) + " | " + enriched["Q코드"].astype(str)
    return enriched


def build_synthetic_p_rows_for_process_scope(
    source_scope: pd.DataFrame,
    p_reference_scope: pd.DataFrame,
    template_columns: list[str],
) -> pd.DataFrame:
    if source_scope.empty or p_reference_scope.empty or "품목코드" not in source_scope.columns:
        return pd.DataFrame(columns=template_columns)
    if "품목코드" not in p_reference_scope.columns:
        return pd.DataFrame(columns=template_columns)

    p_ref = p_reference_scope[p_reference_scope["품목코드"].astype(str).str.upper().str.startswith("P")].copy()
    if p_ref.empty:
        return pd.DataFrame(columns=template_columns)

    p_ref["P코드5"] = p_ref["품목코드"].astype(str).str.upper().str[:5]
    p_ref_name_cols = ["품목코드", "제품명"] if "제품명" in p_ref.columns else ["품목코드"]
    p_by_rq = (
        p_ref.drop_duplicates(subset=["R코드", "Q코드"], keep="first")[["R코드", "Q코드", *p_ref_name_cols]]
        .rename(columns={"품목코드": "_derived_p_code", "제품명": "_derived_p_name"})
        if {"R코드", "Q코드"}.issubset(p_ref.columns)
        else pd.DataFrame()
    )
    p_by_q = (
        p_ref.drop_duplicates(subset=["Q코드"], keep="first")[["Q코드", *p_ref_name_cols]]
        .rename(columns={"품목코드": "_derived_p_code", "제품명": "_derived_p_name"})
        if "Q코드" in p_ref.columns
        else pd.DataFrame()
    )

    actual_r_keys = [c for c in [ORDER_NO_COL, "거래처", "이니셜", "R코드", "Q코드"] if c in p_ref.columns]
    actual_q_keys = [c for c in [ORDER_NO_COL, "거래처", "이니셜", "Q코드"] if c in p_ref.columns]
    actual_r = p_ref[actual_r_keys].drop_duplicates() if actual_r_keys else pd.DataFrame()
    actual_q = p_ref[actual_q_keys].drop_duplicates() if actual_q_keys else pd.DataFrame()

    candidates: list[pd.DataFrame] = []
    source_prefix = source_scope["품목코드"].astype(str).str.upper().str[:1]
    if "사출생산필요수량" in source_scope.columns and {"R코드", "Q코드"}.issubset(source_scope.columns):
        r_source = source_scope[source_prefix == "R"].copy()
        r_source["사출생산필요수량"] = parse_mixed_numeric(r_source["사출생산필요수량"])
        r_source = r_source[r_source["사출생산필요수량"] > 0]
        r_keys = [c for c in [ORDER_NO_COL, "거래처", "이니셜", "R코드", "Q코드"] if c in r_source.columns]
        if r_keys and not r_source.empty:
            r_source = r_source.merge(p_by_rq, on=["R코드", "Q코드"], how="left") if not p_by_rq.empty else r_source
            if "_derived_p_code" not in r_source.columns:
                r_source["_derived_p_code"] = ""
            r_code_text = r_source["_derived_p_code"].astype(str).str.strip()
            r_source = r_source[
                r_source["_derived_p_code"].notna()
                & (r_code_text != "")
                & (~r_code_text.str.lower().isin(INVALID_CATEGORY_VALUES))
            ]
            if not r_source.empty and actual_r_keys == r_keys and not actual_r.empty:
                r_source = r_source.merge(actual_r, on=r_keys, how="left", indicator=True)
                r_source = r_source[r_source["_merge"] == "left_only"].drop(columns=["_merge"])
            candidates.append(r_source)

    if SEPARATION_REQUIRED_QTY_COL in source_scope.columns and "Q코드" in source_scope.columns:
        q_source = source_scope[source_prefix == "Q"].copy()
        q_source[SEPARATION_REQUIRED_QTY_COL] = parse_mixed_numeric(q_source[SEPARATION_REQUIRED_QTY_COL])
        q_source = q_source[q_source[SEPARATION_REQUIRED_QTY_COL] > 0]
        q_keys = [c for c in [ORDER_NO_COL, "거래처", "이니셜", "Q코드"] if c in q_source.columns]
        if q_keys and not q_source.empty:
            q_source = q_source.merge(p_by_q, on=["Q코드"], how="left") if not p_by_q.empty else q_source
            if "_derived_p_code" not in q_source.columns:
                q_source["_derived_p_code"] = ""
            q_code_text = q_source["_derived_p_code"].astype(str).str.strip()
            q_source = q_source[
                q_source["_derived_p_code"].notna()
                & (q_code_text != "")
                & (~q_code_text.str.lower().isin(INVALID_CATEGORY_VALUES))
            ]
            if "R코드" not in q_source.columns:
                q_source["R코드"] = q_source["Q코드"].map(lambda code: map_demand_code_to_process_code(code, "R"))
            if not q_source.empty and actual_q_keys == q_keys and not actual_q.empty:
                q_source = q_source.merge(actual_q, on=q_keys, how="left", indicator=True)
                q_source = q_source[q_source["_merge"] == "left_only"].drop(columns=["_merge"])
            candidates.append(q_source)

    if not candidates:
        return pd.DataFrame(columns=template_columns)

    combined = pd.concat(candidates, ignore_index=True, sort=False)
    if combined.empty:
        return pd.DataFrame(columns=template_columns)

    group_cols = [
        c
        for c in [ORDER_NO_COL, "거래처", "이니셜", "R코드", "Q코드", "_derived_p_code"]
        if c in combined.columns
    ]
    first_rows = combined.sort_values(group_cols).drop_duplicates(subset=group_cols, keep="first").copy()
    synthetic = pd.DataFrame(index=first_rows.index, columns=template_columns)
    for col in template_columns:
        if col in first_rows.columns:
            synthetic[col] = first_rows[col]

    synthetic["품목코드"] = first_rows["_derived_p_code"].values
    if "제품명" in synthetic.columns:
        synthetic["제품명"] = first_rows["_derived_p_name"] if "_derived_p_name" in first_rows.columns else pd.NA
        if "제품명" in first_rows.columns:
            source_name = first_rows["제품명"].astype(str).str.strip()
            prefer_source_name = source_name.str.contains("PIA_KR", case=False, na=False, regex=False)
            synthetic["제품명"] = synthetic["제품명"].where(~prefer_source_name, first_rows["제품명"])
            synthetic["제품명"] = synthetic["제품명"].fillna(first_rows["제품명"])
        synthetic["제품명"] = synthetic["제품명"].fillna("-")

    for col in [
        DEMAND_QTY_COL,
        "부족수량",
        "사출생산필요수량",
        SEPARATION_REQUIRED_QTY_COL,
        LEADJI_REQUIRED_QTY_COL,
        ADHESION_REQUIRED_QTY_COL,
    ]:
        if col in synthetic.columns:
            synthetic[col] = 0.0

    return synthetic.reset_index(drop=True)


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_rcode_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "R코드",
        "R코드 제품명",
        "사출 납기일",
        "사출 생산 필요수량 합계",
        "사출창고 합계",
        "분리창고 합계",
        "공정재고 합계",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)

    r_df = df.copy()
    r_df["R코드"] = r_df["R코드"].astype(str).str.strip()
    r_df = r_df[(r_df["R코드"] != "") & (r_df["R코드"].str.lower() != "nan")]
    if "사출생산필요수량" in r_df.columns:
        r_df["사출생산필요수량"] = parse_mixed_numeric(r_df["사출생산필요수량"])
        r_df = r_df[r_df["사출생산필요수량"] > 0]
    if r_df.empty:
        return pd.DataFrame(columns=columns)

    grouped = (
        r_df.groupby("R코드", as_index=False)
        .agg(
            {
                "R코드 제품명": lambda s: summarize_unique(s, head_count=1),
                "사출납기일": "min",
                "사출생산필요수량": "sum",
                "사출창고": "sum",
                "분리창고": "sum",
                "공정재고 합계": "sum",
            }
        )
        .rename(
            columns={
                "사출납기일": "사출 납기일",
                "사출생산필요수량": "사출 생산 필요수량 합계",
                "사출창고": "사출창고 합계",
                "분리창고": "분리창고 합계",
            }
        )
    )

    grouped = grouped.sort_values(["사출 생산 필요수량 합계", "R코드"], ascending=[False, True])
    return grouped[columns]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_rq_group_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "R코드",
        "Q코드",
        "R코드 제품명",
        "P코드5 수",
        "제품명 예시",
        "P코드 예시",
        "부족수량 합계",
        "사출 생산 필요수량 합계",
        "사출창고 합계",
        "분리창고 합계",
        "공정재고 합계",
        "사출 부족수량",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)
    if "부족수량" not in df.columns:
        return pd.DataFrame(columns=columns)

    df = df.copy()
    df["부족수량"] = parse_mixed_numeric(df["부족수량"])
    if "사출생산필요수량" in df.columns:
        df["사출생산필요수량"] = parse_mixed_numeric(df["사출생산필요수량"])
    else:
        df["사출생산필요수량"] = 0
    df = df[(df["부족수량"] > 0) | (df["사출생산필요수량"] > 0)]
    if df.empty:
        return pd.DataFrame(columns=columns)

    grouped = (
        df.groupby(["R코드5", "Q코드5"], as_index=False)
        .agg(
            {
                "R코드 제품명": lambda s: summarize_unique(s, head_count=1),
                "제품명": lambda s: summarize_unique(s, head_count=3),
                "품목코드": lambda s: summarize_unique(s, head_count=5),
                "부족수량": "sum",
                "사출생산필요수량": "sum",
                "사출창고": "sum",
                "분리창고": "sum",
                "공정재고 합계": "sum",
            }
        )
        .rename(
            columns={
                "R코드5": "R코드",
                "Q코드5": "Q코드",
                "제품명": "제품명 예시",
                "품목코드": "P코드 예시",
                "부족수량": "부족수량 합계",
                "사출생산필요수량": "사출 생산 필요수량 합계",
                "사출창고": "사출창고 합계",
                "분리창고": "분리창고 합계",
            }
        )
    )
    p_count = df.groupby(["R코드5", "Q코드5"])["P코드5"].nunique().rename("P코드5 수").reset_index()
    p_count = p_count.rename(columns={"R코드5": "R코드", "Q코드5": "Q코드"})
    grouped = grouped.merge(p_count, on=["R코드", "Q코드"], how="left")
    grouped["사출 부족수량"] = grouped["사출 생산 필요수량 합계"]
    grouped = grouped.sort_values(["부족수량 합계", "P코드5 수"], ascending=[False, False])
    return grouped[columns]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_initial_injection_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "이니셜",
        "거래처 수",
        "품목코드 수",
        "부족수량 합계",
        "사출 생산 필요수량 합계",
        "사출창고 합계",
        "사출 부족수량",
    ]
    if df.empty or "이니셜" not in df.columns or "품목코드" not in df.columns:
        return pd.DataFrame(columns=columns)

    base = df.copy()
    if "거래처" not in base.columns:
        base["거래처"] = "-"
    if "부족수량" not in base.columns:
        base["부족수량"] = 0
    if "사출생산필요수량" not in base.columns:
        base["사출생산필요수량"] = 0
    if "사출창고" not in base.columns:
        base["사출창고"] = 0

    base["이니셜"] = base["이니셜"].astype(str).str.strip()
    base["이니셜"] = base["이니셜"].replace({"": "(미지정)", "nan": "(미지정)", "None": "(미지정)"})
    base["부족수량"] = parse_mixed_numeric(base["부족수량"])
    base["사출생산필요수량"] = parse_mixed_numeric(base["사출생산필요수량"])
    base["사출창고"] = parse_mixed_numeric(base["사출창고"])
    base = base[(base["부족수량"] > 0) | (base["사출생산필요수량"] > 0)]
    if base.empty:
        return pd.DataFrame(columns=columns)

    # 사출창고는 품목별 고정 재고 성격이라, 이니셜+품목 기준 최대값으로 중복 집계를 완화한다.
    item_level = (
        base.groupby(["이니셜", "품목코드"], as_index=False)
        .agg(
            {
                "부족수량": "sum",
                "사출생산필요수량": "sum",
                "사출창고": "max",
            }
        )
    )

    summary = (
        item_level.groupby("이니셜", as_index=False)
        .agg(
            {
                "품목코드": "nunique",
                "부족수량": "sum",
                "사출생산필요수량": "sum",
                "사출창고": "sum",
            }
        )
        .rename(
            columns={
                "품목코드": "품목코드 수",
                "부족수량": "부족수량 합계",
                "사출생산필요수량": "사출 생산 필요수량 합계",
                "사출창고": "사출창고 합계",
            }
        )
    )

    customer_count = (
        base.groupby("이니셜", as_index=False)["거래처"]
        .nunique()
        .rename(columns={"거래처": "거래처 수"})
    )
    summary = summary.merge(customer_count, on="이니셜", how="left")
    summary["사출 부족수량"] = summary["사출 생산 필요수량 합계"]
    summary = summary[columns].sort_values(
        ["사출 부족수량", "사출 생산 필요수량 합계", "부족수량 합계", "이니셜"],
        ascending=[False, False, False, True],
    )
    return summary


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_qcode_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Q코드",
        "Q기준 제품명",
        "파워",
        "대표 이니셜",
        "대표 P코드",
        "분리 생산 필요수량 합계",
        "부족수량 합계",
        "분리창고",
        "사출창고",
        "공정재고 합계",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)
    if "부족수량" not in df.columns:
        return pd.DataFrame(columns=columns)

    q_df = df.copy()
    q_df["부족수량"] = parse_mixed_numeric(q_df["부족수량"])
    if SEPARATION_REQUIRED_QTY_COL in q_df.columns:
        q_df[SEPARATION_REQUIRED_QTY_COL] = parse_mixed_numeric(q_df[SEPARATION_REQUIRED_QTY_COL])
    else:
        q_df[SEPARATION_REQUIRED_QTY_COL] = 0
    q_df = q_df[(q_df["부족수량"] > 0) | (q_df[SEPARATION_REQUIRED_QTY_COL] > 0)]
    if q_df.empty:
        return pd.DataFrame(columns=columns)

    if "Q코드" in q_df.columns:
        q_df["Q코드"] = q_df["Q코드"].astype(str).str.strip()
    else:
        q_df["Q코드"] = q_df["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "Q"))
    q_df["파워"] = q_df["Q코드"].map(extract_power_from_code)

    summary = (
        q_df.groupby(["Q코드", "파워"], as_index=False)
        .agg(
            {
                "제품명": lambda s: summarize_unique(s, head_count=1),
                "이니셜": lambda s: summarize_unique(s, head_count=1),
                "품목코드": lambda s: summarize_unique(s, head_count=1),
                SEPARATION_REQUIRED_QTY_COL: "sum",
                "부족수량": "sum",
                "분리창고": "max",
                "사출창고": "max",
                "공정재고 합계": "max",
            }
        )
        .rename(
            columns={
                "제품명": "Q기준 제품명",
                "이니셜": "대표 이니셜",
                "품목코드": "대표 P코드",
                SEPARATION_REQUIRED_QTY_COL: "분리 생산 필요수량 합계",
                "부족수량": "부족수량 합계",
            }
        )
        .sort_values(["분리 생산 필요수량 합계", "부족수량 합계", "Q코드"], ascending=[False, False, True])
    )
    return summary[columns]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_summary_group_totals_with_safe_split(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "분류별요약",
        "오더 기준 부족수량",
        "오더 기준 사출부족수량",
        "안전재고 기준 부족수량",
        "안전재고 기준 사출부족수량",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)

    base = df.copy()
    if "분류별요약" not in base.columns:
        base["분류별요약"] = "(미분류)"
    if "이니셜" not in base.columns:
        base["이니셜"] = ""
    for qty_col in ["부족수량", "사출생산필요수량"]:
        if qty_col not in base.columns:
            base[qty_col] = 0
        base[qty_col] = parse_mixed_numeric(base[qty_col])

    group_label = base["분류별요약"].astype(str).str.strip()
    base["분류별요약"] = group_label.replace({"": "(미분류)", "nan": "(미분류)", "None": "(미분류)"})

    initial_text = base["이니셜"].map(clean_text_value)
    safe_mask = initial_text.str.contains("안전", na=False)
    order_mask = ~safe_mask

    order_qty = (
        base[order_mask]
        .groupby("분류별요약", as_index=False)[["부족수량", "사출생산필요수량"]]
        .sum()
        .rename(
            columns={
                "부족수량": "오더 기준 부족수량",
                "사출생산필요수량": "오더 기준 사출부족수량",
            }
        )
    )
    safe_qty = (
        base[safe_mask]
        .groupby("분류별요약", as_index=False)[["부족수량", "사출생산필요수량"]]
        .sum()
        .rename(
            columns={
                "부족수량": "안전재고 기준 부족수량",
                "사출생산필요수량": "안전재고 기준 사출부족수량",
            }
        )
    )

    grouped = order_qty.merge(safe_qty, on="분류별요약", how="outer").fillna(0)
    grouped["_정렬합계"] = (
        grouped["오더 기준 부족수량"]
        + grouped["오더 기준 사출부족수량"]
        + grouped["안전재고 기준 부족수량"]
        + grouped["안전재고 기준 사출부족수량"]
    )
    grouped = grouped.sort_values(["_정렬합계", "분류별요약"], ascending=[False, True]).drop(columns=["_정렬합계"])
    grouped = grouped[columns]

    total_row = pd.DataFrame(
        [
            {
                "분류별요약": "전체",
                "오더 기준 부족수량": grouped["오더 기준 부족수량"].sum(),
                "오더 기준 사출부족수량": grouped["오더 기준 사출부족수량"].sum(),
                "안전재고 기준 부족수량": grouped["안전재고 기준 부족수량"].sum(),
                "안전재고 기준 사출부족수량": grouped["안전재고 기준 사출부족수량"].sum(),
            }
        ]
    )
    return pd.concat([total_row, grouped], ignore_index=True)


def load_raw_data(
    refresh_key: str, base_dir_str: str | None = None
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, object], set[str], dict[str, object], str, str]:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    inv_path, dem_path = find_excel_files(data_base_dir)
    (
        process_code_map,
        warehouse_qty_col_indices,
        qty_col_indices,
        total_qty_col_indices,
        process_qty_col_indices,
        header_labels,
    ) = extract_demand_header_info(dem_path)
    demand_read_plan = build_demand_read_plan(
        header_labels,
        warehouse_qty_col_indices,
        qty_col_indices,
        total_qty_col_indices,
        process_qty_col_indices,
    )
    demand_read_plan["process_code_map"] = process_code_map

    inv = read_inventory_excel_subset(inv_path)
    dem = read_demand_excel_subset(dem_path, demand_read_plan["usecols"])
    rework_item_qty_map, rework_meta = read_rework_item_keys_from_demand_file(dem_path)
    return inv, dem, demand_read_plan, rework_item_qty_map, rework_meta, inv_path.name, dem_path.name


@st.cache_resource(show_spinner=False)
def preprocess_data(refresh_key: str, base_dir_str: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    inv, dem, demand_read_plan, rework_item_qty_map, rework_meta, inv_file_name, dem_file_name = load_raw_data(
        refresh_key, base_dir_str
    )
    reference_refresh_key = build_reference_refresh_key(data_base_dir)
    (
        product_name_map,
        product_group_map,
        sheet2_group_map,
        r_ref_map,
        q_ref_map,
        r_name_map,
        bom_r_base_map,
        bom_q_base_map,
        bom_r_exact_map,
        bom_q_exact_map,
        leadji_r_map,
        leadji_q_map,
        leadji_u_map,
    ) = load_reference_maps_bundle(data_base_dir, reference_refresh_key)
    process_code_map = demand_read_plan.get("process_code_map", {})
    qty_col_indices = list(demand_read_plan.get("qty_col_indices", []))
    total_qty_col_indices = list(demand_read_plan.get("total_qty_col_indices", []))

    def dem_series(idx: object, default: object = "") -> pd.Series:
        if isinstance(idx, int) and idx in dem.columns:
            return dem[idx]
        return pd.Series(default, index=dem.index)

    site_col_idx = demand_read_plan.get("site_col_idx")
    customer_col_idx = demand_read_plan.get("customer_col_idx")
    order_no_col_idx = demand_read_plan.get("order_no_col_idx")
    initial_col_idx = demand_read_plan.get("initial_col_idx")
    demand_item_col_idx = demand_read_plan.get("demand_item_col_idx")
    demand_name_col_idx = demand_read_plan.get("demand_name_col_idx")
    demand_qty_idx = demand_read_plan.get("demand_qty_idx")

    site_series = (
        dem_series(site_col_idx).astype(str).str.strip()
        if site_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    customer_series = (
        dem_series(customer_col_idx).astype(str).str.strip()
        if customer_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    order_no_series = (
        dem_series(order_no_col_idx).astype(str).str.strip()
        if order_no_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    initial_series = (
        dem_series(initial_col_idx).astype(str).str.strip()
        if initial_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    item_series = (
        dem_series(demand_item_col_idx).astype(str).str.strip()
        if demand_item_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    name_series = (
        dem_series(demand_name_col_idx).astype(str).str.strip()
        if demand_name_col_idx is not None
        else pd.Series("", index=dem.index)
    )
    demand_qty = (
        parse_mixed_numeric(dem_series(demand_qty_idx))
        if isinstance(demand_qty_idx, int) and demand_qty_idx in dem.columns
        else pd.Series(0.0, index=dem.index)
    )

    # 기준1) 생산 현황: 누수/규격검사 생산수량 + 납기일
    leak_qty_idx = demand_read_plan.get("leak_qty_idx")
    leak_due_idx = demand_read_plan.get("leak_due_idx")
    if isinstance(leak_qty_idx, int) and leak_qty_idx in dem.columns:
        shortage_qty = parse_mixed_numeric(dem[leak_qty_idx])
    elif total_qty_col_indices:
        total_qty_idx = total_qty_col_indices[-1]
        if not isinstance(total_qty_idx, int) or total_qty_idx not in dem.columns:
            raise ValueError("수요 파일에서 '총합계 생산 수량' 컬럼을 찾지 못했습니다.")
        shortage_qty = parse_mixed_numeric(dem[total_qty_idx])
    else:
        qty_cols = [i for i in qty_col_indices if isinstance(i, int) and i in dem.columns]
        if not qty_cols:
            raise ValueError("수요 파일에서 '생산 수량' 컬럼을 찾지 못했습니다.")
        shortage_qty = dem[qty_cols].apply(parse_mixed_numeric).fillna(0).sum(axis=1)

    if isinstance(leak_due_idx, int) and leak_due_idx in dem.columns:
        leak_due_date = parse_mixed_excel_date(dem[leak_due_idx])
    else:
        leak_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")

    separation_qty_idx = demand_read_plan.get("separation_qty_idx")
    separation_due_idx = demand_read_plan.get("separation_due_idx")
    if isinstance(separation_qty_idx, int) and separation_qty_idx in dem.columns:
        separation_required_qty = parse_mixed_numeric(dem[separation_qty_idx])
        if isinstance(separation_due_idx, int) and separation_due_idx in dem.columns:
            separation_due_date = parse_mixed_excel_date(dem[separation_due_idx])
        else:
            separation_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")
    else:
        separation_required_qty = pd.Series(0.0, index=dem.index)
        separation_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")

    adhesion_qty_idx = demand_read_plan.get("adhesion_qty_idx")
    adhesion_due_idx = demand_read_plan.get("adhesion_due_idx")
    if isinstance(adhesion_qty_idx, int) and adhesion_qty_idx in dem.columns:
        adhesion_required_qty = parse_mixed_numeric(dem[adhesion_qty_idx])
        if isinstance(adhesion_due_idx, int) and adhesion_due_idx in dem.columns:
            adhesion_due_date = parse_mixed_excel_date(dem[adhesion_due_idx])
        else:
            adhesion_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")
    else:
        adhesion_required_qty = pd.Series(0.0, index=dem.index)
        adhesion_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")

    leadji_qty_idx = demand_read_plan.get("leadji_qty_idx")
    leadji_due_idx = demand_read_plan.get("leadji_due_idx")
    if isinstance(leadji_qty_idx, int) and leadji_qty_idx in dem.columns:
        leadji_required_qty = parse_mixed_numeric(dem[leadji_qty_idx])
        if isinstance(leadji_due_idx, int) and leadji_due_idx in dem.columns:
            leadji_due_date = parse_mixed_excel_date(dem[leadji_due_idx])
        else:
            leadji_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")
    else:
        leadji_required_qty = shortage_qty
        leadji_due_date = leak_due_date

    # 기준2) 사출 생산 현황: [10]사출조립 생산수량 + 해당 납기일
    # 파일 구조가 바뀌어도 공정 헤더 기준으로 우선 선택한다.
    selected_qty_idx = demand_read_plan.get("selected_qty_idx")
    if isinstance(selected_qty_idx, int) and selected_qty_idx in dem.columns:
        inj_qty = parse_mixed_numeric(dem[selected_qty_idx])
    else:
        raise ValueError("수요 파일 사출조립 생산수량 컬럼을 찾지 못했습니다.")

    inj_due_idx = demand_read_plan.get("inj_due_idx")
    if isinstance(inj_due_idx, int) and inj_due_idx in dem.columns:
        inj_due_date = parse_mixed_excel_date(dem[inj_due_idx])
    else:
        inj_due_date = pd.Series(pd.NaT, index=dem.index, dtype="datetime64[ns]")

    inv_df = build_inventory_df(inv)

    dem_df = pd.DataFrame(
        {
            "사이트코드": site_series,
            "거래처": customer_series,
            ORDER_NO_COL: order_no_series,
            "이니셜": initial_series,
            "품목코드": item_series,
            "제품명": name_series,
            DEMAND_QTY_COL: demand_qty,
            "납기일": leak_due_date,
            "사출납기일": inj_due_date,
            SEPARATION_REQUIRED_DUE_COL: separation_due_date,
            LEADJI_REQUIRED_DUE_COL: leadji_due_date,
            ADHESION_REQUIRED_DUE_COL: adhesion_due_date,
            "생산수량": shortage_qty,
            "사출생산필요수량": inj_qty,
            SEPARATION_REQUIRED_QTY_COL: separation_required_qty,
            LEADJI_REQUIRED_QTY_COL: leadji_required_qty,
            ADHESION_REQUIRED_QTY_COL: adhesion_required_qty,
        }
    )

    is_summary = (
        (dem_df["사이트코드"] == "총합계")
        | (dem_df["거래처"] == "총합계")
        | (dem_df["이니셜"] == "총합계")
        | (dem_df["품목코드"] == "총합계")
    )
    dem_df = dem_df[~is_summary]
    dem_df = dem_df[(dem_df["사이트코드"] != "") & (dem_df["사이트코드"].str.lower() != "nan")]
    dem_df = dem_df[(dem_df["품목코드"] != "") & (dem_df["품목코드"].str.lower() != "nan")]
    dem_df = dem_df[dem_df["품목코드"].astype(str).str.upper().str.startswith(("P", "Q", "R"))]
    dem_df = dem_df[
        (dem_df[DEMAND_QTY_COL] > 0)
        | (dem_df["생산수량"] > 0)
        | (dem_df["사출생산필요수량"] > 0)
        | (dem_df[SEPARATION_REQUIRED_QTY_COL] > 0)
        | (dem_df[LEADJI_REQUIRED_QTY_COL] > 0)
        | (dem_df[ADHESION_REQUIRED_QTY_COL] > 0)
    ]
    dem_df["제품명"] = dem_df["제품명"].replace({"nan": "", "None": ""})

    grouped_demand = (
        dem_df.groupby(
            [
                "사이트코드",
                "이니셜",
                "거래처",
                ORDER_NO_COL,
                "품목코드",
                "납기일",
                "사출납기일",
                SEPARATION_REQUIRED_DUE_COL,
                LEADJI_REQUIRED_DUE_COL,
                ADHESION_REQUIRED_DUE_COL,
            ],
            as_index=False,
            dropna=False,
        )
        .agg(
            {
                DEMAND_QTY_COL: "sum",
                "생산수량": "sum",
                "사출생산필요수량": "sum",
                SEPARATION_REQUIRED_QTY_COL: "sum",
                LEADJI_REQUIRED_QTY_COL: "sum",
                ADHESION_REQUIRED_QTY_COL: "sum",
                "제품명": lambda s: next((v for v in s if str(v).strip() and str(v).strip().lower() != "nan"), "-"),
            }
        )
        .rename(columns={"생산수량": "부족수량"})
    )
    grouped_demand["코드5"] = grouped_demand["품목코드"].str[:5]
    grouped_demand["제품명"] = grouped_demand["코드5"].map(product_name_map).fillna(grouped_demand["제품명"])
    grouped_demand["제품명"] = grouped_demand["제품명"].replace({"": "-", "nan": "-", "None": "-"}).fillna("-")

    inferred_r = grouped_demand["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "R"))
    inferred_q = grouped_demand["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "Q"))
    mapped_r_base = grouped_demand["코드5"].map(leadji_r_map).fillna(grouped_demand["코드5"].map(r_ref_map))
    mapped_q_base = grouped_demand["코드5"].map(leadji_q_map).fillna(grouped_demand["코드5"].map(q_ref_map))

    merged_r = pd.Series(
        [merge_mapped_base_code(inferred, mapped, "R") for inferred, mapped in zip(inferred_r, mapped_r_base)],
        index=grouped_demand.index,
    )
    merged_q = pd.Series(
        [merge_mapped_base_code(inferred, mapped, "Q") for inferred, mapped in zip(inferred_q, mapped_q_base)],
        index=grouped_demand.index,
    )

    # 리드지정보/분류정보의 공정 base가 있으면 P행에도 우선 적용하고, 없을 때만 P->R/Q 추론을 사용한다.
    grouped_demand["R코드"] = merged_r
    grouped_demand["Q코드"] = merged_q

    # BOM exact mapping has the highest priority when TO_ITEM_ID matches the demand item code.
    if bom_r_exact_map or bom_q_exact_map:
        bom_exact_r = grouped_demand["품목코드"].map(bom_r_exact_map)
        bom_exact_q = grouped_demand["품목코드"].map(bom_q_exact_map)
        exact_r_mask = bom_exact_r.notna() & (bom_exact_r.astype(str).str.strip() != "")
        exact_q_mask = bom_exact_q.notna() & (bom_exact_q.astype(str).str.strip() != "")
        grouped_demand.loc[exact_r_mask, "R코드"] = bom_exact_r.loc[exact_r_mask]
        grouped_demand.loc[exact_q_mask, "Q코드"] = bom_exact_q.loc[exact_q_mask]

    # BOM fallback: only for rows still not mappable as valid R/Q codes.
    if bom_r_base_map or bom_q_base_map:
        bom_r_base = grouped_demand["코드5"].map(bom_r_base_map)
        bom_q_base = grouped_demand["코드5"].map(bom_q_base_map)

        bom_merged_r = pd.Series(
            [merge_mapped_base_code(inferred, mapped, "R") for inferred, mapped in zip(inferred_r, bom_r_base)],
            index=grouped_demand.index,
        )
        bom_merged_q = pd.Series(
            [merge_mapped_base_code(inferred, mapped, "Q") for inferred, mapped in zip(inferred_q, bom_q_base)],
            index=grouped_demand.index,
        )

        r_norm = grouped_demand["R코드"].astype(str).str.strip()
        q_norm = grouped_demand["Q코드"].astype(str).str.strip()
        invalid_r_mask = (r_norm == "") | (r_norm.str.lower() == "nan") | (~r_norm.str.startswith("R"))
        invalid_q_mask = (q_norm == "") | (q_norm.str.lower() == "nan") | (~q_norm.str.startswith("Q"))

        grouped_demand.loc[invalid_r_mask, "R코드"] = bom_merged_r.loc[invalid_r_mask]
        grouped_demand.loc[invalid_q_mask, "Q코드"] = bom_merged_q.loc[invalid_q_mask]

    item_prefix = grouped_demand["품목코드"].astype(str).str.upper().str[:1]
    p_mask = item_prefix == "P"
    r_mask = item_prefix == "R"
    if p_mask.any() and r_mask.any():
        demand_r_codes = grouped_demand.loc[
            r_mask, ["사이트코드", "이니셜", "제품명", "R코드", "Q코드", "사출납기일"]
        ].copy()
        demand_r_codes["파워_매칭"] = demand_r_codes["R코드"].map(extract_power_key_from_code)
        demand_r_codes["제품명_매칭"] = demand_r_codes["제품명"].map(normalize_lookup_key)
        demand_r_codes["납기_매칭"] = (
            pd.to_datetime(demand_r_codes["사출납기일"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
        )
        demand_r_codes = demand_r_codes[
            demand_r_codes["R코드"].astype(str).str.startswith("R")
            & (demand_r_codes["파워_매칭"] != "-")
            & (demand_r_codes["제품명_매칭"] != "")
        ].copy()

        if not demand_r_codes.empty:
            demand_r_codes = demand_r_codes.rename(columns={"R코드": "R코드_수요", "Q코드": "Q코드_수요"})
            base_match_keys = ["사이트코드", "이니셜", "제품명_매칭", "파워_매칭"]
            dated_match_keys = [*base_match_keys, "납기_매칭"]
            demand_r_by_date = demand_r_codes.drop_duplicates(subset=dated_match_keys, keep="first")
            demand_r_by_base = demand_r_codes.drop_duplicates(subset=base_match_keys, keep="first")

            p_match = grouped_demand.loc[p_mask, ["사이트코드", "이니셜", "제품명", "R코드", "납기일"]].copy()
            p_match["_row_id"] = p_match.index
            p_match["파워_매칭"] = p_match["R코드"].map(extract_power_key_from_code)
            p_match["제품명_매칭"] = p_match["제품명"].map(normalize_lookup_key)
            p_match["납기_매칭"] = (
                pd.to_datetime(p_match["납기일"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
            )
            p_match = p_match.merge(
                demand_r_by_date[dated_match_keys + ["R코드_수요", "Q코드_수요"]],
                on=dated_match_keys,
                how="left",
            )

            demand_match_mask = p_match["R코드_수요"].astype(str).str.strip().ne("")
            demand_match_mask &= p_match["R코드_수요"].notna()
            if (~demand_match_mask).any():
                unmatched = p_match.loc[
                    ~demand_match_mask, ["_row_id", *base_match_keys]
                ].merge(
                    demand_r_by_base[base_match_keys + ["R코드_수요", "Q코드_수요"]],
                    on=base_match_keys,
                    how="left",
                )
                fallback_match_mask = unmatched["R코드_수요"].astype(str).str.strip().ne("")
                fallback_match_mask &= unmatched["R코드_수요"].notna()
                fallback_rows = unmatched.loc[fallback_match_mask, "_row_id"]
                grouped_demand.loc[fallback_rows, "R코드"] = unmatched.loc[
                    fallback_match_mask, "R코드_수요"
                ].to_numpy()
                grouped_demand.loc[fallback_rows, "Q코드"] = unmatched.loc[
                    fallback_match_mask, "Q코드_수요"
                ].to_numpy()

            matched_rows = p_match.loc[demand_match_mask, "_row_id"]
            grouped_demand.loc[matched_rows, "R코드"] = p_match.loc[demand_match_mask, "R코드_수요"].to_numpy()
            grouped_demand.loc[matched_rows, "Q코드"] = p_match.loc[demand_match_mask, "Q코드_수요"].to_numpy()

    grouped_demand["R코드5"] = grouped_demand["R코드"].astype(str).str[:5]
    grouped_demand["R코드 제품명"] = grouped_demand["R코드5"].map(r_name_map)
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].fillna(grouped_demand["제품명"])
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].fillna(grouped_demand["R코드5"])
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].replace({"": "-", "nan": "-", "None": "-"}).fillna("-")
    mapped_u_base = grouped_demand["코드5"].map(leadji_u_map)
    grouped_demand["U코드"] = pd.Series(
        [
            merge_mapped_base_code(map_demand_code_to_process_code(q_code, "U"), mapped, "U")
            if str(mapped).strip() and str(mapped).strip().lower() != "nan"
            else ""
            for q_code, mapped in zip(grouped_demand["Q코드"], mapped_u_base)
        ],
        index=grouped_demand.index,
    )
    grouped_demand["분류별요약"] = grouped_demand["코드5"].map(product_group_map).fillna("기타")
    grouped_demand["시트분류"] = grouped_demand["코드5"].map(sheet2_group_map)
    grouped_demand = grouped_demand.drop(columns=["코드5", "R코드5"])

    target_inv = inv_df[inv_df["창고"].isin(TARGET_WAREHOUSES)].copy()
    stock_lookup: dict[str, dict[str, float]] = {}
    for raw_name, display_name in WAREHOUSE_MAP.items():
        stock_lookup[display_name] = (
            target_inv[target_inv["창고"] == raw_name]
            .groupby("품목코드")["재고량"]
            .sum()
            .to_dict()
        )

    code_stock = pd.DataFrame({"품목코드": grouped_demand["품목코드"].drop_duplicates()})
    rq_by_p = grouped_demand.drop_duplicates(subset=["품목코드"], keep="first").set_index("품목코드")[
        ["R코드", "Q코드", "U코드"]
    ]
    r_by_p = {
        item_code: resolve_process_code_for_stock(stock_lookup["사출창고"], process_code)
        for item_code, process_code in rq_by_p["R코드"].to_dict().items()
    }
    q_by_p = {
        item_code: resolve_process_code_for_stock(stock_lookup["분리창고"], process_code)
        for item_code, process_code in rq_by_p["Q코드"].to_dict().items()
    }
    u_by_p = {
        item_code: resolve_process_code_for_stock(stock_lookup["분리창고"], process_code)
        for item_code, process_code in rq_by_p["U코드"].to_dict().items()
    }

    grouped_demand["R코드"] = grouped_demand["품목코드"].map(
        lambda x: r_by_p.get(x, map_demand_code_to_process_code(x, "R"))
    )
    grouped_demand["Q코드"] = grouped_demand["품목코드"].map(
        lambda x: q_by_p.get(x, map_demand_code_to_process_code(x, "Q"))
    )
    grouped_demand["U코드"] = grouped_demand["품목코드"].map(lambda x: u_by_p.get(x, ""))
    grouped_demand["R코드5"] = grouped_demand["R코드"].astype(str).str[:5]
    grouped_demand["R코드 제품명"] = grouped_demand["R코드5"].map(r_name_map)
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].fillna(grouped_demand["제품명"])
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].fillna(grouped_demand["R코드5"])
    grouped_demand["R코드 제품명"] = grouped_demand["R코드 제품명"].replace({"": "-", "nan": "-", "None": "-"}).fillna("-")
    grouped_demand, rework_matched_item_keys, rework_matched_qty_total = apply_rework_flags_to_demand_rows(
        grouped_demand,
        rework_item_qty_map,
        rework_meta,
    )

    code_stock["사출창고"] = code_stock["품목코드"].map(
        lambda x: lookup_stock_qty(stock_lookup["사출창고"], r_by_p.get(x, map_demand_code_to_process_code(x, "R")))
    )
    code_stock["분리창고"] = code_stock["품목코드"].map(
        lambda x: lookup_stock_qty_from_candidates(
            stock_lookup["분리창고"],
            [
                q_by_p.get(x, map_demand_code_to_process_code(x, "Q")),
                u_by_p.get(x, ""),
            ],
        )
    )
    code_stock["검사접착창고"] = code_stock["품목코드"].map(
        lambda x: stock_lookup["검사접착창고"].get(x, 0)
    )
    code_stock["검사접착재작업창고"] = code_stock["품목코드"].map(
        lambda x: stock_lookup["검사접착재작업창고"].get(x, 0)
    )
    code_stock["누수규격검사 창고"] = code_stock["품목코드"].map(
        lambda x: stock_lookup["누수규격검사 창고"].get(x, 0)
    )
    code_stock["공정재고 합계"] = (
        code_stock["사출창고"]
        + code_stock["분리창고"]
        + code_stock["검사접착창고"]
        + code_stock["검사접착재작업창고"]
        + code_stock["누수규격검사 창고"]
    )

    result = grouped_demand.merge(code_stock, on="품목코드", how="left")
    for col in ["사출창고", "분리창고", "검사접착창고", "검사접착재작업창고", "누수규격검사 창고", "공정재고 합계"]:
        result[col] = result[col].fillna(0)

    # 분류 필터 정합성 보정:
    # P코드는 코드5(Pxxxx) 기준 매핑을 그대로 사용하고,
    # R/Q/U 등 비-P코드는 같은 R코드5를 공유하는 P코드의 분류를 이어받는다.
    result["코드5"] = result["품목코드"].astype(str).str[:5]
    result["분류별요약"] = result["코드5"].map(product_group_map)
    result["시트분류"] = result["코드5"].map(sheet2_group_map)
    result["R코드5"] = result["R코드"].astype(str).str[:5]

    item_prefix = result["품목코드"].astype(str).str.upper().str[:1]
    p_scope = result[(item_prefix == "P") & result["R코드5"].str.startswith("R", na=False)].copy()
    if not p_scope.empty:
        p_scope["부족수량_num"] = parse_mixed_numeric(p_scope["부족수량"])
        p_scope = p_scope.sort_values(["부족수량_num", "품목코드"], ascending=[False, True])

        p_sheet_scope = p_scope[p_scope["시트분류"].notna()].copy()
        p_sheet_scope["시트분류"] = p_sheet_scope["시트분류"].astype(str).str.strip()
        p_sheet_scope = p_sheet_scope[
            (p_sheet_scope["시트분류"] != "")
            & (p_sheet_scope["시트분류"].str.lower() != "nan")
            & (p_sheet_scope["시트분류"].str.lower() != "none")
        ]
        r_to_sheet = p_sheet_scope.drop_duplicates(subset=["R코드5"], keep="first").set_index("R코드5")["시트분류"].to_dict()

        p_group_scope = p_scope[p_scope["분류별요약"].notna()].copy()
        p_group_scope["분류별요약"] = p_group_scope["분류별요약"].astype(str).str.strip()
        p_group_scope = p_group_scope[
            (p_group_scope["분류별요약"] != "")
            & (p_group_scope["분류별요약"].str.lower() != "nan")
            & (p_group_scope["분류별요약"].str.lower() != "none")
        ]
        r_to_group = (
            p_group_scope.drop_duplicates(subset=["R코드5"], keep="first").set_index("R코드5")["분류별요약"].to_dict()
        )
    else:
        r_to_sheet = {}
        r_to_group = {}

    non_p_mask = item_prefix != "P"
    result.loc[non_p_mask, "시트분류"] = result.loc[non_p_mask, "시트분류"].fillna(
        result.loc[non_p_mask, "R코드5"].map(r_to_sheet)
    )
    result.loc[non_p_mask, "분류별요약"] = result.loc[non_p_mask, "분류별요약"].fillna(
        result.loc[non_p_mask, "R코드5"].map(r_to_group)
    )

    result["수동시트분류"] = result["시트분류"].map(clean_sheet_category)
    if result.empty:
        result["자동분류결과"] = pd.Series(dtype="object")
        result["분류 판단 근거"] = pd.Series(dtype="object")
    else:
        auto_classification = result.apply(
            lambda row: classify_sheet_with_reason(row),
            axis=1,
            result_type="expand",
        )
        result["자동분류결과"] = auto_classification[0]
        result["분류 판단 근거"] = auto_classification[1]

    manual_mask = result["수동시트분류"].map(clean_sheet_category) != ""
    result["시트분류"] = result["자동분류결과"]
    result.loc[manual_mask, "시트분류"] = result.loc[manual_mask, "수동시트분류"]
    result.loc[manual_mask, "분류 판단 근거"] = "수동 분류값 적용"
    pia_kr_mask = result["제품명"].astype(str).str.contains("PIA_KR", case=False, na=False, regex=False)
    result.loc[pia_kr_mask, "시트분류"] = "국내"
    result.loc[pia_kr_mask, "분류 판단 근거"] = "PIA_KR 제품명 기준 국내 분류"

    result["시트분류"] = result["시트분류"].map(clean_text_value)
    result.loc[result["시트분류"].str.lower().isin(INVALID_CATEGORY_VALUES), "시트분류"] = UNCLASSIFIED_SHEET_CATEGORY
    result["분류별요약"] = result["분류별요약"].astype(str).str.strip()
    result.loc[result["분류별요약"].str.lower().isin({"", "nan", "none"}), "분류별요약"] = "기타"
    result = result.drop(columns=["코드5"], errors="ignore")

    result["파워"] = result["품목코드"].map(extract_power_from_code)
    result["납기일"] = pd.to_datetime(result["납기일"], errors="coerce").dt.strftime("%Y-%m-%d")
    result["납기일"] = result["납기일"].fillna("-")
    if "사출납기일" in result.columns:
        result["사출납기일"] = pd.to_datetime(result["사출납기일"], errors="coerce").dt.strftime("%Y-%m-%d")
        result["사출납기일"] = result["사출납기일"].fillna("-")

    process_map_df = pd.DataFrame(
        {
            "공정창고": ["사출창고", "분리창고", "검사접착창고", "검사접착재작업창고", "누수규격검사 창고"],
            "수요정보 공정코드": [
                process_code_map.get("사출창고", "-"),
                process_code_map.get("분리창고", "-"),
                process_code_map.get("검사접착창고", "-"),
                "-",
                process_code_map.get("누수규격검사 창고", "-"),
            ],
            "재고코드 매핑 규칙": [
                "리드지정보 우선, 없으면 분류정보, 그래도 없으면 P코드->R코드 유추 (BUL1/BUL2는 BUL로 보정)",
                "리드지정보/분류정보 Q코드 우선, 없으면 P코드->Q코드 유추, Q재고가 없으면 리드지정보 외주(U) 코드로 보정",
                "P코드 그대로 사용",
                "WH_NAME=검사접착 중 재공 코드 끝부분 -C 계열은 별도 분류, 재작업가능은 재작업 시트 또는 생산현황 B:J 수요정보 기준",
                "P코드 그대로 사용",
            ],
            "재고>0 품목수": [
                int((code_stock["사출창고"] > 0).sum()),
                int((code_stock["분리창고"] > 0).sum()),
                int((code_stock["검사접착창고"] > 0).sum()),
                int((code_stock["검사접착재작업창고"] > 0).sum()),
                int((code_stock["누수규격검사 창고"] > 0).sum()),
            ],
        }
    )

    file_info_df = pd.DataFrame(
        {
            "재고파일": [inv_file_name],
            "수요파일": [dem_file_name],
            "행수(현황표)": [len(result)],
            "재작업 시트명": [str(rework_meta.get("sheet", "-"))],
            "재작업 기준 컬럼": [
                f"소스={rework_meta.get('match_scope', 'none')}, "
                f"이니셜={rework_meta.get('initial_col', '')}, "
                f"제품코드={rework_meta.get('product_col', '')}, "
                f"수량={rework_meta.get('quantity_col', '')}, "
                f"비고={rework_meta.get('note_col', '')}, "
                f"생산현황 수요키=거래처+이니셜+품목코드+R코드+Q코드+납기일"
            ],
            "재작업 시트 컬럼": [", ".join(rework_meta.get("sheet_columns", []))],
            "재작업 리스트 키 수": [len(rework_item_qty_map)],
            "재작업 리스트 수량 합계": [float(rework_meta.get("source_qty_total", 0.0) or 0.0)],
            "재작업 비고 키 수": [int(rework_meta.get("note_count", 0) or 0)],
            "재작업 매칭 키 수": [len(rework_matched_item_keys)],
            "재작업 매칭 수량 합계": [rework_matched_qty_total],
            "재작업 매칭 키 샘플": [", ".join(rework_matched_item_keys[:10])],
        }
    )

    return result, file_info_df, process_map_df


def load_data(refresh_key: str, base_dir_str: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    return preprocess_data(refresh_key, base_dir_str)


@st.cache_resource(show_spinner=False)
def load_api_shortage_data(
    refresh_key: str,
    base_dir_str: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    raw, error = read_aps_plan_operations_dataframe(APS_PLAN_SHORTAGE_OPERATIONS)
    if error:
        raise ValueError(f"APS API 수요 조회 실패: {error}")
    if raw.empty:
        empty_info = pd.DataFrame(
            {
                "재고파일": [format_reference_timestamp(get_wip_updated_at(data_base_dir))],
                "수요파일": [f"APS API ({format_reference_timestamp(get_plan_api_updated_at())})"],
                "행수(현황표)": [0],
            }
        )
        return pd.DataFrame(), empty_info, pd.DataFrame()

    raw = raw.copy()
    raw.columns = [str(col).strip() for col in raw.columns]
    columns = raw.columns.tolist()

    site_col = pick_api_column(columns, ["res_site_id", "RES_SITE_ID", "사이트코드", "사이트"])
    customer_col = pick_api_column(columns, ["cust_name", "CUST_NAME", "거래처", "거래처명", "고객 이름"])
    order_col = pick_api_column(columns, ["so_id", "SO_ID", "수주번호", "오더번호"])
    initial_col = pick_api_column(columns, ["initial", "INITIAL", "이니셜"])
    demand_type_col = pick_api_column(columns, ["demand_type", "DEMAND_TYPE", "수요유형"])
    demand_id_col = pick_api_column(columns, ["demand_id", "DEMAND_ID", "수요ID"])
    seq_col = pick_api_column(columns, ["seq", "SEQ"])
    item_col = pick_api_column(columns, ["item_id", "ITEM_ID", "제품 코드", "제품코드", "품목코드"])
    demand_item_col = pick_api_column(columns, ["demand_item_id", "DEMAND_ITEM_ID", "수요제품코드"])
    name_col = pick_api_column(columns, ["demand_item_name", "DEMAND_ITEM_NAME", "수요 제품 이름", "제품명"])
    item_name_col = pick_api_column(columns, ["item_name", "ITEM_NAME", "item_name2", "ITEM_NAME2"])
    demand_qty_col = pick_api_column(columns, ["demand_qty", "DEMAND_QTY", "수요 수량", "수요수량"])
    oper_col = pick_api_column(columns, ["oper_id", "OPER_ID", "공정코드", "공정"])
    plan_qty_col = pick_api_column(columns, ["plan_qty", "PLAN_QTY", "생산 수량", "생산수량", "계획수량"])
    due_col = pick_api_column(columns, ["due_date", "DUE_DATE", "납기일"])

    if item_col is None or oper_col is None or plan_qty_col is None:
        raise ValueError("APS API 응답에서 item_id/oper_id/plan_qty 컬럼을 찾지 못했습니다.")

    raw_index = raw.index

    def api_text(column_name: str | None, default: str = "") -> pd.Series:
        if column_name and column_name in raw.columns:
            selected = raw[column_name]
            if isinstance(selected, pd.DataFrame):
                selected = selected.iloc[:, 0]
            return selected.astype(str).str.strip()
        return pd.Series(default, index=raw_index, dtype="object")

    def api_number(column_name: str | None, default: float = 0.0) -> pd.Series:
        if column_name and column_name in raw.columns:
            selected = raw[column_name]
            if isinstance(selected, pd.DataFrame):
                selected = selected.iloc[:, 0]
            return parse_mixed_numeric(selected).fillna(default)
        return pd.Series(default, index=raw_index, dtype="float64")

    def api_date(column_name: str | None) -> pd.Series:
        values = api_text(column_name)
        parsed = pd.to_datetime(values, errors="coerce")
        return parsed.dt.strftime("%Y-%m-%d").fillna(values.where(values.map(clean_text_value).ne(""), ""))

    initial_text = api_text(initial_col)
    fallback_type = api_text(demand_type_col)
    missing_initial = initial_text.map(clean_text_value).eq("")
    initial_text = initial_text.where(~missing_initial, fallback_type)
    initial_text = initial_text.map(clean_text_value)
    initial_text = initial_text.where(initial_text.ne(""), "미지정")

    product_name = api_text(name_col)
    missing_name = product_name.map(clean_text_value).eq("")
    if item_name_col is not None:
        product_name = product_name.where(~missing_name, api_text(item_name_col))
    product_name = product_name.map(clean_text_value).replace({"": "-"})

    process_qty_map = {
        "10": "사출생산필요수량",
        "20": SEPARATION_REQUIRED_QTY_COL,
        "45": LEADJI_REQUIRED_QTY_COL,
        "55": ADHESION_REQUIRED_QTY_COL,
        "80": "생산수량",
    }
    process_due_map = {
        "10": "사출납기일",
        "20": SEPARATION_REQUIRED_DUE_COL,
        "45": LEADJI_REQUIRED_DUE_COL,
        "55": ADHESION_REQUIRED_DUE_COL,
        "80": "납기일",
    }

    work = pd.DataFrame(
        {
            "사이트코드": api_text(site_col, "API").map(normalize_site_group),
            "거래처": api_text(customer_col),
            ORDER_NO_COL: api_text(order_col),
            "이니셜": initial_text,
            "품목코드": api_text(item_col).map(normalize_item_code_value),
            "수요품목코드": api_text(demand_item_col).map(normalize_item_code_value),
            "제품명": product_name,
            DEMAND_QTY_COL: api_number(demand_qty_col),
            "공정": api_text(oper_col).str.extract(r"(\d+)", expand=False).fillna("").str.strip(),
            "생산수량_API": api_number(plan_qty_col),
            "API납기일": api_date(due_col),
            "수요ID": api_text(demand_id_col),
            "SEQ": api_text(seq_col),
        }
    )
    work = work[work["공정"].isin(process_qty_map.keys())].copy()
    work = work[work["품목코드"].ne("") & ~work["품목코드"].str.lower().isin(INVALID_CATEGORY_VALUES)]
    if work.empty:
        empty_info = pd.DataFrame(
            {
                "재고파일": [format_reference_timestamp(get_wip_updated_at(data_base_dir))],
                "수요파일": [f"APS API ({format_reference_timestamp(get_plan_api_updated_at())})"],
                "행수(현황표)": [0],
            }
        )
        return pd.DataFrame(), empty_info, pd.DataFrame()

    group_keys = ["사이트코드", "거래처", ORDER_NO_COL, "이니셜", "품목코드", "제품명", "API납기일"]
    qty_pivot = (
        work.pivot_table(
            index=group_keys,
            columns="공정",
            values="생산수량_API",
            aggfunc="sum",
            fill_value=0,
        )
        .rename(columns=process_qty_map)
        .reset_index()
    )
    demand_qty = (
        work.groupby(group_keys, as_index=False, dropna=False)[DEMAND_QTY_COL]
        .max()
        .rename(columns={DEMAND_QTY_COL: "_수요수량"})
    )
    grouped = qty_pivot.merge(demand_qty, on=group_keys, how="left")
    grouped = grouped.rename(columns={"API납기일": "납기일"})

    for col in ["생산수량", "사출생산필요수량", SEPARATION_REQUIRED_QTY_COL, LEADJI_REQUIRED_QTY_COL, ADHESION_REQUIRED_QTY_COL]:
        if col not in grouped.columns:
            grouped[col] = 0.0
        grouped[col] = parse_mixed_numeric(grouped[col]).fillna(0)
    grouped[DEMAND_QTY_COL] = parse_mixed_numeric(grouped.get("_수요수량", 0)).fillna(0)
    grouped = grouped.drop(columns=["_수요수량"], errors="ignore")
    grouped["부족수량"] = grouped["생산수량"]
    grouped["사출납기일"] = grouped["납기일"]
    grouped[SEPARATION_REQUIRED_DUE_COL] = grouped["납기일"]
    grouped[LEADJI_REQUIRED_DUE_COL] = grouped["납기일"]
    grouped[ADHESION_REQUIRED_DUE_COL] = grouped["납기일"]

    reference_refresh_key = build_reference_refresh_key(data_base_dir)
    (
        product_name_map,
        product_group_map,
        sheet2_group_map,
        r_ref_map,
        q_ref_map,
        r_name_map,
        bom_r_base_map,
        bom_q_base_map,
        bom_r_exact_map,
        bom_q_exact_map,
        leadji_r_map,
        leadji_q_map,
        leadji_u_map,
    ) = load_reference_maps_bundle(data_base_dir, reference_refresh_key)

    inv_df = load_all_item_inventory_file_source(data_base_dir)
    target_inv = inv_df[inv_df["창고"].isin(TARGET_WAREHOUSES)].copy()
    stock_lookup: dict[str, dict[str, float]] = {}
    for raw_name, display_name in WAREHOUSE_MAP.items():
        stock_lookup[display_name] = (
            target_inv[target_inv["창고"] == raw_name]
            .groupby("품목코드")["재고량"]
            .sum()
            .to_dict()
        )

    result = grouped.copy()
    code5 = result["품목코드"].astype(str).str[:5]
    inferred_r = result["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "R"))
    inferred_q = result["품목코드"].map(lambda x: map_demand_code_to_process_code(x, "Q"))
    mapped_r_base = code5.map(leadji_r_map).fillna(code5.map(r_ref_map)).fillna(code5.map(bom_r_base_map))
    mapped_q_base = code5.map(leadji_q_map).fillna(code5.map(q_ref_map)).fillna(code5.map(bom_q_base_map))
    result["R코드"] = [
        merge_mapped_base_code(inferred, mapped, "R") for inferred, mapped in zip(inferred_r, mapped_r_base)
    ]
    result["Q코드"] = [
        merge_mapped_base_code(inferred, mapped, "Q") for inferred, mapped in zip(inferred_q, mapped_q_base)
    ]
    exact_r = result["품목코드"].map(bom_r_exact_map)
    exact_q = result["품목코드"].map(bom_q_exact_map)
    result.loc[exact_r.map(clean_text_value).ne(""), "R코드"] = exact_r[exact_r.map(clean_text_value).ne("")]
    result.loc[exact_q.map(clean_text_value).ne(""), "Q코드"] = exact_q[exact_q.map(clean_text_value).ne("")]
    result["U코드"] = [
        merge_mapped_base_code(map_demand_code_to_process_code(q_code, "U"), mapped, "U")
        if clean_text_value(mapped)
        else ""
        for q_code, mapped in zip(result["Q코드"], code5.map(leadji_u_map))
    ]
    result["R코드 제품명"] = result["R코드"].astype(str).str[:5].map(r_name_map).fillna(result["제품명"])
    result["R코드 제품명"] = result["R코드 제품명"].replace({"": "-", "nan": "-", "None": "-"}).fillna("-")

    item_prefix = result["품목코드"].astype(str).str.upper().str[:1]
    result["사출창고"] = [
        lookup_stock_qty(stock_lookup.get("사출창고", {}), code if prefix == "R" else r_code)
        for code, prefix, r_code in zip(result["품목코드"], item_prefix, result["R코드"])
    ]
    result["분리창고"] = [
        lookup_stock_qty_from_candidates(
            stock_lookup.get("분리창고", {}),
            [code if prefix in {"Q", "U"} else "", q_code, u_code],
        )
        for code, prefix, q_code, u_code in zip(result["품목코드"], item_prefix, result["Q코드"], result["U코드"])
    ]
    result["검사접착창고"] = result["품목코드"].map(lambda x: stock_lookup.get("검사접착창고", {}).get(x, 0))
    result["검사접착재작업창고"] = result["품목코드"].map(
        lambda x: stock_lookup.get("검사접착재작업창고", {}).get(x, 0)
    )
    result["누수규격검사 창고"] = result["품목코드"].map(
        lambda x: stock_lookup.get("누수규격검사 창고", {}).get(x, 0)
    )
    result["공정재고 합계"] = (
        result["사출창고"]
        + result["분리창고"]
        + result["검사접착창고"]
        + result["검사접착재작업창고"]
        + result["누수규격검사 창고"]
    )

    result["분류별요약"] = code5.map(product_group_map)
    result["시트분류"] = code5.map(sheet2_group_map)
    result["수동시트분류"] = result["시트분류"].map(clean_sheet_category)
    if result.empty:
        result["자동분류결과"] = pd.Series(dtype="object")
        result["분류 판단 근거"] = pd.Series(dtype="object")
    else:
        auto_classification = result.apply(
            lambda row: classify_sheet_with_reason(row),
            axis=1,
            result_type="expand",
        )
        result["자동분류결과"] = auto_classification[0]
        result["분류 판단 근거"] = auto_classification[1]

    manual_mask = result["수동시트분류"].map(clean_sheet_category) != ""
    result["시트분류"] = result["자동분류결과"]
    result.loc[manual_mask, "시트분류"] = result.loc[manual_mask, "수동시트분류"]
    result.loc[manual_mask, "분류 판단 근거"] = "수동 분류값 적용"
    result["시트분류"] = result["시트분류"].map(clean_text_value)
    result.loc[result["시트분류"].str.lower().isin(INVALID_CATEGORY_VALUES), "시트분류"] = UNCLASSIFIED_SHEET_CATEGORY
    result["분류별요약"] = result["분류별요약"].astype(str).str.strip()
    result.loc[result["분류별요약"].str.lower().isin(INVALID_CATEGORY_VALUES), "분류별요약"] = "기타"

    result["파워"] = result["품목코드"].map(extract_power_from_code)
    rework_source = find_rework_production_source_file(data_base_dir)
    if rework_source is not None:
        rework_item_qty_map, rework_meta = read_rework_item_keys_from_production_status_file(rework_source)
    else:
        rework_item_qty_map, rework_meta = {}, build_empty_rework_meta()
    result, rework_matched_item_keys, rework_matched_qty_total = apply_rework_flags_to_demand_rows(
        result,
        rework_item_qty_map,
        rework_meta,
    )

    for text_col in [
        "사이트코드",
        "거래처",
        ORDER_NO_COL,
        "이니셜",
        "품목코드",
        "R코드",
        "Q코드",
        "U코드",
        "제품명",
        "납기일",
        "사출납기일",
        "파워",
        "시트분류",
        "분류별요약",
        "R코드 제품명",
    ]:
        result[text_col] = result[text_col].astype(str).replace({"nan": "", "None": ""}).fillna("")

    result = result[
        (result["부족수량"] > 0)
        | (result["사출생산필요수량"] > 0)
        | (result[SEPARATION_REQUIRED_QTY_COL] > 0)
        | (result[LEADJI_REQUIRED_QTY_COL] > 0)
        | (result[ADHESION_REQUIRED_QTY_COL] > 0)
    ].copy()

    process_map_df = pd.DataFrame(
        {
            "공정창고": ["사출창고", "분리창고", "검사접착창고", "검사접착재작업창고", "누수규격검사 창고"],
            "수요정보 공정코드": ["[10]사출조립", "[20]분리", "[55]접착/멸균", "-", "[80]누수/규격검사"],
            "재고코드 매핑 규칙": [
                "APS API oper_id=10 생산수량, WIP 사출창고 매칭",
                "APS API oper_id=20 생산수량, WIP 분리창고 매칭",
                "APS API oper_id=55 생산수량, WIP 검사접착창고 매칭",
                "재작업가능은 생산현황 B:J 수요정보 기준",
                "APS API oper_id=80 생산수량, WIP 누수규격검사 창고 매칭",
            ],
            "재고>0 품목수": [
                int((result["사출창고"] > 0).sum()),
                int((result["분리창고"] > 0).sum()),
                int((result["검사접착창고"] > 0).sum()),
                int((result["검사접착재작업창고"] > 0).sum()),
                int((result["누수규격검사 창고"] > 0).sum()),
            ],
        }
    )

    file_info_df = pd.DataFrame(
        {
            "재고파일": [f"WIP ({format_reference_timestamp(get_wip_updated_at(data_base_dir))})"],
            "수요파일": [f"APS API ({format_reference_timestamp(get_plan_api_updated_at())})"],
            "행수(현황표)": [len(result)],
            "API 처리행수": [len(raw)],
            "재작업 시트명": [str(rework_meta.get("sheet", "-"))],
            "재작업 기준 컬럼": [
                f"소스={rework_meta.get('match_scope', 'none')}, "
                f"이니셜={rework_meta.get('initial_col', '')}, "
                f"제품코드={rework_meta.get('product_col', '')}, "
                f"수량={rework_meta.get('quantity_col', '')}, "
                f"비고={rework_meta.get('note_col', '')}, "
                f"생산현황 수요키=거래처+이니셜+품목코드+R코드+Q코드+납기일"
            ],
            "재작업 시트 컬럼": [", ".join(rework_meta.get("sheet_columns", []))],
            "재작업 리스트 키 수": [len(rework_item_qty_map)],
            "재작업 리스트 수량 합계": [float(rework_meta.get("source_qty_total", 0.0) or 0.0)],
            "재작업 비고 키 수": [int(rework_meta.get("note_count", 0) or 0)],
            "재작업 매칭 키 수": [len(rework_matched_item_keys)],
            "재작업 매칭 수량 합계": [rework_matched_qty_total],
            "재작업 매칭 키 샘플": [", ".join(rework_matched_item_keys[:10])],
        }
    )
    return result.reset_index(drop=True), file_info_df, process_map_df


def normalize_item_code_value(value: object) -> str:
    code = re.sub(r"\s+", "", str(value).strip().upper())
    if not code or code.lower() in INVALID_CATEGORY_VALUES:
        return ""
    return code


def normalize_to_master_p_code(value: object) -> str:
    code = normalize_item_code_value(value)
    if not code:
        return ""
    if code.startswith("P"):
        return code
    if re.match(r"^[A-Z]\d{4}", code):
        return f"P{code[1:]}"
    return ""


def api_text_series(source: pd.DataFrame, column_name: str | None, default: str = "") -> pd.Series:
    if column_name and column_name in source.columns:
        return source[column_name].astype(str).str.strip()
    return pd.Series(default, index=source.index, dtype="object")


def api_numeric_series(source: pd.DataFrame, column_name: str | None, default: float = 0.0) -> pd.Series:
    if column_name and column_name in source.columns:
        return parse_mixed_numeric(source[column_name])
    return pd.Series(default, index=source.index, dtype="float64")


def combine_api_initial_and_type(initial: pd.Series, demand_type: pd.Series) -> pd.Series:
    values: list[str] = []
    for raw_initial, raw_type in zip(initial, demand_type):
        parts: list[str] = []
        for value in (raw_initial, raw_type):
            text = clean_text_value(value)
            if not text or text.lower() in INVALID_CATEGORY_VALUES:
                continue
            if text not in parts:
                parts.append(text)
        values.append(", ".join(parts))
    return pd.Series(values, index=initial.index, dtype="object")


def build_first_occurrence_mask(source: pd.DataFrame, key_columns: list[str | None], fallback_key: pd.Series) -> pd.Series:
    key_frame = pd.DataFrame(index=source.index)
    used = False
    for idx, column_name in enumerate(key_columns):
        if column_name and column_name in source.columns:
            key_frame[f"k{idx}"] = source[column_name].astype(str).str.strip()
            used = True
    key_frame["fallback"] = fallback_key.astype(str).str.strip()
    if not used:
        return ~key_frame["fallback"].duplicated()
    return ~key_frame.astype(str).agg("|".join, axis=1).duplicated()


def load_api_wip_inventory_df() -> pd.DataFrame:
    raw, error = read_plan_api_dataframe(APS_WIP_ENDPOINT, {"limit": PLAN_API_DEFAULT_ROW_LIMIT})
    if error or raw.empty:
        return pd.DataFrame(columns=["품목코드", "창고", "재공코드", "재고량"])
    return build_inventory_df(raw)


def load_api_demand_like_df(site_filter: str = "전체") -> pd.DataFrame:
    raw, error = read_aps_plan_operations_dataframe(APS_PLAN_FLOW_OPERATIONS, site_filter)
    output_columns = [
        "사이트코드",
        "거래처",
        ORDER_NO_COL,
        "이니셜",
        "품목코드",
        "제품명",
        DEMAND_QTY_COL,
        "납기일",
        "사출납기일",
        "부족수량",
        "사출생산필요수량",
        "R코드",
        "Q코드",
        EFFECTIVE_SAMPLE_AVAILABLE_COL,
    ]
    if error or raw.empty:
        return pd.DataFrame(columns=output_columns)

    raw = raw.copy()
    raw.columns = [str(col).strip() for col in raw.columns]
    columns = raw.columns.tolist()
    item_col = pick_api_column(
        columns,
        [
            "DEMAND_ITEM_ID",
            "demand_item_id",
            "수요제품코드",
            "수요 제품코드",
            "품목코드",
            "제품 코드",
            "제품코드",
            "생산코드",
            "P코드",
            "ITEM_ID",
            "item_id",
            "ITEM_CODE",
            "ITEM_CD",
            "item_cd",
        ],
    )
    if item_col is None:
        return pd.DataFrame(columns=output_columns)

    site_col = pick_api_column(columns, ["사이트코드", "사이트", "SITE_CODE", "SITE", "RES_SITE_ID", "res_site_id"])
    customer_col = pick_api_column(
        columns,
        ["거래처", "거래처명", "CUSTOMER", "CUSTOMER_NAME", "CUST_NAME", "cust_name", "고객", "고객명"],
    )
    order_col = pick_api_column(
        columns,
        ["수주번호", "오더번호", "ORDER_NO", "ORDER_ID", "SO_ID", "so_id", "수요ID", "DEMAND_ID", "demand_id"],
    )
    initial_col = pick_api_column(columns, ["이니셜", "INITIAL", "INITIAL_CODE"])
    demand_type_col = pick_api_column(columns, ["수요유형", "DEMAND_TYPE", "demand_type"])
    name_col = pick_api_column(
        columns,
        [
            "DEMAND_ITEM_NAME",
            "demand_item_name",
            "ITEM_NAME2",
            "item_name2",
            "제품명",
            "품명",
            "품목명",
            "ITEM_NAME",
            "item_name",
            "PRODUCT_NAME",
        ],
    )
    qty_col = pick_api_column(
        columns,
        ["수요수량", "오더수량", "수주수량", "ORDER_QTY", "DEMAND_QTY", "demand_qty", "QTY", "수량"],
    )
    due_col = pick_api_column(columns, ["납기일", "요청납기일", "납품일자", "DUE_DATE", "DELIVERY_DATE"])
    inj_due_col = pick_api_column(columns, ["사출납기일", "사출 납기일", "INJECTION_DUE_DATE", "TARGET_DATETIME", "target_datetime"])
    plan_qty_col = pick_api_column(columns, ["PLAN_QTY", "plan_qty", "계획수량"])
    oper_col = pick_api_column(columns, ["OPER_ID", "oper_id", "공정코드", "공정"])
    shortage_col = pick_api_column(
        columns,
        ["부족수량", "생산부족수량", "생산필요수량", "총합계 생산 수량", "총생산필요수량", "REQUIRED_QTY"],
    )
    inj_col = pick_api_column(
        columns,
        ["사출생산필요수량", "사출부족수량", "사출필요수량", "[10]사출조립 생산수량", "INJECTION_REQUIRED_QTY"],
    )
    sample_available_col = pick_api_column(
        columns,
        [
            EFFECTIVE_SAMPLE_AVAILABLE_COL,
            "샘플신청가능수량",
            "샘플 가능 수량",
            "샘플가능수량",
            "SAMPLE_AVAILABLE_QTY",
            "SAMPLE_REQ_AVAILABLE_QTY",
            "SAMPLE_REQUEST_AVAILABLE_QTY",
            "AVAILABLE_SAMPLE_QTY",
            "SAMPLE_QTY_AVAILABLE",
            "sample_available_qty",
            "sample_req_available_qty",
            "sample_request_available_qty",
        ],
    )
    r_col = pick_api_column(columns, ["R코드", "사출코드", "R_CODE", "INJECTION_CODE"])
    q_col = pick_api_column(columns, ["Q코드", "분리코드", "Q_CODE", "SEPARATION_CODE"])

    item_codes = api_text_series(raw, item_col).map(normalize_item_code_value)
    plan_qty = api_numeric_series(raw, plan_qty_col)
    first_order_mask = build_first_occurrence_mask(
        raw,
        [pick_api_column(columns, ["DEMAND_ID", "demand_id"]), order_col, pick_api_column(columns, ["SEQ", "seq"]), item_col],
        item_codes,
    )
    shortage_qty = (
        api_numeric_series(raw, shortage_col)
        if shortage_col is not None
        else plan_qty.where(first_order_mask, 0)
    )
    oper_text = api_text_series(raw, oper_col).str.upper()
    oper_normalized = oper_text.str.replace(r"[^0-9A-Z가-힣]+", "", regex=True)
    injection_mask = (
        item_codes.str.startswith("R", na=False)
        | oper_text.str.contains("사출|INJ|INJECTION", regex=True, na=False)
        | oper_normalized.str.startswith("10", na=False)
    )
    injection_qty = api_numeric_series(raw, inj_col) if inj_col is not None else plan_qty.where(injection_mask, 0)
    order_qty = api_numeric_series(raw, qty_col).where(first_order_mask, 0)
    sample_available_qty = api_numeric_series(raw, sample_available_col) if sample_available_col is not None else 0
    initial_text = combine_api_initial_and_type(api_text_series(raw, initial_col), api_text_series(raw, demand_type_col))
    r_codes = api_text_series(raw, r_col).map(normalize_item_code_value) if r_col else item_codes.where(
        item_codes.str.startswith("R", na=False), ""
    )
    q_codes = api_text_series(raw, q_col).map(normalize_item_code_value) if q_col else item_codes.where(
        item_codes.str.startswith("Q", na=False), ""
    )

    demand = pd.DataFrame(
        {
            "사이트코드": api_text_series(raw, site_col, "API"),
            "거래처": api_text_series(raw, customer_col),
            ORDER_NO_COL: api_text_series(raw, order_col),
            "이니셜": initial_text,
            "품목코드": item_codes,
            "제품명": api_text_series(raw, name_col),
            DEMAND_QTY_COL: order_qty,
            "납기일": api_text_series(raw, due_col),
            "사출납기일": api_text_series(raw, inj_due_col if inj_due_col is not None else due_col),
            "부족수량": shortage_qty,
            "사출생산필요수량": injection_qty,
            "R코드": r_codes,
            "Q코드": q_codes,
            EFFECTIVE_SAMPLE_AVAILABLE_COL: sample_available_qty,
        }
    )
    is_summary = demand[["사이트코드", "거래처", "이니셜", "품목코드"]].eq("총합계").any(axis=1)
    demand = demand[~is_summary].copy()
    demand = demand[demand["품목코드"].ne("") & ~demand["품목코드"].str.lower().isin(INVALID_CATEGORY_VALUES)]
    if demand.empty:
        return pd.DataFrame(columns=output_columns)
    return demand[output_columns]


def load_all_item_shortage_source(
    data_base_dir: Path,
    code_to_p: dict[str, str],
    site_filter: str = "전체",
) -> pd.DataFrame:
    if is_plan_api_enabled():
        return load_api_demand_like_df(site_filter)
    try:
        data_refresh_key = build_data_refresh_key(data_base_dir)
        shortage_df, _, _ = load_data(data_refresh_key, str(data_base_dir))
        normalized_site_filter = clean_text_value(site_filter)
        if normalized_site_filter and normalized_site_filter != "전체" and "사이트코드" in shortage_df.columns:
            shortage_df = shortage_df[shortage_df["사이트코드"].map(normalize_site_group) == normalized_site_filter].copy()
        return shortage_df
    except Exception:
        if is_plan_api_enabled():
            return pd.DataFrame(
                columns=[
                    "품목코드",
                    "R코드",
                    "Q코드",
                    "거래처",
                    "이니셜",
                    "제품명",
                    "납기일",
                    DEMAND_QTY_COL,
                    "부족수량",
                    "사출생산필요수량",
                ]
            )
        raise


def load_all_item_inventory_file_source(data_base_dir: Path) -> pd.DataFrame:
    try:
        inv_path, _ = find_excel_files(data_base_dir)
        inv = read_inventory_excel_subset(inv_path)
        return build_inventory_df(inv)
    except Exception:
        raise


def load_all_item_inventory_source(data_base_dir: Path) -> pd.DataFrame:
    if is_plan_api_enabled():
        api_inv_df = load_api_wip_inventory_df()
        if not api_inv_df.empty:
            return api_inv_df
    try:
        return load_all_item_inventory_file_source(data_base_dir)
    except Exception:
        if is_plan_api_enabled():
            return pd.DataFrame(columns=["품목코드", "창고", "재공코드", "재고량"])
        raise


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def read_all_item_master(master_path_str: str, refresh_key: str) -> pd.DataFrame:
    _ = refresh_key
    master_path = Path(master_path_str)
    if not master_path.exists():
        return pd.DataFrame()

    wanted_columns = {
        "품목코드",
        "코드구분",
        "제품명코드",
        "제품명",
        "제품군",
        "샘플가능수량",
        "신규분류",
    }
    try:
        master = pd.read_excel(
            master_path,
            sheet_name=ALL_ITEM_MASTER_SHEET,
            usecols=lambda c: str(c).strip() in wanted_columns,
        )
    except Exception:
        return pd.DataFrame()

    master.columns = [str(c).strip() for c in master.columns]
    for col in wanted_columns:
        if col not in master.columns:
            master[col] = ""
    master["품목코드"] = master["품목코드"].map(normalize_item_code_value)
    master = master[master["품목코드"].str.startswith("P", na=False)].copy()
    master = master.drop_duplicates(subset=["품목코드"], keep="first")
    return master


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def load_product_info_lookup(base_dir_str: str, reference_refresh_key: str) -> pd.DataFrame:
    _ = reference_refresh_key
    ref_path = find_product_name_reference_file(Path(base_dir_str))
    if ref_path is None:
        return pd.DataFrame(columns=["제품명코드", "제품명_기준", "신규분류_기준", "거래처_기준"])

    wanted_columns = {"제품명코드", "제품명", "분류요약", "거래처명"}
    try:
        info = pd.read_excel(
            ref_path,
            sheet_name="제품명정보",
            usecols=lambda c: str(c).strip() in wanted_columns,
        )
    except Exception:
        return pd.DataFrame(columns=["제품명코드", "제품명_기준", "신규분류_기준", "거래처_기준"])

    info.columns = [str(c).strip() for c in info.columns]
    for col in wanted_columns:
        if col not in info.columns:
            info[col] = ""
    info["제품명코드"] = info["제품명코드"].map(normalize_item_code_value).str[:5]
    info = info[info["제품명코드"].str.startswith("P", na=False)].copy()
    info = info.drop_duplicates(subset=["제품명코드"], keep="first")
    return info.rename(
        columns={
            "제품명": "제품명_기준",
            "분류요약": "신규분류_기준",
            "거래처명": "거래처_기준",
        }
    )[["제품명코드", "제품명_기준", "신규분류_기준", "거래처_기준"]]


def build_target_stock_lookup(inv_df: pd.DataFrame) -> tuple[dict[str, dict[str, float]], pd.DataFrame]:
    if inv_df.empty:
        return {display_name: {} for display_name in WAREHOUSE_MAP.values()}, pd.DataFrame()

    target_inv = inv_df[inv_df["창고"].isin(TARGET_WAREHOUSES)].copy()
    target_inv["품목코드"] = target_inv["품목코드"].map(normalize_item_code_value)
    target_inv = target_inv[target_inv["품목코드"] != ""]
    target_inv["재고량"] = parse_mixed_numeric(target_inv["재고량"])

    stock_lookup: dict[str, dict[str, float]] = {}
    for raw_name, display_name in WAREHOUSE_MAP.items():
        stock_lookup[display_name] = (
            target_inv[target_inv["창고"] == raw_name]
            .groupby("품목코드")["재고량"]
            .sum()
            .to_dict()
        )
    return stock_lookup, target_inv


def build_process_code_scope(
    all_items: pd.DataFrame,
    data_base_dir: Path,
    stock_lookup: dict[str, dict[str, float]],
) -> pd.DataFrame:
    reference_refresh_key = build_reference_refresh_key(data_base_dir)
    (
        _product_name_map,
        _product_group_map,
        _sheet2_group_map,
        r_ref_map,
        q_ref_map,
        _r_name_map,
        bom_r_base_map,
        bom_q_base_map,
        bom_r_exact_map,
        bom_q_exact_map,
        leadji_r_map,
        leadji_q_map,
        leadji_u_map,
    ) = load_reference_maps_bundle(data_base_dir, reference_refresh_key)

    scope = all_items[["생산코드"]].copy()
    scope["코드5"] = scope["생산코드"].str[:5]
    inferred_r = scope["생산코드"].map(lambda x: map_demand_code_to_process_code(x, "R"))
    inferred_q = scope["생산코드"].map(lambda x: map_demand_code_to_process_code(x, "Q"))

    mapped_r_base = scope["코드5"].map(leadji_r_map).fillna(scope["코드5"].map(r_ref_map))
    mapped_q_base = scope["코드5"].map(leadji_q_map).fillna(scope["코드5"].map(q_ref_map))
    scope["사출코드"] = [
        merge_mapped_base_code(inferred, mapped, "R") for inferred, mapped in zip(inferred_r, mapped_r_base)
    ]
    scope["분리코드"] = [
        merge_mapped_base_code(inferred, mapped, "Q") for inferred, mapped in zip(inferred_q, mapped_q_base)
    ]

    if bom_r_exact_map or bom_q_exact_map:
        exact_r = scope["생산코드"].map(bom_r_exact_map)
        exact_q = scope["생산코드"].map(bom_q_exact_map)
        exact_r_mask = exact_r.notna() & (exact_r.astype(str).str.strip() != "")
        exact_q_mask = exact_q.notna() & (exact_q.astype(str).str.strip() != "")
        scope.loc[exact_r_mask, "사출코드"] = exact_r.loc[exact_r_mask]
        scope.loc[exact_q_mask, "분리코드"] = exact_q.loc[exact_q_mask]

    if bom_r_base_map or bom_q_base_map:
        bom_r_base = scope["코드5"].map(bom_r_base_map)
        bom_q_base = scope["코드5"].map(bom_q_base_map)
        bom_r = [merge_mapped_base_code(inferred, mapped, "R") for inferred, mapped in zip(inferred_r, bom_r_base)]
        bom_q = [merge_mapped_base_code(inferred, mapped, "Q") for inferred, mapped in zip(inferred_q, bom_q_base)]
        r_norm = scope["사출코드"].astype(str).str.strip()
        q_norm = scope["분리코드"].astype(str).str.strip()
        invalid_r = (r_norm == "") | (r_norm.str.lower() == "nan") | (~r_norm.str.startswith("R"))
        invalid_q = (q_norm == "") | (q_norm.str.lower() == "nan") | (~q_norm.str.startswith("Q"))
        scope.loc[invalid_r, "사출코드"] = pd.Series(bom_r, index=scope.index).loc[invalid_r]
        scope.loc[invalid_q, "분리코드"] = pd.Series(bom_q, index=scope.index).loc[invalid_q]

    mapped_u_base = scope["코드5"].map(leadji_u_map)
    scope["U코드"] = [
        merge_mapped_base_code(map_demand_code_to_process_code(q_code, "U"), mapped, "U")
        if str(mapped).strip() and str(mapped).strip().lower() != "nan"
        else ""
        for q_code, mapped in zip(scope["분리코드"], mapped_u_base)
    ]

    scope["사출코드"] = scope["사출코드"].map(
        lambda x: resolve_process_code_for_stock(stock_lookup.get("사출창고", {}), x)
    )
    scope["분리코드"] = scope["분리코드"].map(
        lambda x: resolve_process_code_for_stock(stock_lookup.get("분리창고", {}), x)
    )
    scope["U코드"] = scope["U코드"].map(lambda x: resolve_process_code_for_stock(stock_lookup.get("분리창고", {}), x))

    scope["코드매칭상태"] = "매칭"
    valid_mask = (
        scope["생산코드"].astype(str).str.startswith("P")
        & scope["사출코드"].astype(str).str.startswith("R")
        & scope["분리코드"].astype(str).str.startswith("Q")
    )
    scope.loc[~valid_mask, "코드매칭상태"] = "코드미매칭"
    return scope.drop(columns=["코드5"], errors="ignore")


def build_code_to_production_map(code_scope: pd.DataFrame) -> dict[str, str]:
    code_to_p: dict[str, str] = {}
    for row in code_scope[["생산코드", "사출코드", "분리코드", "U코드"]].itertuples(index=False):
        production_code = normalize_item_code_value(row[0])
        if not production_code:
            continue
        for raw_code in row:
            code = normalize_item_code_value(raw_code)
            if not code:
                continue
            for candidate in iter_inventory_code_candidates(code):
                candidate = normalize_item_code_value(candidate)
                if candidate and candidate not in code_to_p:
                    code_to_p[candidate] = production_code
    return code_to_p


def resolve_light_production_code(row: pd.Series) -> str:
    code = normalize_to_master_p_code(row.get("품목코드", ""))
    if code.startswith("P"):
        return code
    for col in ["R코드", "Q코드"]:
        code = normalize_to_master_p_code(row.get(col, ""))
        if code.startswith("P"):
            return code
    return ""


def map_light_demand_code_to_process_code(demand_code: object, process_prefix: str) -> str:
    code = normalize_item_code_value(demand_code)
    if not code:
        return ""
    if code.startswith("P"):
        return map_demand_code_to_process_code(code, process_prefix)
    if code[0] in {"Q", "R"} and len(code) > 1:
        return f"{process_prefix}{code[1:]}"
    if re.match(r"^[A-Z]\d{4}", code):
        return f"{process_prefix}{code[1:]}"
    return map_demand_code_to_process_code(code, process_prefix)


def build_light_code_to_production_map(demand_df: pd.DataFrame) -> dict[str, str]:
    code_to_p: dict[str, str] = {}
    if demand_df.empty:
        return code_to_p
    for _, row in demand_df.iterrows():
        production_code = resolve_light_production_code(row)
        if not production_code:
            continue
        for col in ["품목코드", "R코드", "Q코드"]:
            code = normalize_item_code_value(row.get(col, ""))
            if not code:
                continue
            for candidate in iter_inventory_code_candidates(code):
                if candidate and candidate not in code_to_p:
                    code_to_p[candidate] = production_code
    return code_to_p


def build_light_process_scope(demand_df: pd.DataFrame, stock_lookup: dict[str, dict[str, float]]) -> pd.DataFrame:
    columns = ["생산코드", "사출코드", "분리코드", "U코드", "코드매칭상태"]
    if demand_df.empty:
        return pd.DataFrame(columns=columns)

    scope = demand_df.copy()
    for col in ["품목코드", "R코드", "Q코드"]:
        if col not in scope.columns:
            scope[col] = ""
        scope[col] = scope[col].map(normalize_item_code_value)
    scope["생산코드"] = scope.apply(resolve_light_production_code, axis=1)
    scope = scope[scope["생산코드"].ne("")].copy()
    if scope.empty:
        return pd.DataFrame(columns=columns)

    grouped = (
        scope.groupby("생산코드", as_index=False)
        .agg(
            {
                "R코드": first_nonempty_text,
                "Q코드": first_nonempty_text,
            }
        )
        .rename(columns={"R코드": "사출코드", "Q코드": "분리코드"})
    )
    missing_r = grouped["사출코드"].map(clean_text_value).str.startswith("R").eq(False)
    missing_q = grouped["분리코드"].map(clean_text_value).str.startswith("Q").eq(False)
    grouped.loc[missing_r, "사출코드"] = grouped.loc[missing_r, "생산코드"].map(
        lambda code: map_light_demand_code_to_process_code(code, "R")
    )
    grouped.loc[missing_q, "분리코드"] = grouped.loc[missing_q, "생산코드"].map(
        lambda code: map_light_demand_code_to_process_code(code, "Q")
    )
    grouped["U코드"] = grouped["분리코드"].map(lambda code: map_light_demand_code_to_process_code(code, "U"))
    grouped["사출코드"] = grouped["사출코드"].map(
        lambda code: resolve_process_code_for_stock(stock_lookup.get("사출창고", {}), code)
    )
    grouped["분리코드"] = grouped["분리코드"].map(
        lambda code: resolve_process_code_for_stock(stock_lookup.get("분리창고", {}), code)
    )
    grouped["U코드"] = grouped["U코드"].map(
        lambda code: resolve_process_code_for_stock(stock_lookup.get("분리창고", {}), code)
    )
    valid_mask = (
        grouped["생산코드"].astype(str).str.strip().ne("")
        & grouped["사출코드"].astype(str).str.startswith("R")
        & grouped["분리코드"].astype(str).str.startswith("Q")
    )
    grouped["코드매칭상태"] = "매칭"
    grouped.loc[~valid_mask, "코드매칭상태"] = "코드미매칭"
    return grouped[columns]


def resolve_master_code_from_process_candidates(values: list[object], code_to_p: dict[str, str]) -> str:
    for value in values:
        code = normalize_item_code_value(value)
        if not code:
            continue
        if code in code_to_p:
            return code_to_p[code]
        for candidate in iter_inventory_code_candidates(code):
            candidate = normalize_item_code_value(candidate)
            if candidate in code_to_p:
                return code_to_p[candidate]
    fallback = normalize_to_master_p_code(values[0] if values else "")
    return fallback if fallback.startswith("P") else ""


def build_all_item_demand_summary(shortage_df: pd.DataFrame, code_to_p: dict[str, str]) -> pd.DataFrame:
    if shortage_df.empty or "품목코드" not in shortage_df.columns:
        return pd.DataFrame(
            columns=[
                "생산코드",
                "거래처_수요",
                "이니셜_수요",
                "제품명_수요",
                "오더수량",
                "납기일",
                "생산부족수량",
                "사출부족수량",
                INITIAL_ORDER_MAP_COL,
                DEMAND_DETAIL_ROWS_COL,
            ]
        )

    demand = shortage_df.copy()
    for col in ["품목코드", "R코드", "Q코드", "거래처", "이니셜", "제품명", "납기일", "부족수량", "사출생산필요수량"]:
        if col not in demand.columns:
            demand[col] = ""
    qty_col = DEMAND_QTY_COL if DEMAND_QTY_COL in demand.columns else "수요수량"
    if qty_col not in demand.columns:
        demand[qty_col] = 0
    demand["오더수량"] = parse_mixed_numeric(demand[qty_col])
    demand["생산부족수량"] = parse_mixed_numeric(demand["부족수량"])
    demand["사출부족수량"] = parse_mixed_numeric(demand["사출생산필요수량"])
    demand["생산코드"] = demand.apply(
        lambda row: resolve_master_code_from_process_candidates(
            [row.get("품목코드", ""), row.get("R코드", ""), row.get("Q코드", "")],
            code_to_p,
        ),
        axis=1,
    )
    demand = demand[demand["생산코드"].str.startswith("P", na=False)].copy()
    if demand.empty:
        return pd.DataFrame(
            columns=[
                "생산코드",
                "거래처_수요",
                "이니셜_수요",
                "제품명_수요",
                "오더수량",
                "납기일",
                "생산부족수량",
                "사출부족수량",
                INITIAL_ORDER_MAP_COL,
                DEMAND_DETAIL_ROWS_COL,
            ]
        )

    demand["이니셜"] = demand["이니셜"].map(clean_initial_value)
    demand["_이니셜키"] = demand["이니셜"].map(clean_initial_value)
    demand.loc[demand["_이니셜키"].str.strip().str.lower().isin(INVALID_CATEGORY_VALUES), "_이니셜키"] = "미지정"
    demand["_거래처키"] = demand["거래처"].map(clean_text_value)
    demand.loc[demand["_거래처키"].str.strip().str.lower().isin(INVALID_CATEGORY_VALUES), "_거래처키"] = ""
    demand["_제품명키"] = demand["제품명"].map(clean_text_value)
    demand.loc[demand["_제품명키"].str.strip().str.lower().isin(INVALID_CATEGORY_VALUES), "_제품명키"] = "-"
    demand_detail = (
        demand.groupby(["생산코드", "_거래처키", "_이니셜키", "_제품명키"], as_index=False)
        .agg(
            {
                "오더수량": "sum",
                "납기일": min_date_text,
                "생산부족수량": "sum",
                "사출부족수량": "sum",
            }
        )
    )
    initial_order = demand_detail.groupby(["생산코드", "_이니셜키"], as_index=False)["오더수량"].sum()
    initial_order = initial_order[initial_order["오더수량"].ne(0)].copy()
    initial_map_rows: list[dict[str, object]] = []
    for production_code, group in initial_order.sort_values(["생산코드", "_이니셜키"]).groupby("생산코드"):
        values = {
            str(row["_이니셜키"]): float(row["오더수량"])
            for _, row in group.iterrows()
            if clean_text_value(row["_이니셜키"])
        }
        initial_map_rows.append(
            {
                "생산코드": production_code,
                INITIAL_ORDER_MAP_COL: json.dumps(values, ensure_ascii=False),
            }
        )
    initial_map = pd.DataFrame(initial_map_rows)

    detail_rows: list[dict[str, object]] = []
    for production_code, group in demand_detail.sort_values(
        ["생산코드", "_이니셜키", "_제품명키", "납기일"]
    ).groupby("생산코드"):
        values: list[dict[str, object]] = []
        for _, row in group.iterrows():
            values.append(
                {
                    "거래처": clean_text_value(row["_거래처키"]),
                    "이니셜": clean_text_value(row["_이니셜키"]) or "미지정",
                    "제품명": clean_text_value(row["_제품명키"]) or "-",
                    "오더수량": float(row["오더수량"]),
                    "납기일": clean_text_value(row["납기일"]) or "-",
                    "생산부족수량": float(row["생산부족수량"]),
                    "사출부족수량": float(row["사출부족수량"]),
                }
            )
        detail_rows.append(
            {
                "생산코드": production_code,
                DEMAND_DETAIL_ROWS_COL: json.dumps(values, ensure_ascii=False),
            }
        )
    detail_map = pd.DataFrame(detail_rows)

    summary = (
        demand.groupby("생산코드", as_index=False)
        .agg(
            {
                "거래처": summarize_unique,
                "이니셜": summarize_unique,
                "제품명": summarize_unique,
                "오더수량": "sum",
                "납기일": min_date_text,
                "생산부족수량": "sum",
                "사출부족수량": "sum",
            }
        )
        .rename(columns={"거래처": "거래처_수요", "이니셜": "이니셜_수요", "제품명": "제품명_수요"})
    )
    if initial_map.empty:
        summary[INITIAL_ORDER_MAP_COL] = ""
    else:
        summary = summary.merge(initial_map, on="생산코드", how="left")
        summary[INITIAL_ORDER_MAP_COL] = summary[INITIAL_ORDER_MAP_COL].fillna("")
    if detail_map.empty:
        summary[DEMAND_DETAIL_ROWS_COL] = ""
    else:
        summary = summary.merge(detail_map, on="생산코드", how="left")
        summary[DEMAND_DETAIL_ROWS_COL] = summary[DEMAND_DETAIL_ROWS_COL].fillna("")
    return summary


def apply_nonempty_override(base: pd.Series, override: pd.Series) -> pd.Series:
    output = base.astype(str).replace({"nan": "", "None": "", "NaN": ""}).fillna("")
    values = override.astype(str).replace({"nan": "", "None": "", "NaN": ""}).fillna("")
    valid = ~values.str.strip().str.lower().isin({"", "-", "nan", "none", "nat", "<na>"})
    output.loc[valid] = values.loc[valid]
    return output


def classify_flow_primary_group(*values: object) -> str:
    text = " ".join(clean_text_value(value) for value in values).upper()
    normalized = re.sub(r"[\s_./()\-]+", "", text)
    if "1DAY" in normalized:
        return "1-DAY"
    if any(token in normalized for token in ["FRP", "CONVENTIONAL", "2WEEK", "2WKS", "MONTH", "MONTHLY"]):
        return "FRP"
    if "1D" in normalized and "DAY" in normalized:
        return "1-DAY"
    return "기타"


def normalize_flow_customer_group(value: object) -> str:
    raw = clean_text_value(value)
    if not raw:
        return "거래처 미지정"
    upper = raw.upper().replace(".", " ").replace(",", " ")
    upper = re.sub(r"\s+", " ", upper).strip()
    compact = re.sub(r"[\s_./()\-]+", "", upper)

    if "PIA" in upper or "PIA종합" in raw:
        return "PIA"
    if "OPHTALMIC" in compact:
        return "OPHTALMIC"
    if "SINCERE" in compact:
        return "Sincere"
    if any(token in compact for token in ["HEARTS", "TOPTREND", "BEAUTYICON"]):
        return "HEARTS/TopTrend"
    if "OPTICALSUPPLIES" in compact:
        return "OPTICAL SUPPLIES"
    if any(token in compact for token in ["MAXVUE", "OPTIMAX", "DIGERO"]):
        return "MAXVUE/OPTIMAX"
    if "ALENSA" in compact:
        return "ALENSA"
    if any(token in compact for token in ["FEELGOOD", "CROSSBIRD"]):
        return "FEEL GOOD"
    if "ALCON" in compact:
        return "Alcon"
    if any(token in compact for token in ["CHINA", "IRIS", "WENZHOU"]) or "중국" in raw:
        return "CHINA/IRIS"
    if "MGMEDICAL" in compact:
        return "MG Medical"
    if "TGARDEN" in compact:
        return "T-Garden"
    if any(token in compact for token in ["HAPA", "PPB"]) or "피피비" in raw:
        return "HAPA/PPB"
    if "FROMEYES" in compact:
        return "from-eyes"
    if "EYEQUE" in compact:
        return "EYEQUE"
    if "ESSILOR" in compact:
        return "ESSILOR"
    if any(token in compact for token in ["국내", "KOREA", "CLALEN", "LENSME", "LENSVERY"]) or "렌즈미" in raw:
        return "국내"
    return raw


def min_date_text(values: pd.Series) -> str:
    if values.empty:
        return "-"
    parsed = parse_mixed_excel_date(values)
    valid = parsed.dropna()
    if valid.empty:
        return "-"
    return valid.min().strftime("%Y-%m-%d")


def summarize_signal_values(values: pd.Series) -> str:
    cleaned = [clean_text_value(value) for value in values]
    cleaned = [value for value in cleaned if value]
    if not cleaned:
        return ""
    priority = {"소진": 0, "감소": 1, "신규": 2, "증가": 3, "유지": 4}
    ordered = sorted(dict.fromkeys(cleaned), key=lambda value: (priority.get(value, 9), value))
    if len(ordered) <= 2:
        return ", ".join(ordered)
    return f"{', '.join(ordered[:2])} 외 {len(ordered) - 2}"


def find_finished_goods_stock_sheet(xls: pd.ExcelFile) -> str | None:
    normalized_to_original = {normalize_excel_sheet_name(name): name for name in xls.sheet_names}
    for hint in FINISHED_GOODS_STOCK_SHEET_HINTS:
        matched = normalized_to_original.get(normalize_excel_sheet_name(hint))
        if matched is not None:
            return matched
    return xls.sheet_names[0] if xls.sheet_names else None


def build_finished_goods_stock_disk_cache_path(stock_path: Path) -> Path:
    try:
        stat = stock_path.stat()
        source_key = f"{stock_path.resolve()}:{stat.st_size}:{stat.st_mtime_ns}"
    except OSError:
        source_key = str(stock_path)
    cache_hash = build_local_cache_hash(APP_CACHE_VERSION, "finished_goods_stock_summary_v2", source_key)
    return FINISHED_GOODS_STOCK_DISK_CACHE_DIR / f"{cache_hash}.pkl"


def read_finished_goods_stock_disk_cache(stock_path: Path, output_columns: list[str]) -> pd.DataFrame | None:
    cached = read_pickle_cache(build_finished_goods_stock_disk_cache_path(stock_path))
    if not isinstance(cached, pd.DataFrame):
        return None
    if not set(output_columns).issubset(cached.columns):
        return None
    return cached[output_columns].copy()


def write_finished_goods_stock_disk_cache(stock_path: Path, summary: pd.DataFrame) -> None:
    if summary.empty:
        return
    write_pickle_cache(build_finished_goods_stock_disk_cache_path(stock_path), summary)


def fast_excel_numeric_value(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.lower() in INVALID_CATEGORY_VALUES or text == "-":
        return 0.0

    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace(",", "").replace("\u00a0", "").replace(" ", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        parsed = float(text)
    except ValueError:
        return 0.0
    return -parsed if negative else parsed


def summarize_unique_text_list(values: list[str], head_count: int = 1) -> str:
    uniq: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value).strip()
        if not text or text.lower() in INVALID_CATEGORY_VALUES:
            continue
        if text in seen:
            continue
        seen.add(text)
        uniq.append(text)

    if not uniq:
        return "-"
    if len(uniq) <= head_count:
        return ", ".join(uniq)
    return f"{', '.join(uniq[:head_count])} \uc678 {len(uniq) - head_count}"


def summarize_signal_text_list(values: list[str]) -> str:
    cleaned = [str(value).strip() for value in values if str(value).strip()]
    if not cleaned:
        return ""
    priority = {"\uc18c\uc9c4": 0, "\uac10\uc18c": 1, "\uc2e0\uaddc": 2, "\uc99d\uac00": 3, "\uc720\uc9c0": 4}
    ordered = sorted(dict.fromkeys(cleaned), key=lambda value: (priority.get(value, 9), value))
    if len(ordered) <= 2:
        return ", ".join(ordered)
    return f"{', '.join(ordered[:2])} \uc678 {len(ordered) - 2}"


def median_positive(values: list[float]) -> float:
    positives = sorted(value for value in values if value > 0)
    count = len(positives)
    if count == 0:
        return 0.0
    midpoint = count // 2
    if count % 2:
        return positives[midpoint]
    return (positives[midpoint - 1] + positives[midpoint]) / 2


def row_value_at(row: tuple[object, ...], idx: int | None) -> object:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def read_finished_goods_stock_summary_fast(stock_path: Path, output_columns: list[str]) -> pd.DataFrame | None:
    if len(output_columns) < 11:
        return None

    code_candidates = ["\ud488\ubaa9\ucf54\ub4dc", "\uc81c\ud488\ucf54\ub4dc", "ITEM_ID", "\uc81c\ud488 \ucf54\ub4dc"]
    column_candidates = {
        "customer": ["\uac70\ub798\ucc98", "\uac70\ub798\ucc98\uba85", "CUSTOMER", "CUSTOMER_NAME"],
        "category": [
            "\uc2e0\uaddc\ubd84\ub958\uc694\uc57d",
            "\ubd84\ub958\uc694\uc57d",
            "\uc2e0\uaddc\ubd84\ub958",
            "\ud310\ub9e4\uc81c\ud488\uad70",
            "\uc0dd\uc0b0\uc81c\ud488\uad70",
        ],
        "wear": ["\ucc29\uc6a9\uc8fc\uae30", "\ucc29\uc6a9 \uc8fc\uae30", "WEARING_PERIOD"],
        "end_stock": ["\uc885\ub8cc\uc7ac\uace0", "\uae30\ub9d0\uc7ac\uace0", "\ud604\uc7ac\uc7ac\uace0", "\uc7ac\uace0"],
        "change": ["\ubcc0\ud654", "\uc7ac\uace0\ubcc0\ud654", "\uc99d\uac10", "\uc99d\uac10\uc218\ub7c9"],
        "order": ["\uc624\ub354", "\uc624\ub354\uc218\ub7c9", "\ucd1d\uc624\ub354", "ORDER_QTY"],
        "doi": ["DOI(\uc77c)", "DOI", "DOI\uc77c"],
        "ratio": ["\ube44\uc728", "\uc7ac\uace0\ube44\uc728", "\uc99d\uac10\ube44\uc728"],
        "signal": ["\uc2e0\ud638", "SIGNAL", "Signal"],
        "action": ["\ub300\uc751\ud310\ub2e8", "\ud310\ub2e8", "\uc870\uce58\ud310\ub2e8"],
    }

    wb = None
    try:
        wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
        normalized_to_original = {normalize_excel_sheet_name(name): name for name in wb.sheetnames}
        sheet_name = None
        for hint in FINISHED_GOODS_STOCK_SHEET_HINTS:
            sheet_name = normalized_to_original.get(normalize_excel_sheet_name(hint))
            if sheet_name is not None:
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if sheet_name is None:
            return pd.DataFrame(columns=output_columns)

        ws = wb[sheet_name]
        header_row = None
        column_indices: dict[str, int | None] = {}
        for candidate_header_row in (2, 1, 3):
            header_values = next(
                ws.iter_rows(min_row=candidate_header_row, max_row=candidate_header_row, values_only=True),
                None,
            )
            if not header_values:
                continue
            header_labels = {
                idx: str(value).strip()
                for idx, value in enumerate(header_values)
                if value is not None and str(value).strip()
            }
            code_idx = pick_first_existing_column_index(header_labels, code_candidates)
            if code_idx is None:
                continue
            column_indices = {"code": code_idx}
            for key, candidates in column_candidates.items():
                column_indices[key] = pick_first_existing_column_index(header_labels, candidates)
            header_row = candidate_header_row
            break

        if header_row is None:
            return None

        aggregated: dict[str, dict[str, object]] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            code = normalize_item_code_value(row_value_at(row, column_indices.get("code")))
            if not code.startswith("P"):
                continue

            item = aggregated.get(code)
            if item is None:
                item = {
                    "stock": 0.0,
                    "change": 0.0,
                    "order": 0.0,
                    "doi_values": [],
                    "ratio": [],
                    "signal": [],
                    "action": [],
                    "customer": [],
                    "category": [],
                    "wear": [],
                }
                aggregated[code] = item

            item["stock"] = float(item["stock"]) + fast_excel_numeric_value(
                row_value_at(row, column_indices.get("end_stock"))
            )
            item["change"] = float(item["change"]) + fast_excel_numeric_value(
                row_value_at(row, column_indices.get("change"))
            )
            item["order"] = float(item["order"]) + fast_excel_numeric_value(row_value_at(row, column_indices.get("order")))

            doi_value = fast_excel_numeric_value(row_value_at(row, column_indices.get("doi")))
            if doi_value > 0:
                item["doi_values"].append(doi_value)  # type: ignore[union-attr]

            for key in ("ratio", "signal", "action", "customer", "category", "wear"):
                text = clean_text_value(row_value_at(row, column_indices.get(key)))
                if text:
                    item[key].append(text)  # type: ignore[union-attr]

        records: list[dict[str, object]] = []
        for code in sorted(aggregated):
            item = aggregated[code]
            stock_qty = float(item["stock"])
            order_qty = float(item["order"])
            doi = stock_qty / order_qty * 181 if order_qty > 0 else median_positive(item["doi_values"])  # type: ignore[arg-type]
            records.append(
                {
                    output_columns[0]: code,
                    output_columns[1]: stock_qty,
                    output_columns[2]: float(item["change"]),
                    output_columns[3]: order_qty,
                    output_columns[4]: doi,
                    output_columns[5]: summarize_unique_text_list(item["ratio"]),  # type: ignore[arg-type]
                    output_columns[6]: summarize_signal_text_list(item["signal"]),  # type: ignore[arg-type]
                    output_columns[7]: summarize_unique_text_list(item["action"]),  # type: ignore[arg-type]
                    output_columns[8]: summarize_unique_text_list(item["customer"], head_count=2),  # type: ignore[arg-type]
                    output_columns[9]: summarize_unique_text_list(item["category"], head_count=2),  # type: ignore[arg-type]
                    output_columns[10]: summarize_unique_text_list(item["wear"], head_count=2),  # type: ignore[arg-type]
                }
            )

        return pd.DataFrame.from_records(records, columns=output_columns)
    except Exception:
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def read_finished_goods_stock_summary(stock_path_str: str, refresh_key: str) -> pd.DataFrame:
    _ = refresh_key
    stock_path = Path(stock_path_str)
    output_columns = [
        "생산코드",
        "완제품재고",
        "재고변화",
        "DOI기준오더",
        "DOI",
        "재고비율",
        "신호",
        "재고대응판단",
        "거래처_완제품",
        "신규분류_완제품",
        "착용주기_완제품",
    ]
    if not stock_path.exists():
        return pd.DataFrame(columns=output_columns)

    cached_summary = read_finished_goods_stock_disk_cache(stock_path, output_columns)
    if cached_summary is not None:
        return cached_summary

    fast_summary = read_finished_goods_stock_summary_fast(stock_path, output_columns)
    if fast_summary is not None:
        write_finished_goods_stock_disk_cache(stock_path, fast_summary)
        return fast_summary

    try:
        xls = pd.ExcelFile(stock_path)
    except Exception:
        return pd.DataFrame(columns=output_columns)

    sheet_name = find_finished_goods_stock_sheet(xls)
    if sheet_name is None:
        return pd.DataFrame(columns=output_columns)

    parsed = pd.DataFrame()
    column_candidate_groups = [
        ["품목코드", "제품코드", "ITEM_ID", "제품 코드"],
        ["거래처", "거래처명", "CUSTOMER", "CUSTOMER_NAME"],
        ["신규분류요약", "분류요약", "신규분류", "판매제품군", "생산제품군"],
        ["착용주기", "착용 주기", "WEARING_PERIOD"],
        ["종료재고", "기말재고", "현재재고", "재고"],
        ["변화", "재고변화", "증감", "증감수량"],
        ["오더", "오더수량", "총오더", "ORDER_QTY"],
        ["DOI(일)", "DOI", "DOI일"],
        ["비율", "재고비율", "증감비율"],
        ["신호", "SIGNAL", "Signal"],
        ["대응판단", "판단", "조치판단"],
    ]
    for header_row in (1, 0, 2):
        try:
            header_candidate = pd.read_excel(stock_path, sheet_name=sheet_name, header=header_row, nrows=0)
        except Exception:
            continue
        header_candidate.columns = [str(col).strip() for col in header_candidate.columns]
        columns = header_candidate.columns.tolist()
        code_col = pick_first_existing_column(columns, ["품목코드", "제품코드", "ITEM_ID", "제품 코드"])
        if code_col is None:
            continue
        selected_cols: list[str] = []
        for candidates in column_candidate_groups:
            selected_col = pick_first_existing_column(columns, candidates)
            if selected_col is not None and selected_col not in selected_cols:
                selected_cols.append(selected_col)
        try:
            candidate = pd.read_excel(stock_path, sheet_name=sheet_name, header=header_row, usecols=selected_cols)
        except Exception:
            continue
        candidate.columns = [str(col).strip() for col in candidate.columns]
        parsed = candidate
        break

    if parsed.empty:
        return pd.DataFrame(columns=output_columns)

    columns = parsed.columns.tolist()
    code_col = pick_first_existing_column(columns, ["품목코드", "제품코드", "ITEM_ID", "제품 코드"])
    if code_col is None:
        return pd.DataFrame(columns=output_columns)
    customer_col = pick_first_existing_column(columns, ["거래처", "거래처명", "CUSTOMER", "CUSTOMER_NAME"])
    category_col = pick_first_existing_column(columns, ["신규분류요약", "분류요약", "신규분류", "판매제품군", "생산제품군"])
    wear_col = pick_first_existing_column(columns, ["착용주기", "착용 주기", "WEARING_PERIOD"])
    end_stock_col = pick_first_existing_column(columns, ["종료재고", "기말재고", "현재재고", "재고"])
    change_col = pick_first_existing_column(columns, ["변화", "재고변화", "증감", "증감수량"])
    order_col = pick_first_existing_column(columns, ["오더", "오더수량", "총오더", "ORDER_QTY"])
    doi_col = pick_first_existing_column(columns, ["DOI(일)", "DOI", "DOI일"])
    ratio_col = pick_first_existing_column(columns, ["비율", "재고비율", "증감비율"])
    signal_col = pick_first_existing_column(columns, ["신호", "SIGNAL", "Signal"])
    action_col = pick_first_existing_column(columns, ["대응판단", "판단", "조치판단"])

    stock = pd.DataFrame({"생산코드": parsed[code_col].map(normalize_item_code_value)})
    stock = stock[stock["생산코드"].str.startswith("P", na=False)].copy()
    if stock.empty:
        return pd.DataFrame(columns=output_columns)

    stock["완제품재고"] = parse_mixed_numeric(parsed.loc[stock.index, end_stock_col]) if end_stock_col else 0
    stock["재고변화"] = parse_mixed_numeric(parsed.loc[stock.index, change_col]) if change_col else 0
    stock["DOI기준오더"] = parse_mixed_numeric(parsed.loc[stock.index, order_col]) if order_col else 0
    stock["DOI"] = parse_mixed_numeric(parsed.loc[stock.index, doi_col]) if doi_col else 0
    stock["재고비율"] = parsed.loc[stock.index, ratio_col].map(clean_text_value) if ratio_col else ""
    stock["신호"] = parsed.loc[stock.index, signal_col].map(clean_text_value) if signal_col else ""
    stock["재고대응판단"] = parsed.loc[stock.index, action_col].map(clean_text_value) if action_col else ""
    stock["거래처_완제품"] = parsed.loc[stock.index, customer_col].map(clean_text_value) if customer_col else ""
    stock["신규분류_완제품"] = parsed.loc[stock.index, category_col].map(clean_text_value) if category_col else ""
    stock["착용주기_완제품"] = parsed.loc[stock.index, wear_col].map(clean_text_value) if wear_col else ""

    grouped = (
        stock.groupby("생산코드", as_index=False)
        .agg(
            {
                "완제품재고": "sum",
                "재고변화": "sum",
                "DOI기준오더": "sum",
                "DOI": lambda s: s[s > 0].median() if (s > 0).any() else 0,
                "재고비율": lambda s: summarize_unique(s, head_count=1),
                "신호": summarize_signal_values,
                "재고대응판단": lambda s: summarize_unique(s, head_count=1),
                "거래처_완제품": lambda s: summarize_unique(s, head_count=2),
                "신규분류_완제품": lambda s: summarize_unique(s, head_count=2),
                "착용주기_완제품": lambda s: summarize_unique(s, head_count=2),
            }
        )
    )
    doi_order_mask = grouped["DOI기준오더"] > 0
    grouped.loc[doi_order_mask, "DOI"] = (
        grouped.loc[doi_order_mask, "완제품재고"] / grouped.loc[doi_order_mask, "DOI기준오더"] * 181
    )
    grouped["DOI"] = parse_mixed_numeric(grouped["DOI"]).fillna(0)
    result = grouped[output_columns]
    write_finished_goods_stock_disk_cache(stock_path, result)
    return result


def build_code_mismatch_df(
    all_items: pd.DataFrame,
    target_inv: pd.DataFrame,
    code_to_p: dict[str, str],
) -> pd.DataFrame:
    records: list[dict[str, object]] = []

    if not all_items.empty and "코드매칭상태" in all_items.columns:
        mismatch_items = all_items[all_items["코드매칭상태"] == "코드미매칭"]
        for _, row in mismatch_items.iterrows():
            records.append(
                {
                    "유형": "전체품목 자동매칭 실패",
                    "품목코드": row.get("생산코드", ""),
                    "창고": "",
                    "재고수량": 0,
                    "생산코드": row.get("생산코드", ""),
                    "사출코드": row.get("사출코드", ""),
                    "분리코드": row.get("분리코드", ""),
                    "제품명": row.get("제품명", ""),
                    "사유": "P코드 기준 R/Q 공정코드를 생성하지 못했습니다.",
                }
            )

    if not target_inv.empty:
        inv_group = (
            target_inv.groupby(["창고", "품목코드"], as_index=False)["재고량"]
            .sum()
            .sort_values(["창고", "재고량"], ascending=[True, False])
        )
        for _, row in inv_group.iterrows():
            raw_warehouse = str(row.get("창고", ""))
            item_code = normalize_item_code_value(row.get("품목코드", ""))
            if not item_code:
                continue
            matched_p = resolve_master_code_from_process_candidates([item_code], code_to_p)
            if matched_p:
                continue
            records.append(
                {
                    "유형": "재고원장 단독코드",
                    "품목코드": item_code,
                    "창고": WAREHOUSE_MAP.get(raw_warehouse, raw_warehouse),
                    "재고수량": float(row.get("재고량", 0) or 0),
                    "생산코드": "",
                    "사출코드": item_code if item_code.startswith("R") else "",
                    "분리코드": item_code if item_code.startswith("Q") else "",
                    "제품명": "",
                    "사유": "재고 원장에는 있으나 전체 품목리스트의 P/R/Q 생성 코드와 매칭되지 않습니다.",
                }
            )

    columns = ["유형", "품목코드", "창고", "재고수량", "생산코드", "사출코드", "분리코드", "제품명", "사유"]
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame.from_records(records, columns=columns)


@st.cache_resource(show_spinner=False)
def build_all_item_status_snapshot(refresh_key: str, base_dir_str: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    master_path = find_all_item_master_file(data_base_dir)
    if master_path is None:
        empty_items = pd.DataFrame(columns=ALL_ITEM_DOWNLOAD_COLUMNS)
        empty_mismatch = pd.DataFrame(
            columns=["유형", "품목코드", "창고", "재고수량", "생산코드", "사출코드", "분리코드", "제품명", "사유"]
        )
        return empty_items, empty_mismatch

    master = read_all_item_master(str(master_path), refresh_key)
    if master.empty:
        return pd.DataFrame(columns=ALL_ITEM_DOWNLOAD_COLUMNS), pd.DataFrame()

    inv_df = load_all_item_inventory_source(data_base_dir)
    stock_lookup, target_inv = build_target_stock_lookup(inv_df)

    all_items = pd.DataFrame(
        {
            "생산코드": master["품목코드"].map(normalize_item_code_value),
            "제품명코드": master["제품명코드"].map(normalize_item_code_value).str[:5],
            "제품명": master["제품명"].astype(str).str.strip(),
            "제품군": master["제품군"].astype(str).str.strip(),
            "신규분류": master["신규분류"].astype(str).str.strip(),
            "샘플 신청가능수량": parse_mixed_numeric(master["샘플가능수량"]),
        }
    )

    reference_refresh_key = build_reference_refresh_key(data_base_dir)
    product_info = load_product_info_lookup(str(data_base_dir), reference_refresh_key)
    if not product_info.empty:
        all_items = all_items.merge(product_info, on="제품명코드", how="left")
    else:
        all_items["제품명_기준"] = ""
        all_items["신규분류_기준"] = ""
        all_items["거래처_기준"] = ""

    all_items["제품명"] = apply_nonempty_override(all_items["제품명_기준"], all_items["제품명"])
    all_items["신규분류"] = apply_nonempty_override(all_items["신규분류_기준"], all_items["신규분류"])
    all_items["신규분류"] = apply_nonempty_override(all_items["제품군"], all_items["신규분류"])
    all_items["거래처"] = all_items["거래처_기준"].astype(str).replace({"nan": "", "None": ""}).fillna("")
    all_items["이니셜"] = ""
    all_items["파워"] = all_items["생산코드"].map(extract_power_from_code)

    code_scope = build_process_code_scope(all_items, data_base_dir, stock_lookup)
    all_items = all_items.merge(code_scope, on="생산코드", how="left")
    code_to_p = build_code_to_production_map(code_scope)

    shortage_df = load_all_item_shortage_source(data_base_dir, code_to_p)
    demand_summary = build_all_item_demand_summary(shortage_df, code_to_p)
    all_items = all_items.merge(demand_summary, on="생산코드", how="left")

    all_items["거래처"] = apply_nonempty_override(all_items["거래처"], all_items.get("거래처_수요", pd.Series("", index=all_items.index)))
    all_items["이니셜"] = apply_nonempty_override(all_items["이니셜"], all_items.get("이니셜_수요", pd.Series("", index=all_items.index)))
    all_items["제품명"] = apply_nonempty_override(all_items["제품명"], all_items.get("제품명_수요", pd.Series("", index=all_items.index)))
    all_items["오더수량"] = parse_mixed_numeric(all_items.get("오더수량", pd.Series(0, index=all_items.index))).fillna(0)
    all_items["실수요수량"] = all_items["오더수량"]
    all_items["총수요수량"] = all_items["오더수량"]
    all_items["생산부족수량"] = parse_mixed_numeric(
        all_items.get("생산부족수량", pd.Series(0, index=all_items.index))
    ).fillna(0)
    all_items["사출부족수량"] = parse_mixed_numeric(
        all_items.get("사출부족수량", pd.Series(0, index=all_items.index))
    ).fillna(0)
    all_items[INITIAL_ORDER_MAP_COL] = all_items.get(
        INITIAL_ORDER_MAP_COL,
        pd.Series("", index=all_items.index),
    ).astype(str).replace({"nan": "", "None": ""}).fillna("")
    all_items[DEMAND_DETAIL_ROWS_COL] = all_items.get(
        DEMAND_DETAIL_ROWS_COL,
        pd.Series("", index=all_items.index),
    ).astype(str).replace({"nan": "", "None": ""}).fillna("")
    all_items["납기일"] = all_items.get("납기일", pd.Series("-", index=all_items.index)).astype(str).replace(
        {"nan": "-", "None": "-", "NaT": "-"}
    )

    all_items["사출창고"] = all_items["사출코드"].map(lambda x: lookup_stock_qty(stock_lookup.get("사출창고", {}), x))
    all_items["분리창고"] = all_items.apply(
        lambda row: lookup_stock_qty_from_candidates(
            stock_lookup.get("분리창고", {}),
            [row.get("분리코드", ""), row.get("U코드", "")],
        ),
        axis=1,
    )
    all_items["검사접착창고"] = all_items["생산코드"].map(lambda x: stock_lookup.get("검사접착창고", {}).get(x, 0))
    all_items["누수규격검사"] = all_items["생산코드"].map(lambda x: stock_lookup.get("누수규격검사 창고", {}).get(x, 0))
    all_items["공정재고합계"] = (
        all_items["사출창고"] + all_items["분리창고"] + all_items["검사접착창고"] + all_items["누수규격검사"]
    )
    all_items["초과재고수량"] = (all_items["공정재고합계"] - all_items["총수요수량"]).clip(lower=0)
    all_items["부족수량"] = all_items["생산부족수량"]

    finished_goods_stock_path = find_finished_goods_stock_file(data_base_dir)
    if finished_goods_stock_path is not None:
        finished_goods_summary = read_finished_goods_stock_summary(str(finished_goods_stock_path), refresh_key)
        if not finished_goods_summary.empty:
            all_items = all_items.merge(finished_goods_summary, on="생산코드", how="left")
    for col in ["완제품재고", "재고변화", "DOI기준오더", "DOI"]:
        if col not in all_items.columns:
            all_items[col] = 0
    for col in ["재고비율", "신호", "재고대응판단", "거래처_완제품", "신규분류_완제품", "착용주기_완제품"]:
        if col not in all_items.columns:
            all_items[col] = ""
    all_items["거래처"] = apply_nonempty_override(all_items["거래처"], all_items["거래처_완제품"])
    all_items["신규분류"] = apply_nonempty_override(all_items["신규분류"], all_items["신규분류_완제품"])
    all_items["제품대분류"] = all_items.apply(
        lambda row: classify_flow_primary_group(
            row.get("신규분류", ""),
            row.get("신규분류_완제품", ""),
            row.get("제품명", ""),
            row.get("착용주기_완제품", ""),
        ),
        axis=1,
    )
    all_items["거래처그룹"] = all_items["거래처"].map(normalize_flow_customer_group)

    for col in ALL_ITEM_NUMERIC_COLUMNS:
        if col not in all_items.columns:
            all_items[col] = 0
        all_items[col] = parse_mixed_numeric(all_items[col]).fillna(0)
    all_items["코드매칭상태"] = all_items["코드매칭상태"].fillna("코드미매칭")

    all_items["상태"] = "수요 없음"
    all_items.loc[all_items["샘플 신청가능수량"] > 0, "상태"] = "수요 없음 + 샘플 신청가능수량 있음"
    all_items.loc[all_items["공정재고합계"] > 0, "상태"] = "수요 없음 + 공정재고 있음"
    all_items.loc[all_items["총수요수량"] > 0, "상태"] = "수요 있음"
    all_items.loc[all_items["코드매칭상태"] == "코드미매칭", "상태"] = "코드미매칭"
    all_items["판단"] = "재고 없음"
    all_items.loc[all_items["샘플 신청가능수량"] > 0, "판단"] = "샘플 가능 수량 있음"
    all_items.loc[all_items["총수요수량"] > 0, "판단"] = "수요 있음"
    all_items.loc[
        (all_items["총수요수량"] > 0) & (all_items["초과재고수량"] > 0),
        "판단",
    ] = "수요 대비 재고 초과"
    all_items.loc[
        (all_items["총수요수량"] <= 0) & (all_items["공정재고합계"] > 0),
        "판단",
    ] = "수요 없음 + 공정재고 있음"
    all_items.loc[all_items["코드매칭상태"] == "코드미매칭", "판단"] = "코드 확인 필요"
    all_items["주의정렬순위"] = 5
    all_items.loc[all_items["판단"] == "샘플 가능 수량 있음", "주의정렬순위"] = 4
    all_items.loc[all_items["판단"] == "수요 있음", "주의정렬순위"] = 3
    all_items.loc[all_items["판단"] == "수요 대비 재고 초과", "주의정렬순위"] = 2
    all_items.loc[all_items["판단"] == "수요 없음 + 공정재고 있음", "주의정렬순위"] = 1
    all_items.loc[all_items["판단"] == "코드 확인 필요", "주의정렬순위"] = 0

    for col in [
        "사이트코드",
        "제품대분류",
        "거래처그룹",
        "거래처",
        "이니셜",
        "신규분류",
        "제품명",
        "파워",
        "납기일",
        "사출코드",
        "분리코드",
        "생산코드",
        "재고비율",
        "신호",
        "재고대응판단",
        "판단",
    ]:
        if col not in all_items.columns:
            all_items[col] = ""
        all_items[col] = all_items[col].astype(str).replace({"nan": "", "None": ""}).fillna("")
    all_items.loc[all_items["제품대분류"].str.strip().str.lower().isin({"", "nan", "none"}), "제품대분류"] = "기타"
    all_items.loc[
        all_items["거래처그룹"].str.strip().str.lower().isin({"", "nan", "none"}),
        "거래처그룹",
    ] = "거래처 미지정"
    all_items.loc[all_items["신규분류"].str.strip().str.lower().isin({"", "nan", "none"}), "신규분류"] = "기타"
    all_items.loc[all_items["제품명"].str.strip().str.lower().isin({"", "nan", "none"}), "제품명"] = "-"

    code_mismatch_df = build_code_mismatch_df(all_items, target_inv, code_to_p)
    result = all_items[[*ALL_ITEM_DOWNLOAD_COLUMNS, INITIAL_ORDER_MAP_COL, DEMAND_DETAIL_ROWS_COL, "주의정렬순위"]].sort_values(
        ["주의정렬순위", "초과재고수량", "공정재고합계", "총수요수량", "신규분류", "생산코드"],
        ascending=[True, False, False, False, True, True],
    )
    result = result.drop(columns=["주의정렬순위"], errors="ignore")
    return result.reset_index(drop=True), code_mismatch_df.reset_index(drop=True)


@st.cache_resource(show_spinner=False)
def build_all_item_flow_status_snapshot(
    refresh_key: str,
    base_dir_str: str | None = None,
    site_filter: str = "전체",
) -> pd.DataFrame:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    output_columns = [*ALL_ITEM_DOWNLOAD_COLUMNS, INITIAL_ORDER_MAP_COL, DEMAND_DETAIL_ROWS_COL]

    demand_seed = load_all_item_shortage_source(data_base_dir, {}, site_filter)
    if demand_seed.empty:
        return pd.DataFrame(columns=output_columns)

    demand = demand_seed.copy()
    for col in [
        "사이트코드",
        "품목코드",
        "R코드",
        "Q코드",
        "거래처",
        "이니셜",
        "제품명",
        "납기일",
        "부족수량",
        "사출생산필요수량",
    ]:
        if col not in demand.columns:
            demand[col] = ""
    qty_col = DEMAND_QTY_COL if DEMAND_QTY_COL in demand.columns else "수요수량"
    if qty_col not in demand.columns:
        demand[qty_col] = 0

    demand["생산코드"] = demand.apply(resolve_light_production_code, axis=1)
    demand = demand[demand["생산코드"].map(clean_text_value).ne("")].copy()
    if demand.empty:
        return pd.DataFrame(columns=output_columns)

    demand["사이트코드"] = demand["사이트코드"].map(normalize_site_group)
    demand["거래처"] = demand["거래처"].map(clean_text_value)
    demand["이니셜"] = demand["이니셜"].map(clean_text_value)
    demand.loc[demand["이니셜"].str.lower().isin(INVALID_CATEGORY_VALUES), "이니셜"] = "미지정"
    demand["제품명"] = demand["제품명"].map(clean_text_value)
    demand.loc[demand["제품명"].str.lower().isin(INVALID_CATEGORY_VALUES), "제품명"] = "-"
    demand["오더수량"] = parse_mixed_numeric(demand[qty_col]).fillna(0)
    demand["생산부족수량"] = parse_mixed_numeric(demand["부족수량"]).fillna(0)
    demand["사출부족수량"] = parse_mixed_numeric(demand["사출생산필요수량"]).fillna(0)
    demand["_납기일_dt"] = parse_mixed_excel_date(demand["납기일"])

    group_keys = ["사이트코드", "생산코드", "거래처", "이니셜", "제품명"]
    all_items = (
        demand.groupby(group_keys, as_index=False)
        .agg(
            {
                "오더수량": "sum",
                "_납기일_dt": "min",
                "생산부족수량": "sum",
                "사출부족수량": "sum",
            }
        )
        .copy()
    )
    if all_items.empty:
        return pd.DataFrame(columns=output_columns)
    all_items["납기일"] = pd.to_datetime(all_items["_납기일_dt"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("-")
    all_items = all_items.drop(columns=["_납기일_dt"], errors="ignore")

    try:
        inv_df = load_all_item_inventory_file_source(data_base_dir)
    except Exception:
        inv_df = pd.DataFrame(columns=["품목코드", "창고", "재공코드", "재고량"])
    stock_lookup, _target_inv = build_target_stock_lookup(inv_df)
    process_scope = build_light_process_scope(demand, stock_lookup)

    all_items["신규분류"] = ""
    all_items["파워"] = all_items["생산코드"].map(extract_power_from_code)
    all_items["실수요수량"] = parse_mixed_numeric(all_items["오더수량"]).fillna(0)
    all_items["총수요수량"] = all_items["실수요수량"]
    all_items["부족수량"] = parse_mixed_numeric(all_items["생산부족수량"]).fillna(0)

    if not process_scope.empty:
        all_items = all_items.merge(process_scope, on="생산코드", how="left")
    for col in ["사출코드", "분리코드", "U코드", "코드매칭상태"]:
        if col not in all_items.columns:
            all_items[col] = ""
    all_items["코드매칭상태"] = all_items["코드매칭상태"].fillna("코드미매칭").replace({"": "코드미매칭"})

    all_items["사출창고"] = all_items["사출코드"].map(lambda x: lookup_stock_qty(stock_lookup.get("사출창고", {}), x))
    all_items["분리창고"] = all_items.apply(
        lambda row: lookup_stock_qty_from_candidates(
            stock_lookup.get("분리창고", {}),
            [row.get("분리코드", ""), row.get("U코드", "")],
        ),
        axis=1,
    )
    all_items["검사접착창고"] = all_items["생산코드"].map(lambda x: stock_lookup.get("검사접착창고", {}).get(x, 0))
    all_items["누수규격검사"] = all_items["생산코드"].map(lambda x: stock_lookup.get("누수규격검사 창고", {}).get(x, 0))
    all_items["공정재고합계"] = (
        all_items["사출창고"] + all_items["분리창고"] + all_items["검사접착창고"] + all_items["누수규격검사"]
    )
    all_items["제품대분류"] = all_items.apply(
        lambda row: classify_flow_primary_group(row.get("신규분류", ""), row.get("제품명", "")),
        axis=1,
    )
    all_items["거래처그룹"] = all_items["거래처"].map(normalize_flow_customer_group)
    all_items[INITIAL_ORDER_MAP_COL] = [
        json.dumps({clean_text_value(initial) or "미지정": float(qty)}, ensure_ascii=False) if float(qty or 0) else ""
        for initial, qty in zip(all_items["이니셜"], all_items["오더수량"])
    ]
    all_items[DEMAND_DETAIL_ROWS_COL] = ROW_DETAIL_MARKER
    finished_goods_stock_path = find_finished_goods_stock_file(data_base_dir)
    if finished_goods_stock_path is not None:
        finished_goods_summary = read_finished_goods_stock_summary(str(finished_goods_stock_path), refresh_key)
        if not finished_goods_summary.empty:
            all_items = all_items.merge(finished_goods_summary, on="생산코드", how="left")
    for col in ["완제품재고", "재고변화", "DOI기준오더", "DOI"]:
        if col not in all_items.columns:
            all_items[col] = 0
    for col in ["재고비율", "신호", "재고대응판단"]:
        if col not in all_items.columns:
            all_items[col] = ""
    all_items["초과재고수량"] = (all_items["공정재고합계"] - all_items["총수요수량"]).clip(lower=0)
    all_items["샘플 신청가능수량"] = 0
    all_items["판단"] = "수요 있음"
    all_items["상태"] = "수요 있음"

    for col in ALL_ITEM_DOWNLOAD_COLUMNS:
        if col not in all_items.columns:
            all_items[col] = 0 if col in ALL_ITEM_NUMERIC_COLUMNS else ""
    for col in ALL_ITEM_NUMERIC_COLUMNS:
        all_items[col] = parse_mixed_numeric(all_items[col]).fillna(0)
    for col in [
        "제품대분류",
        "거래처그룹",
        "거래처",
        "이니셜",
        "사이트코드",
        "신규분류",
        "제품명",
        "파워",
        "납기일",
        "사출코드",
        "분리코드",
        "생산코드",
        "재고비율",
        "신호",
        "재고대응판단",
        "판단",
        "상태",
        "코드매칭상태",
        INITIAL_ORDER_MAP_COL,
        DEMAND_DETAIL_ROWS_COL,
    ]:
        if col not in all_items.columns:
            all_items[col] = ""
        all_items[col] = all_items[col].astype(str).replace({"nan": "", "None": ""}).fillna("")

    return all_items[output_columns].sort_values(
        ["사이트코드", "제품대분류", "거래처그룹", "이니셜", "제품명", "납기일", "생산코드"],
        ascending=[True, True, True, True, True, True, True],
    ).reset_index(drop=True)


def normalize_inventory_family_code(item_code: object) -> str:
    code = re.sub(r"\s+", "", str(item_code).strip().upper())
    if not code or code.lower() in {"nan", "none", "-"}:
        return ""
    family_code = POWER_VALUE_PATTERN.sub("", code)
    family_code = re.sub(r"[+-]+$", "", family_code).strip()
    return family_code or code


def format_optional_date_series(series: pd.Series) -> pd.Series:
    parsed = parse_mixed_excel_date(series)
    return pd.to_datetime(parsed, errors="coerce").dt.strftime("%Y-%m-%d").fillna("")


def first_nonempty_text(values: pd.Series) -> str:
    for value in values.astype(str):
        text = value.strip()
        if text and text.lower() not in {"nan", "none", "nat", "-"}:
            return text
    return ""


def build_inventory_risk_source_df(inv: pd.DataFrame) -> pd.DataFrame:
    if inv.empty:
        return pd.DataFrame()

    inv = inv.copy()
    inv.columns = [str(c).strip() for c in inv.columns]
    columns = inv.columns.tolist()
    index = inv.index

    qty_col = pick_api_column(columns, ["총 재공 수량", "재공수량", "재고수량", "WIP_QTY", "WIP수량", "QTY", "재고량"])
    item_col = pick_api_column(columns, ["제품 코드", "ITEM_ID", "ITEM_CODE", "ITEM_CD", "제품코드", "품목코드"])
    warehouse_col = pick_api_column(
        columns,
        ["WH_NAME", "창고명", "공정(버퍼)", "공정", "버퍼 코드", "BUFFER_CODE", "제품위치(창고)", "PROP02", "창고"],
    )
    wip_code_col = pick_api_column(columns, ["재공 코드", "재공코드", "WIP_CODE", "WIP ID", "WIP_ID", "수요ID"])
    lot_col = pick_first_existing_column(columns, ["LOT_NO", "Lot no.", "LOT NO", "LOT"])
    available_col = pick_first_existing_column(columns, ["사용가능한 날짜", "사용가능일", "AVAILABLE_DATE"])
    created_col = pick_first_existing_column(columns, ["생성 일시", "생성일시", "CREATED_AT"])

    if qty_col is None:
        qty_col = columns[6] if len(columns) > 6 else columns[0]
    if item_col is None:
        item_col = columns[8] if len(columns) > 8 else (columns[1] if len(columns) > 1 else columns[0])
    if warehouse_col is None:
        warehouse_col = (
            columns[23]
            if len(columns) > 23
            else (columns[10] if len(columns) > 10 else (columns[5] if len(columns) > 5 else columns[0]))
        )
    if wip_code_col is None:
        wip_code_col = columns[3] if len(columns) > 3 else item_col

    def optional_text(col: str | None) -> pd.Series:
        if col and col in inv.columns:
            return inv[col].astype(str).str.strip()
        return pd.Series("", index=index)

    source = pd.DataFrame(
        {
            "품목코드": inv[item_col].astype(str).str.strip(),
            "창고": inv[warehouse_col].astype(str).str.strip().map(canonicalize_warehouse_label),
            "재공코드": inv[wip_code_col].astype(str).str.strip(),
            "재고수량": parse_mixed_numeric(inv[qty_col]),
            "LOT_NO": optional_text(lot_col),
            "사용가능일": format_optional_date_series(inv[available_col]) if available_col in inv.columns else pd.Series("", index=index),
            "생성일시": format_optional_date_series(inv[created_col]) if created_col in inv.columns else pd.Series("", index=index),
        }
    )

    rework_mask = (source["창고"] == "검사접착") & source["재공코드"].map(is_inspection_rework_wip_code)
    source.loc[rework_mask, "창고"] = "검사접착재작업"
    source = source[(source["품목코드"] != "") & (source["품목코드"].str.lower() != "nan")]
    source = source[source["창고"].isin(TARGET_WAREHOUSES)]
    source = source[source["재고수량"] > 0]
    code_prefix = source["품목코드"].astype(str).str.strip().str.upper().str[:1]
    rq_process_mask = ((source["창고"] == "사출창고") & (code_prefix == "R")) | (
        (source["창고"] == "분리창고") & (code_prefix == "Q")
    )
    source = source[rq_process_mask]
    return source


def build_inventory_demand_code_scope(demand_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if demand_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    base_index = demand_df.index

    def text_col(col: str) -> pd.Series:
        if col in demand_df.columns:
            return demand_df[col].astype(str).str.strip()
        return pd.Series("", index=base_index)

    def numeric_col(col: str) -> pd.Series:
        if col in demand_df.columns:
            return parse_mixed_numeric(demand_df[col])
        return pd.Series(0.0, index=base_index)

    injection_qty = numeric_col("사출생산필요수량")
    separation_qty = numeric_col(SEPARATION_REQUIRED_QTY_COL)

    common = pd.DataFrame(
        {
            "제품명": text_col("제품명"),
            "이니셜": text_col("이니셜"),
            "거래처": text_col("거래처"),
            "납기일": text_col("납기일"),
            "시트분류": text_col("시트분류"),
            "분류별요약": text_col("분류별요약"),
        },
        index=base_index,
    )

    scope_frames: list[pd.DataFrame] = []
    code_specs = [
        ("R코드", injection_qty),
        ("Q코드", separation_qty),
    ]
    for code_col, qty_series in code_specs:
        if code_col not in demand_df.columns:
            continue
        frame = common.copy()
        frame["매칭코드"] = demand_df[code_col].astype(str).str.strip()
        frame["현재수요수량"] = qty_series
        frame = frame[(frame["매칭코드"] != "") & (~frame["매칭코드"].str.lower().isin({"nan", "none", "-"}))]
        if not frame.empty:
            scope_frames.append(frame)

    if not scope_frames:
        return pd.DataFrame(), pd.DataFrame()

    code_scope = pd.concat(scope_frames, ignore_index=True, sort=False)
    code_scope["제품군키"] = code_scope["매칭코드"].map(normalize_inventory_family_code)
    code_summary = (
        code_scope.groupby("매칭코드", as_index=False)
        .agg(
            {
                "현재수요수량": "sum",
                "제품군키": first_nonempty_text,
                "제품명": lambda s: summarize_unique(s, 3),
                "이니셜": lambda s: summarize_unique(s, 3),
                "거래처": lambda s: summarize_unique(s, 2),
                "납기일": first_nonempty_text,
                "시트분류": first_nonempty_text,
                "분류별요약": first_nonempty_text,
            }
        )
        .rename(columns={"제품명": "제품명 예시", "이니셜": "이니셜 예시"})
    )
    family_summary = (
        code_scope.groupby("제품군키", as_index=False)
        .agg(
            {
                "매칭코드": "nunique",
                "현재수요수량": "sum",
                "제품명": lambda s: summarize_unique(s, 3),
                "이니셜": lambda s: summarize_unique(s, 3),
                "거래처": lambda s: summarize_unique(s, 2),
                "납기일": first_nonempty_text,
                "시트분류": first_nonempty_text,
                "분류별요약": first_nonempty_text,
            }
        )
        .rename(columns={"매칭코드": "수요코드수", "제품명": "제품명 예시", "이니셜": "이니셜 예시"})
    )
    return code_summary, family_summary


@st.cache_resource(show_spinner=False)
def build_inventory_risk_snapshot(refresh_key: str, base_dir_str: str | None = None) -> pd.DataFrame:
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    inv_path, _ = find_excel_files(data_base_dir)
    demand_df, _, _ = load_data(refresh_key, base_dir_str)
    inv_raw = read_inventory_excel_subset(inv_path)
    inv_source = build_inventory_risk_source_df(inv_raw)
    if inv_source.empty:
        return pd.DataFrame()

    code_summary, family_summary = build_inventory_demand_code_scope(demand_df)
    code_lookup = code_summary.set_index("매칭코드") if not code_summary.empty else pd.DataFrame()
    family_lookup = family_summary.set_index("제품군키") if not family_summary.empty else pd.DataFrame()

    inventory_group = (
        inv_source.groupby(["품목코드", "창고"], as_index=False)
        .agg(
            {
                "재고수량": "sum",
                "재공코드": lambda s: summarize_unique(s, 3),
                "LOT_NO": lambda s: summarize_unique(s, 3),
                "사용가능일": first_nonempty_text,
                "생성일시": first_nonempty_text,
            }
        )
        .rename(columns={"재공코드": "재공코드 예시", "LOT_NO": "LOT 예시"})
    )
    inventory_group["창고"] = inventory_group["창고"].map(lambda x: WAREHOUSE_MAP.get(x, x))
    inventory_group["제품군키"] = inventory_group["품목코드"].map(normalize_inventory_family_code)
    inventory_group["파워"] = inventory_group["품목코드"].map(extract_power_from_code)

    exact_codes = set(code_summary["매칭코드"].astype(str)) if not code_summary.empty else set()
    family_codes = set(family_summary["제품군키"].astype(str)) if not family_summary.empty else set()
    exact_match = inventory_group["품목코드"].astype(str).isin(exact_codes)
    family_match = inventory_group["제품군키"].astype(str).isin(family_codes)

    def map_from_lookup(lookup: pd.DataFrame, column: str, keys: pd.Series, default: object = "") -> pd.Series:
        if lookup.empty or column not in lookup.columns:
            return pd.Series(default, index=keys.index)
        return keys.map(lookup[column]).fillna(default)

    inventory_group["현재수요수량"] = map_from_lookup(code_lookup, "현재수요수량", inventory_group["품목코드"], 0.0)
    inventory_group["제품명 예시"] = map_from_lookup(code_lookup, "제품명 예시", inventory_group["품목코드"], "")
    inventory_group["이니셜 예시"] = map_from_lookup(code_lookup, "이니셜 예시", inventory_group["품목코드"], "")
    inventory_group["납기일"] = map_from_lookup(code_lookup, "납기일", inventory_group["품목코드"], "")
    inventory_group["시트분류"] = map_from_lookup(code_lookup, "시트분류", inventory_group["품목코드"], "")
    inventory_group["분류별요약"] = map_from_lookup(code_lookup, "분류별요약", inventory_group["품목코드"], "")
    inventory_group["수요코드수"] = 1

    family_keys = inventory_group["제품군키"]
    for target_col in ["제품명 예시", "이니셜 예시", "납기일", "시트분류", "분류별요약"]:
        fallback = map_from_lookup(family_lookup, target_col, family_keys, "")
        missing = inventory_group[target_col].astype(str).str.strip().isin({"", "nan", "None"})
        inventory_group.loc[missing, target_col] = fallback.loc[missing]
    inventory_group["수요코드수"] = map_from_lookup(family_lookup, "수요코드수", family_keys, 0).where(
        ~exact_match, inventory_group["수요코드수"]
    )

    inventory_group["현재수요수량"] = parse_mixed_numeric(inventory_group["현재수요수량"]).fillna(0)
    inventory_group["초과수량"] = (inventory_group["재고수량"] - inventory_group["현재수요수량"]).clip(lower=0)
    inventory_group["리스크구분"] = "현재수요 제품군 없음"
    inventory_group.loc[family_match, "리스크구분"] = "동일제품 타도수 재고"
    inventory_group.loc[exact_match, "리스크구분"] = "수요코드 직접매칭"
    inventory_group.loc[exact_match & (inventory_group["초과수량"] > 0), "리스크구분"] = "수요초과 재고"

    priority = {
        "현재수요 제품군 없음": 1,
        "동일제품 타도수 재고": 2,
        "수요초과 재고": 3,
        "수요코드 직접매칭": 4,
    }
    inventory_group["정렬순위"] = inventory_group["리스크구분"].map(priority).fillna(9)
    for col in ["제품명 예시", "이니셜 예시", "납기일", "시트분류", "분류별요약", "재공코드 예시", "LOT 예시"]:
        inventory_group[col] = inventory_group[col].replace({"": "-", "nan": "-", "None": "-"}).fillna("-")
    return inventory_group.sort_values(["정렬순위", "재고수량", "품목코드"], ascending=[True, False, True])


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_filter_option_maps(
    df: pd.DataFrame, selected_site_option: str = "전체"
) -> tuple[dict[str, float], dict[str, float], dict[str, float]]:
    process_qty_cols = [
        "사출생산필요수량",
        SEPARATION_REQUIRED_QTY_COL,
        LEADJI_REQUIRED_QTY_COL,
        ADHESION_REQUIRED_QTY_COL,
        "부족수량",
    ]
    required_cols = ["사이트코드", "시트분류", "분류별요약", *process_qty_cols]
    option_df = df[[c for c in required_cols if c in df.columns]].copy()
    for qty_col in process_qty_cols:
        if qty_col not in option_df.columns:
            option_df[qty_col] = 0
    if "사이트코드" not in option_df.columns:
        option_df["사이트코드"] = "(미지정)"
    if "시트분류" not in option_df.columns:
        option_df["시트분류"] = "(미분류)"
    if "분류별요약" not in option_df.columns:
        option_df["분류별요약"] = "(미분류)"

    site_label = option_df["사이트코드"].astype(str).str.strip()
    option_df["사이트코드"] = site_label.replace({"": "(미지정)", "nan": "(미지정)", "None": "(미지정)"})
    option_df["필터수량"] = 0.0
    for qty_col in process_qty_cols:
        option_df[qty_col] = parse_mixed_numeric(option_df[qty_col])
        option_df["필터수량"] = option_df["필터수량"] + option_df[qty_col]

    site_sum_map = option_df.groupby("사이트코드", as_index=True)["필터수량"].sum().sort_values(ascending=False).to_dict()

    scoped = option_df
    if selected_site_option and selected_site_option != "전체":
        scoped = scoped[scoped["사이트코드"] == selected_site_option]

    sheet_sum_map = scoped.groupby("시트분류", as_index=True)["필터수량"].sum().sort_values(ascending=False).to_dict()
    summary_sum_map = scoped.groupby("분류별요약", as_index=True)["필터수량"].sum().sort_values(ascending=False).to_dict()
    return site_sum_map, sheet_sum_map, summary_sum_map


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def filter_data(
    df: pd.DataFrame,
    selected_site_option: str,
    unified_query: str,
    exclude_safe_initial: bool,
    selected_sheet_options: tuple[str, ...],
    selected_summary_options: tuple[str, ...],
    only_same_rq_group: bool,
    only_with_stock: bool,
    only_rework_available: bool,
) -> pd.DataFrame:
    base_filtered = df.copy()
    if "사이트코드" not in base_filtered.columns:
        base_filtered["사이트코드"] = "(미지정)"
    site_label = base_filtered["사이트코드"].astype(str).str.strip()
    base_filtered["사이트코드"] = site_label.replace({"": "(미지정)", "nan": "(미지정)", "None": "(미지정)"})

    if selected_site_option and selected_site_option != "전체":
        base_filtered = base_filtered[base_filtered["사이트코드"] == selected_site_option]

    search_cols = [
        c
        for c in [
            "사이트코드",
            ORDER_NO_COL,
            "이니셜",
            "거래처",
            "품목코드",
            "제품명",
            "비고",
            "재작업",
            "R코드 제품명",
            "R코드",
            "Q코드",
            "U코드",
        ]
        if c in base_filtered.columns
    ]
    base_filtered = filter_with_terms_any(base_filtered, search_cols, unified_query)
    if exclude_safe_initial and "이니셜" in base_filtered.columns:
        base_filtered = base_filtered[~base_filtered["이니셜"].astype(str).str.contains("안전", na=False)]
    if is_specific_pill_selection(selected_sheet_options) and "시트분류" in base_filtered.columns:
        base_filtered = base_filtered[base_filtered["시트분류"].isin(selected_sheet_options)]
    if is_specific_pill_selection(selected_summary_options) and "분류별요약" in base_filtered.columns:
        base_filtered = base_filtered[base_filtered["분류별요약"].isin(selected_summary_options)]
    if only_same_rq_group and {"R코드5", "Q코드5", "P코드5"}.issubset(base_filtered.columns):
        p_count_per_group = base_filtered.groupby(["R코드5", "Q코드5"])["P코드5"].transform("nunique")
        base_filtered = base_filtered[p_count_per_group >= 2]
    if only_with_stock and "공정재고 합계" in base_filtered.columns:
        base_filtered = base_filtered[base_filtered["공정재고 합계"] > 0]
    if only_rework_available and "재작업" in base_filtered.columns:
        rework_text = base_filtered["재작업"].astype(str).str.strip()
        base_filtered = base_filtered[rework_text.ne("") & ~rework_text.str.lower().isin(INVALID_CATEGORY_VALUES)]

    return base_filtered.copy()


def apply_filters(
    df: pd.DataFrame,
    updated_at: str,
    data_base_dir: Path | None = None,
    source_label: str = "",
) -> pd.DataFrame:
    with st.sidebar:
        st.markdown('<div class="sidebar-section-title">필터</div>', unsafe_allow_html=True)
        st.caption(f"업데이트: {updated_at}")
        st.caption(f"앱 버전: {APP_CACHE_VERSION}")
        st.caption("기본 적용: 전체 수요")

        site_sum_map, _, _ = build_filter_option_maps(df, "전체")
        site_options = ["전체"] + list(site_sum_map.keys())
        site_count_map = {"전체": float(sum(site_sum_map.values())), **site_sum_map}
        selected_site_option = st.pills(
            "사이트코드",
            options=site_options,
            default="전체",
            key="flt_site_pills",
            format_func=lambda x: format_pill_label(x, site_count_map),
        )

        st.divider()
        unified_query = st.text_input(
            "통합 검색",
            value="",
            key="flt_unified_query",
            placeholder="사이트/거래처/품목/RQ 코드",
            help="콤마(,)로 여러 키워드를 입력하면 OR 조건으로 검색합니다.",
        ).strip()

        only_with_stock = st.checkbox("공정재고만", value=False, key="flt_only_stock")
        only_rework_available = st.checkbox("재작업만", value=False, key="flt_only_rework_available")
        exclude_safe_initial = st.checkbox("안전 이니셜 제외", value=False, key="flt_exclude_safe_initial")
        only_same_rq_group = st.checkbox("동일 RQ그룹만(R5/Q5, P5종류2+)", value=False, key="flt_only_same_rq_group")

        _, sheet_sum_map, summary_sum_map = build_filter_option_maps(df, selected_site_option or "전체")

        sheet_options = ["전체"] + list(sheet_sum_map.keys())
        summary_options = ["전체"] + list(summary_sum_map.keys())
        scoped_total = float(sum(sheet_sum_map.values()))
        sheet_count_map = {"전체": scoped_total, **sheet_sum_map}
        summary_count_map = {"전체": scoped_total, **summary_sum_map}

        st.divider()
        sheet_pills_key = "flt_sheet_pills"
        prepare_multi_pill_state(sheet_pills_key, sheet_options)
        selected_sheet_options = finalize_multi_pill_selection(
            sheet_pills_key,
            st.pills(
                "시트 분류",
                options=sheet_options,
                selection_mode="multi",
                key=sheet_pills_key,
                format_func=lambda x: format_pill_label(x, sheet_count_map),
                on_change=sync_multi_pill_state,
                args=(sheet_pills_key,),
            ),
        )
        summary_pills_key = "flt_summary_pills"
        prepare_multi_pill_state(summary_pills_key, summary_options)
        selected_summary_options = finalize_multi_pill_selection(
            summary_pills_key,
            st.pills(
                "분류별 요약",
                options=summary_options,
                selection_mode="multi",
                key=summary_pills_key,
                format_func=lambda x: format_pill_label(x, summary_count_map),
                on_change=sync_multi_pill_state,
                args=(summary_pills_key,),
            ),
        )
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        if data_base_dir is not None:
            render_sidebar_reference_dates(data_base_dir, source_label)

    return filter_data(
        df,
        selected_site_option or "전체",
        unified_query,
        exclude_safe_initial,
        selected_sheet_options,
        selected_summary_options,
        only_same_rq_group,
        only_with_stock,
        only_rework_available,
    )


@st.cache_resource(show_spinner=False)
def load_leadji_data(refresh_key: str, base_dir_str: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    ref_path = find_product_name_reference_file(data_base_dir)
    if ref_path is None:
        return pd.DataFrame(), pd.DataFrame()

    xls = pd.ExcelFile(ref_path)
    sheet_names = xls.sheet_names
    leadji_info_sheet = next((s for s in sheet_names if s.replace(" ", "") == "리드지정보"), None)
    leadji_stock_sheet = next((s for s in sheet_names if s.replace(" ", "") == "리드지재고"), None)

    leadji_info = pd.DataFrame()
    if leadji_info_sheet:
        preview = xls.parse(sheet_name=leadji_info_sheet, nrows=0)
        preview_cols = [str(c).strip() for c in preview.columns]
        prod_col = pick_first_existing_column(preview_cols, ["생산"]) or (preview_cols[3] if len(preview_cols) > 3 else None)
        b1_col = pick_first_existing_column(preview_cols, ["B1코드"]) or (preview_cols[12] if len(preview_cols) > 12 else None)
        b1_name_col = pick_first_existing_column(preview_cols, ["B1코드명"]) or (
            preview_cols[13] if len(preview_cols) > 13 else None
        )
        selected_cols = [c for c in [prod_col, b1_col, b1_name_col] if c is not None]
        if selected_cols:
            leadji_info = xls.parse(
                sheet_name=leadji_info_sheet,
                usecols=lambda c: str(c).strip() in set(selected_cols),
            )
            leadji_info.columns = [str(c).strip() for c in leadji_info.columns]
            rename_map = {}
            if prod_col is not None:
                rename_map[prod_col] = "생산"
            if b1_col is not None:
                rename_map[b1_col] = "B1코드"
            if b1_name_col is not None:
                rename_map[b1_name_col] = "B1코드명"
            leadji_info = leadji_info.rename(columns=rename_map)

    leadji_stock = pd.DataFrame()
    if leadji_stock_sheet:
        preview = xls.parse(sheet_name=leadji_stock_sheet, nrows=0)
        preview_cols = [str(c).strip() for c in preview.columns]
        code_col = pick_first_existing_column(preview_cols, ["품목코드"])
        warehouse_col = pick_first_existing_column(preview_cols, ["창고"])
        qty_col = pick_first_existing_column(preview_cols, ["재고"])
        selected_cols = [c for c in [code_col, warehouse_col, qty_col] if c is not None]
        if selected_cols:
            leadji_stock = xls.parse(
                sheet_name=leadji_stock_sheet,
                usecols=lambda c: str(c).strip() in set(selected_cols),
            )
            leadji_stock.columns = [str(c).strip() for c in leadji_stock.columns]
            leadji_stock = leadji_stock.rename(
                columns={
                    code_col: "품목코드",
                    warehouse_col: "창고",
                    qty_col: "재고",
                }
            )

    if not leadji_info.empty:
        leadji_info.columns = [str(c).strip() for c in leadji_info.columns]
        for col in leadji_info.columns:
            if "소요량" in col:
                leadji_info[col] = parse_mixed_numeric(leadji_info[col])

    if not leadji_stock.empty:
        leadji_stock.columns = [str(c).strip() for c in leadji_stock.columns]
        for col in ["기초", "입고", "출고", "재고", "검사대기"]:
            if col in leadji_stock.columns:
                leadji_stock[col] = parse_mixed_numeric(leadji_stock[col])

    return leadji_info, leadji_stock


@st.cache_resource(show_spinner=False)
def load_leadji_order_data(refresh_key: str, base_dir_str: str | None = None) -> pd.DataFrame:
    _ = refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR
    order_path = find_leadji_order_status_file(data_base_dir)

    empty = pd.DataFrame(
        columns=[
            "리드지코드",
            "리드지명",
            "발주수량",
            "입고예상일자",
            "입고예상일자_dt",
            "구매발주수량",
            "구매의뢰수량",
        ]
    )
    if order_path is None:
        return empty

    try:
        sheet_names = pd.ExcelFile(order_path).sheet_names
    except Exception:
        return empty

    normalized_sheet_names = {str(name).replace(" ", ""): name for name in sheet_names}
    purchase_order_sheet = normalized_sheet_names.get("구매발주현황", sheet_names[0] if sheet_names else 0)
    purchase_request_sheet = normalized_sheet_names.get("구매의뢰현황")

    summaries: list[pd.DataFrame] = []

    def first_nonempty_text(series: pd.Series) -> str:
        text = series.astype(str).str.strip()
        text = text[(text != "") & (text.str.lower() != "nan") & (text.str.lower() != "none")]
        return text.iloc[0] if not text.empty else "-"

    try:
        # 구매발주현황 기준: J열(품목코드), O열(미입고수량), X열(납기일자).
        # 화면의 "발주수량"은 현재 남아 있는 입고 예정 수량이므로 미입고수량을 사용한다.
        raw_order = pd.read_excel(order_path, sheet_name=purchase_order_sheet, header=0, usecols=[9, 11, 14, 23])
    except Exception:
        raw_order = pd.DataFrame()

    if not raw_order.empty:
        raw_order.columns = ["리드지코드_raw", "리드지명", "구매발주수량", "입고예상일자_raw"]
        raw_order["리드지코드"] = raw_order["리드지코드_raw"].map(normalize_leadji_code_key)
        raw_order = raw_order[raw_order["리드지코드"].str.fullmatch(r"[A-Z]{2}\d{4}", na=False)]
        raw_order["구매발주수량"] = parse_mixed_numeric(raw_order["구매발주수량"])
        raw_order["입고예상일자_dt"] = parse_mixed_excel_date(raw_order["입고예상일자_raw"])
        raw_order = raw_order[raw_order["구매발주수량"] > 0]
        if not raw_order.empty:
            order_summary = raw_order.groupby("리드지코드", as_index=False).agg(
                {"리드지명": first_nonempty_text, "구매발주수량": "sum", "입고예상일자_dt": "min"}
            )
            summaries.append(order_summary)

    if purchase_request_sheet is not None:
        try:
            # 구매의뢰현황 기준: G열(품목코드), U열(발주잔량), Y열(요청일).
            raw_request = pd.read_excel(
                order_path, sheet_name=purchase_request_sheet, header=0, usecols=[6, 7, 20, 24]
            )
        except Exception:
            raw_request = pd.DataFrame()
    else:
        raw_request = pd.DataFrame()

    if not raw_request.empty:
        raw_request.columns = ["리드지코드_raw", "리드지명", "구매의뢰수량", "입고예상일자_raw"]
        raw_request["리드지코드"] = raw_request["리드지코드_raw"].map(normalize_leadji_code_key)
        raw_request = raw_request[raw_request["리드지코드"].str.fullmatch(r"[A-Z]{2}\d{4}", na=False)]
        raw_request["구매의뢰수량"] = parse_mixed_numeric(raw_request["구매의뢰수량"])
        raw_request["입고예상일자_dt"] = parse_mixed_excel_date(raw_request["입고예상일자_raw"])
        raw_request = raw_request[raw_request["구매의뢰수량"] > 0]
        if not raw_request.empty:
            request_summary = raw_request.groupby("리드지코드", as_index=False).agg(
                {"리드지명": first_nonempty_text, "구매의뢰수량": "sum", "입고예상일자_dt": "min"}
            )
            summaries.append(request_summary)

    if not summaries:
        return empty

    summary = pd.concat(summaries, ignore_index=True, sort=False)
    for qty_col in ["구매발주수량", "구매의뢰수량"]:
        if qty_col not in summary.columns:
            summary[qty_col] = 0.0
        summary[qty_col] = parse_mixed_numeric(summary[qty_col])
    summary["입고예상일자_dt"] = pd.to_datetime(summary["입고예상일자_dt"], errors="coerce")
    summary = summary.groupby("리드지코드", as_index=False).agg(
        {
            "리드지명": first_nonempty_text,
            "구매발주수량": "sum",
            "구매의뢰수량": "sum",
            "입고예상일자_dt": "min",
        }
    )
    summary["발주수량"] = summary["구매발주수량"] + summary["구매의뢰수량"]
    summary["입고예상일자"] = summary["입고예상일자_dt"].dt.strftime("%Y-%m-%d").fillna("미확인")
    return summary[
        ["리드지코드", "리드지명", "발주수량", "입고예상일자", "입고예상일자_dt", "구매발주수량", "구매의뢰수량"]
    ]


@st.cache_resource(show_spinner=False)
def load_leadji_status_snapshot(
    leadji_status_refresh_key: str, base_dir_str: str | None = None
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    _ = leadji_status_refresh_key
    data_base_dir = Path(base_dir_str) if base_dir_str else BASE_DIR

    data_refresh_key = build_data_refresh_key(data_base_dir)
    reference_refresh_key = build_reference_refresh_key(data_base_dir)
    leadji_order_refresh_key = build_leadji_order_refresh_key(data_base_dir)

    shortage_df, _, _ = load_data(data_refresh_key, str(data_base_dir))
    leadji_info, leadji_stock = load_leadji_data(reference_refresh_key, str(data_base_dir))
    leadji_order_df = load_leadji_order_data(leadji_order_refresh_key, str(data_base_dir))
    return shortage_df, leadji_info, leadji_stock, leadji_order_df


def render_rework_match_debug(file_info_df: pd.DataFrame | None) -> None:
    if file_info_df is None or file_info_df.empty:
        return

    row = file_info_df.iloc[0]
    source_count = int(row.get("재작업 리스트 키 수", 0) or 0)
    matched_count = int(row.get("재작업 매칭 키 수", 0) or 0)
    note_count = int(row.get("재작업 비고 키 수", 0) or 0)
    source_qty_total = float(row.get("재작업 리스트 수량 합계", 0) or 0)
    matched_qty_total = float(row.get("재작업 매칭 수량 합계", 0) or 0)
    sample_text = str(row.get("재작업 매칭 키 샘플", "") or "").strip()
    sample_codes = [code.strip() for code in sample_text.split(",") if code.strip()]
    rework_sheet = str(row.get("재작업 시트명", "-") or "-")
    basis_columns = str(row.get("재작업 기준 컬럼", "") or "").strip()
    sheet_columns = str(row.get("재작업 시트 컬럼", "") or "").strip()

    with st.expander("재작업 매칭 디버그", expanded=False):
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("재작업 소스 키 수", f"{source_count:,}")
        d2.metric("생산현황 매칭 이니셜+품목코드 수", f"{matched_count:,}")
        d3.metric("재작업 소스 수량 합계", f"{source_qty_total:,.0f}")
        d4.metric("생산현황 매칭 수량 합계", f"{matched_qty_total:,.0f}")
        st.caption(f"재작업 소스: {rework_sheet}")
        st.caption(f"매칭 기준 컬럼: {basis_columns if basis_columns else '없음'}")
        st.caption(f"재작업 비고 입력 키 수: {note_count:,}")
        if sheet_columns:
            st.caption(f"재작업 소스 컬럼: {sheet_columns}")
        if sample_codes:
            st.dataframe(pd.DataFrame({"매칭 이니셜/품목코드 샘플": sample_codes[:10]}), width="stretch", hide_index=True)
        else:
            st.caption("매칭된 재작업 이니셜/품목코드 샘플이 없습니다.")


def filter_all_item_status(df: pd.DataFrame, selected_status: str) -> pd.DataFrame:
    if selected_status == "주의 필요":
        return df[
            df["판단"].isin(["수요 없음 + 공정재고 있음", "수요 대비 재고 초과", "코드 확인 필요"])
        ]
    if selected_status == "수요 있음":
        return df[df["총수요수량"] > 0]
    if selected_status == "수요 없음":
        return df[df["총수요수량"] <= 0]
    if selected_status == "수요 없음 + 공정재고 있음":
        return df[(df["총수요수량"] <= 0) & (df["공정재고합계"] > 0)]
    if selected_status == "수요 대비 재고 초과":
        return df[(df["총수요수량"] > 0) & (df["초과재고수량"] > 0)]
    if selected_status == "수요 없음 + 샘플 신청가능수량 있음":
        return df[(df["총수요수량"] <= 0) & (df["샘플 신청가능수량"] > 0)]
    if selected_status == "재고 있음":
        return df[df["공정재고합계"] > 0]
    if selected_status == "코드미매칭":
        return df[df["코드매칭상태"] == "코드미매칭"]
    return df


def build_new_class_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary_columns = [
        "신규분류",
        "품목 수",
        "실수요수량 합계",
        "총수요수량 합계",
        "사출창고 합계",
        "분리창고 합계",
        "검사접착창고 합계",
        "누수규격검사 합계",
        "공정재고합계",
        "초과재고수량 합계",
        "부족수량 합계",
        "샘플 신청가능수량 합계",
    ]
    if df.empty:
        return pd.DataFrame(columns=summary_columns)

    summary = (
        df.groupby("신규분류", as_index=False)
        .agg(
            {
                "생산코드": "count",
                "실수요수량": "sum",
                "총수요수량": "sum",
                "사출창고": "sum",
                "분리창고": "sum",
                "검사접착창고": "sum",
                "누수규격검사": "sum",
                "공정재고합계": "sum",
                "초과재고수량": "sum",
                "부족수량": "sum",
                "샘플 신청가능수량": "sum",
            }
        )
        .rename(
            columns={
                "생산코드": "품목 수",
                "실수요수량": "실수요수량 합계",
                "총수요수량": "총수요수량 합계",
                "사출창고": "사출창고 합계",
                "분리창고": "분리창고 합계",
                "검사접착창고": "검사접착창고 합계",
                "누수규격검사": "누수규격검사 합계",
                "초과재고수량": "초과재고수량 합계",
                "부족수량": "부족수량 합계",
                "샘플 신청가능수량": "샘플 신청가능수량 합계",
            }
        )
        .sort_values(["총수요수량 합계", "공정재고합계", "품목 수"], ascending=[False, False, False])
    )
    return summary[summary_columns]


def prepare_all_item_flow_data(all_items_df: pd.DataFrame) -> pd.DataFrame:
    working = all_items_df.copy()
    for col in ALL_ITEM_DOWNLOAD_COLUMNS:
        if col not in working.columns:
            working[col] = 0 if col in ALL_ITEM_NUMERIC_COLUMNS else ""
    for col in [INITIAL_ORDER_MAP_COL, DEMAND_DETAIL_ROWS_COL]:
        if col not in working.columns:
            working[col] = ""
    for col in ALL_ITEM_NUMERIC_COLUMNS:
        working[col] = parse_mixed_numeric(working[col]).fillna(0)
    for col in [
        "사이트코드",
        "제품대분류",
        "거래처그룹",
        "거래처",
        "이니셜",
        "신규분류",
        "제품명",
        "파워",
        "납기일",
        "생산코드",
        "신호",
    ]:
        if col not in working.columns:
            working[col] = ""
        working[col] = working[col].map(clean_text_value)
    working["이니셜"] = working["이니셜"].map(clean_initial_value)
    working["사이트코드"] = working["사이트코드"].map(normalize_site_group)

    if working["제품대분류"].str.strip().eq("").all() or "기타" in set(working["제품대분류"]):
        fallback_primary = working.apply(
            lambda row: classify_flow_primary_group(row.get("신규분류", ""), row.get("제품명", "")),
            axis=1,
        )
        missing_primary = working["제품대분류"].str.strip().isin({"", "기타"})
        working.loc[missing_primary, "제품대분류"] = fallback_primary.loc[missing_primary]

    missing_customer_group = working["거래처그룹"].str.strip().isin({"", "거래처 미지정"})
    working.loc[missing_customer_group, "거래처그룹"] = working.loc[missing_customer_group, "거래처"].map(
        normalize_flow_customer_group
    )
    working.loc[working["제품대분류"].str.strip().eq(""), "제품대분류"] = "기타"
    working.loc[working["거래처그룹"].str.strip().eq(""), "거래처그룹"] = "거래처 미지정"
    working.loc[working["제품명"].str.strip().eq(""), "제품명"] = "-"
    working.loc[working["납기일"].str.lower().isin(INVALID_CATEGORY_VALUES), "납기일"] = "-"
    return working


def build_initial_order_map_text(initial: object, qty: object) -> str:
    initial_text = clean_initial_value(initial) or "미지정"
    qty_value = numeric_scalar(qty)
    if not qty_value:
        return ""
    return json.dumps({initial_text: qty_value}, ensure_ascii=False)


def parse_demand_detail_rows(value: object, fallback_source: pd.Series | None = None) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    text = clean_text_value(value)
    if text and text.lower() not in INVALID_CATEGORY_VALUES:
        try:
            payload = json.loads(text)
        except (TypeError, ValueError, json.JSONDecodeError):
            payload = None
        if isinstance(payload, list):
            for item in payload:
                if not isinstance(item, dict):
                    continue
                order_qty = numeric_scalar(item.get("오더수량", 0))
                shortage_qty = numeric_scalar(item.get("생산부족수량", 0))
                injection_qty = numeric_scalar(item.get("사출부족수량", 0))
                rows.append(
                    {
                        "거래처": clean_text_value(item.get("거래처", "")),
                        "사이트코드": normalize_site_group(item.get("사이트코드", "")),
                        "이니셜": clean_initial_value(item.get("이니셜", "")) or "미지정",
                        "제품명": clean_text_value(item.get("제품명", "")) or "-",
                        "오더수량": order_qty,
                        "납기일": clean_text_value(item.get("납기일", "")) or "-",
                        "생산부족수량": shortage_qty,
                        "사출부족수량": injection_qty,
                    }
                )
    if rows or fallback_source is None:
        return rows

    return [
        {
            "거래처": clean_text_value(fallback_source.get("거래처", "")),
            "사이트코드": normalize_site_group(fallback_source.get("사이트코드", "")),
            "이니셜": clean_initial_value(fallback_source.get("이니셜", "")) or "미지정",
            "제품명": clean_text_value(fallback_source.get("제품명", "")) or "-",
            "오더수량": numeric_scalar(fallback_source.get("오더수량", 0)),
            "납기일": clean_text_value(fallback_source.get("납기일", "")) or "-",
            "생산부족수량": numeric_scalar(fallback_source.get("생산부족수량", 0)),
            "사출부족수량": numeric_scalar(fallback_source.get("사출부족수량", 0)),
        }
    ]


def expand_all_item_demand_detail_rows(source_df: pd.DataFrame, include_fallback: bool = True) -> pd.DataFrame:
    if source_df.empty:
        return source_df.copy()

    expanded_rows: list[dict[str, object]] = []
    for _, source in source_df.iterrows():
        fallback_source = source if include_fallback else None
        details = parse_demand_detail_rows(source.get(DEMAND_DETAIL_ROWS_COL, ""), fallback_source)
        for detail in details:
            row = source.to_dict()
            for col in ["거래처", "사이트코드", "이니셜", "제품명", "오더수량", "납기일", "생산부족수량", "사출부족수량"]:
                row[col] = detail.get(col, row.get(col, ""))
            customer_text = clean_text_value(row.get("거래처", ""))
            if customer_text:
                row["거래처그룹"] = normalize_flow_customer_group(customer_text)
            row[INITIAL_ORDER_MAP_COL] = build_initial_order_map_text(row.get("이니셜", ""), row.get("오더수량", 0))
            expanded_rows.append(row)
    if not expanded_rows:
        return source_df.copy() if include_fallback else pd.DataFrame(columns=source_df.columns)
    return pd.DataFrame(expanded_rows)


def filter_rows_with_demand_initial(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "오더수량" not in df.columns or "이니셜" not in df.columns:
        return df.iloc[0:0].copy()

    initial = df["이니셜"].map(clean_initial_value)
    order_qty = parse_mixed_numeric(df["오더수량"]).fillna(0)
    valid_initial = initial.ne("") & ~initial.str.lower().isin(INVALID_CATEGORY_VALUES | {"미지정"})
    return df[(order_qty > 0) & valid_initial].copy()


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_all_item_product_flow_summary(working: pd.DataFrame) -> pd.DataFrame:
    summary_columns = ["사이트코드", "제품대분류", "거래처그룹", "거래처", "생산코드", *ALL_ITEM_FLOW_DISPLAY_COLUMNS]
    if working.empty:
        return pd.DataFrame(columns=summary_columns)

    has_demand_detail = working[DEMAND_DETAIL_ROWS_COL].map(clean_text_value).str.lower()
    demand_mask = has_demand_detail.ne("") & ~has_demand_detail.isin(INVALID_CATEGORY_VALUES)
    base_columns = list(
        dict.fromkeys(
            [
                "제품대분류",
                "사이트코드",
                "거래처그룹",
                "거래처",
                "생산코드",
                "이니셜",
                "제품명",
                "오더수량",
                "납기일",
                "생산부족수량",
                "사출부족수량",
                "공정재고합계",
                "완제품재고",
                "재고변화",
                "DOI기준오더",
                "DOI",
                "재고비율",
                "신호",
                "재고대응판단",
                INITIAL_ORDER_MAP_COL,
                DEMAND_DETAIL_ROWS_COL,
            ]
        )
    )
    base_columns = [col for col in base_columns if col in working.columns]
    base = working.loc[demand_mask, base_columns].copy()
    if base.empty:
        return pd.DataFrame(columns=summary_columns)
    base = expand_all_item_demand_detail_rows(base, include_fallback=True)
    base = filter_rows_with_demand_initial(base)
    if base.empty:
        return pd.DataFrame(columns=summary_columns)
    base["_납기일_dt"] = parse_mixed_excel_date(base["납기일"])

    group_keys = ["사이트코드", "제품대분류", "거래처그룹", "이니셜", "제품명"]
    order_group = (
        base.groupby(group_keys, as_index=False)
        .agg(
            {
                "거래처": summarize_unique,
                "생산코드": summarize_unique,
                "오더수량": "sum",
                "_납기일_dt": "min",
                "생산부족수량": "sum",
                "사출부족수량": "sum",
            }
        )
        .rename(columns={"생산부족수량": "부족수량"})
    )
    stock_group = (
        base.drop_duplicates([*group_keys, "생산코드"])
        .groupby(group_keys, as_index=False)["공정재고합계"]
        .sum()
        .rename(columns={"공정재고합계": "공정재고"})
    )
    finished_stock_group = (
        base.drop_duplicates([*group_keys, "생산코드"])
        .groupby(group_keys, as_index=False)
        .agg(
            {
                "완제품재고": "sum",
                "재고변화": "sum",
                "DOI기준오더": "sum",
                "DOI": "max",
                "재고비율": lambda s: summarize_unique(s, head_count=1),
                "신호": summarize_signal_values,
                "재고대응판단": lambda s: summarize_unique(s, head_count=1),
            }
        )
    )
    grouped = order_group.merge(stock_group, on=group_keys, how="left").merge(finished_stock_group, on=group_keys, how="left")
    grouped["공정재고"] = parse_mixed_numeric(grouped["공정재고"]).fillna(0)
    for col in ["완제품재고", "재고변화", "DOI기준오더", "DOI"]:
        grouped[col] = parse_mixed_numeric(grouped.get(col, pd.Series(0, index=grouped.index))).fillna(0)
    for col in ["재고비율", "신호", "재고대응판단"]:
        grouped[col] = grouped.get(col, pd.Series("", index=grouped.index)).astype(str).replace({"nan": "", "None": ""}).fillna("")
    grouped["납기일"] = pd.to_datetime(grouped["_납기일_dt"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("-")
    grouped = grouped.drop(columns=["_납기일_dt"], errors="ignore")
    doi_order_mask = grouped["DOI기준오더"] > 0
    grouped.loc[doi_order_mask, "DOI"] = (
        grouped.loc[doi_order_mask, "완제품재고"] / grouped.loc[doi_order_mask, "DOI기준오더"] * 181
    )
    grouped["DOI"] = parse_mixed_numeric(grouped["DOI"]).round(1)
    grouped = grouped.sort_values(
        ["이니셜", "제품명", "부족수량", "사출부족수량", "오더수량", "공정재고", "DOI"],
        ascending=[True, True, False, False, False, False, False],
    )
    return grouped[summary_columns]


def filter_all_item_flow_query(df: pd.DataFrame, query: str) -> pd.DataFrame:
    search_columns = [
        col
        for col in ["사이트코드", "제품대분류", "거래처그룹", "거래처", "이니셜", "신규분류", "제품명", "생산코드", "재고비율", "신호", "재고대응판단"]
        if col in df.columns
    ]
    return filter_with_terms_any(df, search_columns, query) if query else df


def build_customer_tab_options(df: pd.DataFrame) -> list[str]:
    present = {clean_text_value(value) for value in df["거래처그룹"].tolist() if clean_text_value(value)}
    present.discard("전체")
    ordered = [option for option in ALL_ITEM_FLOW_CUSTOMER_ORDER if option in present]
    extras = sorted(present - set(ordered))
    return ordered + extras


def build_site_tab_options(df: pd.DataFrame) -> list[str]:
    if df.empty or "사이트코드" not in df.columns:
        return ["전체"]
    present = {normalize_site_group(value) for value in df["사이트코드"].tolist()}
    present.discard("")
    ordered = [site for site in SITE_GROUP_ORDER if site in present]
    extras = sorted(present - set(ordered) - {"전체"})
    return ["전체", *ordered, *extras]


def numeric_scalar(value: object) -> float:
    if value is None:
        return 0.0
    try:
        if pd.isna(value):
            return 0.0
    except (TypeError, ValueError):
        pass
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text.lower() in INVALID_CATEGORY_VALUES:
        return 0.0
    text = re.sub(r"^\((.*)\)$", r"-\1", text)
    text = text.replace(",", "").replace("\u00a0", "").replace(" ", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_initial_order_qty_map(
    value: object,
    fallback_initial: object = "",
    fallback_qty: object = 0,
) -> dict[str, float]:
    result: dict[str, float] = {}
    payload: object = None
    if isinstance(value, dict):
        payload = value
    else:
        text = clean_text_value(value)
        if text and text.lower() not in INVALID_CATEGORY_VALUES:
            try:
                payload = json.loads(text)
            except (TypeError, ValueError, json.JSONDecodeError):
                payload = None

    if isinstance(payload, dict):
        for raw_initial, raw_qty in payload.items():
            initial = clean_initial_value(raw_initial) or "미지정"
            qty = numeric_scalar(raw_qty)
            if qty:
                result[initial] = result.get(initial, 0) + qty

    if not result:
        qty = numeric_scalar(fallback_qty)
        if qty:
            initial = clean_initial_value(fallback_initial) or "미지정"
            result[initial] = qty
    return result


def filter_all_item_excel_activity_rows(working: pd.DataFrame) -> pd.DataFrame:
    activity_mask = (
        (working["오더수량"] > 0)
        | (working["생산부족수량"] > 0)
        | (working["사출부족수량"] > 0)
        | (working["공정재고합계"] > 0)
        | (working["완제품재고"] > 0)
        | (working["재고변화"] != 0)
        | (working["DOI"] > 0)
        | working["신호"].astype(str).str.strip().ne("")
        | working["재고대응판단"].astype(str).str.strip().ne("")
    )
    return working[activity_mask].copy()


def build_all_item_customer_excel_sheet(scope: pd.DataFrame) -> pd.DataFrame:
    if scope.empty:
        return pd.DataFrame()
    scope = expand_all_item_demand_detail_rows(scope)
    scope["생산코드"] = scope["생산코드"].map(normalize_to_master_p_code)
    scope = scope[scope["생산코드"].str.startswith("P", na=False)].copy()
    if scope.empty:
        return pd.DataFrame()

    base_columns = ["생산코드", "사출코드", "분리코드", "제품명", "파워"]
    rows: dict[tuple[str, ...], dict[str, object]] = {}
    initial_totals: dict[str, float] = {}
    for _, source in scope.iterrows():
        key = tuple(clean_text_value(source.get(col, "")) for col in base_columns)
        if key not in rows:
            rows[key] = {
                "생산코드": key[0],
                "사출코드": key[1],
                "분리코드": key[2],
                "제품명": key[3],
                "파워": key[4],
                "오더합계": 0.0,
                "부족수량합계": 0.0,
                "사출부족수량합계": 0.0,
                "사출재고": 0.0,
                "분리재고": 0.0,
                "검사접착재고": 0.0,
                "누수규격재고": 0.0,
                "공정재고 합계": 0.0,
                "완제품재고": 0.0,
                "재고변화": 0.0,
                "DOI기준오더": 0.0,
                "DOI": 0.0,
                "재고비율": "",
                "_상태값": [],
                "_재고신호값": [],
                "_재고대응판단값": [],
                "_이니셜오더": {},
            }

        row = rows[key]
        order_map = parse_initial_order_qty_map(
            source.get(INITIAL_ORDER_MAP_COL, ""),
            source.get("이니셜", ""),
            source.get("오더수량", 0),
        )
        for initial, qty in order_map.items():
            orders = row["_이니셜오더"]
            orders[initial] = orders.get(initial, 0.0) + qty
            initial_totals[initial] = initial_totals.get(initial, 0.0) + qty

        row["오더합계"] = numeric_scalar(row["오더합계"]) + numeric_scalar(source.get("오더수량", 0))
        row["부족수량합계"] = numeric_scalar(row["부족수량합계"]) + numeric_scalar(source.get("생산부족수량", 0))
        row["사출부족수량합계"] = numeric_scalar(row["사출부족수량합계"]) + numeric_scalar(source.get("사출부족수량", 0))
        row["사출재고"] = max(numeric_scalar(row["사출재고"]), numeric_scalar(source.get("사출창고", 0)))
        row["분리재고"] = max(numeric_scalar(row["분리재고"]), numeric_scalar(source.get("분리창고", 0)))
        row["검사접착재고"] = max(numeric_scalar(row["검사접착재고"]), numeric_scalar(source.get("검사접착창고", 0)))
        row["누수규격재고"] = max(numeric_scalar(row["누수규격재고"]), numeric_scalar(source.get("누수규격검사", 0)))
        row["공정재고 합계"] = max(numeric_scalar(row["공정재고 합계"]), numeric_scalar(source.get("공정재고합계", 0)))
        row["완제품재고"] = max(numeric_scalar(row["완제품재고"]), numeric_scalar(source.get("완제품재고", 0)))
        row["재고변화"] = numeric_scalar(source.get("재고변화", row["재고변화"]))
        row["DOI기준오더"] = max(numeric_scalar(row["DOI기준오더"]), numeric_scalar(source.get("DOI기준오더", 0)))
        row["DOI"] = max(numeric_scalar(row["DOI"]), numeric_scalar(source.get("DOI", 0)))
        stock_ratio = clean_text_value(source.get("재고비율", ""))
        if stock_ratio:
            row["재고비율"] = stock_ratio
        stock_signal = clean_text_value(source.get("신호", ""))
        if stock_signal and stock_signal not in row["_재고신호값"]:
            row["_재고신호값"].append(stock_signal)
        stock_action = clean_text_value(source.get("재고대응판단", ""))
        if stock_action and stock_action not in row["_재고대응판단값"]:
            row["_재고대응판단값"].append(stock_action)
        status = clean_text_value(source.get("상태", ""))
        if status and status not in row["_상태값"]:
            row["_상태값"].append(status)

    initial_values = sorted(initial_totals)
    order_columns = {initial: f"오더수량({initial})" for initial in initial_values}
    output_rows: list[dict[str, object]] = []
    for row in rows.values():
        output = {col: row[col] for col in base_columns}
        orders = row["_이니셜오더"]
        for initial in initial_values:
            output[order_columns[initial]] = orders.get(initial, 0.0)
        output.update(
            {
                "오더합계": row["오더합계"],
                "부족수량합계": row["부족수량합계"],
                "사출부족수량합계": row["사출부족수량합계"],
                "사출재고": row["사출재고"],
                "분리재고": row["분리재고"],
                "검사접착재고": row["검사접착재고"],
                "누수규격재고": row["누수규격재고"],
                "공정재고 합계": row["공정재고 합계"],
                "완제품재고": row["완제품재고"],
                "재고변화": row["재고변화"],
                "DOI기준오더": row["DOI기준오더"],
                "DOI": row["DOI"],
                "재고비율": row["재고비율"],
                "신호": ", ".join(row["_재고신호값"]),
                "재고대응판단": ", ".join(row["_재고대응판단값"]),
                "상태": ", ".join(row["_상태값"]),
            }
        )
        output_rows.append(output)

    result = pd.DataFrame(output_rows)
    if result.empty:
        return result
    sort_columns = [col for col in ["제품명", "파워", "생산코드"] if col in result.columns]
    return result.sort_values(sort_columns).reset_index(drop=True)


def build_all_item_customer_excel_sheets(working: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    scoped = working.copy()
    if scoped.empty:
        return []
    scoped["생산코드"] = scoped["생산코드"].map(normalize_to_master_p_code)
    scoped = scoped[scoped["생산코드"].str.startswith("P", na=False)].copy()
    if scoped.empty:
        return []
    customer_options = build_customer_tab_options(scoped)
    sheets: list[tuple[str, pd.DataFrame]] = []
    for customer_group in customer_options:
        customer_scope = scoped[scoped["거래처그룹"] == customer_group].copy()
        sheet_df = build_all_item_customer_excel_sheet(customer_scope)
        if not sheet_df.empty:
            sheets.append((customer_group, sheet_df))
    return sheets


def render_all_item_customer_excel_download(
    working: pd.DataFrame,
    download_stamp: str,
    source_key: str,
    full_working_builder=None,
) -> None:
    prepare_key = "all_item_customer_excel_prepare_v2"
    data_key = f"{prepare_key}_data"
    meta_key = f"{prepare_key}_meta"
    source_state_key = f"{prepare_key}_source"
    if st.session_state.get(source_state_key) != source_key:
        st.session_state.pop(prepare_key, None)
        st.session_state.pop(data_key, None)
        st.session_state.pop(meta_key, None)
        st.session_state[source_state_key] = source_key

    if st.button("거래처별 전체 엑셀 파일 생성", key=f"{prepare_key}_button", width="content"):
        st.session_state[prepare_key] = True
        source_working = full_working_builder() if full_working_builder is not None else working
        sheets = build_all_item_customer_excel_sheets(source_working)
        if sheets:
            sheet_count = len(sheets)
            row_count = sum(len(df) for _, df in sheets)
            st.session_state[data_key] = dataframes_to_excel_bytes(sheets)
            st.session_state[meta_key] = {
                "sheet_count": sheet_count,
                "row_count": row_count,
                "file_name": f"all_item_customer_flow_{download_stamp}.xlsx",
            }
        else:
            st.session_state.pop(data_key, None)
            st.session_state[meta_key] = {"sheet_count": 0, "row_count": 0, "file_name": ""}
    st.caption("엑셀은 화면의 관/제품분류/거래처 선택과 관계없이 전체 P코드 품목 기준으로 생성됩니다.")

    if st.session_state.get(prepare_key, False) and meta_key in st.session_state:
        meta = st.session_state[meta_key]
        if not st.session_state.get(data_key):
            st.info("엑셀로 생성할 거래처별 데이터가 없습니다.")
            return
        st.caption(f"거래처 시트 {meta['sheet_count']:,}개 / 행 {meta['row_count']:,}건")
        st.download_button(
            "거래처별 전체 엑셀 다운로드",
            data=st.session_state[data_key],
            file_name=meta["file_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_all_item_customer_flow_v2",
            width="content",
        )
        return

    if not st.session_state.get(prepare_key, False):
        st.caption("거래처별 전체 엑셀은 수요가 없는 품목의 공정재고까지 확인할 수 있도록 전체 P코드 기준으로 생성합니다.")
        return

    if data_key not in st.session_state:
        st.info("엑셀로 생성할 거래처별 데이터가 없습니다.")
        return


def render_all_item_flow_kpis(scope: pd.DataFrame) -> None:
    stock_scope = scope.drop_duplicates("제품명") if "제품명" in scope.columns else scope
    k1, k2, k3, k4, k5 = st.columns(5, gap="medium")
    with k1:
        render_dashboard_kpi("제품 수", f"{scope['제품명'].nunique():,}", "stock")
    with k2:
        render_dashboard_kpi("오더수량", f"{scope['오더수량'].sum():,.0f}", "stock")
    with k3:
        render_dashboard_kpi("부족수량", f"{scope['부족수량'].sum():,.0f}", "risk")
    with k4:
        render_dashboard_kpi("사출부족수량", f"{scope['사출부족수량'].sum():,.0f}", "risk")
    with k5:
        render_dashboard_kpi("공정재고", f"{stock_scope['공정재고'].sum():,.0f}", "stock")


def build_all_item_flow_power_detail(
    working: pd.DataFrame,
    site_group: str,
    primary_group: str,
    customer_group: str,
    query: str,
    selected_signals: tuple[str, ...],
) -> pd.DataFrame:
    detail = working[working["제품대분류"] == primary_group].copy()
    if site_group != "전체" and "사이트코드" in detail.columns:
        detail = detail[detail["사이트코드"] == site_group].copy()
    if customer_group != "전체":
        detail = detail[detail["거래처그룹"] == customer_group].copy()
    detail = expand_all_item_demand_detail_rows(detail, include_fallback=True)
    detail = filter_rows_with_demand_initial(detail)
    if detail.empty:
        return pd.DataFrame(columns=[col for col in ALL_ITEM_FLOW_POWER_DETAIL_COLUMNS if col in working.columns])
    detail = filter_all_item_flow_query(detail, query).copy()
    if selected_signals:
        signal_mask = detail["신호"].astype(str).map(lambda text: any(signal in text for signal in selected_signals))
        detail = detail[signal_mask].copy()
    if detail.empty:
        return pd.DataFrame(columns=[col for col in ALL_ITEM_FLOW_POWER_DETAIL_COLUMNS if col in working.columns])
    detail = detail.sort_values(
        ["이니셜", "제품명", "파워", "생산부족수량", "사출부족수량", "오더수량", "공정재고합계"],
        ascending=[True, True, True, False, False, False, False],
    )
    columns = [col for col in ALL_ITEM_FLOW_POWER_DETAIL_COLUMNS if col in detail.columns]
    return detail[columns]


def render_all_item_flow_table(
    scope: pd.DataFrame,
    working: pd.DataFrame,
    site_group: str,
    primary_group: str,
    customer_group: str,
    query: str,
    selected_signals: tuple[str, ...],
    download_stamp: str,
    key_suffix: str,
) -> None:
    scoped = filter_all_item_flow_query(scope, query).copy()
    if scoped.empty:
        st.info("표시할 품목이 없습니다.")
        return

    render_all_item_flow_kpis(scoped)
    display = scoped[ALL_ITEM_FLOW_DISPLAY_COLUMNS].copy()
    st.caption(f"제품 {display['제품명'].nunique():,}개 / 이니셜-제품명 {len(display):,}건")

    def build_power_detail() -> pd.DataFrame:
        return build_all_item_flow_power_detail(working, site_group, primary_group, customer_group, query, selected_signals)

    render_lazy_excel_download_builder(
        "파워별 상세 엑셀 다운로드",
        build_power_detail,
        "파워별상세",
        f"all_item_flow_{primary_group}_{key_suffix}_{download_stamp}.xlsx",
        f"download_all_item_flow_{primary_group}_{key_suffix}",
        build_local_cache_hash(
            APP_CACHE_VERSION,
            site_group,
            primary_group,
            customer_group,
            query,
            selected_signals,
            len(scoped),
            scoped["오더수량"].sum(),
            scoped["부족수량"].sum(),
            scoped["사출부족수량"].sum(),
        ),
    )

    display_source, _ = limit_dataframe_for_display(display)
    display_formatted = format_numeric_columns_for_display(display_source)
    column_config = build_auto_column_config(display_formatted, display_formatted.columns.tolist(), source_df=display_source)
    st.dataframe(
        style_operational_table(display_formatted, display_source),
        width="stretch",
        height=700,
        column_order=ALL_ITEM_FLOW_DISPLAY_COLUMNS,
        column_config=column_config,
        hide_index=True,
        key=f"all_item_flow_table_{primary_group}_{key_suffix}",
    )


def render_all_item_alert_panel(working: pd.DataFrame) -> None:
    no_demand_stock = working[(working["총수요수량"] <= 0) & (working["공정재고합계"] > 0)].copy()
    excess_stock = working[(working["총수요수량"] > 0) & (working["초과재고수량"] > 0)].copy()
    code_mismatch = working[working["코드매칭상태"] == "코드미매칭"].copy()

    st.markdown("#### 주의 필요")
    st.caption("현재 수요가 연결되지 않았거나 수요보다 재고가 많은 공정재고입니다. 생산/이동 전 확인이 필요합니다.")
    a1, a2, a3, a4 = st.columns(4, gap="medium")
    with a1:
        render_dashboard_kpi("수요 없는 공정재고 품목", f"{len(no_demand_stock):,}", "risk")
    with a2:
        render_dashboard_kpi("수요 없는 공정재고 수량", f"{no_demand_stock['공정재고합계'].sum():,.0f}", "risk")
    with a3:
        render_dashboard_kpi("수요 대비 초과재고", f"{excess_stock['초과재고수량'].sum():,.0f}", "risk")
    with a4:
        render_dashboard_kpi("코드 확인 필요", f"{len(code_mismatch):,}", "risk")

    status_key = "all_item_status_filter_v2"
    b1, b2, b3, b4 = st.columns([1.35, 1.25, 1.0, 1.0])
    with b1:
        if st.button("수요 없는 공정재고만 보기", key="all_item_quick_no_demand_stock_v1", width="stretch"):
            st.session_state[status_key] = "수요 없음 + 공정재고 있음"
    with b2:
        if st.button("수요 대비 초과재고만 보기", key="all_item_quick_excess_stock_v1", width="stretch"):
            st.session_state[status_key] = "수요 대비 재고 초과"
    with b3:
        if st.button("주의 필요만 보기", key="all_item_quick_attention_v1", width="stretch"):
            st.session_state[status_key] = "주의 필요"
    with b4:
        if st.button("전체 보기", key="all_item_quick_all_v1", width="stretch"):
            st.session_state[status_key] = "전체"

    top_columns = [
        "판단",
        "신규분류",
        "제품명",
        "파워",
        "생산코드",
        "사출코드",
        "분리코드",
        "공정재고합계",
        "사출창고",
        "분리창고",
        "검사접착창고",
        "누수규격검사",
        "샘플 신청가능수량",
    ]
    top_columns = [col for col in top_columns if col in no_demand_stock.columns]
    top_no_demand_stock = no_demand_stock.sort_values(
        ["공정재고합계", "샘플 신청가능수량", "신규분류", "생산코드"],
        ascending=[False, False, True, True],
    ).head(50)
    with st.expander("수요 없는 공정재고 TOP 50", expanded=True):
        if top_no_demand_stock.empty:
            st.success("수요 없는 공정재고가 없습니다.")
        else:
            top_view = top_no_demand_stock[top_columns].copy()
            top_display = format_numeric_columns_for_display(top_view)
            top_config = build_auto_column_config(top_display, top_display.columns.tolist(), source_df=top_view)
            st.dataframe(
                style_operational_table(top_display, top_view),
                width="stretch",
                height=300,
                column_config=top_config,
                hide_index=True,
                key="all_item_no_demand_stock_top50_v1",
            )


def render_all_item_site_customer_summary(view_flow: pd.DataFrame, selected_site_group: str) -> None:
    if view_flow.empty or "거래처그룹" not in view_flow.columns:
        return

    initial_col = "이니셜"
    product_col = "제품명"
    order_col = "오더수량"
    shortage_col = "부족수량"
    injection_col = "사출부족수량"
    stock_col = "공정재고"
    finished_stock_col = "완제품재고"
    stock_change_col = "재고변화"
    doi_col = "DOI"
    required_columns = [
        "제품대분류",
        "거래처그룹",
        initial_col,
        product_col,
        order_col,
        shortage_col,
        injection_col,
        stock_col,
    ]
    if any(col not in view_flow.columns for col in required_columns):
        return

    summary_source = view_flow.copy()
    optional_numeric_columns = [finished_stock_col, stock_change_col, doi_col]
    for qty_col in [order_col, shortage_col, injection_col, stock_col, *optional_numeric_columns]:
        if qty_col in summary_source.columns:
            summary_source[qty_col] = parse_mixed_numeric(summary_source[qty_col]).fillna(0)

    agg_map = {
        "제품분류": ("제품대분류", summarize_unique),
        "제품수": (product_col, "nunique"),
        "이니셜수": (initial_col, "nunique"),
        "오더수량": (order_col, "sum"),
        "부족수량": (shortage_col, "sum"),
        "사출부족수량": (injection_col, "sum"),
        "공정재고": (stock_col, "sum"),
    }
    if finished_stock_col in summary_source.columns:
        agg_map["완제품재고"] = (finished_stock_col, "sum")
    if stock_change_col in summary_source.columns:
        agg_map["재고변화"] = (stock_change_col, "sum")
    if doi_col in summary_source.columns:
        agg_map["최대DOI"] = (doi_col, "max")

    summary = (
        summary_source.groupby("거래처그룹", as_index=False)
        .agg(**agg_map)
        .sort_values(["오더수량", "부족수량", "사출부족수량"], ascending=[False, False, False])
        .reset_index(drop=True)
    )
    if summary.empty:
        return

    with st.expander(f"{selected_site_group} 전체 거래처 요약", expanded=True):
        st.caption("제품분류 선택과 관계없이 현재 관에 반영된 전체 거래처입니다.")
        customer_names = [clean_text_value(value) for value in summary["거래처그룹"].tolist()]
        customer_names = [value for value in customer_names if value]
        if customer_names:
            st.caption(f"전체 거래처: {', '.join(customer_names)}")
        summary_display = format_numeric_columns_for_display(summary)
        summary_column_config = build_auto_column_config(
            summary_display,
            summary_display.columns.tolist(),
            source_df=summary,
        )
        st.dataframe(
            style_operational_table(summary_display, summary),
            width="stretch",
            height=min(460, max(180, 48 + len(summary) * 35)),
            column_config=summary_column_config,
            hide_index=True,
            key=f"all_item_site_customer_summary_{selected_site_group}",
        )


def render_all_items_dashboard(
    all_items_df: pd.DataFrame,
    updated_at: str,
    full_all_items_builder=None,
    selected_site_group: str = "전체",
) -> None:
    st.subheader("전체 품목 현황")
    st.caption(f"업데이트: {updated_at}")
    st.caption("화면은 선택한 관/제품분류/거래처 안에서 수요가 있는 이니셜-제품명만 표시하고, 전체 품목 상세는 엑셀에서 확인합니다.")

    if all_items_df.empty:
        st.warning("전체 품목 현황을 계산할 데이터가 없습니다. 전체 품목리스트 파일을 확인해주세요.")
        return

    download_stamp = datetime.now(DISPLAY_TZ).strftime("%Y%m%d_%H%M%S")
    working = prepare_all_item_flow_data(all_items_df)
    product_flow = build_all_item_product_flow_summary(working)

    if product_flow.empty:
        st.info("표시할 품목 흐름 데이터가 없습니다.")
        return

    query = st.text_input(
        "통합 검색",
        value="",
        key="all_item_flow_query_v1",
        placeholder="이니셜, 제품명, 거래처, 생산코드 검색",
        help="콤마(,)로 여러 키워드를 입력하면 OR 조건으로 검색합니다.",
    ).strip()
    selected_signals: tuple[str, ...] = tuple()
    selected_site_group = clean_text_value(selected_site_group) or "전체"

    def build_full_working_for_excel() -> pd.DataFrame:
        if full_all_items_builder is None:
            return working
        return prepare_all_item_flow_data(full_all_items_builder())

    render_all_item_customer_excel_download(
        working,
        download_stamp,
        f"{APP_CACHE_VERSION}:{updated_at}:{selected_site_group}",
        build_full_working_for_excel if full_all_items_builder is not None else None,
    )

    view_flow = product_flow.copy()
    if selected_site_group != "전체":
        view_flow = view_flow[view_flow["사이트코드"] == selected_site_group].copy()
    if view_flow.empty:
        st.info("선택한 관에 표시할 품목이 없습니다.")
        return
    st.caption(f"관 필터: {selected_site_group}")
    render_all_item_site_customer_summary(view_flow, selected_site_group)

    customer_options = build_customer_tab_options(view_flow)
    if not customer_options:
        st.info("표시할 거래처가 없습니다.")
        return

    customer_group = st.pills(
        "거래처",
        options=customer_options,
        default=customer_options[0],
        key=f"all_item_flow_customer_group_site_{selected_site_group}_v3",
    )
    if customer_group is None:
        st.info("거래처를 선택해 주세요.")
        return

    customer_scope = view_flow[view_flow["거래처그룹"] == customer_group].copy()
    if customer_scope.empty:
        st.info("표시할 품목이 없습니다.")
        return

    key_token = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", customer_group).strip("_") or "customer"
    primary_order = ["1-DAY", "FRP"]
    existing_primary = [group for group in primary_order if (customer_scope["제품대분류"] == group).any()]
    extra_primary = sorted(set(customer_scope["제품대분류"].dropna().tolist()) - set(existing_primary) - {"기타"})
    if (customer_scope["제품대분류"] == "기타").any():
        extra_primary.append("기타")
    primary_tabs = existing_primary + extra_primary
    if not primary_tabs:
        st.info("선택한 거래처에 표시할 1-DAY/FRP/기타 품목이 없습니다.")
        return

    primary_group = st.pills(
        "제품 분류 (선택 거래처 기준)",
        options=primary_tabs,
        default=primary_tabs[0],
        key=f"all_item_flow_primary_group_{selected_site_group}_{key_token}_v3",
    )
    if primary_group is None:
        st.info("제품 분류를 선택해 주세요.")
        return

    scoped = customer_scope[customer_scope["제품대분류"] == primary_group].copy()
    st.caption(f"현재 화면 필터: {selected_site_group} / {primary_group} / {customer_group}")
    st.caption("거래처 탭은 현재 관 전체 기준이고, 제품분류는 선택한 거래처 안에 수요가 있는 분류만 표시합니다.")
    render_all_item_flow_table(
        scoped,
        working,
        selected_site_group,
        primary_group,
        customer_group,
        query,
        tuple(selected_signals),
        download_stamp,
        key_token,
    )


def render_inventory_risk_dashboard(risk_df: pd.DataFrame, updated_at: str) -> None:
    st.subheader("공정재고 리스크")
    st.caption(f"업데이트: {updated_at}")
    st.caption("ODV_WIP 원장 중 사출창고 R코드와 분리창고 Q코드만 기준으로 현재 수요코드 직접매칭, 동일제품 타도수, 현재수요 제품군 없음, 수요초과 재고를 분리합니다.")

    if risk_df.empty:
        st.warning("공정재고 리스크를 계산할 데이터가 없습니다.")
        return

    download_stamp = datetime.now(DISPLAY_TZ).strftime("%Y%m%d_%H%M%S")
    working = risk_df.copy()

    risk_options = ["전체", "현재수요 제품군 없음", "동일제품 타도수 재고", "수요초과 재고", "수요코드 직접매칭"]
    selected_risk = st.segmented_control(
        "리스크 구분",
        options=risk_options,
        default="전체",
        key="inventory_risk_selector_v1",
        width="stretch",
    )

    filter_col, search_col = st.columns([1.35, 2.65])
    with filter_col:
        warehouse_options = sorted(working["창고"].astype(str).dropna().unique().tolist())
        selected_warehouses = st.multiselect(
            "공정창고",
            options=warehouse_options,
            default=warehouse_options,
            key="inventory_risk_warehouse_filter_v1",
        )
    with search_col:
        direct_query = st.text_input(
            "직접 검색",
            value="",
            key="inventory_risk_direct_query_v1",
            placeholder="품목코드, 제품군, 제품명, LOT, 재공코드 검색",
            help="콤마(,)로 여러 키워드를 입력하면 OR 조건으로 검색합니다.",
        ).strip()

    if selected_risk != "전체":
        working = working[working["리스크구분"] == selected_risk]
    if selected_warehouses:
        working = working[working["창고"].isin(selected_warehouses)]
    if direct_query:
        working = filter_display_table_with_query(working, direct_query).copy()

    total_stock = parse_mixed_numeric(working["재고수량"]).sum()
    no_demand_stock = parse_mixed_numeric(
        working.loc[working["리스크구분"].isin(["현재수요 제품군 없음", "동일제품 타도수 재고"]), "재고수량"]
    ).sum()
    same_family_stock = parse_mixed_numeric(
        working.loc[working["리스크구분"] == "동일제품 타도수 재고", "재고수량"]
    ).sum()
    no_family_stock = parse_mixed_numeric(
        working.loc[working["리스크구분"] == "현재수요 제품군 없음", "재고수량"]
    ).sum()
    excess_stock = parse_mixed_numeric(
        working.loc[working["리스크구분"] == "수요초과 재고", "초과수량"]
    ).sum()

    c1, c2, c3, c4, c5 = st.columns(5, gap="medium")
    with c1:
        render_dashboard_kpi("R/Q 공정재고", f"{total_stock:,.0f}", "stock")
    with c2:
        render_dashboard_kpi("R/Q 수요외 재고", f"{no_demand_stock:,.0f}", "risk")
    with c3:
        render_dashboard_kpi("동일제품 타도수", f"{same_family_stock:,.0f}", "risk")
    with c4:
        render_dashboard_kpi("현재수요 없음", f"{no_family_stock:,.0f}", "risk")
    with c5:
        render_dashboard_kpi("수요초과", f"{excess_stock:,.0f}", "risk")

    summary = (
        working.groupby(["리스크구분", "창고"], as_index=False)
        .agg(
            {
                "품목코드": "nunique",
                "재고수량": "sum",
                "현재수요수량": "sum",
                "초과수량": "sum",
            }
        )
        .rename(columns={"품목코드": "품목수"})
        .sort_values(["리스크구분", "재고수량"], ascending=[True, False])
    )
    with st.expander("리스크 구분/공정창고 요약", expanded=False):
        if summary.empty:
            st.info("요약할 데이터가 없습니다.")
        else:
            summary_display = format_numeric_columns_for_display(summary)
            summary_column_config = build_auto_column_config(
                summary_display, summary_display.columns.tolist(), source_df=summary
            )
            st.dataframe(
                style_operational_table(summary_display, summary),
                width="stretch",
                height=320,
                column_config=summary_column_config,
                hide_index=True,
            )

    detail_columns = [
        "리스크구분",
        "창고",
        "품목코드",
        "제품군키",
        "파워",
        "재고수량",
        "현재수요수량",
        "초과수량",
        "수요코드수",
        "제품명 예시",
        "이니셜 예시",
        "납기일",
        "분류별요약",
        "시트분류",
        "재공코드 예시",
        "LOT 예시",
        "사용가능일",
        "생성일시",
    ]
    detail_columns = [col for col in detail_columns if col in working.columns]
    detail = working.sort_values(["정렬순위", "재고수량", "품목코드"], ascending=[True, False, True])[detail_columns]

    st.caption(f"표시 {len(detail):,}건 / 전체 {len(risk_df):,}건")
    render_lazy_excel_download_button(
        "엑셀 다운로드",
        detail,
        "공정재고리스크",
        f"inventory_risk_{download_stamp}.xlsx",
        "download_inventory_risk_v1",
    )

    detail_display_source, _ = limit_dataframe_for_display(detail)
    detail_display = format_numeric_columns_for_display(detail_display_source)
    detail_column_config = build_auto_column_config(
        detail_display, detail_display.columns.tolist(), source_df=detail_display_source
    )
    st.dataframe(
        style_operational_table(detail_display, detail_display_source),
        width="stretch",
        height=720,
        column_config=detail_column_config,
        hide_index=True,
        key="inventory_risk_table_v1",
    )


def render_shortage_dashboard(
    df: pd.DataFrame,
    updated_at: str,
    file_info_df: pd.DataFrame | None = None,
    data_base_dir: Path | None = None,
    source_label: str = "",
) -> None:
    enriched_df = add_rq_group_columns(df)
    filtered = apply_filters(enriched_df, updated_at, data_base_dir, source_label)
    download_stamp = datetime.now(DISPLAY_TZ).strftime("%Y%m%d_%H%M%S")

    detail_columns = [
        "거래처",
        "이니셜",
        "품목코드",
        "R코드",
        "Q코드",
        "제품명",
        "파워",
        "납기일",
        "부족수량",
        "사출창고",
        "분리창고",
        "검사접착창고",
        "검사접착재작업창고",
        "누수규격검사 창고",
        "공정재고 합계",
        "비고",
    ]

    shortage_views = ["생산 현황", "사출 현황", "공용 품목 현황"]
    selected_shortage_view = st.segmented_control(
        "공정별 현황",
        options=shortage_views,
        default=shortage_views[0],
        key="shortage_view_selector_v4",
        width="stretch",
    )
    search_col, result_col = st.columns([3.2, 1.0])
    with search_col:
        direct_query = st.text_input(
            "직접 검색",
            value="",
            key="shortage_direct_query_v2",
            placeholder="합계와 표에 적용할 검색어",
            help="현재 필터 범위 안에서 합계와 표에 함께 적용됩니다. 콤마(,)로 여러 키워드를 입력하면 OR 조건으로 검색합니다.",
        ).strip()
    with result_col:
        result_caption = st.empty()
    if direct_query:
        filtered = filter_display_table_with_query(filtered, direct_query).copy()
    link_mapping_scope = enriched_df.copy()

    render_rework_match_debug(file_info_df)
    if "재작업" in filtered.columns and "품목코드" in filtered.columns:
        rework_text = filtered["재작업"].astype(str).str.strip()
        rework_scope = filtered[rework_text.ne("") & ~rework_text.str.lower().isin(INVALID_CATEGORY_VALUES)]
        st.caption(
            f"현재 필터 범위 재작업: {len(rework_scope):,}행 / "
            f"{rework_scope['품목코드'].astype(str).str.strip().str.upper().nunique():,}개 품목코드"
        )

    if selected_shortage_view == "생산 현황":
        full_demand_summary = build_summary_group_totals_with_safe_split(filtered)
        with st.expander("분류별요약 기준 부족수량 요약", expanded=False):
            st.caption(
                "오더 기준 = 이니셜에 안전 미포함, 안전재고 기준 = 이니셜에 안전 포함, "
                "사출부족수량 = 사출생산필요수량 합계"
            )
            if full_demand_summary.empty:
                st.info("분류별요약 기준 부족수량 요약을 계산할 데이터가 없습니다.")
            else:
                full_demand_summary_display = format_numeric_columns_for_display(full_demand_summary)
                full_demand_summary_column_config = build_auto_column_config(
                    full_demand_summary_display,
                    full_demand_summary_display.columns.tolist(),
                    source_df=full_demand_summary,
                )
                st.dataframe(
                    full_demand_summary_display,
                    width="stretch",
                    height=320,
                    column_config=full_demand_summary_column_config,
                    hide_index=True,
                )

        c1, c2, c3 = st.columns(3, gap="medium")
        with c1:
            shortage_kpi_slot = st.empty()
        with c2:
            injection_kpi_slot = st.empty()
        with c3:
            process_stock_kpi_slot = st.empty()

        c1, c2, c3, c4, c5 = st.columns(5, gap="medium")
        with c1:
            injection_stock_kpi_slot = st.empty()
        with c2:
            separation_stock_kpi_slot = st.empty()
        with c3:
            adhesion_stock_kpi_slot = st.empty()
        with c4:
            adhesion_rework_stock_kpi_slot = st.empty()
        with c5:
            leakage_stock_kpi_slot = st.empty()

        initial_inj_summary = build_initial_injection_summary(filtered)
        with st.expander("이니셜별 사출부족수량 요약", expanded=False):
            st.caption("사출 부족수량 = 이니셜별(품목코드 단위) 사출 생산 필요수량 합계")
            if initial_inj_summary.empty:
                st.info("이니셜별 사출부족수량 요약을 계산할 데이터가 없습니다.")
            else:
                initial_inj_summary_display = format_numeric_columns_for_display(initial_inj_summary)
                initial_inj_summary_column_config = build_auto_column_config(
                    initial_inj_summary_display,
                    initial_inj_summary_display.columns.tolist(),
                    source_df=initial_inj_summary,
                )
                st.dataframe(
                    initial_inj_summary_display,
                    width="stretch",
                    height=320,
                    column_config=initial_inj_summary_column_config,
                    hide_index=True,
                )

        p_view = filtered.copy()
        p_view["부족수량"] = parse_mixed_numeric(p_view["부족수량"])
        if "사출생산필요수량" in p_view.columns:
            p_view["사출생산필요수량"] = parse_mixed_numeric(p_view["사출생산필요수량"])
        else:
            p_view["사출생산필요수량"] = 0

        mapped_inj_total = 0.0
        unmatched_inj_total = 0.0
        mapped_sep_total = 0.0
        unmatched_sep_total = 0.0
        if "품목코드" in p_view.columns:
            item_prefix = p_view["품목코드"].astype(str).str.upper().str[:1]
            p_rows = p_view[item_prefix == "P"].copy()
            r_rows = p_view[item_prefix == "R"].copy()
        else:
            p_rows = p_view.copy()
            r_rows = p_view.iloc[0:0].copy()

        synthetic_full_rows = build_synthetic_p_rows_for_process_scope(
            link_mapping_scope,
            link_mapping_scope,
            p_view.columns.tolist(),
        )
        synthetic_display_rows = build_synthetic_p_rows_for_process_scope(
            p_view,
            link_mapping_scope,
            p_view.columns.tolist(),
        )
        if not synthetic_display_rows.empty:
            p_rows = pd.concat([p_rows, synthetic_display_rows], ignore_index=True, sort=False)

        if p_rows.empty:
            p_view["사출 부족수량"] = p_view["사출생산필요수량"]
            key_cols = [
                c
                for c in [ORDER_NO_COL, "거래처", "이니셜", "R코드", "Q코드"]
                if c in p_view.columns
            ]
            if key_cols:
                p_view = normalize_flow_link_key_columns(p_view, key_cols)
            if key_cols and not r_rows.empty and "품목코드" in enriched_df.columns:
                # Fallback: when current filters leave only R rows, recover representative P codes
                # from the full scope using (사이트코드+이니셜+R코드+Q코드) keys.
                p_universe = enriched_df.copy()
                universe_prefix = p_universe["품목코드"].astype(str).str.upper().str[:1]
                p_universe = p_universe[universe_prefix == "P"]
                if not p_universe.empty and all(c in p_universe.columns for c in key_cols):
                    p_universe = normalize_flow_link_key_columns(p_universe, key_cols)
                    if "부족수량" in p_universe.columns:
                        p_universe["부족수량_num"] = parse_mixed_numeric(p_universe["부족수량"])
                    else:
                        p_universe["부족수량_num"] = 0
                    if "제품명" not in p_universe.columns:
                        p_universe["제품명"] = "-"

                    p_lookup_columns = [*key_cols, "품목코드", "제품명", "부족수량_num"]
                    if "납기일" in p_universe.columns:
                        p_lookup_columns.append("납기일")
                    p_key_map = (
                        p_universe.sort_values(["부족수량_num", "품목코드"], ascending=[False, True])
                        .drop_duplicates(subset=key_cols, keep="first")[p_lookup_columns]
                        .rename(
                            columns={
                                "품목코드": "매핑P코드",
                                "제품명": "매핑제품명",
                                "부족수량_num": "매핑P부족수량",
                                "납기일": "매핑P납기일",
                            }
                        )
                    )
                    p_view = p_view.merge(p_key_map, on=key_cols, how="left")
                    mapped_mask = p_view["매핑P코드"].astype(str).str.strip().str.lower().ne("nan")
                    mapped_mask = mapped_mask & p_view["매핑P코드"].astype(str).str.strip().ne("")
                    p_view.loc[mapped_mask, "품목코드"] = p_view.loc[mapped_mask, "매핑P코드"]
                    if "제품명" in p_view.columns:
                        p_view.loc[mapped_mask, "제품명"] = p_view.loc[mapped_mask, "매핑제품명"]
                    if "부족수량" in p_view.columns and "매핑P부족수량" in p_view.columns:
                        current_shortage = parse_mixed_numeric(p_view["부족수량"])
                        mapped_shortage = parse_mixed_numeric(p_view["매핑P부족수량"])
                        p_view["부족수량"] = current_shortage.where(current_shortage > 0, mapped_shortage).fillna(0)
                    if "납기일" in p_view.columns and "매핑P납기일" in p_view.columns:
                        due_text = p_view["납기일"].astype(str).str.strip()
                        mapped_due_text = p_view["매핑P납기일"].astype(str).str.strip()
                        due_missing = due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
                        mapped_due_valid = ~mapped_due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
                        p_view.loc[due_missing & mapped_due_valid, "납기일"] = mapped_due_text[
                            due_missing & mapped_due_valid
                        ]
                    p_view = p_view.drop(
                        columns=["매핑P코드", "매핑제품명", "매핑P부족수량", "매핑P납기일"],
                        errors="ignore",
                    )
        else:
            if SEPARATION_REQUIRED_QTY_COL not in p_rows.columns:
                p_rows[SEPARATION_REQUIRED_QTY_COL] = 0
            p_rows[SEPARATION_REQUIRED_QTY_COL] = parse_mixed_numeric(p_rows[SEPARATION_REQUIRED_QTY_COL])
            if ADHESION_REQUIRED_QTY_COL not in p_rows.columns:
                p_rows[ADHESION_REQUIRED_QTY_COL] = 0
            p_rows[ADHESION_REQUIRED_QTY_COL] = parse_mixed_numeric(p_rows[ADHESION_REQUIRED_QTY_COL])
            p_rows["사출 부족수량"] = p_rows["사출생산필요수량"]
            r_scope = link_mapping_scope.copy()
            if "품목코드" in r_scope.columns:
                r_scope = r_scope[r_scope["품목코드"].astype(str).str.upper().str.startswith("R")]
            if "사출생산필요수량" in r_scope.columns:
                r_scope["사출생산필요수량"] = parse_mixed_numeric(r_scope["사출생산필요수량"])
                r_scope = r_scope[r_scope["사출생산필요수량"] > 0]
            else:
                r_scope = r_scope.iloc[0:0]
            key_cols = [
                c
                for c in [ORDER_NO_COL, "거래처", "이니셜", "R코드", "Q코드"]
                if c in p_rows.columns and c in r_scope.columns
            ]

            if key_cols and not r_scope.empty:
                p_rows = normalize_flow_link_key_columns(p_rows, key_cols)
                r_scope = normalize_flow_link_key_columns(r_scope, key_cols)
                r_key_inj_all = (
                    r_scope.groupby(key_cols, as_index=False, dropna=False)["사출생산필요수량"]
                    .sum()
                    .rename(columns={"사출생산필요수량": "연결R 사출수량"})
                )
                p_keys = p_rows[key_cols].drop_duplicates()
                full_p_keys = pd.DataFrame(columns=key_cols)
                if "품목코드" in link_mapping_scope.columns and all(c in link_mapping_scope.columns for c in key_cols):
                    full_p_scope = link_mapping_scope[
                        link_mapping_scope["품목코드"].astype(str).str.upper().str.startswith("P")
                    ]
                    if not synthetic_full_rows.empty:
                        full_p_scope = pd.concat([full_p_scope, synthetic_full_rows], ignore_index=True, sort=False)
                    full_p_scope = normalize_flow_link_key_columns(full_p_scope, key_cols)
                    full_p_keys = full_p_scope[key_cols].drop_duplicates()
                if not full_p_keys.empty:
                    unmatched_r = r_key_inj_all.merge(full_p_keys, on=key_cols, how="left", indicator=True)
                    unmatched_inj_total = float(
                        unmatched_r.loc[unmatched_r["_merge"] == "left_only", "연결R 사출수량"].sum()
                    )
                r_key_inj = (
                    r_key_inj_all.merge(p_keys, on=key_cols, how="inner")
                    if not p_keys.empty
                    else r_key_inj_all.iloc[0:0].copy()
                )

                p_rows = p_rows.merge(r_key_inj, on=key_cols, how="left")
                p_key_short_sum = p_rows.groupby(key_cols, dropna=False)["부족수량"].transform("sum")
                p_key_count = p_rows.groupby(key_cols, dropna=False)["품목코드"].transform("count")

                mapped_by_short = (
                    parse_mixed_numeric(p_rows["연결R 사출수량"])
                    * p_rows["부족수량"]
                    / p_key_short_sum.replace(0, pd.NA)
                )
                mapped_by_split = (
                    parse_mixed_numeric(p_rows["연결R 사출수량"])
                    / p_key_count.replace(0, pd.NA)
                )
                p_rows["사출 부족수량(연결R)"] = mapped_by_short.where(p_key_short_sum > 0, mapped_by_split).fillna(0)
                p_rows["사출 부족수량"] = p_rows["사출 부족수량"].where(
                    p_rows["사출 부족수량"] > 0, p_rows["사출 부족수량(연결R)"]
                )
                mapped_inj_total = float(p_rows["사출 부족수량(연결R)"].sum())
            else:
                unmatched_inj_total = float(parse_mixed_numeric(r_scope["사출생산필요수량"]).sum())

            q_link_cols = [
                c
                for c in [ORDER_NO_COL, "거래처", "이니셜", "Q코드"]
                if c in p_rows.columns and c in link_mapping_scope.columns
            ]
            if q_link_cols and SEPARATION_REQUIRED_QTY_COL in link_mapping_scope.columns:
                q_scope = link_mapping_scope.copy()
                q_scope[SEPARATION_REQUIRED_QTY_COL] = parse_mixed_numeric(q_scope[SEPARATION_REQUIRED_QTY_COL])
                if "품목코드" in q_scope.columns:
                    q_scope = q_scope[q_scope["품목코드"].astype(str).str.upper().str.startswith("Q")]
                q_scope = q_scope[q_scope[SEPARATION_REQUIRED_QTY_COL] > 0]
                if not q_scope.empty:
                    p_rows = normalize_flow_link_key_columns(p_rows, q_link_cols)
                    q_scope = normalize_flow_link_key_columns(q_scope, q_link_cols)
                    q_key_sep_all = (
                        q_scope.groupby(q_link_cols, as_index=False, dropna=False)[SEPARATION_REQUIRED_QTY_COL]
                        .sum()
                        .rename(columns={SEPARATION_REQUIRED_QTY_COL: "연결Q 분리수량"})
                    )
                    q_keys = p_rows[q_link_cols].drop_duplicates()
                    full_p_q_keys = pd.DataFrame(columns=q_link_cols)
                    if "품목코드" in link_mapping_scope.columns and all(c in link_mapping_scope.columns for c in q_link_cols):
                        full_p_scope = link_mapping_scope[
                            link_mapping_scope["품목코드"].astype(str).str.upper().str.startswith("P")
                        ]
                        if not synthetic_full_rows.empty:
                            full_p_scope = pd.concat([full_p_scope, synthetic_full_rows], ignore_index=True, sort=False)
                        full_p_scope = normalize_flow_link_key_columns(full_p_scope, q_link_cols)
                        full_p_q_keys = full_p_scope[q_link_cols].drop_duplicates()
                    if not full_p_q_keys.empty:
                        unmatched_q = q_key_sep_all.merge(full_p_q_keys, on=q_link_cols, how="left", indicator=True)
                        unmatched_sep_total = float(
                            unmatched_q.loc[unmatched_q["_merge"] == "left_only", "연결Q 분리수량"].sum()
                        )
                    q_key_sep = (
                        q_key_sep_all.merge(q_keys, on=q_link_cols, how="inner")
                        if not q_keys.empty
                        else q_key_sep_all.iloc[0:0].copy()
                    )
                    mapped_sep_total = float(q_key_sep["연결Q 분리수량"].sum())
                    p_rows = p_rows.merge(q_key_sep, on=q_link_cols, how="left")
                    linked_sep = parse_mixed_numeric(p_rows["연결Q 분리수량"])
                    current_sep = parse_mixed_numeric(p_rows[SEPARATION_REQUIRED_QTY_COL])
                    p_rows[SEPARATION_REQUIRED_QTY_COL] = current_sep.where(current_sep > 0, linked_sep).fillna(0)
                    p_rows = p_rows.drop(columns=["연결Q 분리수량"], errors="ignore")

            p_view = p_rows.copy()

        if SEPARATION_REQUIRED_QTY_COL not in p_view.columns:
            p_view[SEPARATION_REQUIRED_QTY_COL] = 0
        p_view[SEPARATION_REQUIRED_QTY_COL] = parse_mixed_numeric(p_view[SEPARATION_REQUIRED_QTY_COL])
        if LEADJI_REQUIRED_QTY_COL not in p_view.columns:
            p_view[LEADJI_REQUIRED_QTY_COL] = 0
        p_view[LEADJI_REQUIRED_QTY_COL] = parse_mixed_numeric(p_view[LEADJI_REQUIRED_QTY_COL])
        if ADHESION_REQUIRED_QTY_COL not in p_view.columns:
            p_view[ADHESION_REQUIRED_QTY_COL] = 0
        p_view[ADHESION_REQUIRED_QTY_COL] = parse_mixed_numeric(p_view[ADHESION_REQUIRED_QTY_COL])
        if DEMAND_QTY_COL not in p_view.columns:
            p_view[DEMAND_QTY_COL] = 0
        p_view[DEMAND_QTY_COL] = parse_mixed_numeric(p_view[DEMAND_QTY_COL])
        p_view = p_view[
            (p_view["부족수량"] > 0)
            | (p_view["사출 부족수량"] > 0)
            | (p_view[SEPARATION_REQUIRED_QTY_COL] > 0)
            | (p_view[LEADJI_REQUIRED_QTY_COL] > 0)
            | (p_view[ADHESION_REQUIRED_QTY_COL] > 0)
            | (p_view[DEMAND_QTY_COL] > 0)
        ]
        p_view["표시부족수량"] = (
            p_view["부족수량"]
            + p_view["사출 부족수량"]
            + p_view[SEPARATION_REQUIRED_QTY_COL]
            + p_view[LEADJI_REQUIRED_QTY_COL]
            + p_view[ADHESION_REQUIRED_QTY_COL]
        )
        stock_total = (
            parse_mixed_numeric(p_view["공정재고 합계"])
            if "공정재고 합계" in p_view.columns
            else pd.Series(0.0, index=p_view.index)
        )
        injection_shortage = parse_mixed_numeric(p_view["사출 부족수량"])
        final_shortage = parse_mixed_numeric(p_view["부족수량"])
        p_view["확인구분"] = ""
        p_view.loc[(final_shortage > 0) & (stock_total > 0), "확인구분"] = "공정재고 확인"
        p_view.loc[injection_shortage > 0, "확인구분"] = "사출필요"
        p_view.loc[
            (final_shortage > 0) & (injection_shortage <= 0) & (stock_total <= 0),
            "확인구분",
        ] = "최종부족/재고없음"
        if "재작업" in p_view.columns:
            rework_text = p_view["재작업"].astype(str).str.strip()
            rework_available = rework_text.ne("") & ~rework_text.str.lower().isin(INVALID_CATEGORY_VALUES)
            p_view.loc[rework_available & (p_view["확인구분"].astype(str).str.strip() == ""), "확인구분"] = "재작업가능"
        if "납기일" not in p_view.columns:
            p_view["납기일"] = "-"
        if "사출납기일" in p_view.columns:
            due_text = p_view["납기일"].astype(str).str.strip()
            inj_due_text = p_view["사출납기일"].astype(str).str.strip()
            due_missing = due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
            inj_due_valid = ~inj_due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
            p_view.loc[due_missing & inj_due_valid, "납기일"] = inj_due_text[due_missing & inj_due_valid]
        for fallback_due_col in [ADHESION_REQUIRED_DUE_COL, LEADJI_REQUIRED_DUE_COL, SEPARATION_REQUIRED_DUE_COL]:
            if fallback_due_col not in p_view.columns:
                continue
            due_text = p_view["납기일"].astype(str).str.strip()
            fallback_due_text = p_view[fallback_due_col].astype(str).str.strip()
            due_missing = due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
            fallback_due_valid = ~fallback_due_text.str.lower().isin({"", "-", "nan", "nat", "none"})
            p_view.loc[due_missing & fallback_due_valid, "납기일"] = fallback_due_text[
                due_missing & fallback_due_valid
            ]

        p_detail_columns = detail_columns.copy()
        if "사출 부족수량" not in p_detail_columns:
            insert_idx = p_detail_columns.index("부족수량") + 1 if "부족수량" in p_detail_columns else len(p_detail_columns)
            p_detail_columns.insert(insert_idx, "사출 부족수량")
        p_detail_columns = move_columns_to_end(p_detail_columns, ["비고"])
        p_table = p_view.sort_values(
            ["표시부족수량", "부족수량", "사출 부족수량", "이니셜", "거래처"],
            ascending=[False, False, False, True, True],
        )[p_detail_columns]
        p_table_ui = p_table.drop(columns=["상태"], errors="ignore")
        kpi_source = p_table_ui.copy()
        kpi_totals = {
            "부족수량": parse_mixed_numeric(kpi_source["부족수량"]).sum() if "부족수량" in kpi_source.columns else 0,
            "사출 부족수량": (
                parse_mixed_numeric(kpi_source["사출 부족수량"]).sum()
                if "사출 부족수량" in kpi_source.columns
                else 0
            ),
            "공정재고 합계": (
                parse_mixed_numeric(kpi_source["공정재고 합계"]).sum()
                if "공정재고 합계" in kpi_source.columns
                else 0
            ),
            "사출창고": parse_mixed_numeric(kpi_source["사출창고"]).sum() if "사출창고" in kpi_source.columns else 0,
            "분리창고": parse_mixed_numeric(kpi_source["분리창고"]).sum() if "분리창고" in kpi_source.columns else 0,
            "검사접착창고": (
                parse_mixed_numeric(kpi_source["검사접착창고"]).sum()
                if "검사접착창고" in kpi_source.columns
                else 0
            ),
            "검사접착재작업창고": (
                parse_mixed_numeric(kpi_source["검사접착재작업창고"]).sum()
                if "검사접착재작업창고" in kpi_source.columns
                else 0
            ),
            "누수규격검사 창고": (
                parse_mixed_numeric(kpi_source["누수규격검사 창고"]).sum()
                if "누수규격검사 창고" in kpi_source.columns
                else 0
            ),
        }
        with shortage_kpi_slot.container():
            render_dashboard_kpi("부족수량 합계", f"{kpi_totals['부족수량']:,.0f}", "risk")
        with injection_kpi_slot.container():
            render_dashboard_kpi("사출부족수량 합계", f"{kpi_totals['사출 부족수량']:,.0f}", "risk")
        with process_stock_kpi_slot.container():
            render_dashboard_kpi("공정재고 합계", f"{kpi_totals['공정재고 합계']:,.0f}", "stock")
        with injection_stock_kpi_slot.container():
            render_dashboard_kpi("사출 재고", f"{kpi_totals['사출창고']:,.0f}", "stock")
        with separation_stock_kpi_slot.container():
            render_dashboard_kpi("분리 재고", f"{kpi_totals['분리창고']:,.0f}", "stock")
        with adhesion_stock_kpi_slot.container():
            render_dashboard_kpi("검사접착 재고", f"{kpi_totals['검사접착창고']:,.0f}", "stock")
        with adhesion_rework_stock_kpi_slot.container():
            render_dashboard_kpi("검사접착재작업 재고", f"{kpi_totals['검사접착재작업창고']:,.0f}", "stock")
        with leakage_stock_kpi_slot.container():
            render_dashboard_kpi("누수규격 재고", f"{kpi_totals['누수규격검사 창고']:,.0f}", "stock")
        p_table_total_count = len(p_table_ui)
        result_caption.caption(f"표시 {len(p_table_ui):,}건 / 전체 {p_table_total_count:,}건")
        p_table_display_source, _ = limit_dataframe_for_display(p_table_ui)
        caption_limited_rows(len(p_table_ui), len(p_table_display_source))
        p_display_columns = p_table_display_source.columns.tolist()
        p_table_display = format_numeric_columns_for_display(p_table_display_source)
        p_detail_column_config = build_auto_column_config(
            p_table_display, p_display_columns, source_df=p_table_display_source
        )
        render_lazy_excel_download_button(
            "엑셀 다운로드",
            p_table_ui,
            "생산현황",
            f"shortage_production_{download_stamp}.xlsx",
            "download_shortage_tab_p",
        )

        st.dataframe(
            style_operational_table(p_table_display, p_table_display_source),
            width="stretch",
            height=700,
            column_order=p_display_columns,
            column_config=p_detail_column_config,
            hide_index=True,
            key="shortage_p_table_v2",
        )

    elif selected_shortage_view == "사출 현황":
        r_summary = build_rcode_summary(filtered)
        r_summary_ui = r_summary.drop(columns=["상태"], errors="ignore")
        r_summary_total_count = len(r_summary_ui)
        r_summary = r_summary_ui
        result_caption.caption(f"표시 {len(r_summary_ui):,}건 / 전체 {r_summary_total_count:,}건")
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("R코드 수", f"{len(r_summary):,}")
        r2.metric(
            "R기준 사출 생산 필요수량 합계",
            f"{r_summary['사출 생산 필요수량 합계'].sum():,.0f}" if not r_summary.empty else "0",
        )
        r3.metric("R기준 사출 재고", f"{r_summary['사출창고 합계'].sum():,.0f}" if not r_summary.empty else "0")
        r4.metric("R기준 분리 재고", f"{r_summary['분리창고 합계'].sum():,.0f}" if not r_summary.empty else "0")
        r_summary_display_source, _ = limit_dataframe_for_display(r_summary_ui)
        caption_limited_rows(len(r_summary_ui), len(r_summary_display_source))
        r_summary_display = format_numeric_columns_for_display(r_summary_display_source)
        r_summary_column_config = build_auto_column_config(
            r_summary_display, r_summary_display.columns.tolist(), source_df=r_summary_display_source
        )
        render_lazy_excel_download_button(
            "엑셀 다운로드",
            r_summary,
            "사출생산현황",
            f"shortage_injection_summary_{download_stamp}.xlsx",
            "download_shortage_tab_r",
        )

        st.dataframe(
            style_operational_table(r_summary_display, r_summary_display_source),
            width="stretch",
            height=700,
            column_config=r_summary_column_config,
            hide_index=True,
            key="shortage_r_table_v2",
        )

    else:
        st.caption("사출 부족수량 = 수요정보 사출 생산 필요수량 합계")
        st.caption("표시 기준: 품목코드는 P코드만, 납기일/부족수량은 누수규격검사 기준")
        rq_filtered = filtered.copy()
        if "품목코드" in rq_filtered.columns:
            rq_filtered = rq_filtered[rq_filtered["품목코드"].astype(str).str.upper().str.startswith("P")]
        if "부족수량" in rq_filtered.columns:
            rq_filtered["부족수량"] = parse_mixed_numeric(rq_filtered["부족수량"])
        else:
            rq_filtered["부족수량"] = 0
        if "사출생산필요수량" in rq_filtered.columns:
            rq_filtered["사출생산필요수량"] = parse_mixed_numeric(rq_filtered["사출생산필요수량"])
        else:
            rq_filtered["사출생산필요수량"] = 0
        rq_filtered = rq_filtered[(rq_filtered["부족수량"] > 0) | (rq_filtered["사출생산필요수량"] > 0)]

        multi_p_r_codes: set[str] = set()
        if {"R코드5", "P코드5", "부족수량"}.issubset(rq_filtered.columns):
            rq_mapping_scope = rq_filtered.copy()
            rq_mapping_scope["R코드5"] = rq_mapping_scope["R코드5"].astype(str).str.strip()
            rq_mapping_scope["P코드5"] = rq_mapping_scope["P코드5"].astype(str).str.strip()
            rq_mapping_scope = rq_mapping_scope[
                rq_mapping_scope["R코드5"].str.startswith("R") & rq_mapping_scope["P코드5"].str.startswith("P")
            ]

            r_to_p_count = rq_mapping_scope.groupby("R코드5")["P코드5"].nunique()
            multi_p_r_codes = set(r_to_p_count[r_to_p_count >= 2].index.tolist())
            if multi_p_r_codes:
                rq_filtered = rq_filtered[rq_filtered["R코드5"].isin(multi_p_r_codes)]
            else:
                rq_filtered = rq_filtered.iloc[0:0]

        if "R코드 제품명" in rq_filtered.columns:
            rq_product_scope = rq_filtered.copy()
            rq_product_scope["부족수량"] = parse_mixed_numeric(rq_product_scope["부족수량"])
            rq_product_scope["사출생산필요수량"] = parse_mixed_numeric(rq_product_scope["사출생산필요수량"])
            rq_product_scope = rq_product_scope[
                (rq_product_scope["부족수량"] > 0) | (rq_product_scope["사출생산필요수량"] > 0)
            ]
            rq_product_scope["표시부족수량"] = rq_product_scope["부족수량"] + rq_product_scope["사출생산필요수량"]
            rq_product_sum_map = (
                rq_product_scope.groupby("R코드 제품명", as_index=True)["표시부족수량"].sum().sort_values(ascending=False).to_dict()
                if not rq_product_scope.empty
                else {}
            )
            rq_product_options = ["전체"] + [
                p for p in list(rq_product_sum_map.keys()) if str(p).strip() not in {"", "-", "nan", "None"}
            ]
            rq_product_count_map = {
                "전체": float(rq_product_scope["표시부족수량"].sum()) if not rq_product_scope.empty else 0.0,
                **rq_product_sum_map,
            }
            rq_selected_product = st.pills(
                "사출 제품명 (R코드5 1개당 P코드5 2+)",
                options=rq_product_options,
                default="전체",
                key="rq_tab_r_product_pills",
                format_func=lambda x: format_pill_label(x, rq_product_count_map),
            )
            if rq_selected_product != "전체":
                rq_filtered = rq_filtered[rq_filtered["R코드 제품명"] == rq_selected_product]

        rq_summary_tab = build_rq_group_summary(rq_filtered)
        rq_shortage_total = (
            parse_mixed_numeric(rq_filtered["부족수량"]).sum()
            if "부족수량" in rq_filtered.columns
            else 0
        )
        rq_inj_shortage_total = (
            parse_mixed_numeric(rq_summary_tab["사출 부족수량"]).sum()
            if not rq_summary_tab.empty and "사출 부족수량" in rq_summary_tab.columns
            else 0
        )
        r1, r2 = st.columns(2)
        r1.metric("사출 부족수량 합계", f"{rq_inj_shortage_total:,.0f}")
        r2.metric("부족수량 합계", f"{rq_shortage_total:,.0f}")

        if rq_summary_tab.empty:
            result_caption.caption("표시 0건 / 전체 0건")
            st.info("표시할 RQ 그룹 데이터가 없습니다.")
            render_lazy_excel_download_button(
                "엑셀 다운로드",
                pd.DataFrame(columns=detail_columns),
                "사출분리공용",
                f"shortage_shared_rq_{download_stamp}.xlsx",
                "download_shortage_tab_rq_empty",
            )
        else:
            rq_sort_cols = ["R코드5", "Q코드5", "부족수량"] if {"R코드5", "Q코드5", "부족수량"}.issubset(rq_filtered.columns) else ["R코드", "Q코드", "부족수량"]
            rq_sort_asc = [True, True, False]
            rq_view = rq_filtered.copy()
            if "납기일" not in rq_view.columns:
                rq_view["납기일"] = "-"
            if "부족수량" not in rq_view.columns:
                rq_view["부족수량"] = 0
            if "사출생산필요수량" in rq_view.columns:
                rq_view["사출부족수량"] = parse_mixed_numeric(rq_view["사출생산필요수량"])
            else:
                rq_view["사출부족수량"] = 0

            rq_detail_columns = [c for c in detail_columns if c in rq_view.columns]
            if "납기일" not in rq_detail_columns:
                insert_idx = rq_detail_columns.index("파워") + 1 if "파워" in rq_detail_columns else len(rq_detail_columns)
                rq_detail_columns.insert(insert_idx, "납기일")
            if "부족수량" not in rq_detail_columns:
                insert_idx = rq_detail_columns.index("납기일") + 1 if "납기일" in rq_detail_columns else len(rq_detail_columns)
                rq_detail_columns.insert(insert_idx, "부족수량")
            if "사출부족수량" not in rq_detail_columns:
                insert_idx = rq_detail_columns.index("부족수량") + 1 if "부족수량" in rq_detail_columns else len(rq_detail_columns)
                rq_detail_columns.insert(insert_idx, "사출부족수량")
            rq_detail_columns = move_columns_to_end(rq_detail_columns, ["비고"])

            rq_table = rq_view.sort_values(rq_sort_cols, ascending=rq_sort_asc)[rq_detail_columns]
            rq_table_ui = rq_table.drop(columns=["상태"], errors="ignore")
            rq_table_total_count = len(rq_table_ui)
            result_caption.caption(f"표시 {len(rq_table_ui):,}건 / 전체 {rq_table_total_count:,}건")
            rq_table_display_source, _ = limit_dataframe_for_display(rq_table_ui)
            caption_limited_rows(len(rq_table_ui), len(rq_table_display_source))
            rq_display_columns = rq_table_display_source.columns.tolist()
            rq_table_display = format_numeric_columns_for_display(rq_table_display_source)
            rq_detail_column_config = build_auto_column_config(
                rq_table_display, rq_display_columns, source_df=rq_table_display_source
            )
            render_lazy_excel_download_button(
                "엑셀 다운로드",
                rq_table_ui,
                "사출분리공용",
                f"shortage_shared_rq_{download_stamp}.xlsx",
                "download_shortage_tab_rq",
            )
            st.dataframe(
                style_operational_table(rq_table_display, rq_table_display_source),
                width="stretch",
                height=700,
                column_order=rq_display_columns,
                column_config=rq_detail_column_config,
                hide_index=True,
                key="shortage_rq_table_v2",
            )


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_leadji_code_mapping(leadji_info: pd.DataFrame) -> pd.DataFrame:
    columns = ["P코드5", "리드지코드", "리드지명"]
    if leadji_info.empty:
        return pd.DataFrame(columns=columns)

    info_cols = leadji_info.columns.tolist()
    prod_col = pick_first_existing_column(info_cols, ["생산"])
    b1_col = pick_first_existing_column(info_cols, ["B1코드"])
    b1_name_col = pick_first_existing_column(info_cols, ["B1코드명"])

    if prod_col is None and len(info_cols) > 3:
        prod_col = info_cols[3]
    if b1_col is None and len(info_cols) > 12:
        b1_col = info_cols[12]
    if b1_name_col is None and len(info_cols) > 13:
        b1_name_col = info_cols[13]
    if prod_col is None or b1_col is None:
        return pd.DataFrame(columns=columns)

    selected_cols = [prod_col, b1_col] + ([b1_name_col] if b1_name_col is not None else [])
    mapping = leadji_info[selected_cols].copy()
    for col in selected_cols:
        mapping[col] = mapping[col].astype(str).str.strip().replace({"nan": "", "None": ""})

    mapping["P코드5"] = mapping[prod_col].str.upper().str[:5]
    mapping = mapping[(mapping["P코드5"].str.startswith("P")) & (mapping[b1_col] != "")]
    mapping = mapping.rename(columns={b1_col: "리드지코드"})
    if b1_name_col is not None:
        mapping = mapping.rename(columns={b1_name_col: "리드지명"})
    else:
        mapping["리드지명"] = "-"
    return mapping[columns].drop_duplicates(subset=["P코드5", "리드지코드"], keep="first")


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_leadji_stock_pivot(leadji_stock: pd.DataFrame) -> pd.DataFrame:
    if leadji_stock.empty:
        return pd.DataFrame(columns=["리드지코드"])

    stock_cols = leadji_stock.columns.tolist()
    code_col = pick_first_existing_column(stock_cols, ["품목코드"])
    warehouse_col = pick_first_existing_column(stock_cols, ["창고"])
    qty_col = pick_first_existing_column(stock_cols, ["재고"])
    if not code_col or not warehouse_col or not qty_col:
        return pd.DataFrame(columns=["리드지코드"])

    stock = leadji_stock[[code_col, warehouse_col, qty_col]].copy()
    stock[code_col] = stock[code_col].astype(str).str.strip()
    stock[warehouse_col] = stock[warehouse_col].astype(str).str.strip()
    stock[qty_col] = parse_mixed_numeric(stock[qty_col])
    stock = stock[(stock[code_col] != "") & (stock[warehouse_col] != "") & (stock[qty_col] > 0)]
    if stock.empty:
        return pd.DataFrame(columns=["리드지코드"])

    stock = stock.groupby([code_col, warehouse_col], as_index=False)[qty_col].sum()
    pivot = stock.pivot_table(
        index=code_col,
        columns=warehouse_col,
        values=qty_col,
        aggfunc="sum",
        fill_value=0,
    )
    if pivot.empty:
        return pd.DataFrame(columns=["리드지코드"])

    excluded_warehouse_columns = {"L관창고(자재불량)"}
    warehouse_columns = [
        str(c) for c in pivot.sum(axis=0).sort_values(ascending=False).index.tolist() if str(c) not in excluded_warehouse_columns
    ]
    pivot = pivot.reindex(columns=warehouse_columns).reset_index().rename(columns={code_col: "리드지코드"})
    for w_col in warehouse_columns:
        pivot[w_col] = parse_mixed_numeric(pivot[w_col])
    return pivot


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_leadji_requirement_summary(
    shortage_df: pd.DataFrame, leadji_info: pd.DataFrame, leadji_stock: pd.DataFrame
) -> pd.DataFrame:
    fixed_columns = [
        "리드지코드",
        "리드지명",
        "수요사이트",
        "생산필요수량",
        "리드지필요수량",
        "리드지부족",
        "리드지부족수량",
        "최소납기일",
    ]
    if shortage_df.empty or "품목코드" not in shortage_df.columns:
        return pd.DataFrame(columns=fixed_columns)

    p_shortage = build_leadji_p_shortage(shortage_df)
    if p_shortage.empty:
        return pd.DataFrame(columns=fixed_columns)

    mapping = build_leadji_code_mapping(leadji_info)
    if mapping.empty:
        return pd.DataFrame(columns=fixed_columns)

    bs_base = p_shortage.merge(mapping, on="P코드5", how="left")
    unmatched_map = bs_base["리드지코드"].isna() | (bs_base["리드지코드"].astype(str).str.strip() == "")
    bs_base.loc[unmatched_map, "리드지코드"] = "매칭필요:" + bs_base.loc[unmatched_map, "P코드5"].astype(str)
    bs_base.loc[unmatched_map, "리드지명"] = "리드지정보 B1코드 없음"
    bs_base["리드지코드"] = bs_base["리드지코드"].fillna("-")
    bs_base["리드지명"] = bs_base["리드지명"].fillna("-")

    summary = (
        bs_base.groupby(["리드지코드", "리드지명"], as_index=False)
        .agg({"수요사이트": join_unique_text_values, "생산필요수량": "sum", "최소납기일": "min"})
        .sort_values(["생산필요수량", "리드지코드"], ascending=[False, True])
    )
    warehouse_columns: list[str] = []
    stock_pivot = build_leadji_stock_pivot(leadji_stock)
    if not stock_pivot.empty:
        warehouse_columns = [c for c in stock_pivot.columns if c != "리드지코드"]
        summary = summary.merge(stock_pivot, on="리드지코드", how="left")
        for w_col in warehouse_columns:
            summary[w_col] = parse_mixed_numeric(summary[w_col])

    active_warehouse_columns: list[str] = []
    for w_col in warehouse_columns:
        col_sum = parse_mixed_numeric(summary[w_col]).sum() if w_col in summary.columns else 0
        if col_sum > 0:
            active_warehouse_columns.append(w_col)

    summary["리드지필요수량"] = (parse_mixed_numeric(summary["생산필요수량"]) * 1.3).round(0)
    leadji_target_warehouses = ["L관창고(자재)", "C관 공정부자재", "S관 공정부자재", "A관 공정부자재"]
    leadji_stock_total = pd.Series(0.0, index=summary.index)
    for warehouse_name in leadji_target_warehouses:
        matched_col = find_warehouse_column(summary.columns.tolist(), [warehouse_name])
        if matched_col is None:
            continue
        leadji_stock_total = leadji_stock_total + parse_mixed_numeric(summary[matched_col])

    shortage_qty = leadji_stock_total - summary["리드지필요수량"]
    summary["리드지부족"] = ""
    summary.loc[shortage_qty < 0, "리드지부족"] = "🔴"
    summary["리드지부족수량"] = shortage_qty.where(shortage_qty < 0)
    summary["최소납기일"] = pd.to_datetime(summary["최소납기일"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("-")
    return summary[
        [
            "리드지코드",
            "리드지명",
            "수요사이트",
            "생산필요수량",
            "리드지필요수량",
            "리드지부족",
            "리드지부족수량",
            *active_warehouse_columns,
            "최소납기일",
        ]
    ]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def compute_leadji_source_total(shortage_df: pd.DataFrame) -> float:
    p_shortage = build_leadji_p_shortage(shortage_df)
    if p_shortage.empty:
        return 0.0
    return float(parse_mixed_numeric(p_shortage["생산필요수량"]).sum())


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_leadji_p_shortage(shortage_df: pd.DataFrame) -> pd.DataFrame:
    if shortage_df.empty or "품목코드" not in shortage_df.columns:
        return pd.DataFrame(columns=["P코드5", "수요사이트", "생산필요수량", "최소납기일"])

    qty_source_col = LEADJI_REQUIRED_QTY_COL if LEADJI_REQUIRED_QTY_COL in shortage_df.columns else "부족수량"
    if qty_source_col not in shortage_df.columns:
        return pd.DataFrame(columns=["P코드5", "수요사이트", "생산필요수량", "최소납기일"])

    if qty_source_col == LEADJI_REQUIRED_QTY_COL and LEADJI_REQUIRED_DUE_COL in shortage_df.columns:
        due_source_col = LEADJI_REQUIRED_DUE_COL
    else:
        due_source_col = "최소납기일" if "최소납기일" in shortage_df.columns else "납기일"

    base = shortage_df.copy()
    base["품목코드"] = base["품목코드"].astype(str).str.strip().str.upper()
    base["P코드5"] = base["품목코드"].str[:5]
    base = base[base["P코드5"].str.startswith("P")]
    if "사이트코드" in base.columns:
        base["수요사이트"] = base["사이트코드"].astype(str).str.strip()
        base["수요사이트"] = base["수요사이트"].where(
            ~base["수요사이트"].str.lower().isin({"", "nan", "none", "null", "nat", "<na>"}),
            "(미지정)",
        )
    else:
        base["수요사이트"] = "(미지정)"
    base["생산필요수량"] = parse_mixed_numeric(base[qty_source_col])
    base["완료재고수량"] = (
        parse_mixed_numeric(base[LEADJI_COMPLETED_STOCK_COL])
        if LEADJI_COMPLETED_STOCK_COL in base.columns
        else 0
    )
    if due_source_col in base.columns:
        base["납기일_dt"] = pd.to_datetime(base[due_source_col], errors="coerce")
    else:
        base["납기일_dt"] = pd.NaT

    item_shortage = (
        base.groupby("품목코드", as_index=False)
        .agg(
            {
                "P코드5": "first",
                "수요사이트": join_unique_text_values,
                "생산필요수량": "sum",
                "완료재고수량": "max",
                "납기일_dt": "min",
            }
        )
    )
    # 품목코드 단위로 필요수량 합산 후 완료재고(누수규격검사 창고)를 1회 차감한다.
    item_shortage["생산필요수량"] = (item_shortage["생산필요수량"] - item_shortage["완료재고수량"]).clip(lower=0)
    item_shortage = item_shortage[item_shortage["생산필요수량"] > 0]

    p_shortage = (
        item_shortage.groupby("P코드5", as_index=False)
        .agg({"수요사이트": join_unique_text_values, "생산필요수량": "sum", "납기일_dt": "min"})
        .rename(columns={"납기일_dt": "최소납기일"})
    )
    return p_shortage


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def build_pcode5_leadji_requirement_summary(
    shortage_df: pd.DataFrame, leadji_info: pd.DataFrame, leadji_stock: pd.DataFrame
) -> pd.DataFrame:
    fixed_columns = ["생산코드", "수요사이트", "리드지코드", "리드지명", "생산필요수량", "최소납기일"]
    if shortage_df.empty or "품목코드" not in shortage_df.columns:
        return pd.DataFrame(columns=fixed_columns)

    p_shortage = build_leadji_p_shortage(shortage_df)
    if p_shortage.empty:
        return pd.DataFrame(columns=fixed_columns)

    mapping = build_leadji_code_mapping(leadji_info)
    if not mapping.empty:
        detail = p_shortage.merge(mapping, on="P코드5", how="left")
    else:
        detail = p_shortage.copy()
        detail["리드지코드"] = "-"
        detail["리드지명"] = "-"

    detail["리드지코드"] = detail["리드지코드"].fillna("-")
    detail["리드지명"] = detail["리드지명"].fillna("-")

    summary = (
        detail.groupby(["P코드5", "리드지코드", "리드지명"], as_index=False)
        .agg({"수요사이트": join_unique_text_values, "생산필요수량": "sum", "최소납기일": "min"})
        .sort_values(["생산필요수량", "P코드5", "리드지코드"], ascending=[False, True, True])
    )

    warehouse_columns: list[str] = []
    stock_pivot = build_leadji_stock_pivot(leadji_stock)
    if not stock_pivot.empty:
        warehouse_columns = [c for c in stock_pivot.columns if c != "리드지코드"]
        summary = summary.merge(stock_pivot, on="리드지코드", how="left")
        for w_col in warehouse_columns:
            summary[w_col] = parse_mixed_numeric(summary[w_col])

    active_warehouse_columns: list[str] = []
    for w_col in warehouse_columns:
        col_sum = parse_mixed_numeric(summary[w_col]).sum() if w_col in summary.columns else 0
        if col_sum > 0:
            active_warehouse_columns.append(w_col)

    summary["최소납기일"] = pd.to_datetime(summary["최소납기일"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("-")
    summary = summary.rename(columns={"P코드5": "생산코드"})
    return summary[["생산코드", "수요사이트", "리드지코드", "리드지명", "생산필요수량", *active_warehouse_columns, "최소납기일"]]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def merge_leadji_with_order_status(summary_df: pd.DataFrame, leadji_order_df: pd.DataFrame) -> pd.DataFrame:
    merged = summary_df.copy()
    merged["리드지코드_join"] = merged["리드지코드"].map(normalize_leadji_code_key)

    def first_nonempty_text(series: pd.Series) -> str:
        text = series.astype(str).str.strip()
        text = text[(text != "") & (text.str.lower() != "nan") & (text.str.lower() != "none")]
        return text.iloc[0] if not text.empty else "-"

    if leadji_order_df.empty:
        merged["발주수량"] = 0.0
        merged["구매발주수량"] = 0.0
        merged["구매의뢰수량"] = 0.0
        merged["입고예상일자_dt"] = pd.NaT
    else:
        order = leadji_order_df.copy()
        order["리드지코드_join"] = order["리드지코드"].map(normalize_leadji_code_key)
        for qty_col in ["구매발주수량", "구매의뢰수량"]:
            if qty_col not in order.columns:
                order[qty_col] = 0.0
            order[qty_col] = parse_mixed_numeric(order[qty_col])
        if "리드지명" not in order.columns:
            order["리드지명"] = "-"
        order = (
            order.groupby("리드지코드_join", as_index=False)
            .agg(
                {
                    "리드지코드": first_nonempty_text,
                    "리드지명": first_nonempty_text,
                    "발주수량": "sum",
                    "구매발주수량": "sum",
                    "구매의뢰수량": "sum",
                    "입고예상일자_dt": "min",
                }
            )
            .rename(
                columns={
                    "리드지코드": "리드지코드_order",
                    "리드지명": "리드지명_order",
                    "발주수량": "발주수량_join",
                }
            )
        )
        summary_keys = set(merged["리드지코드_join"].dropna().astype(str))
        order_only = order[~order["리드지코드_join"].astype(str).isin(summary_keys)].copy()
        merged = merged.merge(order, on="리드지코드_join", how="left")
        merged["발주수량"] = parse_mixed_numeric(merged["발주수량_join"])
        merged["구매발주수량"] = parse_mixed_numeric(merged["구매발주수량"])
        merged["구매의뢰수량"] = parse_mixed_numeric(merged["구매의뢰수량"])

        if not order_only.empty:
            extra = pd.DataFrame(index=order_only.index)
            for col in summary_df.columns:
                if col in {"리드지코드", "리드지명", "리드지부족", "최소납기일"}:
                    extra[col] = "-"
                elif col == "리드지부족수량":
                    extra[col] = pd.NA
                else:
                    extra[col] = 0.0

            if "리드지코드" in extra.columns:
                extra["리드지코드"] = order_only["리드지코드_order"].where(
                    order_only["리드지코드_order"].astype(str).str.strip() != "",
                    order_only["리드지코드_join"],
                )
            if "리드지명" in extra.columns:
                extra["리드지명"] = order_only["리드지명_order"].fillna("-")
            if "리드지부족" in extra.columns:
                extra["리드지부족"] = ""
            if "최소납기일" in extra.columns:
                extra["최소납기일"] = "-"

            extra["리드지코드_join"] = order_only["리드지코드_join"]
            extra["발주수량"] = parse_mixed_numeric(order_only["발주수량_join"])
            extra["구매발주수량"] = parse_mixed_numeric(order_only["구매발주수량"])
            extra["구매의뢰수량"] = parse_mixed_numeric(order_only["구매의뢰수량"])
            extra["입고예상일자_dt"] = pd.to_datetime(order_only["입고예상일자_dt"], errors="coerce")
            merged = pd.concat([merged, extra], ignore_index=True, sort=False)

    merged["입고예상일자_dt"] = pd.to_datetime(merged["입고예상일자_dt"], errors="coerce")
    merged["입고예상일자"] = merged["입고예상일자_dt"].dt.strftime("%Y-%m-%d").fillna("미확인")

    shortage_raw = parse_mixed_numeric(merged["리드지부족수량"])
    shortage_qty = shortage_raw.where(shortage_raw > 0, -shortage_raw).clip(lower=0)
    shortage_mask = shortage_qty > 0
    missing_due_mask = merged["입고예상일자_dt"].isna()
    enough_order_mask = merged["발주수량"] >= shortage_qty
    has_purchase_order_mask = parse_mixed_numeric(merged["구매발주수량"]) > 0
    has_purchase_request_mask = parse_mixed_numeric(merged["구매의뢰수량"]) > 0
    order_only_mask = parse_mixed_numeric(merged["생산필요수량"]) <= 0

    merged["상태"] = "부족 없음"
    merged.loc[
        order_only_mask & ~missing_due_mask & has_purchase_order_mask & ~has_purchase_request_mask,
        "상태",
    ] = "입고 예정"
    merged.loc[
        order_only_mask & ~missing_due_mask & ~has_purchase_order_mask & has_purchase_request_mask,
        "상태",
    ] = "구매의뢰"
    merged.loc[
        order_only_mask & ~missing_due_mask & has_purchase_order_mask & has_purchase_request_mask,
        "상태",
    ] = "입고 예정+의뢰"
    merged.loc[order_only_mask & missing_due_mask & (has_purchase_order_mask | has_purchase_request_mask), "상태"] = (
        "입고일 미확인"
    )
    merged.loc[shortage_mask & missing_due_mask, "상태"] = "입고일 미확인"
    merged.loc[
        shortage_mask & ~missing_due_mask & enough_order_mask & has_purchase_order_mask & ~has_purchase_request_mask,
        "상태",
    ] = "입고 예정"
    merged.loc[
        shortage_mask & ~missing_due_mask & enough_order_mask & ~has_purchase_order_mask & has_purchase_request_mask,
        "상태",
    ] = "구매의뢰"
    merged.loc[
        shortage_mask & ~missing_due_mask & enough_order_mask & has_purchase_order_mask & has_purchase_request_mask,
        "상태",
    ] = "입고 예정+의뢰"
    merged.loc[shortage_mask & ~missing_due_mask & ~enough_order_mask, "상태"] = "발주부족"

    ordered_cols: list[str] = []
    for col in summary_df.columns:
        ordered_cols.append(col)
        if col == "리드지부족수량":
            ordered_cols.extend(["발주수량", "입고예상일자", "상태"])

    keep_cols = [c for c in ordered_cols if c in merged.columns] + ["입고예상일자_dt"]
    return merged[keep_cols]


def render_leadji_dashboard(
    updated_at: str,
    shortage_df: pd.DataFrame,
    leadji_info: pd.DataFrame,
    leadji_stock: pd.DataFrame,
    leadji_order_df: pd.DataFrame,
) -> None:
    st.subheader("리드지 현황")
    st.caption(f"업데이트: {updated_at}")
    download_stamp = datetime.now(DISPLAY_TZ).strftime("%Y%m%d_%H%M%S")

    summary_df = build_leadji_requirement_summary(shortage_df, leadji_info, leadji_stock)
    if summary_df.empty and leadji_order_df.empty:
        st.warning("리드지재고현황을 계산할 데이터가 없습니다.")
    else:
        summary_df = merge_leadji_with_order_status(summary_df, leadji_order_df)
        st.warning("입고예정일자는 구매의뢰 기준입니다. 실제 입고 일정은 구매팀 확인이 필요합니다.")

        stock_target_names = ["L관창고(자재)", "C관 공정부자재", "S관 공정부자재", "A관 공정부자재"]
        stock_detail_columns: list[str] = []
        summary_df["재고합계"] = 0.0
        for warehouse_name in stock_target_names:
            matched_col = find_warehouse_column(summary_df.columns.tolist(), [warehouse_name])
            display_col = warehouse_name
            if matched_col is not None:
                summary_df[display_col] = parse_mixed_numeric(summary_df[matched_col])
            elif display_col not in summary_df.columns:
                summary_df[display_col] = 0.0
            summary_df["재고합계"] = summary_df["재고합계"] + parse_mixed_numeric(summary_df[display_col])
            stock_detail_columns.append(display_col)

        summary_df["생산 최소 납기일"] = summary_df["최소납기일"] if "최소납기일" in summary_df.columns else "-"
        shortage_numeric = parse_mixed_numeric(summary_df["리드지부족수량"])
        shortage_abs = (-shortage_numeric).clip(lower=0)
        inbound_date = pd.to_datetime(summary_df["입고예상일자_dt"], errors="coerce")
        summary_df["부족수량_abs"] = shortage_abs
        summary_df["우선순위"] = "정상"
        summary_df.loc[(shortage_numeric < 0) & inbound_date.isna(), "우선순위"] = "긴급"
        summary_df.loc[(shortage_numeric < 0) & inbound_date.notna(), "우선순위"] = "확인필요"
        priority_order = {"긴급": 0, "확인필요": 1, "정상": 2}
        summary_df["우선순위정렬"] = summary_df["우선순위"].map(priority_order).fillna(9)
        summary_df = summary_df.sort_values(
            ["우선순위정렬", "부족수량_abs", "리드지코드"],
            ascending=[True, False, True],
        )

        total_codes = summary_df["리드지코드"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
        shortage_count = int((shortage_numeric < 0).sum())
        inbound_planned_count = int(summary_df["상태"].astype(str).str.contains("입고 예정", regex=False).sum())
        total_shortage_qty = float(shortage_abs.sum())
        st.markdown(
            f"""
            <style>
            .leadji-kpi-grid {{
                display: grid;
                grid-template-columns: repeat(4, minmax(0, 1fr));
                gap: 14px;
                margin: 16px 0 20px;
            }}
            .leadji-kpi-card {{
                border: 1px solid #E5E7EB;
                border-left: 4px solid #1A2B5E;
                border-radius: 12px;
                background: #FFFFFF;
                padding: 16px 18px;
                min-height: 104px;
                box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
            }}
            .leadji-kpi-card strong {{
                display: block;
                color: #64748B;
                font-size: 13px;
                font-weight: 700;
                margin-bottom: 8px;
            }}
            .leadji-kpi-card span {{
                display: block;
                color: #374151;
                font-size: 30px;
                font-weight: 850;
                line-height: 1.15;
            }}
            .leadji-kpi-card.risk {{
                background: #FFFFFF;
                border-color: #E5E7EB;
                border-left-color: #DC2626;
            }}
            .leadji-kpi-card.risk strong,
            .leadji-kpi-card.risk span {{
                color: #DC2626;
            }}
            .leadji-kpi-card.inbound {{
                background: #FFFFFF;
                border-color: #E5E7EB;
                border-left-color: #1A2B5E;
            }}
            .leadji-kpi-card.inbound span {{
                color: #1A2B5E;
            }}
            @media (max-width: 900px) {{
                .leadji-kpi-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
            }}
            </style>
            <div class="leadji-kpi-grid">
                <div class="leadji-kpi-card"><strong>전체 리드지코드</strong><span>{total_codes:,.0f}</span></div>
                <div class="leadji-kpi-card risk"><strong>부족 리드지</strong><span>{shortage_count:,.0f}</span></div>
                <div class="leadji-kpi-card inbound"><strong>입고예정</strong><span>{inbound_planned_count:,.0f}</span></div>
                <div class="leadji-kpi-card risk"><strong>총 리드지 부족수량</strong><span>{total_shortage_qty:,.0f}</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        qcol, _ = st.columns([1.35, 2.65])
        with qcol:
            summary_query = st.text_input(
                "검색",
                value="",
                key="leadji_summary_query",
                placeholder="리드지코드, 리드지명, 창고로 검색하세요",
            ).strip()

        hidden_cols = ["입고예상일자_dt", "우선순위정렬", "부족수량_abs", "최소납기일", "상태"]
        summary_visible = summary_df.drop(columns=hidden_cols, errors="ignore")
        summary_search_cols = [c for c in summary_visible.columns if c not in ["생산필요수량"]]
        filtered_visible = filter_with_terms_any(summary_visible, summary_search_cols, summary_query)
        filtered_summary = summary_df.loc[filtered_visible.index].copy()

        ordered_columns = [
            "리드지코드",
            "리드지명",
            "발주수량",
            "구매발주수량",
            "구매의뢰수량",
            "입고예상일자",
            "리드지부족수량",
            "생산 최소 납기일",
        ]
        ordered_rows = filtered_summary[parse_mixed_numeric(filtered_summary["발주수량"]) > 0].copy()
        ordered_rows = ordered_rows.sort_values(
            ["입고예상일자_dt", "부족수량_abs", "리드지코드"],
            ascending=[True, False, True],
            na_position="last",
        )
        st.markdown(
            f"""
            <div class="dashboard-section-header">
                <h3>발주 반영 품목 리스트</h3>
                <span class="dashboard-count-badge">전체 {len(ordered_rows):,}건</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if ordered_rows.empty:
            st.info("발주 또는 구매의뢰가 반영된 리드지가 없습니다.")
        else:
            ordered_table = ordered_rows[[c for c in ordered_columns if c in ordered_rows.columns]]
            ordered_display_source, _ = limit_dataframe_for_display(ordered_table)
            caption_limited_rows(len(ordered_table), len(ordered_display_source))
            ordered_display = format_numeric_columns_for_display(ordered_display_source)
            ordered_column_config = build_auto_column_config(
                ordered_display, ordered_display.columns.tolist(), source_df=ordered_display_source
            )
            ordered_styled = style_leadji_shortage_table(ordered_display, ordered_display_source)
            st.dataframe(
                ordered_styled,
                width="stretch",
                height=min(520, 78 + len(ordered_display_source) * 38),
                column_config=ordered_column_config,
                hide_index=True,
            )

        st.markdown(
            f"""
            <div class="dashboard-section-header">
                <h3>리드지 목록</h3>
                <span class="dashboard-count-badge">전체 {len(filtered_summary):,}건</span>
            </div>
            <div class="dashboard-section-subtle">핵심 운영 컬럼만 기본 표시합니다. 창고별 수량은 아래 재고 상세 컬럼에서 확인하세요.</div>
            """,
            unsafe_allow_html=True,
        )
        basic_columns = [
            "리드지코드",
            "리드지명",
            "리드지필요수량",
            "재고합계",
            "리드지부족수량",
            "발주수량",
            "입고예상일자",
            "생산 최소 납기일",
        ]
        table_df = filtered_summary[[c for c in basic_columns if c in filtered_summary.columns]]
        table_df = table_df.drop(columns=["상태"], errors="ignore")

        download_df = filtered_summary.drop(columns=hidden_cols, errors="ignore")
        st.download_button(
            "엑셀 다운로드",
            data=dataframe_to_excel_bytes(download_df, sheet_name="리드지현황"),
            file_name=f"leadji_status_{download_stamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_leadji_summary",
            width="content",
        )

        table_display_source, _ = limit_dataframe_for_display(table_df)
        caption_limited_rows(len(table_df), len(table_display_source))
        leadji_display = format_numeric_columns_for_display(table_display_source)
        leadji_column_config = build_auto_column_config(
            leadji_display, leadji_display.columns.tolist(), source_df=table_display_source
        )
        leadji_column_config.pop("리드지부족", None)
        leadji_styled = style_leadji_shortage_table(leadji_display, table_display_source)
        st.dataframe(
            leadji_styled,
            width="stretch",
            height=620,
            column_config=leadji_column_config,
            hide_index=True,
        )

        with st.expander("재고 상세 컬럼"):
            stock_table_columns = [
                "리드지코드",
                "리드지명",
                "재고합계",
                *stock_detail_columns,
            ]
            stock_table = filtered_summary[[c for c in stock_table_columns if c in filtered_summary.columns]]
            stock_display_source, _ = limit_dataframe_for_display(stock_table)
            caption_limited_rows(len(stock_table), len(stock_display_source))
            stock_display = format_numeric_columns_for_display(stock_display_source)
            stock_column_config = build_auto_column_config(
                stock_display, stock_display.columns.tolist(), source_df=stock_display_source
            )
            st.dataframe(
                stock_display,
                width="stretch",
                height=420,
                column_config=stock_column_config,
                hide_index=True,
            )

        source_total = compute_leadji_source_total(shortage_df)
        summary_total = float(parse_mixed_numeric(summary_df["생산필요수량"]).sum())
        verify_diff = summary_total - source_total
        with st.expander("데이터 검증 정보"):
            st.caption(
                f"검증: 품목코드별 ({LEADJI_REQUIRED_QTY_COL} - {LEADJI_COMPLETED_STOCK_COL}) 합계 {source_total:,.0f} / "
                f"리드지 합계 {summary_total:,.0f} / 차이 {verify_diff:,.0f}"
            )


def render_leadji_pcode5_dashboard(
    updated_at: str, shortage_df: pd.DataFrame, leadji_info: pd.DataFrame, leadji_stock: pd.DataFrame
) -> None:
    st.subheader("생산코드별 리드지 현황")
    st.caption(f"업데이트: {updated_at}")
    st.caption(
        f"집계 기준: 품목코드별 ({LEADJI_REQUIRED_QTY_COL} - {LEADJI_COMPLETED_STOCK_COL})를 0 미만 0으로 만든 뒤 P코드 단위 합산(sum)"
    )
    st.caption("기준: 동일 생산코드에 여러 리드지가 매핑되면 생산필요수량이 각 리드지 행에 반복 표시됩니다.")
    download_stamp = datetime.now(DISPLAY_TZ).strftime("%Y%m%d_%H%M%S")

    summary_df = build_pcode5_leadji_requirement_summary(shortage_df, leadji_info, leadji_stock)
    if summary_df.empty:
        st.warning("생산코드별 리드지 현황을 계산할 데이터가 없습니다.")
        return

    qcol, _ = st.columns([3.0, 1.0])
    with qcol:
        summary_query = st.text_input(
            "통합 검색 (생산코드/리드지코드/리드지명/창고)",
            value="",
            key="leadji_pcode5_summary_query",
            placeholder="예: P1234, BS0314, 블리스터케이스, 원료창고",
        ).strip()

    summary_search_cols = [c for c in summary_df.columns if c not in ["생산필요수량"]]
    filtered_summary = filter_with_terms_any(summary_df, summary_search_cols, summary_query)

    c1, c2, c3 = st.columns(3)
    c1.metric("생산코드 수", f"{filtered_summary['생산코드'].astype(str).nunique():,}")
    p_qty_total = (
        filtered_summary.drop_duplicates(subset=["생산코드"], keep="first")["생산필요수량"].sum()
        if not filtered_summary.empty
        else 0
    )
    c2.metric("생산코드 기준 생산필요수량 합계", f"{p_qty_total:,.0f}")
    min_due_dt = pd.to_datetime(filtered_summary["최소납기일"], errors="coerce").min()
    c3.metric("생산 최소 납기일", "-" if pd.isna(min_due_dt) else min_due_dt.strftime("%Y-%m-%d"))

    st.download_button(
        "엑셀 다운로드",
        data=dataframe_to_excel_bytes(filtered_summary, sheet_name="생산코드기준리드지현황"),
        file_name=f"leadji_status_by_production_code_{download_stamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_leadji_pcode5_summary",
        width="content",
    )

    leadji_p_display_source, _ = limit_dataframe_for_display(filtered_summary)
    caption_limited_rows(len(filtered_summary), len(leadji_p_display_source))
    leadji_p_display = format_numeric_columns_for_display(leadji_p_display_source)
    leadji_p_column_config = build_auto_column_config(
        leadji_p_display, leadji_p_display.columns.tolist(), source_df=leadji_p_display_source
    )
    st.dataframe(
        leadji_p_display,
        width="stretch",
        height=700,
        column_config=leadji_p_column_config,
        hide_index=True,
    )


def effective_column_series(df: pd.DataFrame, column_name: str | None, default: object = "") -> pd.Series:
    if column_name and column_name in df.columns:
        selected = df[column_name]
        if isinstance(selected, pd.DataFrame):
            selected = selected.iloc[:, 0]
        return selected
    return pd.Series(default, index=df.index, dtype="object")


def effective_numeric_series(df: pd.DataFrame, column_name: str | None, default: float = 0.0) -> pd.Series:
    if column_name and column_name in df.columns:
        return parse_mixed_numeric(effective_column_series(df, column_name)).fillna(default)
    return pd.Series(default, index=df.index, dtype="float64")


def effective_date_text_series(df: pd.DataFrame, column_name: str | None) -> pd.Series:
    if column_name is None or column_name not in df.columns:
        return pd.Series("", index=df.index, dtype="object")
    parsed = parse_mixed_excel_date(effective_column_series(df, column_name))
    return parsed.dt.strftime("%Y-%m-%d").fillna("")


def normalize_effective_process(value: object) -> str:
    text = clean_text_value(value)
    digit_match = re.search(r"\d+", text)
    code = digit_match.group(0).lstrip("0") if digit_match else ""
    if code == "10" or "사출" in text:
        return "[10]사출조립"
    if code == "80" or "누수" in text or "규격" in text:
        return "[80]누수/규격검사"
    return text


def normalize_effective_site(value: object) -> str:
    text = clean_text_value(value)
    if not text:
        return ""
    compact = re.sub(r"[\s_./()\-]+", "", text.upper())
    for site in SITE_GROUP_ORDER:
        site_key = site.upper()
        site_letter = site_key[0]
        if site_key in compact or re.match(rf"^{site_letter}(관|동|공장|$)", compact):
            return site
    digit_match = re.search(r"\d+", text)
    code = digit_match.group(0).zfill(2) if digit_match else ""
    site_by_code = {
        "01": "A관",
        "02": "C관",
        "03": "S관",
        "05": "5공장",
    }
    if code in site_by_code:
        return site_by_code[code]
    site_group = normalize_site_group(text)
    return site_group if site_group in SITE_GROUP_ORDER else text


def calculate_effective_rate(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    numerator_values = parse_mixed_numeric(numerator)
    denominator_values = parse_mixed_numeric(denominator)
    return (numerator_values / denominator_values.where(denominator_values.ne(0)) * 100).fillna(0)


def format_effective_int(value: object) -> str:
    try:
        return f"{float(value):,.0f}"
    except (TypeError, ValueError):
        return "0"


def format_effective_pct(value: object) -> str:
    try:
        return f"{float(value):,.1f}%"
    except (TypeError, ValueError):
        return "0.0%"


def format_effective_decimal(value: object, digits: int = 1) -> str:
    try:
        return f"{float(value):,.{digits}f}"
    except (TypeError, ValueError):
        return f"{0:.{digits}f}"


def classify_effective_major_category(row: pd.Series) -> str:
    customer = clean_text_value(row.get("거래처", ""))
    product_name = clean_text_value(row.get("제품명", ""))
    combined = f"{customer} {product_name}"
    combined_key = normalize_keyword_key(combined)

    if "pia_kr" in product_name.lower() or "piakr" in combined_key:
        return "국내"
    if any(token in combined_key for token in ("국내", "korea", "clalen", "렌즈미", "lensme", "lensvery")):
        return "국내"

    category = clean_text_value(classify_sheet(pd.Series({"거래처": customer, "제품명": product_name})))
    category_key = normalize_keyword_key(category)
    if "국내" in category_key:
        return "국내"
    if "pia" in category_key or "pia" in combined_key:
        return "PIA"
    return "기타해외"


def fetch_effective_plan_operations_dataframe(
    site_filter: str,
    start_date_text: str,
    end_date_text: str,
    api_key_hash: str,
    source_updated_at: str,
) -> tuple[pd.DataFrame, str]:
    api_key = get_plan_api_key()
    base_url = get_plan_api_base_url()
    site_param = build_plan_api_site_param(site_filter)
    frames: list[pd.DataFrame] = []
    errors: list[str] = []

    def fetch_operation(operation: str) -> tuple[str, pd.DataFrame, str]:
        params: dict[str, object] = {
            "limit": PLAN_API_DEFAULT_ROW_LIMIT,
            "oper": operation,
            "plan_from": start_date_text,
            "plan_to": end_date_text,
        }
        if site_param:
            params["site"] = site_param
        params_tuple = tuple(sorted(params.items(), key=lambda item: item[0]))
        frame, error = fetch_plan_api_dataframe_direct(
            base_url,
            APS_PLAN_ENDPOINT,
            params_tuple,
            api_key,
            api_key_hash,
            source_updated_at,
        )
        return operation, frame, error

    with ThreadPoolExecutor(max_workers=len(EFFECTIVE_PRODUCTION_OPERATIONS)) as executor:
        futures = [executor.submit(fetch_operation, operation) for operation in EFFECTIVE_PRODUCTION_OPERATIONS]
        for future in as_completed(futures):
            operation, frame, error = future.result()
            if error:
                errors.append(f"APS oper={operation}: {error}")
            elif not frame.empty:
                frames.append(frame)

    if frames:
        return pd.concat(frames, ignore_index=True, sort=False), ""
    return pd.DataFrame(), "; ".join(errors)


def fetch_effective_production_performance_chunk(
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    api_key_hash: str,
    source_updated_at: str,
) -> tuple[pd.DataFrame, str]:
    api_key = get_plan_api_key()
    date_from_param = get_streamlit_or_env_secret("PRODUCTION_API_DATE_FROM_PARAM", "date_from")
    date_to_param = get_streamlit_or_env_secret("PRODUCTION_API_DATE_TO_PARAM", "date_to")
    params: dict[str, object] = {"limit": PLAN_API_DEFAULT_ROW_LIMIT}
    if date_from_param:
        params[date_from_param] = start_date.strftime("%Y-%m-%d")
    if date_to_param:
        params[date_to_param] = end_date.strftime("%Y-%m-%d")

    frame, error = fetch_plan_api_dataframe_direct(
        get_plan_api_base_url(),
        PRODUCTION_PERFORMANCE_ENDPOINT,
        tuple(sorted(params.items(), key=lambda item: item[0])),
        api_key,
        api_key_hash,
        source_updated_at,
    )
    if error and "truncated" in error.lower() and start_date < end_date:
        midpoint = start_date + pd.Timedelta(days=(end_date - start_date).days // 2)
        left, left_error = fetch_effective_production_performance_chunk(
            start_date,
            midpoint,
            api_key_hash,
            source_updated_at,
        )
        right, right_error = fetch_effective_production_performance_chunk(
            midpoint + pd.Timedelta(days=1),
            end_date,
            api_key_hash,
            source_updated_at,
        )
        errors = "; ".join(error_text for error_text in (left_error, right_error) if error_text)
        frames = [part for part in (left, right) if not part.empty]
        return (pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()), errors
    return frame, error


def fetch_effective_production_performance_dataframe(
    start_date_text: str,
    end_date_text: str,
    api_key_hash: str,
    source_updated_at: str,
) -> tuple[pd.DataFrame, str]:
    start_date = pd.to_datetime(start_date_text, errors="coerce")
    end_date = pd.to_datetime(end_date_text, errors="coerce")
    if pd.isna(start_date) or pd.isna(end_date):
        return pd.DataFrame(), "생산실적 API 조회 기간이 올바르지 않습니다."

    frames: list[pd.DataFrame] = []
    errors: list[str] = []
    current = start_date.normalize()
    end_date = end_date.normalize()
    while current <= end_date:
        chunk_end = min(current + pd.Timedelta(days=6), end_date)
        frame, error = fetch_effective_production_performance_chunk(
            current,
            chunk_end,
            api_key_hash,
            source_updated_at,
        )
        if error:
            errors.append(f"{current:%Y-%m-%d}~{chunk_end:%Y-%m-%d}: {error}")
        elif not frame.empty:
            frames.append(frame)
        current = chunk_end + pd.Timedelta(days=1)

    if frames:
        return pd.concat(frames, ignore_index=True, sort=False).drop_duplicates(), "; ".join(errors)
    return pd.DataFrame(), "; ".join(errors)


def normalize_effective_plan_frame(
    raw: pd.DataFrame,
    start_date_text: str,
    end_date_text: str,
    site_filter: str,
) -> tuple[pd.DataFrame, str]:
    output_columns = [
        "일자",
        "공정",
        "제품코드",
        "제품명",
        "거래처",
        "이니셜",
        "수요수량",
        "필요수량",
        "_계획존재",
    ]
    if raw.empty:
        return pd.DataFrame(columns=output_columns), ""

    raw = raw.copy()
    raw.columns = [str(col).strip() for col in raw.columns]
    columns = raw.columns.tolist()
    date_col = pick_api_column(
        columns,
        [
            "plan_date",
            "PLAN_DATE",
            "target_datetime",
            "TARGET_DATETIME",
            "target_date",
            "TARGET_DATE",
            "production_date",
            "PRODUCTION_DATE",
            "work_date",
            "WORK_DATE",
            "due_date",
            "DUE_DATE",
            "납기일",
        ],
    )
    site_col = pick_api_column(columns, ["res_site_id", "RES_SITE_ID", "사이트코드", "사이트", "SITE"])
    customer_col = pick_api_column(columns, ["cust_name", "CUST_NAME", "거래처", "거래처명", "CUSTOMER_NAME"])
    initial_col = pick_api_column(columns, ["initial", "INITIAL", "이니셜", "INITIAL_CODE"])
    demand_type_col = pick_api_column(columns, ["demand_type", "DEMAND_TYPE", "수요유형"])
    item_col = pick_api_column(
        columns,
        [
            "item_id",
            "ITEM_ID",
            "ITEM_CODE",
            "ITEM_CD",
            "품목코드",
            "제품코드",
            "생산코드",
            "demand_item_id",
            "DEMAND_ITEM_ID",
        ],
    )
    product_name_col = pick_api_column(
        columns,
        [
            "item_name",
            "ITEM_NAME",
            "item_name2",
            "ITEM_NAME2",
            "demand_item_name",
            "DEMAND_ITEM_NAME",
            "제품명",
            "품명",
        ],
    )
    demand_qty_col = pick_api_column(columns, ["demand_qty", "DEMAND_QTY", "수요수량", "오더수량", "ORDER_QTY"])
    operation_col = pick_api_column(columns, ["oper_id", "OPER_ID", "공정코드", "공정"])
    plan_qty_col = pick_api_column(columns, ["plan_qty", "PLAN_QTY", "생산수량", "계획수량", "필요수량"])

    missing = []
    if date_col is None:
        missing.append("날짜")
    if item_col is None:
        missing.append("제품코드")
    if operation_col is None:
        missing.append("공정")
    if plan_qty_col is None:
        missing.append("계획수량")
    if missing:
        return pd.DataFrame(columns=output_columns), f"APS API 응답에서 {', '.join(missing)} 컬럼을 찾지 못했습니다."

    initial = effective_column_series(raw, initial_col).astype(str).str.strip()
    demand_type = effective_column_series(raw, demand_type_col).astype(str).str.strip()
    missing_initial = initial.map(clean_text_value).eq("")
    initial = initial.where(~missing_initial, demand_type).map(clean_initial_value)
    initial = initial.where(initial.ne(""), "미지정")

    work = pd.DataFrame(
        {
            "일자": effective_date_text_series(raw, date_col),
            "사이트": effective_column_series(raw, site_col).map(normalize_effective_site),
            "공정": effective_column_series(raw, operation_col).map(normalize_effective_process),
            "제품코드": effective_column_series(raw, item_col).map(normalize_item_code_value),
            "제품명": effective_column_series(raw, product_name_col).map(clean_text_value),
            "거래처": effective_column_series(raw, customer_col).map(clean_text_value),
            "이니셜": initial,
            "수요수량": effective_numeric_series(raw, demand_qty_col),
            "필요수량": effective_numeric_series(raw, plan_qty_col),
        }
    )
    start_date = pd.to_datetime(start_date_text, errors="coerce")
    end_date = pd.to_datetime(end_date_text, errors="coerce")
    work_date = pd.to_datetime(work["일자"], errors="coerce")
    work = work[
        work["공정"].isin(EFFECTIVE_PRODUCTION_PROCESS_ORDER)
        & work["제품코드"].ne("")
        & work["필요수량"].gt(0)
        & work_date.between(start_date, end_date, inclusive="both")
    ].copy()
    if site_filter != "전체" and site_col is not None:
        work = work[work["사이트"].eq(normalize_effective_site(site_filter))].copy()
    if work.empty:
        return pd.DataFrame(columns=output_columns), ""

    demand = (
        work.groupby(["일자", "공정", "제품코드"], as_index=False)
        .agg(
            제품명=("제품명", first_nonempty_text),
            거래처=("거래처", join_unique_text_values),
            이니셜=("이니셜", join_unique_text_values),
            수요수량=("수요수량", "sum"),
            필요수량=("필요수량", "sum"),
        )
        .sort_values(["일자", "공정", "제품코드"])
    )
    demand["_계획존재"] = True
    return demand[output_columns], ""


def normalize_effective_production_frame(
    raw: pd.DataFrame,
    start_date_text: str,
    end_date_text: str,
    site_filter: str,
) -> tuple[pd.DataFrame, str]:
    output_columns = ["일자", "공정", "제품코드", "_생산제품명", "실적수량", "_실적존재"]
    if raw.empty:
        return pd.DataFrame(columns=output_columns), ""

    raw = raw.copy()
    raw.columns = [str(col).strip() for col in raw.columns]
    columns = raw.columns.tolist()
    date_col = pick_api_column(columns, ["pr_dt", "PR_DT", "생산일자", "기간", "date", "DATE", "work_date"])
    site_col = pick_api_column(columns, ["fac_nm", "FAC_NM", "fac_cd", "FAC_CD", "공장", "사이트", "SITE"])
    process_col = pick_api_column(columns, ["gong_cd", "GONG_CD", "공정코드", "공정", "PROCESS"])
    sku_col = pick_api_column(columns, ["gd_cd", "GD_CD", "품목코드", "제품코드", "ITEM_ID", "ITEM_CODE"])
    product_name_col = pick_api_column(columns, ["gd_nm", "GD_NM", "품명", "제품명", "ITEM_NAME"])
    actual_col = pick_api_column(
        columns,
        ["pr_qty", "PR_QTY", "샘플제외 양품수량", "총_양품수량", "실적수량", "생산수량", "GOOD_QTY"],
    )
    sample_col = pick_api_column(columns, ["sample_qty", "SAMPLE_QTY", "샘플수량"])

    missing = []
    if date_col is None:
        missing.append("생산일자")
    if site_col is None:
        missing.append("공장")
    if process_col is None:
        missing.append("공정")
    if sku_col is None:
        missing.append("제품코드")
    if actual_col is None:
        missing.append("실적수량")
    if missing:
        return pd.DataFrame(columns=output_columns), f"생산실적 API 응답에서 {', '.join(missing)} 컬럼을 찾지 못했습니다."

    actual_qty = effective_numeric_series(raw, actual_col)
    if sample_col is not None and normalize_api_column_key(actual_col) in {"prqty", "생산수량", "실적수량"}:
        actual_qty = (actual_qty - effective_numeric_series(raw, sample_col)).clip(lower=0)

    work = pd.DataFrame(
        {
            "일자": effective_date_text_series(raw, date_col),
            "사이트": effective_column_series(raw, site_col).map(normalize_effective_site),
            "공정": effective_column_series(raw, process_col).map(normalize_effective_process),
            "제품코드": effective_column_series(raw, sku_col).map(normalize_item_code_value),
            "_생산제품명": effective_column_series(raw, product_name_col).map(clean_text_value),
            "실적수량": actual_qty,
        }
    )
    start_date = pd.to_datetime(start_date_text, errors="coerce")
    end_date = pd.to_datetime(end_date_text, errors="coerce")
    work_date = pd.to_datetime(work["일자"], errors="coerce")
    work = work[
        work["공정"].isin(EFFECTIVE_PRODUCTION_PROCESS_ORDER)
        & work["제품코드"].ne("")
        & work["실적수량"].gt(0)
        & work_date.between(start_date, end_date, inclusive="both")
    ].copy()
    if site_filter != "전체":
        work = work[work["사이트"].eq(normalize_effective_site(site_filter))].copy()
    if work.empty:
        return pd.DataFrame(columns=output_columns), ""

    production = (
        work.groupby(["일자", "공정", "제품코드"], as_index=False)
        .agg(
            _생산제품명=("_생산제품명", first_nonempty_text),
            실적수량=("실적수량", "sum"),
        )
        .sort_values(["일자", "공정", "제품코드"])
    )
    production["_실적존재"] = True
    return production[output_columns], ""


def build_effective_production_detail(demand: pd.DataFrame, production: pd.DataFrame) -> pd.DataFrame:
    detail = demand.merge(production, on=["일자", "공정", "제품코드"], how="outer")
    detail["_계획존재"] = detail["_계획존재"].eq(True)
    detail["_실적존재"] = detail["_실적존재"].eq(True)
    detail["제품명"] = detail["제품명"].fillna("").astype(str)
    production_name = detail.get("_생산제품명", pd.Series("", index=detail.index)).fillna("").astype(str)
    missing_product_name = detail["제품명"].map(clean_text_value).eq("")
    detail.loc[missing_product_name, "제품명"] = production_name[missing_product_name]

    for column in ["거래처", "이니셜"]:
        detail[column] = detail[column].fillna("").astype(str)
    for column in ["수요수량", "필요수량", "실적수량"]:
        detail[column] = parse_mixed_numeric(detail.get(column, pd.Series(0, index=detail.index))).fillna(0)

    detail["유효생산량"] = detail[["필요수량", "실적수량"]].min(axis=1)
    detail["비유효생산량"] = (detail["실적수량"] - detail["유효생산량"]).clip(lower=0)
    detail["잔여필요수량"] = (detail["필요수량"] - detail["실적수량"]).clip(lower=0)
    detail["생산유효도(%)"] = calculate_effective_rate(detail["유효생산량"], detail["실적수량"])
    detail["매칭상태"] = "둘다없음"
    detail.loc[detail["_계획존재"] & detail["_실적존재"], "매칭상태"] = "정상매칭"
    detail.loc[detail["_계획존재"] & ~detail["_실적존재"], "매칭상태"] = "실적없음"
    detail.loc[~detail["_계획존재"] & detail["_실적존재"], "매칭상태"] = "계획없음"
    detail = detail[detail["매칭상태"].ne("둘다없음")].copy()
    detail["분류"] = detail.apply(classify_effective_major_category, axis=1)

    columns = [
        "일자",
        "공정",
        "분류",
        "제품코드",
        "제품명",
        "거래처",
        "이니셜",
        "수요수량",
        "필요수량",
        "실적수량",
        "유효생산량",
        "비유효생산량",
        "잔여필요수량",
        "생산유효도(%)",
        "매칭상태",
    ]
    return detail[columns].sort_values(["일자", "공정", "제품코드"]).reset_index(drop=True)


@st.cache_data(show_spinner=False, ttl=PLAN_API_CACHE_TTL_SECONDS, max_entries=CACHE_MAX_ENTRIES)
def load_effective_production_dashboard_data(
    start_date_text: str,
    end_date_text: str,
    site_filter: str,
    api_key_hash: str,
    source_updated_at: str,
    refresh_nonce: int,
) -> tuple[pd.DataFrame, dict[str, object], str]:
    _ = refresh_nonce
    raw_plan, plan_fetch_error = fetch_effective_plan_operations_dataframe(
        site_filter,
        start_date_text,
        end_date_text,
        api_key_hash,
        source_updated_at,
    )
    demand, plan_normalize_error = normalize_effective_plan_frame(
        raw_plan,
        start_date_text,
        end_date_text,
        site_filter,
    )
    valid_plan_dates = sorted(demand["일자"].dropna().astype(str).unique()) if not demand.empty else []
    production_start_date = valid_plan_dates[0] if valid_plan_dates else start_date_text
    production_end_date = valid_plan_dates[-1] if valid_plan_dates else end_date_text
    raw_production = pd.DataFrame()
    production_fetch_error = ""
    if valid_plan_dates:
        raw_production, production_fetch_error = fetch_effective_production_performance_dataframe(
            production_start_date,
            production_end_date,
            api_key_hash,
            source_updated_at,
        )
    production, production_normalize_error = normalize_effective_production_frame(
        raw_production,
        production_start_date,
        production_end_date,
        site_filter,
    )
    if valid_plan_dates:
        production = production[production["일자"].isin(valid_plan_dates)].copy()
    detail = build_effective_production_detail(demand, production)
    metadata = {
        "plan_raw_rows": len(raw_plan),
        "production_raw_rows": len(raw_production),
        "plan_rows": len(demand),
        "production_rows": len(production),
    }
    errors = "; ".join(
        error
        for error in (
            plan_fetch_error,
            production_fetch_error,
            plan_normalize_error,
            production_normalize_error,
        )
        if error
    )
    return detail, metadata, errors


def summarize_effective_total(frame: pd.DataFrame) -> dict[str, float]:
    if frame.empty:
        return {"need": 0.0, "actual": 0.0, "effective": 0.0, "ineffective": 0.0, "remaining": 0.0, "rate": 0.0}
    need = float(frame["필요수량"].sum())
    actual = float(frame["실적수량"].sum())
    effective = float(frame["유효생산량"].sum())
    ineffective = float(frame["비유효생산량"].sum())
    remaining = float(frame["잔여필요수량"].sum())
    rate = 0.0 if actual == 0 else effective / actual * 100
    return {
        "need": need,
        "actual": actual,
        "effective": effective,
        "ineffective": ineffective,
        "remaining": remaining,
        "rate": rate,
    }


def aggregate_effective_quantities(detail: pd.DataFrame, dimensions: list[str]) -> pd.DataFrame:
    output_columns = [
        *dimensions,
        "필요수량 합계",
        "실적수량 합계",
        "유효생산량 합계",
        "비유효생산량 합계",
        "잔여필요수량 합계",
        "생산 SKU수",
        "유효 SKU수",
        "생산유효도(%)",
    ]
    if detail.empty:
        return pd.DataFrame(columns=output_columns)

    summary = (
        detail.groupby(dimensions, as_index=False)
        .agg(
            **{
                "필요수량 합계": ("필요수량", "sum"),
                "실적수량 합계": ("실적수량", "sum"),
                "유효생산량 합계": ("유효생산량", "sum"),
                "비유효생산량 합계": ("비유효생산량", "sum"),
                "잔여필요수량 합계": ("잔여필요수량", "sum"),
                "생산 SKU수": ("제품코드", lambda values: values[detail.loc[values.index, "실적수량"].gt(0)].nunique()),
                "유효 SKU수": ("제품코드", lambda values: values[detail.loc[values.index, "유효생산량"].gt(0)].nunique()),
            }
        )
        .sort_values(dimensions)
    )
    summary["생산유효도(%)"] = calculate_effective_rate(summary["유효생산량 합계"], summary["실적수량 합계"])
    return summary[output_columns]


def build_effective_daily_trend(detail: pd.DataFrame) -> pd.DataFrame:
    overall = aggregate_effective_quantities(detail, ["일자"])
    if not overall.empty:
        overall["비교분류"] = "전체"
    category = aggregate_effective_quantities(detail, ["일자", "분류"])
    if not category.empty:
        category = category.rename(columns={"분류": "비교분류"})
    trend = pd.concat([overall, category], ignore_index=True, sort=False)
    if trend.empty:
        return trend
    trend["일자_dt"] = pd.to_datetime(trend["일자"], errors="coerce")
    return trend.sort_values(["일자_dt", "비교분류"]).reset_index(drop=True)


def build_effective_trend_figure(trend: pd.DataFrame) -> go.Figure:
    colors = {
        "전체": "#111827",
        "국내": "#2563eb",
        "PIA": "#f97316",
        "기타해외": "#0f766e",
    }
    fig = go.Figure()
    if trend.empty:
        fig.add_annotation(text="표시할 데이터가 없습니다.", showarrow=False, x=0.5, y=0.5)
    else:
        for label in ("전체", *EFFECTIVE_PRODUCTION_CATEGORY_ORDER):
            rows = trend[trend["비교분류"].eq(label)]
            if rows.empty:
                continue
            fig.add_trace(
                go.Scatter(
                    x=rows["일자"],
                    y=rows["생산유효도(%)"],
                    mode="lines+markers",
                    name=label,
                    line={"color": colors.get(label, "#64748b"), "width": 2},
                    marker={"size": 7},
                    customdata=rows[["실적수량 합계", "유효생산량 합계", "비유효생산량 합계"]].to_numpy(),
                    hovertemplate=(
                        "%{x}<br>%{fullData.name} 생산유효도 %{y:.1f}%<br>"
                        "실적 %{customdata[0]:,.0f}<br>"
                        "유효 %{customdata[1]:,.0f}<br>"
                        "비유효 %{customdata[2]:,.0f}<extra></extra>"
                    ),
                )
            )
    fig.update_layout(
        height=420,
        margin={"l": 10, "r": 10, "t": 24, "b": 10},
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        hovermode="x unified",
    )
    fig.update_xaxes(title_text="", type="category")
    fig.update_yaxes(title_text="생산유효도", ticksuffix="%", range=[0, 105])
    return fig


def render_effective_metric(label: str, frame: pd.DataFrame) -> None:
    summary = summarize_effective_total(frame)
    st.metric(label, format_effective_pct(summary["rate"]))
    st.caption(
        f"실적 {format_effective_int(summary['actual'])} / "
        f"유효 {format_effective_int(summary['effective'])} / "
        f"비유효 {format_effective_int(summary['ineffective'])}"
    )


def prepare_effective_detail_table(detail: pd.DataFrame) -> pd.DataFrame:
    if detail.empty:
        return detail
    table = detail.copy()
    table["일자"] = pd.to_datetime(table["일자"], errors="coerce").dt.strftime("%Y-%m-%d").fillna(table["일자"])
    return table.sort_values(
        ["일자", "비유효생산량", "잔여필요수량", "실적수량"],
        ascending=[False, False, False, False],
    )


def filter_effective_detail_table(
    detail: pd.DataFrame,
    process_filter: list[str],
    status_filter: list[str],
    query: str,
) -> pd.DataFrame:
    filtered = detail.copy()
    if process_filter:
        filtered = filtered[filtered["공정"].isin(process_filter)]
    if status_filter:
        filtered = filtered[filtered["매칭상태"].isin(status_filter)]
    query_text = query.strip()
    if query_text:
        search_columns = ["제품코드", "제품명", "거래처", "이니셜", "분류", "매칭상태"]
        haystack = filtered[search_columns].fillna("").astype(str).agg(" ".join, axis=1)
        for term in query_text.split():
            haystack = filtered[search_columns].fillna("").astype(str).agg(" ".join, axis=1)
            filtered = filtered[haystack.str.contains(re.escape(term), case=False, na=False)]
    return filtered


def build_effective_source_signature(source_dir: Path) -> tuple[tuple[str, str, int, int], ...]:
    if effective_report is None or not source_dir.exists():
        return tuple()
    source_paths = [path for _date, path in effective_report.find_input_files(source_dir)]
    source_paths.append(source_dir / effective_report.CLASSIFICATION_FILE)
    source_paths.append(source_dir / EFFECTIVE_PRODUCTION_DOI_CRITERIA_FILE)
    signature: list[tuple[str, str, int, int]] = []
    for path in source_paths:
        if not path.exists():
            continue
        stat = path.stat()
        signature.append((path.name, str(path), stat.st_size, stat.st_mtime_ns))
    return tuple(sorted(signature))


def restore_environment_values(previous_values: dict[str, str | None]) -> None:
    for name, value in previous_values.items():
        if value is None:
            os.environ.pop(name, None)
        else:
            os.environ[name] = value


def load_effective_report_with_main_api_settings(
    source_dir: Path,
    valid_dates: set[str],
    base_url: str,
) -> pd.DataFrame:
    if effective_report is None:
        raise ValueError("생산유효도 계산 모듈을 찾지 못했습니다.")

    env_updates = {
        effective_report.PRODUCTION_API_KEY_ENV: get_plan_api_key(),
        effective_report.PRODUCTION_API_URL_ENV: f"{base_url.rstrip('/')}/{PRODUCTION_PERFORMANCE_ENDPOINT.lstrip('/')}",
        effective_report.PRODUCTION_API_SAVE_RAW_ENV: "0",
    }
    previous_values = {name: os.environ.get(name) for name in env_updates}
    try:
        for name, value in env_updates.items():
            if value:
                os.environ[name] = value
        return effective_report.load_production_api_data(valid_dates, None)
    finally:
        restore_environment_values(previous_values)


@st.cache_data(show_spinner=False, ttl=PLAN_API_CACHE_TTL_SECONDS, max_entries=CACHE_MAX_ENTRIES)
def load_original_effective_dashboard_data(
    source_dir_text: str,
    display_cutoff_date: str,
    source_signature: tuple[tuple[str, str, int, int], ...],
    base_url: str,
    api_key_hash: str,
    refresh_nonce: int,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    _ = source_signature, api_key_hash, refresh_nonce
    if effective_report is None:
        raise ValueError("생산유효도 계산 모듈을 찾지 못했습니다.")

    source_dir = Path(source_dir_text)
    input_files = [
        (date, path)
        for date, path in effective_report.find_input_files(source_dir)
        if str(date) <= str(display_cutoff_date)
    ]
    if not input_files:
        raise ValueError(f"{format_effective_date_option(display_cutoff_date)}까지 표시할 수요정보 파일이 없습니다.")

    input_dates = {date for date, _path in input_files}
    production_df = load_effective_report_with_main_api_settings(source_dir, input_dates, base_url)
    if production_df.empty:
        raise ValueError("생산실적 API에서 조회된 데이터가 없습니다.")

    match_on_product_name = effective_report.production_uses_product_name_key_from_df("production API", production_df)
    demand = effective_report.build_demand(input_files, match_on_product_name)
    valid_dates = set(demand[effective_report.DATE_COL].astype(str).unique())
    production = effective_report.prepare_production_data(production_df, "production API", valid_dates)
    detail = effective_report.build_detail(demand, production)
    detail = effective_report.apply_classification(detail, source_dir)
    summary = effective_report.summarize_by_process(detail)
    change_analysis = effective_report.build_change_analysis(summary, detail)
    sheet_name_analysis = effective_report.attach_major_category(
        effective_report.summarize_by_classification(detail, effective_report.SHEET_NAME_COL),
        detail,
    )
    major_category_analysis = effective_report.summarize_by_classification(detail, effective_report.MAJOR_CATEGORY_COL)
    metadata = {
        "source": "기존 생산유효도 계산 모듈",
        "demand_start": min(input_dates),
        "demand_end": max(input_dates),
        "demand_file_count": len(input_files),
        "display_cutoff_date": display_cutoff_date,
        "raw_rows": len(production_df),
        "production_rows": len(production),
        "detail_rows": len(detail),
        "match_key": "제품명" if match_on_product_name else "제품코드 전체 문자열",
        "loaded_at": datetime.now(DISPLAY_TZ).strftime("%Y-%m-%d %H:%M:%S"),
    }
    return summary, change_analysis, sheet_name_analysis, major_category_analysis, detail, metadata


def format_effective_date_option(value: object) -> str:
    text = str(value)
    parsed = pd.to_datetime(text, format="%Y%m%d", errors="coerce")
    return text if pd.isna(parsed) else parsed.strftime("%Y-%m-%d")


def format_effective_month_label(month_key: str) -> str:
    if len(month_key) == 6 and month_key.isdigit():
        return f"{int(month_key[:4])}년 {int(month_key[4:])}월"
    return month_key


def format_effective_month_short_label(month_key: str) -> str:
    if len(month_key) == 6 and month_key.isdigit():
        return f"{int(month_key[4:])}월 종합"
    return f"{month_key} 종합"


def build_effective_display_cutoff_date() -> str:
    return (pd.Timestamp.now(tz=DISPLAY_TZ).normalize() - pd.Timedelta(days=1)).strftime("%Y%m%d")


def effective_report_rate(frame: pd.DataFrame, numerator_col: str, denominator_col: str) -> float:
    if frame.empty:
        return 0.0
    numerator = float(parse_mixed_numeric(frame[numerator_col]).sum())
    denominator = float(parse_mixed_numeric(frame[denominator_col]).sum())
    return 0.0 if denominator == 0 else numerator / denominator * 100


def filter_effective_report_month(frame: pd.DataFrame, date_col: str, month_key: str) -> pd.DataFrame:
    if frame.empty or date_col not in frame.columns:
        return pd.DataFrame()
    return frame[frame[date_col].astype(str).str.startswith(month_key)].copy()


def build_original_effective_daily_trend(summary: pd.DataFrame, major_category: pd.DataFrame) -> pd.DataFrame:
    pieces: list[pd.DataFrame] = []
    actual_col = effective_report.SUMMARY_ACTUAL_COL
    effective_col = effective_report.SUMMARY_EFFECTIVE_COL
    if not summary.empty:
        overall = summary.groupby(effective_report.DATE_COL, as_index=False).agg(
            **{
                actual_col: (actual_col, "sum"),
                effective_col: (effective_col, "sum"),
            }
        )
        overall["비교분류"] = "전체"
        pieces.append(overall)
    if not major_category.empty and effective_report.MAJOR_CATEGORY_COL in major_category.columns:
        categories = major_category[major_category[effective_report.MAJOR_CATEGORY_COL].isin(EFFECTIVE_PRODUCTION_CATEGORY_ORDER)]
        if not categories.empty:
            categories = categories.groupby(
                [effective_report.DATE_COL, effective_report.MAJOR_CATEGORY_COL],
                as_index=False,
            ).agg(
                **{
                    actual_col: (actual_col, "sum"),
                    effective_col: (effective_col, "sum"),
                }
            )
            categories = categories.rename(columns={effective_report.MAJOR_CATEGORY_COL: "비교분류"})
            pieces.append(categories)
    if not pieces:
        return pd.DataFrame()
    trend = pd.concat(pieces, ignore_index=True, sort=False)
    trend = trend[trend[actual_col].gt(0)].copy()
    trend[effective_report.PRODUCTION_EFFECTIVENESS_COL] = calculate_effective_rate(trend[effective_col], trend[actual_col])
    trend["표시일자"] = trend[effective_report.DATE_COL].map(format_effective_date_option)
    return trend.sort_values([effective_report.DATE_COL, "비교분류"])


def build_original_effective_trend_figure(summary: pd.DataFrame, major_category: pd.DataFrame) -> go.Figure:
    trend = build_original_effective_daily_trend(summary, major_category)
    fig = go.Figure()
    colors = {
        "전체": "#111827",
        "국내": "#2563eb",
        "PIA": "#f97316",
        "기타해외": "#0f766e",
    }
    if trend.empty:
        fig.add_annotation(text="표시할 데이터가 없습니다.", showarrow=False, x=0.5, y=0.5)
    else:
        for label in ("전체", *EFFECTIVE_PRODUCTION_CATEGORY_ORDER):
            rows = trend[trend["비교분류"].eq(label)]
            if rows.empty:
                continue
            fig.add_trace(
                go.Scatter(
                    x=rows["표시일자"],
                    y=rows[effective_report.PRODUCTION_EFFECTIVENESS_COL],
                    mode="lines+markers",
                    name=label,
                    line={"color": colors.get(label, "#64748b"), "width": 4 if label == "전체" else 2.5},
                    marker={"size": 8},
                    customdata=rows[
                        [
                            effective_report.DATE_COL,
                            effective_report.SUMMARY_ACTUAL_COL,
                            effective_report.SUMMARY_EFFECTIVE_COL,
                        ]
                    ].to_numpy(),
                    hovertemplate=(
                        "%{fullData.name}<br>"
                        "일자 %{customdata[0]}<br>"
                        "생산유효도 %{y:.1f}%<br>"
                        "실적 %{customdata[1]:,.0f}<br>"
                        "유효 %{customdata[2]:,.0f}<extra></extra>"
                    ),
                )
            )
    fig.update_layout(
        height=390,
        margin={"l": 10, "r": 90, "t": 24, "b": 10},
        legend_title_text="",
        hovermode="x unified",
    )
    fig.update_xaxes(title_text="일자", type="category")
    fig.update_yaxes(title_text=effective_report.PRODUCTION_EFFECTIVENESS_COL, ticksuffix="%", range=[0, 100])
    return fig


def prepare_original_effective_detail_table(detail: pd.DataFrame, month_key: str) -> pd.DataFrame:
    table = filter_effective_report_month(detail, effective_report.DATE_COL, month_key)
    if table.empty:
        return table
    table = table.copy()
    table[effective_report.DATE_COL] = table[effective_report.DATE_COL].map(format_effective_date_option)
    preferred_columns = [
        effective_report.DATE_COL,
        effective_report.OUTPUT_PROCESS_COL,
        effective_report.MAJOR_CATEGORY_COL,
        effective_report.SHEET_NAME_COL,
        effective_report.OUTPUT_SKU_COL,
        effective_report.OUTPUT_PRODUCT_NAME_COL,
        effective_report.CUSTOMER_COL,
        effective_report.INITIAL_COL,
        effective_report.NEED_QTY_COL,
        effective_report.ACTUAL_QTY_COL,
        effective_report.EFFECTIVE_PRODUCTION_COL,
        effective_report.INEFFECTIVE_PRODUCTION_COL,
        effective_report.REMAINING_NEED_COL,
        effective_report.PRODUCTION_EFFECTIVENESS_COL,
        effective_report.MATCH_STATUS_COL,
    ]
    table = table[[column for column in preferred_columns if column in table.columns]].copy()
    sort_columns = [column for column in [effective_report.DATE_COL, effective_report.INEFFECTIVE_PRODUCTION_COL] if column in table.columns]
    if sort_columns:
        table = table.sort_values(sort_columns, ascending=[False] * len(sort_columns))
    return table


def build_top_production_products_table(detail: pd.DataFrame, top_n: int = 15) -> pd.DataFrame:
    required_columns = [
        effective_report.OUTPUT_SKU_COL,
        effective_report.OUTPUT_PRODUCT_NAME_COL,
        effective_report.OUTPUT_PROCESS_COL,
        effective_report.ACTUAL_QTY_COL,
        effective_report.EFFECTIVE_PRODUCTION_COL,
        effective_report.INEFFECTIVE_PRODUCTION_COL,
    ]
    if detail.empty or any(column not in detail.columns for column in required_columns):
        return pd.DataFrame()

    scope = detail.copy()
    for column in [
        effective_report.ACTUAL_QTY_COL,
        effective_report.EFFECTIVE_PRODUCTION_COL,
        effective_report.INEFFECTIVE_PRODUCTION_COL,
    ]:
        scope[column] = parse_mixed_numeric(scope[column])
    scope = scope[scope[effective_report.ACTUAL_QTY_COL].gt(0)].copy()
    if scope.empty:
        return pd.DataFrame()

    group_columns = [effective_report.OUTPUT_SKU_COL, effective_report.OUTPUT_PRODUCT_NAME_COL]
    optional_columns = [
        effective_report.MAJOR_CATEGORY_COL,
        effective_report.SHEET_NAME_COL,
    ]
    for column in optional_columns:
        if column in scope.columns:
            group_columns.append(column)

    summary = (
        scope.groupby(group_columns, as_index=False)
        .agg(
            **{
                "생산량": (effective_report.ACTUAL_QTY_COL, "sum"),
                "유효생산량": (effective_report.EFFECTIVE_PRODUCTION_COL, "sum"),
                "비유효생산량": (effective_report.INEFFECTIVE_PRODUCTION_COL, "sum"),
                "생산공정": (effective_report.OUTPUT_PROCESS_COL, join_unique_text_values),
            }
        )
        .sort_values("생산량", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )
    summary["순위"] = range(1, len(summary) + 1)
    summary["유효비중(%)"] = calculate_effective_rate(summary["유효생산량"], summary["생산량"])
    return summary[
        [
            "순위",
            effective_report.OUTPUT_SKU_COL,
            effective_report.OUTPUT_PRODUCT_NAME_COL,
            *[column for column in optional_columns if column in summary.columns],
            "생산공정",
            "생산량",
            "유효생산량",
            "비유효생산량",
            "유효비중(%)",
        ]
    ]


def build_top_production_products_figure(detail: pd.DataFrame, top_n: int = 12) -> go.Figure:
    top_table = build_top_production_products_table(detail, top_n)
    fig = go.Figure()
    if top_table.empty:
        fig.add_annotation(text="표시할 생산량 데이터가 없습니다.", showarrow=False, x=0.5, y=0.5)
        fig.update_layout(height=420, margin={"l": 10, "r": 10, "t": 20, "b": 10})
        return fig

    scope = detail.copy()
    scope[effective_report.ACTUAL_QTY_COL] = parse_mixed_numeric(scope[effective_report.ACTUAL_QTY_COL])
    top_codes = top_table[effective_report.OUTPUT_SKU_COL].astype(str).tolist()
    scope = scope[
        scope[effective_report.OUTPUT_SKU_COL].astype(str).isin(top_codes)
        & scope[effective_report.ACTUAL_QTY_COL].gt(0)
    ].copy()

    label_by_code: dict[str, str] = {}
    for _, row in top_table.iterrows():
        code = str(row[effective_report.OUTPUT_SKU_COL])
        name = clean_text_value(row.get(effective_report.OUTPUT_PRODUCT_NAME_COL, ""))
        short_name = name[:22] + "..." if len(name) > 22 else name
        label_by_code[code] = f"{int(row['순위'])}. {code}" + (f" · {short_name}" if short_name else "")
    ordered_labels = [label_by_code[code] for code in top_codes][::-1]

    process_colors = {
        "[10]사출조립": "#2563eb",
        "[80]누수/규격검사": "#f97316",
    }
    for process in effective_report.TARGET_PROCESSES:
        process_rows = scope[scope[effective_report.OUTPUT_PROCESS_COL].eq(process)].copy()
        if process_rows.empty:
            continue
        grouped = (
            process_rows.groupby(effective_report.OUTPUT_SKU_COL, as_index=False)[effective_report.ACTUAL_QTY_COL]
            .sum()
        )
        grouped["표시품목"] = grouped[effective_report.OUTPUT_SKU_COL].astype(str).map(label_by_code)
        grouped = grouped[grouped["표시품목"].isin(ordered_labels)].copy()
        grouped["표시품목"] = pd.Categorical(grouped["표시품목"], categories=ordered_labels, ordered=True)
        grouped = grouped.sort_values("표시품목")
        fig.add_trace(
            go.Bar(
                x=grouped[effective_report.ACTUAL_QTY_COL],
                y=grouped["표시품목"],
                orientation="h",
                name=process,
                marker={"color": process_colors.get(process, "#64748b")},
                customdata=grouped[[effective_report.OUTPUT_SKU_COL]].to_numpy(),
                hovertemplate="%{y}<br>%{fullData.name}<br>생산량 %{x:,.0f}<extra></extra>",
            )
        )

    fig.update_layout(
        barmode="stack",
        height=max(420, 32 * len(ordered_labels) + 120),
        margin={"l": 12, "r": 24, "t": 20, "b": 36},
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        hovermode="closest",
    )
    fig.update_xaxes(title_text="실적수량", rangemode="tozero", tickformat=",")
    fig.update_yaxes(title_text="", categoryorder="array", categoryarray=ordered_labels)
    return fig


def normalize_effective_doi_reference(source: pd.DataFrame, site_filter: str = EFFECTIVE_PRODUCTION_DEFAULT_SITE) -> pd.DataFrame:
    output_columns = [
        "제품코드",
        "제품명_DOI",
        "거래처그룹_DOI",
        "거래처_DOI",
        "이니셜목록_DOI",
        "현재DOI",
        "기준DOI하한",
        "기준DOI상한",
        "DOI상태",
        "기준등급",
        "DOI판단상태",
        "DOI판단메모",
        "신제품여부",
        "완제품재고",
        "공정재고합계",
        "DOI기준오더",
        "DOI기준오더수량",
        "DOI기준부족수량",
        "DOI기준사출부족수량",
        "DOI우선순위점수",
    ]
    if source.empty:
        return pd.DataFrame(columns=output_columns)

    work = source.copy()
    work.columns = [str(column).strip() for column in work.columns]
    if "사이트코드" in work.columns:
        scoped_site = clean_text_value(site_filter)
        if scoped_site and scoped_site != "전체":
            work = work[work["사이트코드"].map(normalize_site_group).eq(scoped_site)].copy()
    if work.empty:
        return pd.DataFrame(columns=output_columns)

    columns = work.columns.tolist()
    code_col = pick_first_existing_column(columns, ["생산코드", "제품코드", "품목코드"])
    if code_col is None:
        return pd.DataFrame(columns=output_columns)

    result = pd.DataFrame({"제품코드": work[code_col].map(normalize_item_code_value)})
    result = result[result["제품코드"].str.startswith("P", na=False)].copy()
    if result.empty:
        return pd.DataFrame(columns=output_columns)

    text_candidates = {
        "제품명_DOI": ["제품명", "제품명_DOI"],
        "거래처그룹_DOI": ["거래처그룹", "거래처그룹_DOI"],
        "거래처_DOI": ["거래처", "거래처_DOI"],
        "이니셜목록_DOI": ["이니셜목록", "이니셜", "이니셜목록_DOI"],
        "DOI상태": ["DOI상태", "DOI 상태"],
        "기준등급": ["기준등급", "품목등급"],
        "DOI판단상태": ["상태", "DOI판단상태", "판단"],
        "DOI판단메모": ["판단메모", "DOI판단메모", "재고대응판단"],
        "신제품여부": ["신제품여부"],
    }
    numeric_candidates = {
        "현재DOI": ["현재DOI", "DOI", "DOI_가중", "DOI_중앙값"],
        "기준DOI하한": ["기준DOI하한"],
        "기준DOI상한": ["기준DOI상한", "DOI_상한"],
        "완제품재고": ["완제품재고"],
        "공정재고합계": ["공정재고합계"],
        "DOI기준오더": ["DOI기준오더"],
        "DOI기준오더수량": ["오더수량", "DOI기준오더수량"],
        "DOI기준부족수량": ["생산부족수량", "부족수량", "DOI기준부족수량"],
        "DOI기준사출부족수량": ["사출부족수량", "DOI기준사출부족수량"],
        "DOI우선순위점수": ["우선순위점수"],
    }

    for output_col, candidates in text_candidates.items():
        source_col = pick_first_existing_column(columns, candidates)
        result[output_col] = (
            work.loc[result.index, source_col].map(clean_text_value)
            if source_col is not None
            else ""
        )
    for output_col, candidates in numeric_candidates.items():
        source_col = pick_first_existing_column(columns, candidates)
        result[output_col] = (
            parse_mixed_numeric(work.loc[result.index, source_col])
            if source_col is not None
            else 0.0
        )

    grouped = (
        result[output_columns]
        .groupby("제품코드", as_index=False)
        .agg(
            제품명_DOI=("제품명_DOI", lambda s: summarize_unique(s, head_count=1)),
            거래처그룹_DOI=("거래처그룹_DOI", lambda s: summarize_unique(s, head_count=1)),
            거래처_DOI=("거래처_DOI", lambda s: summarize_unique(s, head_count=1)),
            이니셜목록_DOI=("이니셜목록_DOI", lambda s: summarize_unique(s, head_count=1)),
            현재DOI=("현재DOI", "max"),
            기준DOI하한=("기준DOI하한", "max"),
            기준DOI상한=("기준DOI상한", "max"),
            DOI상태=("DOI상태", lambda s: summarize_unique(s, head_count=1)),
            기준등급=("기준등급", lambda s: summarize_unique(s, head_count=1)),
            DOI판단상태=("DOI판단상태", lambda s: summarize_unique(s, head_count=1)),
            DOI판단메모=("DOI판단메모", lambda s: summarize_unique(s, head_count=1)),
            신제품여부=("신제품여부", lambda s: summarize_unique(s, head_count=1)),
            완제품재고=("완제품재고", "max"),
            공정재고합계=("공정재고합계", "max"),
            DOI기준오더=("DOI기준오더", "max"),
            DOI기준오더수량=("DOI기준오더수량", "sum"),
            DOI기준부족수량=("DOI기준부족수량", "sum"),
            DOI기준사출부족수량=("DOI기준사출부족수량", "sum"),
            DOI우선순위점수=("DOI우선순위점수", "max"),
        )
    )
    return grouped[output_columns]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def load_effective_doi_reference(
    source_dir_str: str,
    source_signature: tuple[tuple[str, str, int, int], ...],
) -> tuple[pd.DataFrame, str]:
    _ = source_signature
    source_dir = Path(source_dir_str)
    criteria_path = source_dir / EFFECTIVE_PRODUCTION_DOI_CRITERIA_FILE
    if criteria_path.exists():
        try:
            criteria = pd.read_csv(criteria_path, encoding="utf-8-sig")
        except Exception:
            criteria = pd.DataFrame()
        doi_reference = normalize_effective_doi_reference(criteria, EFFECTIVE_PRODUCTION_DEFAULT_SITE)
        if not doi_reference.empty:
            return doi_reference, f"{EFFECTIVE_PRODUCTION_DEFAULT_SITE} DOI 기준: {criteria_path.name}"

    try:
        all_items, _code_mismatch = load_cloud_all_item_status_snapshot()
    except Exception:
        all_items = pd.DataFrame()
    doi_reference = normalize_effective_doi_reference(all_items, EFFECTIVE_PRODUCTION_DEFAULT_SITE)
    if not doi_reference.empty:
        return doi_reference, f"{EFFECTIVE_PRODUCTION_DEFAULT_SITE} 전체 품목 스냅샷 DOI"
    return doi_reference, "DOI 기준 없음"


def build_effective_sample_available_signature(source_dir: Path) -> tuple[tuple[str, str, int, int], ...]:
    sample_path = source_dir / EFFECTIVE_SAMPLE_AVAILABLE_REFERENCE_FILE
    if not sample_path.exists():
        return (("effective-sample", str(sample_path), -1, 0),)
    stat = sample_path.stat()
    return (("effective-sample", str(sample_path), stat.st_size, stat.st_mtime_ns),)


def normalize_effective_sample_available_reference(source: pd.DataFrame, source_label: str) -> pd.DataFrame:
    output_columns = ["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"]
    if source.empty:
        return pd.DataFrame(columns=output_columns)

    work = source.copy()
    work.columns = [str(column).strip() for column in work.columns]
    code_col = pick_first_existing_column(work.columns.tolist(), ["제품코드", "생산코드", "품목코드"])
    sample_col = pick_first_existing_column(
        work.columns.tolist(),
        [EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플신청가능수량", "샘플가능수량"],
    )
    if code_col is None or sample_col is None:
        return pd.DataFrame(columns=output_columns)

    result = pd.DataFrame(
        {
            "제품코드": work[code_col].map(normalize_item_code_value),
            EFFECTIVE_SAMPLE_AVAILABLE_COL: parse_mixed_numeric(work[sample_col]).fillna(0),
        }
    )
    result = result[
        result["제품코드"].str.startswith("P", na=False)
        & result[EFFECTIVE_SAMPLE_AVAILABLE_COL].gt(0)
    ].copy()
    if result.empty:
        return pd.DataFrame(columns=output_columns)

    result = (
        result.groupby("제품코드", as_index=False)[EFFECTIVE_SAMPLE_AVAILABLE_COL]
        .max()
        .sort_values("제품코드")
        .reset_index(drop=True)
    )
    result["샘플수량출처"] = source_label
    return result[output_columns]


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def load_effective_sample_available_reference_file(
    source_dir_str: str,
    sample_signature: tuple[tuple[str, str, int, int], ...],
) -> pd.DataFrame:
    _ = sample_signature
    sample_path = Path(source_dir_str) / EFFECTIVE_SAMPLE_AVAILABLE_REFERENCE_FILE
    if not sample_path.exists():
        return pd.DataFrame(columns=["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"])
    try:
        source = pd.read_csv(sample_path, encoding="utf-8-sig")
    except Exception:
        return pd.DataFrame(columns=["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"])
    return normalize_effective_sample_available_reference(source, "전체품목 샘플신청가능수량")


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def load_effective_sample_available_master_reference(master_path_str: str, master_signature: tuple[str, int, int]) -> pd.DataFrame:
    _ = master_signature
    master_path = Path(master_path_str)
    if not master_path.exists():
        return pd.DataFrame(columns=["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"])
    try:
        source = pd.read_excel(
            master_path,
            sheet_name=ALL_ITEM_MASTER_SHEET,
            usecols=lambda column: str(column).strip() in {"품목코드", "샘플가능수량"},
        )
    except Exception:
        return pd.DataFrame(columns=["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"])
    return normalize_effective_sample_available_reference(source, "전체품목 샘플신청가능수량")


@st.cache_data(show_spinner=False, max_entries=CACHE_MAX_ENTRIES)
def load_effective_sample_available_reference(
    source_dir_str: str,
    sample_signature: tuple[tuple[str, str, int, int], ...],
    base_dir_str: str,
    plan_sample_refresh_key: str,
) -> pd.DataFrame:
    _ = sample_signature, plan_sample_refresh_key
    output_columns = ["제품코드", EFFECTIVE_SAMPLE_AVAILABLE_COL, "샘플수량출처"]
    references: list[pd.DataFrame] = []

    file_reference = load_effective_sample_available_reference_file(source_dir_str, sample_signature)
    if not file_reference.empty:
        references.append(file_reference)

    if not references and is_plan_api_enabled():
        demand = load_api_demand_like_df(EFFECTIVE_PRODUCTION_DEFAULT_SITE)
        if not demand.empty and EFFECTIVE_SAMPLE_AVAILABLE_COL in demand.columns:
            work = demand.copy()
            work["제품코드"] = work.apply(resolve_light_production_code, axis=1)
            work[EFFECTIVE_SAMPLE_AVAILABLE_COL] = parse_mixed_numeric(work[EFFECTIVE_SAMPLE_AVAILABLE_COL]).fillna(0)
            work = work[work["제품코드"].str.startswith("P", na=False)].copy()
            if not work.empty and work[EFFECTIVE_SAMPLE_AVAILABLE_COL].gt(0).any():
                api_reference = (
                    work.groupby("제품코드", as_index=False)[EFFECTIVE_SAMPLE_AVAILABLE_COL]
                    .max()
                    .sort_values("제품코드")
                    .reset_index(drop=True)
                )
                api_reference = api_reference[api_reference[EFFECTIVE_SAMPLE_AVAILABLE_COL].gt(0)].copy()
                api_reference["샘플수량출처"] = "APS 샘플신청가능수량"
                references.append(api_reference)

    if not references:
        master_path = find_all_item_master_file(Path(base_dir_str))
        if master_path is not None:
            stat = master_path.stat()
            master_reference = load_effective_sample_available_master_reference(
                str(master_path),
                (master_path.name, stat.st_size, stat.st_mtime_ns),
            )
            if not master_reference.empty:
                references.append(master_reference)

    if not references:
        return pd.DataFrame(columns=output_columns)

    combined = pd.concat(references, ignore_index=True, sort=False)
    combined["출처우선순위"] = combined["샘플수량출처"].eq("APS 샘플신청가능수량").map({True: 0, False: 1})
    combined = combined.sort_values(["제품코드", "출처우선순위", EFFECTIVE_SAMPLE_AVAILABLE_COL], ascending=[True, True, False])
    combined = combined.drop_duplicates(subset=["제품코드"], keep="first")
    return combined[output_columns].reset_index(drop=True)


def apply_effective_sample_available_stock_basis(
    doi_reference: pd.DataFrame,
    sample_reference: pd.DataFrame,
    doi_source_label: str,
) -> tuple[pd.DataFrame, str]:
    if doi_reference.empty:
        return doi_reference, doi_source_label
    if sample_reference.empty or EFFECTIVE_SAMPLE_AVAILABLE_COL not in sample_reference.columns:
        return doi_reference, f"{doi_source_label} · APS 샘플신청가능수량 없음"

    updated = doi_reference.copy()
    sample_lookup = sample_reference.set_index("제품코드")[EFFECTIVE_SAMPLE_AVAILABLE_COL]
    sample_stock = updated["제품코드"].map(sample_lookup).fillna(0)
    updated["완제품재고"] = parse_mixed_numeric(sample_stock).fillna(0)

    if "DOI기준오더" in updated.columns:
        doi_order = parse_mixed_numeric(updated["DOI기준오더"]).fillna(0)
        valid_order = doi_order.gt(0)
        updated.loc[valid_order, "현재DOI"] = updated.loc[valid_order, "완제품재고"] / doi_order.loc[valid_order] * 181
        updated.loc[~valid_order, "현재DOI"] = 0

    if "샘플수량출처" in sample_reference.columns:
        source_lookup = sample_reference.set_index("제품코드")["샘플수량출처"]
        source_label = summarize_unique(sample_reference["샘플수량출처"], head_count=2) or "샘플신청가능수량"
        updated["재고기준"] = updated["제품코드"].map(source_lookup).fillna(source_label)
    else:
        updated["재고기준"] = "샘플신청가능수량"
        source_label = "샘플신청가능수량"
    return updated, f"{doi_source_label} · 완제품재고={source_label}"


def classify_effective_overproduction_action(row: pd.Series) -> str:
    doi_status = clean_text_value(row.get("DOI판단상태", ""))
    current_doi = numeric_scalar(row.get("현재DOI", 0))
    doi_upper = numeric_scalar(row.get("기준DOI상한", 0))
    doi_excess = numeric_scalar(row.get("DOI초과일", 0))
    ineffective_qty = numeric_scalar(row.get("비유효생산량", 0))
    remaining_need = numeric_scalar(row.get("잔여필요수량", 0))

    if doi_status == "생산지양":
        return "생산자제"
    if doi_status == "생산조정":
        return "생산조정"
    if doi_upper > 0 and current_doi >= doi_upper * 1.5:
        return "생산자제"
    if doi_excess > 0:
        return "생산조정"
    if ineffective_qty > 0 and remaining_need <= 0:
        return "계획초과 확인"
    return "모니터링"


def build_effective_overproduction_candidates(
    detail: pd.DataFrame,
    doi_reference: pd.DataFrame,
    top_n: int = 50,
) -> pd.DataFrame:
    required_columns = [
        effective_report.OUTPUT_SKU_COL,
        effective_report.OUTPUT_PRODUCT_NAME_COL,
        effective_report.OUTPUT_PROCESS_COL,
        effective_report.NEED_QTY_COL,
        effective_report.ACTUAL_QTY_COL,
        effective_report.EFFECTIVE_PRODUCTION_COL,
        effective_report.INEFFECTIVE_PRODUCTION_COL,
        effective_report.REMAINING_NEED_COL,
    ]
    if detail.empty or any(column not in detail.columns for column in required_columns):
        return pd.DataFrame()

    scope = detail.copy()
    for column in [
        effective_report.NEED_QTY_COL,
        effective_report.ACTUAL_QTY_COL,
        effective_report.EFFECTIVE_PRODUCTION_COL,
        effective_report.INEFFECTIVE_PRODUCTION_COL,
        effective_report.REMAINING_NEED_COL,
    ]:
        scope[column] = parse_mixed_numeric(scope[column])
    scope = scope[scope[effective_report.ACTUAL_QTY_COL].gt(0)].copy()
    if scope.empty:
        return pd.DataFrame()

    group_columns = [effective_report.OUTPUT_SKU_COL, effective_report.OUTPUT_PRODUCT_NAME_COL]
    optional_dimensions = [
        effective_report.MAJOR_CATEGORY_COL,
        effective_report.SHEET_NAME_COL,
    ]
    for column in optional_dimensions:
        if column in scope.columns:
            group_columns.append(column)

    summary = (
        scope.groupby(group_columns, as_index=False)
        .agg(
            **{
                "생산공정": (effective_report.OUTPUT_PROCESS_COL, join_unique_text_values),
                "계획수량": (effective_report.NEED_QTY_COL, "sum"),
                "생산량": (effective_report.ACTUAL_QTY_COL, "sum"),
                "유효생산량": (effective_report.EFFECTIVE_PRODUCTION_COL, "sum"),
                "비유효생산량": (effective_report.INEFFECTIVE_PRODUCTION_COL, "sum"),
                "잔여필요수량": (effective_report.REMAINING_NEED_COL, "sum"),
                "거래처": (effective_report.CUSTOMER_COL, lambda s: summarize_unique(s, head_count=2))
                if effective_report.CUSTOMER_COL in scope.columns
                else (effective_report.OUTPUT_PRODUCT_NAME_COL, lambda _s: ""),
                "이니셜": (effective_report.INITIAL_COL, lambda s: summarize_unique(s, head_count=2))
                if effective_report.INITIAL_COL in scope.columns
                else (effective_report.OUTPUT_PRODUCT_NAME_COL, lambda _s: ""),
            }
        )
        .copy()
    )
    summary["제품코드"] = summary[effective_report.OUTPUT_SKU_COL].map(normalize_item_code_value)
    summary["비유효비중(%)"] = calculate_effective_rate(summary["비유효생산량"], summary["생산량"])

    if doi_reference.empty:
        merged = summary.copy()
        for column in [
            "현재DOI",
            "기준DOI하한",
            "기준DOI상한",
            "완제품재고",
            "공정재고합계",
            "DOI기준오더",
            "DOI기준오더수량",
            "DOI기준부족수량",
            "DOI기준사출부족수량",
            "DOI우선순위점수",
        ]:
            merged[column] = 0.0
        for column in ["DOI상태", "기준등급", "DOI판단상태", "DOI판단메모", "신제품여부", "재고기준"]:
            merged[column] = ""
    else:
        merged = summary.merge(doi_reference, on="제품코드", how="left")
        numeric_columns = [
            "현재DOI",
            "기준DOI하한",
            "기준DOI상한",
            "완제품재고",
            "공정재고합계",
            "DOI기준오더",
            "DOI기준오더수량",
            "DOI기준부족수량",
            "DOI기준사출부족수량",
            "DOI우선순위점수",
        ]
        for column in numeric_columns:
            merged[column] = parse_mixed_numeric(merged.get(column, pd.Series(0, index=merged.index))).fillna(0)
        for column in ["DOI상태", "기준등급", "DOI판단상태", "DOI판단메모", "신제품여부", "재고기준"]:
            merged[column] = merged.get(column, pd.Series("", index=merged.index)).map(clean_text_value)

    merged["DOI초과일"] = (merged["현재DOI"] - merged["기준DOI상한"]).where(merged["기준DOI상한"].gt(0), 0).clip(lower=0)
    merged["조치"] = merged.apply(classify_effective_overproduction_action, axis=1)
    action_rank = {action: index for index, action in enumerate(EFFECTIVE_OVERPRODUCTION_ACTION_ORDER)}
    merged["조치순위"] = merged["조치"].map(action_rank).fillna(len(action_rank))
    merged["생산자제점수"] = (
        merged["비유효생산량"] * 10
        + merged["DOI초과일"] * 10
        + merged["현재DOI"]
        + merged["완제품재고"] * 0.02
    )
    candidate_mask = (
        merged["비유효생산량"].gt(0)
        | merged["DOI초과일"].gt(0)
        | merged["DOI판단상태"].isin(["생산지양", "생산조정"])
    )
    candidates = merged[candidate_mask].copy()
    if candidates.empty:
        return candidates

    preferred_columns = [
        "조치",
        "제품코드",
        effective_report.OUTPUT_PRODUCT_NAME_COL,
        "생산공정",
        "계획수량",
        "생산량",
        "유효생산량",
        "비유효생산량",
        "비유효비중(%)",
        "잔여필요수량",
        "현재DOI",
        "기준DOI상한",
        "DOI초과일",
        "완제품재고",
        "재고기준",
        "공정재고합계",
        "DOI판단상태",
        "DOI판단메모",
        "DOI상태",
        "기준등급",
        "거래처",
        "이니셜",
        *[column for column in optional_dimensions if column in candidates.columns],
        "생산자제점수",
        "조치순위",
    ]
    return candidates[preferred_columns].sort_values(
        ["조치순위", "비유효생산량", "DOI초과일", "현재DOI", "생산자제점수"],
        ascending=[True, False, False, False, False],
    ).head(top_n).reset_index(drop=True)


def build_effective_overproduction_figure(candidates: pd.DataFrame, top_n: int = 12) -> go.Figure:
    fig = go.Figure()
    if candidates.empty:
        fig.add_annotation(text="표시할 생산자제 후보가 없습니다.", showarrow=False, x=0.5, y=0.5)
        fig.update_layout(height=300, margin={"l": 10, "r": 10, "t": 16, "b": 10})
        return fig

    chart = candidates.head(top_n).copy()
    labels: list[str] = []
    for idx, row in chart.iterrows():
        code = clean_text_value(row.get("제품코드", ""))
        name = clean_text_value(row.get(effective_report.OUTPUT_PRODUCT_NAME_COL, ""))
        short_name = name[:24] + "..." if len(name) > 24 else name
        labels.append(f"{idx + 1}. {code}" + (f" · {short_name}" if short_name else ""))
    chart["표시품목"] = labels
    ordered_labels = labels[::-1]
    chart["표시품목"] = pd.Categorical(chart["표시품목"], categories=ordered_labels, ordered=True)

    action_colors = {
        "생산자제": "#dc2626",
        "생산조정": "#f97316",
        "계획초과 확인": "#d97706",
        "모니터링": "#64748b",
    }
    for action in EFFECTIVE_OVERPRODUCTION_ACTION_ORDER:
        rows = chart[chart["조치"].eq(action)].sort_values("표시품목")
        if rows.empty:
            continue
        fig.add_trace(
            go.Bar(
                x=rows["비유효생산량"],
                y=rows["표시품목"],
                orientation="h",
                name=action,
                marker={"color": action_colors.get(action, "#64748b")},
                customdata=rows[
                    [
                        "현재DOI",
                        "기준DOI상한",
                        "완제품재고",
                        "생산량",
                        "비유효비중(%)",
                    ]
                ].to_numpy(),
                hovertemplate=(
                    "%{y}<br>%{fullData.name}<br>"
                    "비유효생산량 %{x:,.0f}<br>"
                    "생산량 %{customdata[3]:,.0f} / 비유효비중 %{customdata[4]:.1f}%<br>"
                    "DOI %{customdata[0]:.1f} / 기준상한 %{customdata[1]:.1f}<br>"
                    "샘플신청가능수량 %{customdata[2]:,.0f}<extra></extra>"
                ),
            )
        )
    fig.update_layout(
        barmode="stack",
        height=max(360, 32 * len(ordered_labels) + 110),
        margin={"l": 12, "r": 24, "t": 18, "b": 36},
        legend={"orientation": "h", "yanchor": "bottom", "y": 1.02, "xanchor": "right", "x": 1},
        hovermode="closest",
    )
    fig.update_xaxes(title_text="비유효생산량", rangemode="tozero", tickformat=",")
    fig.update_yaxes(title_text="", categoryorder="array", categoryarray=ordered_labels)
    return fig


def render_effective_overproduction_panel(
    candidates: pd.DataFrame,
    doi_source_label: str,
    month_label: str,
) -> None:
    st.markdown("#### 생산자제 신호")
    st.caption(
        f"{EFFECTIVE_PRODUCTION_DEFAULT_SITE} 기준 DOI와 생산유효도를 함께 봅니다. "
        "DOI 기준을 초과했거나 계획 대비 비유효생산량이 큰 품목을 먼저 표시합니다."
    )
    st.caption(f"DOI 소스: {doi_source_label}")

    if candidates.empty:
        st.success(f"{month_label} 기준으로 생산자제 후보가 없습니다.")
        return

    action_scope = candidates[candidates["조치"].isin(["생산자제", "생산조정"])].copy()
    if action_scope.empty:
        action_scope = candidates.copy()
    doi_excess_count = int(candidates["DOI초과일"].gt(0).sum()) if "DOI초과일" in candidates.columns else 0
    max_doi = float(parse_mixed_numeric(candidates["현재DOI"]).max()) if "현재DOI" in candidates.columns else 0.0

    kpi_cols = st.columns(4, gap="medium")
    with kpi_cols[0]:
        render_dashboard_kpi("생산자제/조정 품목", f"{len(action_scope):,}", "risk")
    with kpi_cols[1]:
        render_dashboard_kpi("비유효생산량", f"{action_scope['비유효생산량'].sum():,.0f}", "risk")
    with kpi_cols[2]:
        render_dashboard_kpi("DOI 기준초과 품목", f"{doi_excess_count:,}", "risk")
    with kpi_cols[3]:
        render_dashboard_kpi("최대 DOI", format_effective_decimal(max_doi), "stock")

    top = candidates.iloc[0]
    top_code = clean_text_value(top.get("제품코드", ""))
    top_name = clean_text_value(top.get(effective_report.OUTPUT_PRODUCT_NAME_COL, ""))
    st.warning(
        f"우선 확인: {top_code} {top_name} · {top['조치']} · "
        f"비유효 {format_effective_int(top['비유효생산량'])} / "
        f"DOI {format_effective_decimal(top['현재DOI'])}"
    )

    st.plotly_chart(build_effective_overproduction_figure(candidates, top_n=12), width="stretch")

    table_columns = [
        "조치",
        "제품코드",
        effective_report.OUTPUT_PRODUCT_NAME_COL,
        "생산공정",
        "생산량",
        "비유효생산량",
        "비유효비중(%)",
        "잔여필요수량",
        "현재DOI",
        "기준DOI상한",
        "DOI초과일",
        "완제품재고",
        "재고기준",
        "DOI판단상태",
        "DOI판단메모",
    ]
    table_columns = [column for column in table_columns if column in candidates.columns]
    with st.expander("생산자제 후보 상세", expanded=True):
        st.dataframe(
            candidates[table_columns],
            width="stretch",
            hide_index=True,
            column_config={
                "생산량": st.column_config.NumberColumn("생산량", format="%d"),
                "비유효생산량": st.column_config.NumberColumn("비유효생산량", format="%d"),
                "비유효비중(%)": st.column_config.ProgressColumn(
                    "비유효비중(%)",
                    format="%.1f%%",
                    min_value=0,
                    max_value=100,
                ),
                "잔여필요수량": st.column_config.NumberColumn("잔여필요수량", format="%d"),
                "현재DOI": st.column_config.NumberColumn("현재DOI", format="%.1f"),
                "기준DOI상한": st.column_config.NumberColumn("기준DOI상한", format="%.1f"),
                "DOI초과일": st.column_config.NumberColumn("DOI초과일", format="%.1f"),
                "완제품재고": st.column_config.NumberColumn("샘플신청가능수량", format="%d"),
            },
        )


def render_effective_production_dashboard() -> None:
    st.subheader("생산유효도 분석")
    st.caption("기존 생산유효도 앱의 계산 정의를 그대로 사용합니다.")

    if effective_report is None:
        st.error("생산유효도 계산 모듈을 찾지 못했습니다.")
        return
    if not EFFECTIVE_PRODUCTION_SOURCE_DIR.exists():
        st.error(f"생산유효도 기준 데이터 폴더가 없습니다: {EFFECTIVE_PRODUCTION_SOURCE_DIR.name}")
        return
    if not sync_plan_api_data_mode():
        st.warning(f"{PLAN_API_KEY_ENV}가 설정되어 있지 않아 생산유효도 API 데이터를 조회할 수 없습니다.")
        return

    source_signature = build_effective_source_signature(EFFECTIVE_PRODUCTION_SOURCE_DIR)
    api_key_hash = hashlib.sha256(get_plan_api_key().encode("utf-8")).hexdigest()[:12]
    display_cutoff_date = build_effective_display_cutoff_date()
    filter_cols = st.columns([2.0, 1.0])
    with filter_cols[1]:
        if st.button("새로고침", key="refresh_effective_production_api", use_container_width=True):
            set_session_value("plan_api_refresh_nonce", get_plan_api_refresh_nonce() + 1)
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()

    with st.spinner("생산유효도 데이터를 계산하는 중입니다..."):
        summary_df, change_df, sheet_name_df, major_category_df, detail_df, metadata = load_original_effective_dashboard_data(
            str(EFFECTIVE_PRODUCTION_SOURCE_DIR),
            display_cutoff_date,
            source_signature,
            get_plan_api_base_url(),
            api_key_hash,
            get_plan_api_refresh_nonce(),
        )

    if summary_df.empty:
        st.warning("선택한 기간과 관에 표시할 생산유효도 데이터가 없습니다.")
        return

    month_options = sorted(summary_df[effective_report.DATE_COL].astype(str).str[:6].unique(), reverse=True)
    with filter_cols[0]:
        selected_month = st.selectbox(
            "월",
            month_options,
            format_func=format_effective_month_label,
            key="effective_original_month_filter",
        )
    month_summary = filter_effective_report_month(summary_df, effective_report.DATE_COL, selected_month)
    month_major = filter_effective_report_month(major_category_df, effective_report.DATE_COL, selected_month)
    month_detail = prepare_original_effective_detail_table(detail_df, selected_month)
    month_label = format_effective_month_short_label(selected_month)
    doi_reference, doi_source_label = load_effective_doi_reference(str(EFFECTIVE_PRODUCTION_SOURCE_DIR), source_signature)
    sample_signature = build_effective_sample_available_signature(EFFECTIVE_PRODUCTION_SOURCE_DIR)
    plan_sample_refresh_key = "" if sample_signature[0][2] >= 0 else build_plan_api_refresh_key()
    sample_reference = load_effective_sample_available_reference(
        str(EFFECTIVE_PRODUCTION_SOURCE_DIR),
        sample_signature,
        str(BASE_DIR),
        plan_sample_refresh_key,
    )
    doi_reference, doi_source_label = apply_effective_sample_available_stock_basis(
        doi_reference,
        sample_reference,
        doi_source_label,
    )
    overproduction_candidates = build_effective_overproduction_candidates(month_detail, doi_reference, top_n=50)

    st.caption(
        f"데이터 소스: {metadata['source']} · 수요기간: "
        f"{format_effective_date_option(metadata['demand_start'])} ~ {format_effective_date_option(metadata['demand_end'])} · "
        f"표시 기준: {format_effective_date_option(metadata['display_cutoff_date'])}까지 · "
        f"수요파일 {metadata['demand_file_count']:,}개 · API 조회시각: {metadata['loaded_at']}"
    )
    st.caption(
        f"API 원천 {format_effective_int(metadata['raw_rows'])}행 / "
        f"집계 대상 {format_effective_int(metadata['production_rows'])}행 / "
        f"매칭 기준: {metadata['match_key']}"
    )
    st.info("같은 일자의 계획수량과 샘플제외 양품수량을 제품코드 전체 문자열 기준으로 매칭합니다. 공정별 KPI만 표시하며 두 공정 합산 KPI는 표시하지 않습니다.")

    render_effective_overproduction_panel(overproduction_candidates, doi_source_label, month_label)

    process_kpi_cols = st.columns(2)
    for column, process in zip(process_kpi_cols, effective_report.TARGET_PROCESSES):
        rows = month_summary[month_summary[effective_report.OUTPUT_PROCESS_COL].eq(process)]
        with column:
            rate = effective_report_rate(
                rows,
                effective_report.SUMMARY_EFFECTIVE_COL,
                effective_report.SUMMARY_ACTUAL_COL,
            )
            label = process.replace("[10]", "").replace("[80]", "").replace("/", "")
            st.metric(f"{month_label} {label} 생산유효도", format_effective_pct(rate))
            st.caption(process)

    if not month_major.empty:
        category_cols = st.columns(3)
        for column, category in zip(category_cols, EFFECTIVE_PRODUCTION_CATEGORY_ORDER):
            rows = month_major[month_major[effective_report.MAJOR_CATEGORY_COL].eq(category)]
            actual = float(parse_mixed_numeric(rows[effective_report.SUMMARY_ACTUAL_COL]).sum()) if not rows.empty else 0.0
            effective = float(parse_mixed_numeric(rows[effective_report.SUMMARY_EFFECTIVE_COL]).sum()) if not rows.empty else 0.0
            rate = 0.0 if actual == 0 else effective / actual * 100
            with column:
                st.metric(f"{month_label} {category} 생산유효도", format_effective_pct(rate))
                st.caption(f"실적 {format_effective_int(actual)} / 유효 {format_effective_int(effective)}")

    top_products = build_top_production_products_table(month_detail, top_n=15)
    if not top_products.empty:
        st.markdown("#### 생산량 Top 품목")
        st.caption("실적수량 기준으로 많이 생산한 제품을 공정별 색상으로 나눠 표시합니다.")
        highlight_cols = st.columns(min(3, len(top_products)))
        for column, (_, row) in zip(highlight_cols, top_products.head(3).iterrows()):
            code = clean_text_value(row.get(effective_report.OUTPUT_SKU_COL, ""))
            name = clean_text_value(row.get(effective_report.OUTPUT_PRODUCT_NAME_COL, ""))
            with column:
                st.metric(f"{int(row['순위'])}. {code}", format_effective_int(row["생산량"]))
                st.caption(name if name else "제품명 미지정")
        st.plotly_chart(build_top_production_products_figure(month_detail, top_n=12), width="stretch")
        with st.expander("생산량 Top 품목 상세"):
            st.dataframe(
                top_products,
                width="stretch",
                hide_index=True,
                column_config={
                    "순위": st.column_config.NumberColumn("순위", format="%d"),
                    "생산량": st.column_config.NumberColumn("생산량", format="%d"),
                    "유효생산량": st.column_config.NumberColumn("유효생산량", format="%d"),
                    "비유효생산량": st.column_config.NumberColumn("비유효생산량", format="%d"),
                    "유효비중(%)": st.column_config.ProgressColumn(
                        "유효비중(%)",
                        format="%.1f%%",
                        min_value=0,
                        max_value=100,
                    ),
                },
            )

    st.markdown("#### 일자별 생산유효도 추이")
    st.plotly_chart(build_original_effective_trend_figure(month_summary, month_major), width="stretch")

    st.markdown("#### 공정별 요약")
    process_summary = month_summary.copy()
    process_summary[effective_report.DATE_COL] = process_summary[effective_report.DATE_COL].map(format_effective_date_option)
    st.dataframe(
        process_summary,
        width="stretch",
        hide_index=True,
        column_config={
            effective_report.SUMMARY_NEED_COL: st.column_config.NumberColumn(effective_report.SUMMARY_NEED_COL, format="%d"),
            effective_report.SUMMARY_ACTUAL_COL: st.column_config.NumberColumn(effective_report.SUMMARY_ACTUAL_COL, format="%d"),
            effective_report.SUMMARY_EFFECTIVE_COL: st.column_config.NumberColumn(effective_report.SUMMARY_EFFECTIVE_COL, format="%d"),
            effective_report.SUMMARY_INEFFECTIVE_COL: st.column_config.NumberColumn(effective_report.SUMMARY_INEFFECTIVE_COL, format="%d"),
            effective_report.SUMMARY_REMAINING_COL: st.column_config.NumberColumn(effective_report.SUMMARY_REMAINING_COL, format="%d"),
            effective_report.PRODUCTION_EFFECTIVENESS_COL: st.column_config.ProgressColumn(
                effective_report.PRODUCTION_EFFECTIVENESS_COL,
                format="%.1f%%",
                min_value=0,
                max_value=100,
            ),
        },
    )

    st.markdown("#### 상세표")
    table_filter_cols = st.columns([1.2, 1.4, 1.4, 2.2])
    with table_filter_cols[0]:
        date_options = (
            sorted(month_detail[effective_report.DATE_COL].dropna().astype(str).unique(), reverse=True)
            if not month_detail.empty and effective_report.DATE_COL in month_detail.columns
            else []
        )
        date_filter = st.multiselect(
            "일자",
            options=date_options,
            default=date_options,
            key="effective_detail_date_filter",
        )
    with table_filter_cols[1]:
        process_filter = st.multiselect(
            "공정",
            options=list(effective_report.TARGET_PROCESSES),
            default=list(effective_report.TARGET_PROCESSES),
            key="effective_detail_process_filter",
        )
    with table_filter_cols[2]:
        status_options = sorted(month_detail[effective_report.MATCH_STATUS_COL].dropna().astype(str).unique()) if not month_detail.empty else []
        status_filter = st.multiselect(
            "매칭상태",
            options=status_options,
            default=status_options,
            key="effective_detail_status_filter",
        )
    with table_filter_cols[3]:
        query = st.text_input(
            "검색",
            value="",
            key="effective_detail_search",
            placeholder="일자, 제품코드, 제품명, 거래처, 이니셜",
        )

    detail_table = month_detail.copy()
    if date_filter and effective_report.DATE_COL in detail_table.columns:
        detail_table = detail_table[detail_table[effective_report.DATE_COL].isin(date_filter)]
    if process_filter and effective_report.OUTPUT_PROCESS_COL in detail_table.columns:
        detail_table = detail_table[detail_table[effective_report.OUTPUT_PROCESS_COL].isin(process_filter)]
    if status_filter and effective_report.MATCH_STATUS_COL in detail_table.columns:
        detail_table = detail_table[detail_table[effective_report.MATCH_STATUS_COL].isin(status_filter)]
    if query.strip():
        search_columns = [
            column
            for column in [
                effective_report.DATE_COL,
                effective_report.OUTPUT_SKU_COL,
                effective_report.OUTPUT_PRODUCT_NAME_COL,
                effective_report.CUSTOMER_COL,
                effective_report.INITIAL_COL,
                effective_report.MAJOR_CATEGORY_COL,
                effective_report.SHEET_NAME_COL,
            ]
            if column in detail_table.columns
        ]
        for term in query.split():
            haystack = detail_table[search_columns].fillna("").astype(str).agg(" ".join, axis=1)
            detail_table = detail_table[haystack.str.contains(re.escape(term), case=False, na=False)]

    display_limit = 1000
    if len(detail_table) > display_limit:
        st.caption(f"상세표는 화면 성능을 위해 {display_limit:,}행까지만 표시합니다. 현재 필터 결과 {len(detail_table):,}행")
    st.dataframe(
        detail_table.head(display_limit),
        width="stretch",
        height=640,
        hide_index=True,
        column_config={
            effective_report.NEED_QTY_COL: st.column_config.NumberColumn(effective_report.NEED_QTY_COL, format="%d"),
            effective_report.ACTUAL_QTY_COL: st.column_config.NumberColumn(effective_report.ACTUAL_QTY_COL, format="%d"),
            effective_report.EFFECTIVE_PRODUCTION_COL: st.column_config.NumberColumn(effective_report.EFFECTIVE_PRODUCTION_COL, format="%d"),
            effective_report.INEFFECTIVE_PRODUCTION_COL: st.column_config.NumberColumn(effective_report.INEFFECTIVE_PRODUCTION_COL, format="%d"),
            effective_report.REMAINING_NEED_COL: st.column_config.NumberColumn(effective_report.REMAINING_NEED_COL, format="%d"),
            effective_report.PRODUCTION_EFFECTIVENESS_COL: st.column_config.ProgressColumn(
                effective_report.PRODUCTION_EFFECTIVENESS_COL,
                format="%.1f%%",
                min_value=0,
                max_value=100,
            ),
        },
    )



def main() -> None:
    inject_dashboard_theme()
    st.markdown(
        """
        <div class="dashboard-hero">
            <div class="dashboard-hero-title">생산현황</div>
            <p class="dashboard-hero-subtitle">생산 부족 리스크, 공정 재고, 자재 입고 현황을<br>실시간으로 모니터링합니다.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    top_views = ["생산 부족 현황", "리드지 현황", "생산코드별 리드지", "전체 품목 현황", "생산유효도 분석"]
    all_item_site_filter = "전체"
    with st.sidebar:
        st.markdown(
            """
            <div class="sidebar-brand">
                <span class="sidebar-brand-icon">
                    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" aria-hidden="true">
                        <path d="M4 19V5" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                        <path d="M4 19H20" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                        <path d="M8 16V11" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                        <path d="M12 16V7" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                        <path d="M16 16V9" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
                    </svg>
                </span>
                <span class="sidebar-brand-title">생산현황</span>
            </div>
            <div class="sidebar-section-title">메뉴</div>
            """,
            unsafe_allow_html=True,
        )
        if st.session_state.get("top_view_radio_v2") not in top_views:
            st.session_state["top_view_radio_v2"] = top_views[0]
        selected_top_view = st.radio(
            "메뉴",
            options=top_views,
            index=0,
            key="top_view_radio_v2",
            label_visibility="collapsed",
        )
        if selected_top_view == "생산유효도 분석":
            sync_plan_api_data_mode()
            data_base_dir = BASE_DIR
            source_label = "생산유효도 수요 스냅샷 + 생산실적 API"
            updated_at = get_plan_api_updated_at() if is_plan_api_configured() else ""
            cloud_snapshots_available = False
            data_live_updated_at = ""
            sidebar_status_caption = "API 모드: 기존 생산유효도 계산 기준"
        else:
            sync_plan_api_data_mode()
            if selected_top_view == "전체 품목 현황":
                default_all_item_site_filter = EFFECTIVE_PRODUCTION_DEFAULT_SITE if is_plan_api_enabled() else "전체"
                all_item_site_filter = st.pills(
                    "전체품목 관",
                    options=[*SITE_GROUP_ORDER, "전체"],
                    default=default_all_item_site_filter,
                    key="all_item_flow_site_prefilter_v1",
                    help="선택한 관만 APS API로 조회합니다. 전체를 선택하면 모든 관을 한 번에 조회합니다.",
                ) or default_all_item_site_filter
            reference_dates_slot = st.empty()
            data_base_dir, source_label, updated_at = resolve_data_source_from_state(BASE_DIR)
            cloud_snapshots_available = should_use_cloud_snapshots(data_base_dir)
            data_live_updated_at = get_data_updated_at(data_base_dir)
            if selected_top_view == "전체 품목 현황":
                sidebar_meta_key = "all_item_updated_at"
                sidebar_live_updated_at = get_aps_or_file_updated_at(get_all_item_updated_at(data_base_dir))
            elif selected_top_view in {"리드지 현황", "생산코드별 리드지"}:
                sidebar_meta_key = "leadji_updated_at"
                sidebar_live_updated_at = get_leadji_status_updated_at(data_base_dir)
            else:
                sidebar_meta_key = "data_updated_at"
                sidebar_live_updated_at = get_plan_api_updated_at() if is_plan_api_enabled() else data_live_updated_at

            if cloud_snapshots_available and is_cloud_snapshot_fresh(sidebar_meta_key, sidebar_live_updated_at):
                updated_at = get_cloud_snapshot_meta_value(sidebar_meta_key, sidebar_live_updated_at)
                sidebar_status_caption = "Cloud 모드: 사전 계산 스냅샷 사용"
            elif cloud_snapshots_available:
                updated_at = sidebar_live_updated_at
                sidebar_status_caption = "Cloud 모드: 원본 엑셀 자동 반영"
            elif is_plan_api_enabled() and selected_top_view in {"전체 품목 현황", "생산 부족 현황"}:
                updated_at = sidebar_live_updated_at
                sidebar_status_caption = "API 모드: APS 수요 + WIP 기준"
            elif is_plan_api_enabled():
                updated_at = sidebar_live_updated_at
                sidebar_status_caption = "파일 계산 모드: WIP/로컬 수요 파일 기준"
            else:
                sidebar_status_caption = "업로드 모드: 업로드 파일 직접 계산"
            with reference_dates_slot.container():
                if selected_top_view != "생산 부족 현황":
                    st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
                    render_sidebar_reference_dates(data_base_dir, source_label)

    try:
        df = pd.DataFrame()
        file_info_df = pd.DataFrame()
        all_items_df = pd.DataFrame()
        code_mismatch_df = pd.DataFrame()
        all_items_full_builder = None
        if selected_top_view == "전체 품목 현황":
            all_item_live_updated_at = get_aps_or_file_updated_at(get_all_item_updated_at(data_base_dir))
            use_all_item_cloud_snapshot = (not is_plan_api_enabled()) and cloud_snapshots_available and is_cloud_snapshot_fresh(
                "all_item_updated_at", all_item_live_updated_at
            )
            if use_all_item_cloud_snapshot:
                all_items_df, code_mismatch_df = load_cloud_all_item_status_snapshot()
                if DEMAND_DETAIL_ROWS_COL in all_items_df.columns:
                    updated_at = get_cloud_snapshot_meta_value("all_item_updated_at", all_item_live_updated_at)
                else:
                    use_all_item_cloud_snapshot = False
            if not use_all_item_cloud_snapshot:
                try:
                    all_item_refresh_key = build_all_item_refresh_key(data_base_dir)

                    def build_full_all_items_for_download(
                        refresh_key: str = all_item_refresh_key,
                        base_dir_str: str = str(data_base_dir),
                    ) -> pd.DataFrame:
                        cached_full = read_all_item_status_disk_cache(refresh_key)
                        if cached_full is not None:
                            return cached_full[0]
                        full_items, full_mismatch = build_all_item_status_snapshot(refresh_key, base_dir_str)
                        write_all_item_status_disk_cache(refresh_key, full_items, full_mismatch)
                        return full_items

                    if is_plan_api_enabled():
                        cached_flow_items = read_all_item_flow_status_disk_cache(all_item_refresh_key, all_item_site_filter)
                        if cached_flow_items is not None:
                            all_items_df = cached_flow_items
                        else:
                            all_items_df = build_all_item_flow_status_snapshot(
                                all_item_refresh_key,
                                str(data_base_dir),
                                all_item_site_filter,
                            )
                            write_all_item_flow_status_disk_cache(all_item_refresh_key, all_item_site_filter, all_items_df)
                        code_mismatch_df = pd.DataFrame()
                        all_items_full_builder = build_full_all_items_for_download
                    else:
                        cached_all_items = read_all_item_status_disk_cache(all_item_refresh_key)
                        if cached_all_items is not None:
                            all_items_df, code_mismatch_df = cached_all_items
                        else:
                            all_items_df, code_mismatch_df = build_all_item_status_snapshot(
                                all_item_refresh_key,
                                str(data_base_dir),
                            )
                            write_all_item_status_disk_cache(all_item_refresh_key, all_items_df, code_mismatch_df)
                    updated_at = all_item_live_updated_at
                except Exception as live_exc:
                    if not cloud_snapshots_available:
                        raise
                    st.warning(f"원본 엑셀 자동 반영 실패로 기존 스냅샷을 표시합니다: {live_exc}")
                    all_items_df, code_mismatch_df = load_cloud_all_item_status_snapshot()
                    updated_at = get_cloud_snapshot_meta_value("all_item_updated_at", all_item_live_updated_at)

        if selected_top_view == "생산 부족 현황":
            if is_plan_api_enabled():
                try:
                    refresh_key = build_api_shortage_refresh_key(data_base_dir)
                    df, file_info_df, _ = load_api_shortage_data(refresh_key, str(data_base_dir))
                    updated_at = get_plan_api_updated_at()
                except Exception as live_exc:
                    st.error(f"APS API 수요 기준 생산 부족 현황 계산 실패: {live_exc}")
                    st.stop()
            else:
                use_data_cloud_snapshot = cloud_snapshots_available and is_cloud_snapshot_fresh(
                    "data_updated_at", data_live_updated_at
                )
                if use_data_cloud_snapshot:
                    df, file_info_df, _ = load_cloud_shortage_snapshot()
                    updated_at = get_cloud_snapshot_meta_value("data_updated_at", data_live_updated_at)
                else:
                    try:
                        refresh_key = build_data_refresh_key(data_base_dir)
                        df, file_info_df, _ = load_data(refresh_key, str(data_base_dir))
                        updated_at = data_live_updated_at
                    except Exception as live_exc:
                        if not cloud_snapshots_available:
                            raise
                        st.warning(f"원본 엑셀 자동 반영 실패로 기존 스냅샷을 표시합니다: {live_exc}")
                        df, file_info_df, _ = load_cloud_shortage_snapshot()
                        updated_at = get_cloud_snapshot_meta_value("data_updated_at", data_live_updated_at)

        leadji_info = pd.DataFrame()
        leadji_stock = pd.DataFrame()
        leadji_order_df = pd.DataFrame()
        if selected_top_view in {"리드지 현황", "생산코드별 리드지"}:
            leadji_live_updated_at = get_leadji_status_updated_at(data_base_dir)
            use_leadji_cloud_snapshot = cloud_snapshots_available and is_cloud_snapshot_fresh(
                "leadji_updated_at", leadji_live_updated_at
            )
            if use_leadji_cloud_snapshot:
                df, leadji_info, leadji_stock, leadji_order_df = load_cloud_leadji_status_snapshot()
                updated_at = get_cloud_snapshot_meta_value("leadji_updated_at", leadji_live_updated_at)
            else:
                try:
                    leadji_status_refresh_key = build_leadji_status_refresh_key(data_base_dir)
                    df, leadji_info, leadji_stock, leadji_order_df = load_leadji_status_snapshot(
                        leadji_status_refresh_key, str(data_base_dir)
                    )
                    updated_at = leadji_live_updated_at
                except Exception as live_exc:
                    if not cloud_snapshots_available:
                        raise
                    st.warning(f"원본 엑셀 자동 반영 실패로 기존 스냅샷을 표시합니다: {live_exc}")
                    df, leadji_info, leadji_stock, leadji_order_df = load_cloud_leadji_status_snapshot()
                    updated_at = get_cloud_snapshot_meta_value("leadji_updated_at", leadji_live_updated_at)
    except Exception as exc:
        st.error(f"데이터 로드 실패: {exc}")
        st.stop()

    if selected_top_view == "전체 품목 현황":
        render_all_items_dashboard(all_items_df, updated_at, all_items_full_builder, all_item_site_filter)
    elif selected_top_view == "생산 부족 현황":
        render_shortage_dashboard(df, updated_at, file_info_df, data_base_dir, source_label)
    elif selected_top_view == "리드지 현황":
        render_leadji_dashboard(updated_at, df, leadji_info, leadji_stock, leadji_order_df)
    elif selected_top_view == "생산코드별 리드지":
        render_leadji_pcode5_dashboard(updated_at, df, leadji_info, leadji_stock)
    elif selected_top_view == "생산유효도 분석":
        render_effective_production_dashboard()

    with st.sidebar:
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
        st.caption(sidebar_status_caption)
        if selected_top_view != "생산유효도 분석":
            rendered_data_base_dir, rendered_source_label, _ = select_data_source(BASE_DIR)
            st.caption(f"적용 데이터: {rendered_source_label}")
            if rendered_data_base_dir.resolve() != BASE_DIR.resolve():
                st.caption(f"업로드 작업폴더: {rendered_data_base_dir.name}")


if __name__ == "__main__":
    main()
